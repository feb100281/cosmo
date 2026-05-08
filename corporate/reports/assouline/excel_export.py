# # corporate/reports/assouline/excel_export.py
# """Экспорт в Excel для Assouline"""
# from django.http import HttpResponse
# from datetime import datetime
# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# def export_assouline_to_excel(recommendations, region):
#     """Экспорт рекомендаций в Excel"""
    
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Заказ ASSOULINE"
    
#     # Стили
#     header_font = Font(bold=True, color="FFFFFF")
#     header_fill = PatternFill(start_color="2E6B2E", end_color="2E6B2E", fill_type="solid")
#     border = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )
    
#     # Заголовки
#     headers = [
#         'Статус', 'Приоритет', 'ISBN', 'Название', 'Коллекция',
#         'Цена ($)', 'Продано (шт)', 'Выручка (₽)', 'Остаток EU',
#         'Рекомендуемое КОЛИЧЕСТВО', 'Сумма заказа ($)', 'Примечание'
#     ]
    
#     for col, header in enumerate(headers, 1):
#         cell = ws.cell(row=1, column=col, value=header)
#         cell.font = header_font
#         cell.fill = header_fill
#         cell.alignment = Alignment(horizontal='center')
#         cell.border = border
    
#     # Данные
#     order_items = [r for r in recommendations if r.get('recommended_qty', 0) > 0]
    
#     for row, rec in enumerate(order_items, 2):
#         if rec.get('priority', 0) >= 4:
#             status_text = '🔥 MUST HAVE'
#         elif rec.get('priority', 0) >= 3:
#             status_text = '⭐ РЕКОМЕНДУЕТСЯ'
#         elif rec.get('priority', 0) >= 2:
#             status_text = '📘 НА ТЕСТ'
#         else:
#             status_text = 'ℹ️ ПРОЧЕЕ'
        
#         ws.cell(row=row, column=1, value=status_text)
#         ws.cell(row=row, column=2, value=rec.get('priority', 0))
#         ws.cell(row=row, column=3, value=rec.get('isbn', ''))
#         ws.cell(row=row, column=4, value=rec.get('title', ''))
#         ws.cell(row=row, column=5, value=rec.get('collection', ''))
#         ws.cell(row=row, column=6, value=float(rec.get('price_usd', 0)))
#         ws.cell(row=row, column=7, value=rec.get('qty_sold', 0))
#         ws.cell(row=row, column=8, value=round(rec.get('revenue', 0), 2))
#         ws.cell(row=row, column=9, value=rec.get('eu_stock', 0))
#         ws.cell(row=row, column=10, value=rec.get('recommended_qty', 0))
#         ws.cell(row=row, column=11, value=round(rec.get('estimated_cost_usd', 0), 2))
#         ws.cell(row=row, column=12, value=rec.get('notes', ''))
        
#         if rec.get('priority', 0) >= 4:
#             for col in range(1, 13):
#                 ws.cell(row=row, column=col).fill = PatternFill(
#                     start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"
#                 )
    
#     # Ширина колонок
#     column_widths = {
#         'A': 18, 'B': 10, 'C': 15, 'D': 45, 'E': 20, 'F': 12,
#         'G': 12, 'H': 15, 'I': 12, 'J': 25, 'K': 15, 'L': 20
#     }
    
#     for col_letter, width in column_widths.items():
#         ws.column_dimensions[col_letter].width = width
    
#     # Итоги
#     last_row = len(order_items) + 2
#     total_qty = sum(r.get('recommended_qty', 0) for r in order_items)
#     total_cost = sum(r.get('estimated_cost_usd', 0) for r in order_items)
    
#     ws.cell(row=last_row, column=9, value="ИТОГО:")
#     ws.cell(row=last_row, column=9).font = Font(bold=True)
#     ws.cell(row=last_row, column=10, value=total_qty)
#     ws.cell(row=last_row, column=10).font = Font(bold=True)
#     ws.cell(row=last_row, column=11, value=round(total_cost, 2))
#     ws.cell(row=last_row, column=11).font = Font(bold=True)
    
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = f'attachment; filename="assouline_order_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
#     wb.save(response)
    
#     return response


# corporate/reports/assouline/excel_export.py
"""Экспорт в Excel для Assouline"""
from django.http import HttpResponse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Импортируем стили из твоей темы
from corporate.reports.excel_report.styles.theme import (
    COLORS, FILLS, FONTS, ALIGNMENTS, FORMATS, BORDERS
)


def export_assouline_to_excel(recommendations, region):
    """Экспорт рекомендаций в Excel с красивыми стилями"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказ ASSOULINE"
    
    # ============================================================
    # ЗАГОЛОВОК И ИНФОРМАЦИЯ
    # ============================================================
    ws.merge_cells('A1:L1')
    title_cell = ws.cell(row=1, column=1, value=f"ЗАКАЗ ASSOULINE")
    title_cell.font = Font(name="Roboto", size=16, bold=True, color=COLORS["black"])
    title_cell.alignment = ALIGNMENTS["center"]
    
    ws.merge_cells('A2:L2')
    subtitle_cell = ws.cell(row=2, column=1, value=f"Регион: {region} | Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    subtitle_cell.font = Font(name="Roboto", size=10, color=COLORS["text_gray"])
    subtitle_cell.alignment = ALIGNMENTS["center"]
    
    # Пустая строка
    ws.row_dimensions[3].height = 10
    
    # ============================================================
    # ЗАГОЛОВКИ ТАБЛИЦЫ
    # ============================================================
    headers = [
        ('A', 'Статус'),
        ('B', 'Приор.'),
        ('C', 'ISBN'),
        ('D', 'Название'),
        ('E', 'Коллекция'),
        ('F', 'Цена $'),
        ('G', 'Продано'),
        ('H', 'Выручка ₽'),
        ('I', 'Остаток'),
        ('J', 'Кол-во к заказу'),
        ('K', 'Сумма заказа $'),
        ('L', 'Примечание'),
    ]
    
    header_row = 4
    for col, header in headers:
        cell = ws.cell(row=header_row, column=ord(col) - 64, value=header)
        cell.font = FONTS["header_white"]
        cell.fill = FILLS["header"]
        cell.alignment = ALIGNMENTS["center"]
        cell.border = BORDERS["thin"]
    
    # ============================================================
    # ДАННЫЕ
    # ============================================================
    order_items = [r for r in recommendations if r.get('recommended_qty', 0) > 0]
    
    for idx, rec in enumerate(order_items, start=header_row + 1):
        priority = rec.get('priority', 0)
        
        # Определяем статус
        if priority >= 4:
            status_text = '🔥 MUST HAVE'
            status_fill = FILLS.get("heat_light", PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"))
        elif priority >= 3:
            status_text = '⭐ РЕКОМЕНДУЕТСЯ'
            status_fill = FILLS.get("section", PatternFill(start_color="FFF8E8", end_color="FFF8E8", fill_type="solid"))
        elif priority >= 2:
            status_text = '📘 НА ТЕСТ'
            status_fill = None
        else:
            status_text = 'ℹ️ ПРОЧЕЕ'
            status_fill = None
        
        # Статус
        cell = ws.cell(row=idx, column=1, value=status_text)
        cell.alignment = ALIGNMENTS["center"]
        if status_fill:
            cell.fill = status_fill
        
        # Приоритет
        cell = ws.cell(row=idx, column=2, value=priority)
        cell.alignment = ALIGNMENTS["center"]
        if status_fill:
            cell.fill = status_fill
        
        # ISBN
        cell = ws.cell(row=idx, column=3, value=rec.get('isbn', ''))
        cell.alignment = ALIGNMENTS["left"]
        if status_fill:
            cell.fill = status_fill
        
        # Название
        cell = ws.cell(row=idx, column=4, value=rec.get('title', '')[:80])
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if status_fill:
            cell.fill = status_fill
        
        # Коллекция
        cell = ws.cell(row=idx, column=5, value=rec.get('collection', '')[:40])
        cell.alignment = ALIGNMENTS["left"]
        if status_fill:
            cell.fill = status_fill
        
        # Цена $
        cell = ws.cell(row=idx, column=6, value=float(rec.get('price_usd', 0)))
        cell.number_format = FORMATS["money"]
        cell.alignment = ALIGNMENTS["right"]
        if status_fill:
            cell.fill = status_fill
        
        # Продано
        cell = ws.cell(row=idx, column=7, value=rec.get('qty_sold', 0))
        cell.number_format = FORMATS["int"]
        cell.alignment = ALIGNMENTS["center"]
        if status_fill:
            cell.fill = status_fill
        
        # Выручка ₽
        cell = ws.cell(row=idx, column=8, value=round(rec.get('revenue', 0), 2))
        cell.number_format = FORMATS["money"]
        cell.alignment = ALIGNMENTS["right"]
        if status_fill:
            cell.fill = status_fill
        
        # Остаток EU
        cell = ws.cell(row=idx, column=9, value=rec.get('eu_stock', 0))
        cell.number_format = FORMATS["int"]
        cell.alignment = ALIGNMENTS["center"]
        if status_fill:
            cell.fill = status_fill
        
        # Рекомендуемое количество (ГЛАВНОЕ!)
        cell = ws.cell(row=idx, column=10, value=rec.get('recommended_qty', 0))
        cell.font = FONTS["bold"]
        cell.number_format = FORMATS["int"]
        cell.alignment = ALIGNMENTS["center"]
        if rec.get('recommended_qty', 0) > 0:
            cell.fill = FILLS.get("back", PatternFill(start_color="E7F1ED", end_color="E7F1ED", fill_type="solid"))
        
        # Сумма заказа $
        cell = ws.cell(row=idx, column=11, value=round(rec.get('estimated_cost_usd', 0), 2))
        cell.font = FONTS["bold"]
        cell.number_format = FORMATS["money"]
        cell.alignment = ALIGNMENTS["right"]
        if rec.get('estimated_cost_usd', 0) > 0:
            cell.fill = FILLS.get("back", PatternFill(start_color="E7F1ED", end_color="E7F1ED", fill_type="solid"))
        
        # Примечание
        cell = ws.cell(row=idx, column=12, value=rec.get('notes', ''))
        cell.alignment = ALIGNMENTS["left"]
        if status_fill:
            cell.fill = status_fill
        
        # Добавляем границы для всех ячеек строки
        for col in range(1, 13):
            cell = ws.cell(row=idx, column=col)
            cell.border = BORDERS["thin"]
    
    # ============================================================
    # ИТОГОВАЯ СТРОКА
    # ============================================================
    last_row = header_row + len(order_items) + 1
    
    if order_items:
        total_qty = sum(r.get('recommended_qty', 0) for r in order_items)
        total_cost = sum(r.get('estimated_cost_usd', 0) for r in order_items)
        
        # Объединяем ячейки для надписи "ИТОГО"
        ws.merge_cells(f'I{last_row}:I{last_row}')
        cell = ws.cell(row=last_row, column=9, value="ИТОГО:")
        cell.font = FONTS["total"]
        cell.fill = FILLS["total"]
        cell.alignment = ALIGNMENTS["right"]
        cell.border = BORDERS["thin"]
        
        # Итоговое количество
        cell = ws.cell(row=last_row, column=10, value=total_qty)
        cell.font = FONTS["total"]
        cell.fill = FILLS["total"]
        cell.number_format = FORMATS["int"]
        cell.alignment = ALIGNMENTS["center"]
        cell.border = BORDERS["thin"]
        
        # Итоговая сумма
        cell = ws.cell(row=last_row, column=11, value=round(total_cost, 2))
        cell.font = FONTS["total"]
        cell.fill = FILLS["total"]
        cell.number_format = FORMATS["money"]
        cell.alignment = ALIGNMENTS["right"]
        cell.border = BORDERS["thin"]
        
        # Пустая ячейка для последнего столбца
        cell = ws.cell(row=last_row, column=12, value="")
        cell.fill = FILLS["total"]
        cell.border = BORDERS["thin"]
    
    # ============================================================
    # НАСТРОЙКА ШИРИНЫ КОЛОНОК
    # ============================================================
    column_widths = {
        'A': 20,   # Статус
        'B': 8,    # Приор.
        'C': 15,   # ISBN
        'D': 45,   # Название
        'E': 25,   # Коллекция
        'F': 12,   # Цена $
        'G': 10,   # Продано
        'H': 15,   # Выручка ₽
        'I': 10,   # Остаток
        'J': 20,   # Кол-во
        'K': 14,   # Сумма $
        'L': 25,   # Примечание
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # ============================================================
    # НАСТРОЙКА ВЫСОТЫ СТРОК
    # ============================================================
    ws.row_dimensions[header_row].height = 30
    for row in range(header_row + 1, last_row):
        ws.row_dimensions[row].height = 45
    
    # ============================================================
    # ЗАМОРОЗКА ПАНЕЛИ
    # ============================================================
    ws.freeze_panes = 'D5'  # Замораживаем до названия
    
    # ============================================================
    # ОТКЛЮЧАЕМ СЕТКУ
    # ============================================================
    ws.sheet_view.showGridLines = False
    
    # ============================================================
    # ФОРМИРУЕМ ОТВЕТ
    # ============================================================
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="assouline_order_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    wb.save(response)
    
    return response