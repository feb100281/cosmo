# corporate/reports/excel_report/styles/style_helpers.py
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from .theme import FILLS, FONTS, BORDERS, ALIGNMENTS, FORMATS, COLORS


def set_column_widths(ws, widths: dict):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def set_row_heights(ws, heights: dict):
    for row_idx, height in heights.items():
        ws.row_dimensions[row_idx].height = height


def draw_toc_button(ws, cell="A1", text="← Оглавление", target_sheet="TOC"):
    ws[cell] = text
    ws[cell].hyperlink = f"#'{target_sheet}'!A1"
    ws[cell].font = FONTS["back"]
    ws[cell].alignment = ALIGNMENTS["left"]
    ws[cell].fill = FILLS["back"]
    ws[cell].border = Border(
        left=Side(style="thin", color=COLORS["border_gray"]),
        right=Side(style="thin", color=COLORS["border_gray"]),
        top=Side(style="thin", color=COLORS["border_gray"]),
        bottom=Side(style="thin", color=COLORS["border_gray"]),
    )


def draw_nav_button(ws, cell="A1", text="Перейти", target_sheet="TOC", target_cell="A1"):
    ws[cell] = text
    ws[cell].hyperlink = f"#'{target_sheet}'!{target_cell}"
    ws[cell].font = FONTS["back"]
    ws[cell].alignment = ALIGNMENTS["left"]
    ws[cell].fill = FILLS["back"]
    ws[cell].border = Border(
        left=Side(style="thin", color=COLORS["border_gray"]),
        right=Side(style="thin", color=COLORS["border_gray"]),
        top=Side(style="thin", color=COLORS["border_gray"]),
        bottom=Side(style="thin", color=COLORS["border_gray"]),
    )


def draw_sheet_header(ws, title, subtitle="", note="", line_to_col=8):
    ws["A2"] = title
    ws["A3"] = subtitle
    ws["A4"] = note

    ws["A2"].font = FONTS["title"]
    ws["A3"].font = FONTS["subtitle"]
    ws["A4"].font = FONTS["subtitle"]

    ws["A2"].alignment = ALIGNMENTS["left"]
    ws["A3"].alignment = ALIGNMENTS["left"]
    ws["A4"].alignment = ALIGNMENTS["left"]

    for col in range(1, line_to_col + 1):
        ws.cell(row=6, column=col).border = BORDERS["bottom_medium"]


def draw_section_title(ws, row, col_start, col_end, title):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = FILLS["section"]
        cell.border = BORDERS["bottom_thin"]
        if col == col_start:
            cell.value = title
            cell.font = FONTS["section"]
            cell.alignment = ALIGNMENTS["left"]
        else:
            cell.value = None


def draw_table_header(ws, row, headers, start_col=1, wrap=False):
    alignment = ALIGNMENTS["center_wrap"] if wrap else ALIGNMENTS["center"]

    for i, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=row, column=i)
        cell.value = header
        cell.fill = FILLS["header"]
        cell.font = FONTS["header_white"]
        cell.alignment = alignment
        cell.border = BORDERS["thin"]


def draw_table_header_with_gaps(ws, row, headers, start_col=1, wrap=False, gap_cols=None):
    """
    Рисует header, но для gap-колонок не применяет заливку/границы/шрифт хедера.
    """
    alignment = ALIGNMENTS["center_wrap"] if wrap else ALIGNMENTS["center"]
    gap_cols = set(gap_cols or [])

    for i, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=row, column=i)
        cell.value = header

        if i in gap_cols:
            cell.fill = FILLS["none"]
            cell.font = FONTS["normal"]
            cell.alignment = ALIGNMENTS["center"]
            cell.border = BORDERS["none"]
        else:
            cell.fill = FILLS["header"]
            cell.font = FONTS["header_white"]
            cell.alignment = alignment
            cell.border = BORDERS["thin"]


def style_data_row(ws, row, values, start_col=1, fills=None, number_formats=None):
    for i, value in enumerate(values, start=start_col):
        cell = ws.cell(row=row, column=i, value=value)
        cell.font = FONTS["normal"]
        cell.alignment = ALIGNMENTS["left"] if i == start_col else ALIGNMENTS["right"]
        cell.border = BORDERS["thin"]

        if fills and i in fills:
            cell.fill = fills[i]
        else:
            cell.fill = FILLS["alt"] if row % 2 == 0 else FILLS["none"]

        if number_formats and i in number_formats:
            cell.number_format = number_formats[i]


def style_total_row(ws, row, values, start_col=1, number_formats=None):
    for i, value in enumerate(values, start=start_col):
        cell = ws.cell(row=row, column=i, value=value)
        cell.font = FONTS["total"]
        cell.border = BORDERS["top_bottom_medium"]
        cell.fill = FILLS["total"]
        cell.alignment = ALIGNMENTS["left"] if i == start_col else ALIGNMENTS["right"]

        if number_formats and i in number_formats:
            cell.number_format = number_formats[i]


def draw_blank_row(ws, row, col_start=1, col_end=10, height=6):
    ws.row_dimensions[row].height = height
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.value = None
        cell.fill = FILLS["none"]
        cell.border = BORDERS["none"]


def clear_column(ws, col_idx, row_start, row_end):
    """
    Делает колонку визуальным разделителем:
    без заливки, без границ, без текста.
    """
    for row in range(row_start, row_end + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = None
        cell.fill = FILLS["none"]
        cell.border = BORDERS["none"]
        cell.number_format = "General"
        cell.font = FONTS["normal"]
        cell.alignment = ALIGNMENTS["center"]


def set_money_format(ws, col_idx, row_start, row_end):
    for row in range(row_start, row_end + 1):
        ws.cell(row=row, column=col_idx).number_format = FORMATS["money"]


def set_pct_format(ws, col_idx, row_start, row_end):
    for row in range(row_start, row_end + 1):
        ws.cell(row=row, column=col_idx).number_format = FORMATS["pct"]


def hide_grid_and_freeze(ws, freeze_cell="A8"):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze_cell


def autosize_by_map(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        length = 0
        col_idx = col_cells[0].column
        for cell in col_cells:
            try:
                value = "" if cell.value is None else str(cell.value)
                length = max(length, len(value))
            except Exception:
                pass
        width = min(max(length + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def apply_3color_heatmap(ws, start_row, end_row, start_col, end_col):
    if end_row < start_row or end_col < start_col:
        return

    start_cell = f"{get_column_letter(start_col)}{start_row}"
    end_cell = f"{get_column_letter(end_col)}{end_row}"
    cell_range = f"{start_cell}:{end_cell}"

    rule = ColorScaleRule(
        start_type="min",
        start_color="F7FBF9",
        mid_type="percentile",
        mid_value=50,
        mid_color="DDEDE5",
        end_type="max",
        end_color="8FB9A8",
    )
    ws.conditional_formatting.add(cell_range, rule)
    
    

def draw_table_header_with_gaps(ws, row, headers, start_col=1, wrap=False, gap_cols=None):
    """
    Рисует header, но gap-колонки оставляет пустыми:
    без заливки, без границ, без жирного header-стиля.
    """
    alignment = ALIGNMENTS["center_wrap"] if wrap else ALIGNMENTS["center"]
    gap_cols = set(gap_cols or [])

    for i, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=row, column=i)
        cell.value = header

        if i in gap_cols:
            cell.fill = FILLS["none"]
            cell.font = FONTS["normal"]
            cell.alignment = ALIGNMENTS["center"]
            cell.border = BORDERS["none"]
        else:
            cell.fill = FILLS["header"]
            cell.font = FONTS["header_white"]
            cell.alignment = alignment
            cell.border = BORDERS["thin"]


def clear_column(ws, col_idx, row_start, row_end):
    """
    Делает колонку визуальным разделителем:
    пустой текст, без заливки, без бордеров.
    """
    for row in range(row_start, row_end + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = None
        cell.fill = FILLS["none"]
        cell.border = BORDERS["none"]
        cell.number_format = "General"
        cell.font = FONTS["normal"]
        cell.alignment = ALIGNMENTS["center"]
        
        
        
        
        
def style_sub_data_row(ws, row, values, start_col=1, number_formats=None):
    for i, value in enumerate(values, start=start_col):
        cell = ws.cell(row=row, column=i, value=value)
        cell.border = BORDERS["thin"]
        cell.fill = FILLS["alt"]

        if i == start_col:
            cell.font = FONTS["normal"]
            cell.alignment = ALIGNMENTS["left"]
        else:
            cell.font = FONTS["normal"]
            cell.alignment = ALIGNMENTS["right"]

        if number_formats and i in number_formats:
            cell.number_format = number_formats[i]
            
            
            

def draw_subcategory_divider(ws, row, col_start=1, col_end=19, height=6):
    """
    Тонкий декоративный разделитель между логическими блоками таблицы.
    """
    ws.row_dimensions[row].height = height

    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.value = None
        cell.fill = FILLS["divider"]
        cell.border = BORDERS["divider"]