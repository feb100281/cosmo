from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

THIN_GRAY = Side(style="thin", color="D9E1F2")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")
ACCENT_FILL = PatternFill("solid", fgColor="FCE4D6")

HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD_FONT = Font(bold=True)
SMALL_BOLD_FONT = Font(bold=True, size=10)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

BORDER_BOTTOM = Border(bottom=THIN_GRAY)
BORDER_ALL = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)

MONEY_FMT = '#,##0.00₽;[Red]-#,##0.00₽'
INT_FMT = '#,##0'
PCT_FMT = '0.0%'