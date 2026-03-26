# corporate/reports/excel_report/sheets/sheet_index.py

# corporate/reports/excel_report/sheets/sheet_index.py

from openpyxl.styles import Border, Side, Font

from ..styles.style_helpers import (
    draw_sheet_header,
    hide_grid_and_freeze,
    set_row_heights,
    set_column_widths,
)
from ..styles.theme import FONTS, ALIGNMENTS, FILLS, BORDERS, COLORS


def build_index_sheet(wb, generated_at=None):
    ws = wb.create_sheet("TOC", 0)

    # -------------------------------------------------
    # Header
    # -------------------------------------------------
    generated_text = (
        f"Сформировано: {generated_at.strftime('%d.%m.%Y %H:%M')}"
        if generated_at
        else "Сформировано:"
    )

    draw_sheet_header(
        ws,
        title="Оглавление Excel-отчета",
        subtitle="Аналитика продаж",
        note=generated_text,
        line_to_col=3,
    )

    # -------------------------------------------------
    # Заголовок блока
    # -------------------------------------------------
    ws["A8"] = "Навигация по отчету"
    ws["A8"].font = Font(name="Roboto", size=12, bold=True, color=COLORS["black"])
    ws["A8"].alignment = ALIGNMENTS["left"]

    ws["A9"] = "Основные и аналитические разделы отчета"
    ws["A9"].font = Font(name="Roboto", size=10, italic=True, color=COLORS["text_gray"])
    ws["A9"].alignment = ALIGNMENTS["left"]

    # -------------------------------------------------
    # Таблица оглавления
    # -------------------------------------------------
    header_row = 11
    start_row = 12

    headers = ["№ листа", "Раздел", "Описание"]
    for col_idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=value)
        cell.font = FONTS["header_white"]
        cell.alignment = ALIGNMENTS["center"]
        cell.fill = FILLS["header"]
        cell.border = BORDERS["thin"]

    links = [
        ("1", "Общая сводка по продажам", "Ключевые показатели по годам и каналам", "key"),
        # ("2", "Производители", "Выручка производителей в разрезе лет", "key"),
        ("2", "YOY анализ", "Динамика и изменения продаж по производителям год к году", "key"),
        ("3", "Просадка производителей", "Производители с наибольшим снижением", "key"),
        ("4", "Рост производителей", "Производители с наибольшим ростом", "key"),
        ("5", "Доли производителей", "Структура выручки по производителям", "key"),
        ("6", "Категории", "Выручка категорий в разрезе лет", "key"),
        # ("8", "Падающие категории", "Категории с отрицательной динамикой", "key"),
        ("7", "Группы категорий", "Сводка по укрупненным группам", "key"),
        ("8", "Зависимость категории от производителя", "Концентрация выручки по производителям", "key"),
        ("9", "Матрица производитель × категория", "Перекрестный анализ производителей и категорий", "key"),
        ("10", "Ассортимент vs выручка", "Сопоставление ассортимента и результата", "key"),
        ("11", "ABC по производителям", "ABC-анализ производителей", "key"),
        ("12", "ABC по категориям", "ABC-анализ категорий", "key"),
        # ("15", "Новые vs старые производители", "Сравнение новых и действующих производителей", "key"),
        ("13", "Магазин × производитель", "Матрица каналов/магазинов и производителей", "key"),
        ("14", "Топ товаров", "Лидеры продаж по товарам", "key"),
    ]

    thin_gray = Side(style="thin", color=COLORS["border_gray"])
    bottom_green = Side(style="medium", color=COLORS["dark_green"])

    key_fill = "EEF4F1"      # мягкое выделение ключевых разделов
    normal_alt_fill = FILLS["alt"]
    normal_fill = FILLS["none"]

    row = start_row
    for idx, (sheet_name, label, description, level) in enumerate(links, start=1):
        if level == "key":
            fill_a = key_fill
            fill_b = key_fill
            fill_c = key_fill
        else:
            fill_obj = normal_alt_fill if idx % 2 == 0 else normal_fill
            fill_a = fill_obj.fgColor.rgb if fill_obj.fill_type == "solid" and fill_obj.fgColor and fill_obj.fgColor.rgb else None
            fill_b = fill_a
            fill_c = fill_a

        # A: номер листа
        c1 = ws.cell(row=row, column=1, value=sheet_name)
        c1.font = FONTS["bold"]
        c1.alignment = ALIGNMENTS["center"]
        c1.border = BORDERS["thin"]
        c1.hyperlink = f"#'{sheet_name}'!A1"

        # B: раздел
        c2 = ws.cell(row=row, column=2, value=label)
        c2.font = FONTS["bold"] if level == "key" else FONTS["normal"]
        c2.alignment = ALIGNMENTS["left"]
        c2.border = BORDERS["thin"]
        c2.hyperlink = f"#'{sheet_name}'!A1"

        # C: описание
        c3 = ws.cell(row=row, column=3, value=description)
        c3.font = Font(
            name="Roboto",
            size=10,
            italic=True,
            color=COLORS["text_gray"],
        )
        c3.alignment = ALIGNMENTS["left"]
        c3.border = BORDERS["thin"]

        # fills
        if level == "key":
            from openpyxl.styles import PatternFill
            fill = PatternFill("solid", fgColor=key_fill)
            c1.fill = fill
            c2.fill = fill
            c3.fill = fill
        else:
            fill = normal_alt_fill if idx % 2 == 0 else normal_fill
            c1.fill = fill
            c2.fill = fill
            c3.fill = fill

        row += 1

    last_row = row - 1

    # -------------------------------------------------
    # Нижняя граница таблицы
    # -------------------------------------------------
    for col in range(1, 4):
        ws.cell(row=last_row, column=col).border = Border(
            left=thin_gray,
            right=thin_gray,
            top=thin_gray,
            bottom=bottom_green,
        )

    # -------------------------------------------------
    # Размеры
    # -------------------------------------------------
    set_column_widths(
        ws,
        {
            "A": 10,
            "B": 40,
            "C": 54,
        },
    )

    heights = {
        1: 20,
        2: 26,
        3: 18,
        4: 18,
        8: 22,
        9: 18,
        11: 24,
    }

    for r in range(start_row, last_row + 1):
        heights[r] = 24

    set_row_heights(ws, heights)

    hide_grid_and_freeze(ws, "A12")