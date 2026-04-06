# corporate/reports/excel_report/sheets/sheet_matrix.py

# from collections import defaultdict

# from ..styles.theme import FORMATS
# from ..styles.style_helpers import (
#     draw_toc_button,
#     draw_sheet_header,
#     draw_table_header,
#     hide_grid_and_freeze,
#     set_column_widths,
#     set_row_heights,
#     style_data_row,
#     style_total_row,
# )


# def build_matrix_sheet(wb, rows):
#     ws = wb.create_sheet("Matrix MF x Cat")

#     draw_toc_button(ws, cell="A1", target_sheet="TOC")
#     draw_sheet_header(
#         ws,
#         title="Matrix MF x Cat",
#         subtitle="Матрица выручки: производитель × категория",
#         note="По строкам — производители, по колонкам — категории, в ячейках — выручка",
#     )

#     manufacturers = sorted({r.get("manufacturer") for r in rows})
#     categories = sorted({r.get("category") for r in rows})

#     data = defaultdict(dict)
#     for r in rows:
#         manufacturer = r.get("manufacturer")
#         category = r.get("category")
#         amount = float(r.get("net_amount") or 0)
#         data[manufacturer][category] = amount

#     headers = ["Производитель"] + categories + ["Итого"]

#     header_row = 8
#     data_start_row = 9

#     draw_table_header(ws, row=header_row, headers=headers, wrap=True)

#     cur_row = data_start_row
#     totals_by_category = {cat: 0 for cat in categories}
#     grand_total = 0

#     for manufacturer in manufacturers:
#         values_by_cat = [float(data[manufacturer].get(cat, 0) or 0) for cat in categories]
#         row_total = sum(values_by_cat)

#         row_values = [manufacturer] + values_by_cat + [row_total]

#         number_formats = {
#             col_idx: FORMATS["money"] for col_idx in range(2, len(headers) + 1)
#         }

#         style_data_row(
#             ws,
#             row=cur_row,
#             values=row_values,
#             number_formats=number_formats,
#         )

#         for cat, val in zip(categories, values_by_cat):
#             totals_by_category[cat] += val
#         grand_total += row_total

#         cur_row += 1

#     total_values = ["ИТОГО"] + [totals_by_category[cat] for cat in categories] + [grand_total]

#     total_number_formats = {
#         col_idx: FORMATS["money"] for col_idx in range(2, len(headers) + 1)
#     }

#     style_total_row(
#         ws,
#         row=cur_row,
#         values=total_values,
#         number_formats=total_number_formats,
#     )

#     widths = {"A": 28}
#     for idx in range(2, len(headers)):
#         widths[ws.cell(1, idx).column_letter] = 14
#     widths[ws.cell(1, len(headers)).column_letter] = 16

#     set_column_widths(ws, widths)

#     set_row_heights(
#         ws,
#         {
#             2: 24,
#             3: 18,
#             4: 18,
#             8: 26,
#         },
#     )

#     hide_grid_and_freeze(ws, "B9")





# corporate/reports/excel_report/sheets/sheet_matrix.py

from collections import defaultdict

from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter

from ..styles.theme import FORMATS
from ..styles.style_helpers import (
    draw_toc_button,
    draw_sheet_header,
    draw_table_header,
    hide_grid_and_freeze,
    set_column_widths,
    set_row_heights,
    style_data_row,
    style_total_row,
)


def build_matrix_sheet(wb, rows):
    ws = wb.create_sheet("Matrix MF x Cat")
    ws.sheet_properties.outlinePr.summaryBelow = False

    draw_toc_button(ws, cell="A1", target_sheet="TOC")
    draw_sheet_header(
        ws,
        title="Матрица производитель × категория",
        subtitle="Матрица выручки: производитель × категория",
        note=(
            "По строкам — производители, по колонкам — категории, "
            "в ячейках — выручка. Строки отсортированы по убыванию общей выручки."
        ),
        line_to_col=8,
    )

    categories = sorted({(r.get("category") or "—") for r in rows})
    manufacturers = sorted({(r.get("manufacturer") or "—") for r in rows})

    data = defaultdict(lambda: defaultdict(float))
    for r in rows:
        manufacturer = r.get("manufacturer") or "—"
        category = r.get("category") or "—"
        amount = float(r.get("net_amount") or 0)
        data[manufacturer][category] += amount

    manufacturer_totals = {
        manufacturer: sum(float(data[manufacturer].get(cat, 0) or 0) for cat in categories)
        for manufacturer in manufacturers
    }

    manufacturers = sorted(
        manufacturers,
        key=lambda x: manufacturer_totals.get(x, 0),
        reverse=True,
    )

    layout = [("manufacturer", "Производитель")]
    for i, category in enumerate(categories, start=1):
        layout.append((f"cat_{i}", category))
    layout.append(("total", "Итого"))

    headers = [label for _, label in layout]

    col_map = {}
    for idx, (key, _) in enumerate(layout, start=1):
        col_map[key] = idx

    header_row = 8
    data_start_row = 9

    draw_table_header(ws, row=header_row, headers=headers, wrap=True)

    cur_row = data_start_row
    totals_by_category = {cat: 0 for cat in categories}
    grand_total = 0

    for manufacturer in manufacturers:
        values_by_cat = [float(data[manufacturer].get(cat, 0) or 0) for cat in categories]
        row_total = sum(values_by_cat)

        row_values = [manufacturer] + values_by_cat + [row_total]

        number_formats = {
            col_idx: FORMATS["money"]
            for col_idx in range(2, len(headers) + 1)
        }

        style_data_row(
            ws,
            row=cur_row,
            values=row_values,
            number_formats=number_formats,
        )

        for cat, val in zip(categories, values_by_cat):
            totals_by_category[cat] += val
        grand_total += row_total

        cur_row += 1

    total_values = ["ИТОГО"] + [totals_by_category[cat] for cat in categories] + [grand_total]

    total_number_formats = {
        col_idx: FORMATS["money"]
        for col_idx in range(2, len(headers) + 1)
    }

    style_total_row(
        ws,
        row=cur_row,
        values=total_values,
        number_formats=total_number_formats,
    )

    total_row = cur_row

    # Фильтр на таблицу
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{total_row}"

    # Data bars только по колонкам категорий, без колонки "Итого"
    for col_idx in range(2, len(headers)):
        col_letter = get_column_letter(col_idx)
        ws.conditional_formatting.add(
            f"{col_letter}{data_start_row}:{col_letter}{total_row - 1}",
            DataBarRule(
                start_type="num",
                start_value=0,
                end_type="max",
                end_value=0,
                color="A9CBB7",
                showValue=True,
            ),
        )

    widths = {"A": 30}

    for idx in range(2, len(headers)):
        widths[get_column_letter(idx)] = 14

    widths[get_column_letter(len(headers))] = 16

    set_column_widths(ws, widths)

    set_row_heights(
        ws,
        {
            2: 24,
            3: 18,
            4: 30,
            8: 30,
        },
    )

    hide_grid_and_freeze(ws, "B9")