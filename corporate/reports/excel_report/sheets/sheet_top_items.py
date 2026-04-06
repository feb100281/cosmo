# corporate/reports/excel_report/sheets/sheet_top_items.py

from collections import defaultdict

from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter

from ..styles.theme import FORMATS
from ..styles.style_helpers import (
    draw_toc_button,
    draw_sheet_header,
    draw_table_header_with_gaps,
    hide_grid_and_freeze,
    set_column_widths,
    set_row_heights,
    style_data_row,
    style_total_row,
    clear_column,
)


def build_top_items_sheet(wb, rows, qty_by_year_rows):
    ws = wb.create_sheet("Top Items")
    ws.sheet_properties.outlinePr.summaryBelow = False

    draw_toc_button(ws, cell="A1", target_sheet="TOC")
    draw_sheet_header(
        ws,
        title="Топ товаров",
        subtitle="Топ товаров по чистой выручке",
        note=(
            "Показаны ключевые атрибуты товара, проданное количество, "
            "разбивка количества по годам, средняя цена за единицу и чистая выручка. "
            "Количество и выручка рассчитываются с учетом возвратов."
        ),
        line_to_col=15,
    )

    years = sorted({int(r.get("year")) for r in qty_by_year_rows if r.get("year") is not None})

    qty_by_item_year = defaultdict(dict)
    for r in qty_by_year_rows:
        item_id = r.get("item_id")
        year = int(r.get("year"))
        qty = float(r.get("qty_sold_year") or 0)
        qty_by_item_year[item_id][year] = qty

    layout = [
        ("rank", "Место"),
        ("item_id", "ID"),
        ("item_name", "Номенклатура"),
        ("article", "Артикль"),
        ("manufacturer", "Производитель"),
        ("category", "Категория"),
        ("qty_total", "Кол-во\nитого"),
        ("gap_before_years", ""),
    ]

    for year in years:
        layout.append((f"qty_{year}", f"Кол-во\n{year}"))

    layout.extend([
        ("gap_after_years", ""),
        ("avg_price", "Средняя цена\nза ед."),
        ("net_amount", "Чистая выручка"),
    ])

    headers = [label for _, label in layout]

    gap_cols = set()
    col_map = {}
    for idx, (key, _) in enumerate(layout, start=1):
        col_map[key] = idx
        if key.startswith("gap_"):
            gap_cols.add(idx)

    header_row = 8
    data_start_row = 9

    draw_table_header_with_gaps(
        ws,
        row=header_row,
        headers=headers,
        wrap=True,
        gap_cols=gap_cols,
    )

    sorted_rows = sorted(
        rows,
        key=lambda r: float(r.get("net_amount") or 0),
        reverse=True,
    )

    cur_row = data_start_row
    total_qty = 0
    total_revenue = 0
    totals_by_year = {year: 0 for year in years}

    for idx, r in enumerate(sorted_rows, start=1):
        item_id = r.get("item_id")
        qty_sold = float(r.get("qty_sold") or 0)
        net_amount = float(r.get("net_amount") or 0)
        avg_price = 0 if qty_sold == 0 else net_amount / qty_sold

        row_values = [
            idx,
            item_id,
            r.get("item_name") or "—",
            r.get("article") or "—",
            r.get("manufacturer") or "—",
            r.get("category") or "—",
            qty_sold,
            None,  # gap before years
        ]

        for year in years:
            qty_year = float(qty_by_item_year.get(item_id, {}).get(year, 0) or 0)
            row_values.append(qty_year)
            totals_by_year[year] += qty_year

        row_values.append(None)  # gap after years

        row_values.extend([
            avg_price,
            net_amount,
        ])

        number_formats = {
            col_map["rank"]: FORMATS["int"],
            col_map["item_id"]: FORMATS["int"],
            col_map["qty_total"]: FORMATS["int"],
            col_map["avg_price"]: FORMATS["money"],
            col_map["net_amount"]: FORMATS["money"],
        }

        for year in years:
            number_formats[col_map[f"qty_{year}"]] = FORMATS["int"]

        style_data_row(
            ws,
            row=cur_row,
            values=row_values,
            number_formats=number_formats,
        )

        total_qty += qty_sold
        total_revenue += net_amount
        cur_row += 1

    total_avg_price = 0 if total_qty == 0 else total_revenue / total_qty

    total_values = [
        "",
        "",
        "ИТОГО",
        "",
        "",
        "",
        total_qty,
        None,  # gap before years
    ]

    for year in years:
        total_values.append(totals_by_year[year])

    total_values.append(None)  # gap after years

    total_values.extend([
        total_avg_price,
        total_revenue,
    ])

    total_row = cur_row

    total_number_formats = {
        col_map["qty_total"]: FORMATS["int"],
        col_map["avg_price"]: FORMATS["money"],
        col_map["net_amount"]: FORMATS["money"],
    }

    for year in years:
        total_number_formats[col_map[f"qty_{year}"]] = FORMATS["int"]

    style_total_row(
        ws,
        row=total_row,
        values=total_values,
        number_formats=total_number_formats,
    )

    for gap_col in gap_cols:
        clear_column(ws, col_idx=gap_col, row_start=header_row, row_end=total_row)

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{total_row}"

    revenue_col_letter = get_column_letter(col_map["net_amount"])
    ws.conditional_formatting.add(
        f"{revenue_col_letter}{data_start_row}:{revenue_col_letter}{total_row - 1}",
        DataBarRule(
            start_type="num",
            start_value=0,
            end_type="max",
            end_value=0,
            color="A9CBB7",
            showValue=True,
        ),
    )

    widths = {
        get_column_letter(col_map["rank"]): 10,
        get_column_letter(col_map["item_id"]): 10,
        get_column_letter(col_map["item_name"]): 38,
        get_column_letter(col_map["article"]): 18,
        get_column_letter(col_map["manufacturer"]): 24,
        get_column_letter(col_map["category"]): 22,
        get_column_letter(col_map["qty_total"]): 12,
        get_column_letter(col_map["gap_before_years"]): 3.5,
        get_column_letter(col_map["gap_after_years"]): 3.5,
        get_column_letter(col_map["avg_price"]): 16,
        get_column_letter(col_map["net_amount"]): 18,
    }

    for year in years:
        widths[get_column_letter(col_map[f"qty_{year}"])] = 12

    set_column_widths(ws, widths)

    set_row_heights(
        ws,
        {
            2: 24,
            3: 18,
            4: 32,
            8: 30,
        },
    )

    hide_grid_and_freeze(ws, "D9")