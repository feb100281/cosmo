# # corporate/reports/excel_report/sheets/sheet_price_segmentation.py
# from collections import defaultdict

# from openpyxl.styles import Font, PatternFill

# from ..styles.theme import FORMATS, ALIGNMENTS
# from ..styles.style_helpers import (
#     draw_toc_button,
#     draw_sheet_header,
#     draw_table_header_with_gaps,
#     hide_grid_and_freeze,
#     set_column_widths,
#     set_row_heights,
#     style_data_row,
#     style_total_row,
#     draw_subcategory_divider,
# )

# KEY_FILL = PatternFill("solid", fgColor="FFF3CD")
# KEY_FONT = Font(bold=True, color="7A4E00")


# def _safe_float(v) -> float:
#     try:
#         return float(v or 0)
#     except Exception:
#         return 0.0


# def _trend_status(delta_qty: float, delta_rev: float) -> str:
#     if delta_qty > 0 and delta_rev > 0:
#         return "Спрос растет"
#     if delta_qty < 0 and delta_rev < 0:
#         return "Спрос снижается"
#     if delta_qty > 0 and delta_rev <= 0:
#         return "Объем растет, выручка нет"
#     if delta_qty < 0 and delta_rev > 0:
#         return "Выручка держится за счет цены"
#     return "Стабильно"


# def _get_year_weights(subcat_qty_prev: float, subcat_qty_last: float) -> tuple[float, float]:
#     """
#     Если подкатегория сильно просела по qty,
#     больше весим прошлый год, чтобы ключевой диапазон не прыгал.
#     """
#     if subcat_qty_prev > 0 and subcat_qty_last < subcat_qty_prev * 0.85:
#         return 0.7, 0.3
#     return 0.4, 0.6


# def _calc_recommended_band(subcat_rows: list[dict]) -> str | None:
#     """
#     Ключевой диапазон = диапазон, где концентрируется ядро спроса.
#     Основа = qty, revenue как доп. сигнал.
#     """
#     if not subcat_rows:
#         return None

#     total_qty_prev = sum(r["qty_prev"] for r in subcat_rows)
#     total_qty_last = sum(r["qty_last"] for r in subcat_rows)
#     total_rev_prev = sum(r["rev_prev"] for r in subcat_rows)
#     total_rev_last = sum(r["rev_last"] for r in subcat_rows)

#     w_prev, w_last = _get_year_weights(total_qty_prev, total_qty_last)

#     weighted_qty_total = total_qty_prev * w_prev + total_qty_last * w_last
#     weighted_rev_total = total_rev_prev * w_prev + total_rev_last * w_last

#     best_band = None
#     best_score = -1

#     for r in subcat_rows:
#         weighted_qty = r["qty_prev"] * w_prev + r["qty_last"] * w_last
#         weighted_rev = r["rev_prev"] * w_prev + r["rev_last"] * w_last

#         qty_share = weighted_qty / weighted_qty_total if weighted_qty_total else 0
#         rev_share = weighted_rev / weighted_rev_total if weighted_rev_total else 0

#         score = qty_share * 0.9 + rev_share * 0.1
#         r["recommended_score"] = score

#         if score > best_score:
#             best_score = score
#             best_band = r["price_band"]

#     return best_band


# def _calc_anchor_price(
#     price_prev: float,
#     price_last: float,
#     qty_prev: float,
#     qty_last: float,
#     w_prev: float,
#     w_last: float,
# ) -> float:
#     """
#     Опорная цена — сглаженная цена для ориентира по продаже.
#     """
#     if price_prev > 0 and price_last > 0:
#         return price_prev * w_prev + price_last * w_last
#     if price_last > 0:
#         return price_last
#     if price_prev > 0:
#         return price_prev
#     return 0.0


# def _build_corridor_text(anchor_price: float, width: float = 0.10) -> str:
#     """
#     Рекомендованный коридор вокруг опорной цены.
#     По умолчанию ±10%.
#     """
#     if anchor_price <= 0:
#         return ""

#     low = anchor_price * (1 - width)
#     high = anchor_price * (1 + width)

#     return f"{low:,.0f} – {high:,.0f}"


# def _apply_key_highlight(ws, row_num: int):
#     """
#     Подсветка ключевого диапазона.
#     """
#     for col in (3, 6, 7, 19):  # Диапазон, Опорная цена, Коридор, Статус
#         cell = ws.cell(row=row_num, column=col)
#         cell.fill = KEY_FILL
#         cell.font = KEY_FONT


# def build_price_segmentation_sheet(wb, rows):
#     ws = wb.create_sheet("Price segmentation")
#     ws.sheet_properties.outlinePr.summaryBelow = False

#     draw_toc_button(ws, cell="A1", target_sheet="TOC")

#     headers = [
#         "Категория",
#         "Подкатегория",
#         "Диапазон",
#         "Мин.\nSKU",
#         "Макс.\nSKU",
#         "Опорная\nцена",
#         "Реком.\nкоридор",
#         "Выручка\nпред. год",
#         "Выручка\nтек. год",
#         "Δ Выручка",
#         "Qty\nпред. год",
#         "Qty\nтек. год",
#         "Δ Qty",
#         "Ср. цена\nпред. год",
#         "Ср. цена\nтек. год",
#         "Δ Цена",
#         "Доля Qty\nтек. год",
#         "Доля выручки\nтек. год",
#         "Статус",
#     ]

#     header_row = 8
#     data_start_row = 9

#     draw_table_header_with_gaps(
#         ws,
#         row=header_row,
#         headers=headers,
#         wrap=True,
#     )

#     grouped = defaultdict(dict)
#     years = set()

#     for r in rows:
#         category = r.get("category") or "Без категории"
#         subcategory = r.get("subcategory") or "Нет подкатегории"
#         price_band = r.get("price_band") or "Без диапазона"
#         year = int(r.get("year"))
#         years.add(year)

#         grouped[(category, subcategory, price_band)][year] = {
#             "revenue": _safe_float(r.get("revenue")),
#             "qty": _safe_float(r.get("qty")),
#             "revenue_ytd": _safe_float(r.get("revenue_ytd")),
#             "qty_ytd": _safe_float(r.get("qty_ytd")),
#             "min_price": _safe_float(r.get("min_price")),
#             "max_price": _safe_float(r.get("max_price")),
#             "cutoff_year": r.get("cutoff_year"),
#             "cutoff_month": r.get("cutoff_month"),
#             "cutoff_day": r.get("cutoff_day"),
#         }

#     years = sorted(years)
#     prev_year = years[-2] if len(years) >= 2 else None
#     last_year = years[-1] if len(years) >= 1 else None

#     is_ytd_comparison = False
#     cutoff_year = None
#     cutoff_month = None
#     cutoff_day = None

#     if prev_year and last_year and grouped:
#         sample_vals = next(iter(grouped.values()))
#         sample_last = sample_vals.get(last_year, {})

#         cutoff_year = sample_last.get("cutoff_year")
#         cutoff_month = sample_last.get("cutoff_month")
#         cutoff_day = sample_last.get("cutoff_day")

#         if cutoff_year and cutoff_month and cutoff_day:
#             is_ytd_comparison = bool(
#                 int(last_year) == int(cutoff_year)
#                 and (int(cutoff_month) < 12 or int(cutoff_day) < 31)
#             )

#     if is_ytd_comparison:
#         note = (
#             f"Сравнение YTD: {prev_year} vs {last_year} "
#             f"на {int(cutoff_day):02d}.{int(cutoff_month):02d}.{int(cutoff_year)}. "
#             f"Мин/Макс SKU — это технические границы бакета, "
#             f"для ориентира по продаже см. колонки 'Опорная цена' и 'Реком. коридор'."
#         )
#     else:
#         note = (
#             "Сравнение по полным годам. "
#             "Мин/Макс SKU — это технические границы бакета, "
#             "для ориентира по продаже см. колонки 'Опорная цена' и 'Реком. коридор'."
#         )

#     draw_sheet_header(
#         ws,
#         title="Спрос по ценовым диапазонам",
#         subtitle="Автоматические диапазоны внутри каждой подкатегории",
#         note=note,
#         line_to_col=19,
#     )

#     prepared_rows = []
#     subcat_map = defaultdict(list)

#     band_order = {
#         "Низкий диапазон": 1,
#         "Средний диапазон": 2,
#         "Высокий диапазон": 3,
#     }

#     for (category, subcategory, price_band), vals in grouped.items():
#         prev_vals = vals.get(prev_year, {}) if prev_year else {}
#         last_vals = vals.get(last_year, {}) if last_year else {}

#         if is_ytd_comparison:
#             rev_prev = _safe_float(prev_vals.get("revenue_ytd"))
#             rev_last = _safe_float(last_vals.get("revenue_ytd"))
#             qty_prev = _safe_float(prev_vals.get("qty_ytd"))
#             qty_last = _safe_float(last_vals.get("qty_ytd"))
#         else:
#             rev_prev = _safe_float(prev_vals.get("revenue"))
#             rev_last = _safe_float(last_vals.get("revenue"))
#             qty_prev = _safe_float(prev_vals.get("qty"))
#             qty_last = _safe_float(last_vals.get("qty"))

#         # Полностью пустые строки не показываем
#         if rev_prev == 0 and rev_last == 0 and qty_prev == 0 and qty_last == 0:
#             continue

#         row_data = {
#             "category": category,
#             "subcategory": subcategory,
#             "price_band": price_band,
#             "min_price": _safe_float(last_vals.get("min_price")) or _safe_float(prev_vals.get("min_price")),
#             "max_price": _safe_float(last_vals.get("max_price")) or _safe_float(prev_vals.get("max_price")),
#             "rev_prev": rev_prev,
#             "rev_last": rev_last,
#             "qty_prev": qty_prev,
#             "qty_last": qty_last,
#         }
#         subcat_map[(category, subcategory)].append(row_data)

#     for (category, subcategory), items in subcat_map.items():
#         recommended_band = _calc_recommended_band(items)

#         subcat_total_rev_last = sum(r["rev_last"] for r in items)
#         subcat_total_qty_last = sum(r["qty_last"] for r in items)
#         subcat_total_qty_prev = sum(r["qty_prev"] for r in items)

#         w_prev, w_last = _get_year_weights(subcat_total_qty_prev, subcat_total_qty_last)

#         for r in items:
#             delta_rev = r["rev_last"] - r["rev_prev"]
#             delta_qty = r["qty_last"] - r["qty_prev"]

#             price_prev = r["rev_prev"] / r["qty_prev"] if r["qty_prev"] else 0
#             price_last = r["rev_last"] / r["qty_last"] if r["qty_last"] else 0
#             delta_price = price_last - price_prev

#             share_qty_last = r["qty_last"] / subcat_total_qty_last if subcat_total_qty_last else 0
#             share_rev_last = r["rev_last"] / subcat_total_rev_last if subcat_total_rev_last else 0

#             anchor_price = _calc_anchor_price(
#                 price_prev=price_prev,
#                 price_last=price_last,
#                 qty_prev=r["qty_prev"],
#                 qty_last=r["qty_last"],
#                 w_prev=w_prev,
#                 w_last=w_last,
#             )
#             corridor_text = _build_corridor_text(anchor_price, width=0.10)

#             if r["price_band"] == recommended_band:
#                 status = "Ключевой диапазон"
#             else:
#                 status = _trend_status(delta_qty, delta_rev)

#             prepared_rows.append(
#                 [
#                     category,
#                     subcategory,
#                     r["price_band"],
#                     r["min_price"],
#                     r["max_price"],
#                     anchor_price,
#                     corridor_text,
#                     r["rev_prev"],
#                     r["rev_last"],
#                     delta_rev,
#                     r["qty_prev"],
#                     r["qty_last"],
#                     delta_qty,
#                     price_prev,
#                     price_last,
#                     delta_price,
#                     share_qty_last,
#                     share_rev_last,
#                     status,
#                 ]
#             )

#     prepared_rows.sort(
#         key=lambda x: (
#             x[0],
#             x[1],
#             band_order.get(x[2], 999),
#         )
#     )

#     cur_row = data_start_row

#     total_rev_prev = 0.0
#     total_rev_last = 0.0
#     total_qty_prev = 0.0
#     total_qty_last = 0.0

#     prev_group = None

#     for row in prepared_rows:
#         current_group = (row[0], row[1])  # category, subcategory

#         if prev_group is not None and current_group != prev_group:
#             draw_subcategory_divider(ws, cur_row, col_start=1, col_end=19, height=5)
#             cur_row += 1

#         style_data_row(
#             ws,
#             row=cur_row,
#             values=row,
#             number_formats={
#                 4: FORMATS["money"],
#                 5: FORMATS["money"],
#                 6: FORMATS["money"],
#                 8: FORMATS["money"],
#                 9: FORMATS["money"],
#                 10: FORMATS["money"],
#                 11: "#,##0",
#                 12: "#,##0",
#                 13: "#,##0;[Red]-#,##0",
#                 14: FORMATS["money"],
#                 15: FORMATS["money"],
#                 16: FORMATS["money"],
#                 17: "0.0%",
#                 18: "0.0%",
#             },
#         )

#         for col_idx in range(4, 19):
#             ws.cell(row=cur_row, column=col_idx).alignment = ALIGNMENTS["center"]

#         ws.cell(row=cur_row, column=19).alignment = ALIGNMENTS["left_wrap"]

#         if row[18] == "Ключевой диапазон":
#             _apply_key_highlight(ws, cur_row)

#         total_rev_prev += row[7]
#         total_rev_last += row[8]
#         total_qty_prev += row[10]
#         total_qty_last += row[11]

#         prev_group = current_group
#         cur_row += 1

#     total_delta_rev = total_rev_last - total_rev_prev
#     total_delta_qty = total_qty_last - total_qty_prev

#     total_price_prev = total_rev_prev / total_qty_prev if total_qty_prev else 0
#     total_price_last = total_rev_last / total_qty_last if total_qty_last else 0
#     total_delta_price = total_price_last - total_price_prev

#     total_anchor_price = _calc_anchor_price(
#         price_prev=total_price_prev,
#         price_last=total_price_last,
#         qty_prev=total_qty_prev,
#         qty_last=total_qty_last,
#         w_prev=0.4,
#         w_last=0.6,
#     )
#     total_corridor = _build_corridor_text(total_anchor_price, width=0.10)

#     style_total_row(
#         ws,
#         row=cur_row,
#         values=[
#             "ИТОГО",
#             "",
#             "",
#             "",
#             "",
#             total_anchor_price,
#             total_corridor,
#             total_rev_prev,
#             total_rev_last,
#             total_delta_rev,
#             total_qty_prev,
#             total_qty_last,
#             total_delta_qty,
#             total_price_prev,
#             total_price_last,
#             total_delta_price,
#             1 if total_qty_last else 0,
#             1 if total_rev_last else 0,
#             "",
#         ],
#         number_formats={
#             6: FORMATS["money"],
#             8: FORMATS["money"],
#             9: FORMATS["money"],
#             10: FORMATS["money"],
#             11: "#,##0",
#             12: "#,##0",
#             13: "#,##0;[Red]-#,##0",
#             14: FORMATS["money"],
#             15: FORMATS["money"],
#             16: FORMATS["money"],
#             17: "0.0%",
#             18: "0.0%",
#         },
#     )

#     set_column_widths(
#         ws,
#         {
#             "A": 20,
#             "B": 26,
#             "C": 18,
#             "D": 12,
#             "E": 12,
#             "F": 14,
#             "G": 16,
#             "H": 16,
#             "I": 16,
#             "J": 14,
#             "K": 11,
#             "L": 11,
#             "M": 10,
#             "N": 14,
#             "O": 14,
#             "P": 12,
#             "Q": 13,
#             "R": 15,
#             "S": 22,
#         },
#     )

#     set_row_heights(
#         ws,
#         {
#             2: 24,
#             3: 18,
#             4: 36,
#             8: 42,
#         },
#     )

#     hide_grid_and_freeze(ws, "H9")






# corporate/reports/excel_report/sheets/sheet_price_segmentation.py
# corporate/reports/excel_report/sheets/sheet_price_segmentation.py
from collections import defaultdict

from openpyxl.styles import Font, PatternFill

from ..styles.theme import FORMATS, ALIGNMENTS
from ..styles.style_helpers import (
    draw_toc_button,
    draw_sheet_header,
    draw_table_header_with_gaps,
    hide_grid_and_freeze,
    set_column_widths,
    set_row_heights,
    style_data_row,
    style_total_row,
    draw_subcategory_divider,
)

KEY_FILL = PatternFill("solid", fgColor="FFF3CD")
KEY_FONT = Font(bold=True, color="7A4E00")


def _safe_float(v) -> float:
    try:
        return float(v or 0)
    except Exception:
        return 0.0


def _trend_status(delta_qty: float, delta_rev: float) -> str:
    if delta_qty > 0 and delta_rev > 0:
        return "Спрос растет"
    if delta_qty < 0 and delta_rev < 0:
        return "Спрос снижается"
    if delta_qty > 0 and delta_rev <= 0:
        return "Объем растет, выручка нет"
    if delta_qty < 0 and delta_rev > 0:
        return "Выручка держится за счет цены"
    return "Стабильно"


def _get_year_weights(subcat_qty_prev: float, subcat_qty_last: float) -> tuple[float, float]:
    """
    Если подкатегория резко просела по qty, чуть больше весим прошлый период,
    чтобы управленческий ориентир не прыгал.
    """
    if subcat_qty_prev > 0 and subcat_qty_last < subcat_qty_prev * 0.85:
        return 0.65, 0.35
    return 0.40, 0.60


def _calc_recommended_band(subcat_rows: list[dict]) -> str | None:
    """
    Управленческий выбор ключевого диапазона.

    Логика score:
    - 55%: доля qty текущего периода
    - 20%: устойчивая доля qty (с учетом прошлого периода)
    - 15%: доля выручки текущего периода
    - 10%: бонус/штраф за динамику qty
    """
    if not subcat_rows:
        return None

    total_qty_prev = sum(r["qty_prev"] for r in subcat_rows)
    total_qty_last = sum(r["qty_last"] for r in subcat_rows)
    total_rev_last = sum(r["rev_last"] for r in subcat_rows)

    w_prev, w_last = _get_year_weights(total_qty_prev, total_qty_last)
    weighted_qty_total = total_qty_prev * w_prev + total_qty_last * w_last

    best_band = None
    best_score = -10**9

    for r in subcat_rows:
        weighted_qty = r["qty_prev"] * w_prev + r["qty_last"] * w_last

        qty_share_last = r["qty_last"] / total_qty_last if total_qty_last else 0
        weighted_qty_share = weighted_qty / weighted_qty_total if weighted_qty_total else 0
        rev_share_last = r["rev_last"] / total_rev_last if total_rev_last else 0

        if r["qty_prev"] > 0:
            qty_growth = (r["qty_last"] - r["qty_prev"]) / r["qty_prev"]
        else:
            qty_growth = 1.0 if r["qty_last"] > 0 else 0.0

        trend_component = max(min(qty_growth, 1.0), -1.0)

        score = (
            qty_share_last * 0.55
            + weighted_qty_share * 0.20
            + rev_share_last * 0.15
            + trend_component * 0.10
        )

        r["recommended_score"] = score

        if score > best_score:
            best_score = score
            best_band = r["price_band"]

    return best_band


def _calc_anchor_price(
    price_prev: float,
    price_last: float,
    qty_prev: float,
    qty_last: float,
    w_prev: float,
    w_last: float,
) -> float:
    """
    Опорная цена = сглаженная средняя цена сегмента
    с учетом и веса периодов, и объемов qty.
    """
    weighted_prev = qty_prev * w_prev if price_prev > 0 and qty_prev > 0 else 0
    weighted_last = qty_last * w_last if price_last > 0 and qty_last > 0 else 0

    denom = weighted_prev + weighted_last
    if denom > 0:
        return (price_prev * weighted_prev + price_last * weighted_last) / denom

    if price_last > 0:
        return price_last
    if price_prev > 0:
        return price_prev
    return 0.0


def _build_corridor_text(
    anchor_price: float,
    band_low: float,
    band_high: float,
    width: float = 0.08,
) -> str:
    """
    Коридор строится вокруг опорной цены,
    но обязательно остается внутри фактических границ сегмента.
    """
    if anchor_price <= 0:
        return ""

    low = anchor_price * (1 - width)
    high = anchor_price * (1 + width)

    if band_low > 0:
        low = max(low, band_low)
    if band_high > 0:
        high = min(high, band_high)

    if band_low > 0 and band_high > 0 and low >= high:
        low = band_low
        high = band_high

    return f"{low:,.0f} – {high:,.0f}"


def _apply_key_highlight(ws, row_num: int):
    for col in (3, 6, 7, 19):  # Диапазон, Опорная цена, Коридор, Статус
        cell = ws.cell(row=row_num, column=col)
        cell.fill = KEY_FILL
        cell.font = KEY_FONT


def build_price_segmentation_sheet(wb, rows):
    ws = wb.create_sheet("Price segmentation")
    ws.sheet_properties.outlinePr.summaryBelow = False

    draw_toc_button(ws, cell="A1", target_sheet="TOC")

    headers = [
        "Категория",
        "Подкатегория",
        "Диапазон",
        "Нижняя\nграница",
        "Верхняя\nграница",
        "Опорная\nцена",
        "Реком.\nкоридор",
        "Выручка\nпред. период",
        "Выручка\nтек. период",
        "Δ Выручка",
        "Qty\nпред. период",
        "Qty\nтек. период",
        "Δ Qty",
        "Ср. цена\nпред. период",
        "Ср. цена\nтек. период",
        "Δ Цена",
        "Доля Qty\nтек. период",
        "Доля выручки\nтек. период",
        "Статус",
    ]

    header_row = 8
    data_start_row = 9

    draw_table_header_with_gaps(
        ws,
        row=header_row,
        headers=headers,
        wrap=True,
    )

    grouped = defaultdict(dict)
    years = set()
    global_is_partial_year = 0

    for r in rows:
        category = r.get("category") or "Без категории"
        subcategory = r.get("subcategory") or "Нет подкатегории"
        price_band = r.get("price_band") or "Без диапазона"
        year = int(r.get("year"))
        years.add(year)

        global_is_partial_year = int(r.get("is_partial_year") or 0)

        grouped[(category, subcategory, price_band)][year] = {
            "period_revenue": _safe_float(r.get("period_revenue")),
            "period_qty": _safe_float(r.get("period_qty")),
            "segment_low_bound": _safe_float(r.get("segment_low_bound")),
            "segment_high_bound": _safe_float(r.get("segment_high_bound")),
            "cutoff_year": r.get("cutoff_year"),
            "cutoff_month": r.get("cutoff_month"),
            "cutoff_day": r.get("cutoff_day"),
        }

    years = sorted(years)
    prev_year = years[-2] if len(years) >= 2 else None
    last_year = years[-1] if len(years) >= 1 else None

    cutoff_year = None
    cutoff_month = None
    cutoff_day = None

    if grouped and last_year is not None:
        sample_vals = next(iter(grouped.values()))
        sample_last = sample_vals.get(last_year, {})
        cutoff_year = sample_last.get("cutoff_year")
        cutoff_month = sample_last.get("cutoff_month")
        cutoff_day = sample_last.get("cutoff_day")

    is_ytd_comparison = bool(
        prev_year is not None
        and last_year is not None
        and cutoff_year
        and int(last_year) == int(cutoff_year)
        and int(global_is_partial_year) == 1
    )

    if is_ytd_comparison:
        note = (
            f"Сравнение YTD: {prev_year} vs {last_year} "
            f"на {int(cutoff_day):02d}.{int(cutoff_month):02d}.{int(cutoff_year)}. "
            f"То есть {prev_year} сравнивается по периоду "
            f"01.01–{int(cutoff_day):02d}.{int(cutoff_month):02d}, а не с полным годом. "
            f"Ценовые сегменты также зафиксированы по YTD последних 2 лет, "
            f"поэтому YoY сопоставим. "
            f"'Нижняя/Верхняя граница' — это границы сегмента, "
            f"'Опорная цена' — сглаженный ориентир внутри сегмента."
        )
    else:
        note = (
            "Сравнение по полным годам. "
            "Ценовые сегменты фиксированы по последним 2 полным годам, "
            "поэтому YoY сопоставим. "
            "'Нижняя/Верхняя граница' — это границы сегмента, "
            "'Опорная цена' — сглаженный ориентир внутри сегмента."
        )

    draw_sheet_header(
        ws,
        title="Спрос по ценовым диапазонам",
        subtitle="Фиксированные ценовые сегменты внутри каждой подкатегории",
        note=note,
        line_to_col=19,
    )

    prepared_rows = []
    subcat_map = defaultdict(list)

    band_order = {
        "Низкий диапазон": 1,
        "Средний диапазон": 2,
        "Высокий диапазон": 3,
    }

    for (category, subcategory, price_band), vals in grouped.items():
        prev_vals = vals.get(prev_year, {}) if prev_year else {}
        last_vals = vals.get(last_year, {}) if last_year else {}

        rev_prev = _safe_float(prev_vals.get("period_revenue"))
        rev_last = _safe_float(last_vals.get("period_revenue"))
        qty_prev = _safe_float(prev_vals.get("period_qty"))
        qty_last = _safe_float(last_vals.get("period_qty"))

        if rev_prev == 0 and rev_last == 0 and qty_prev == 0 and qty_last == 0:
            continue

        row_data = {
            "category": category,
            "subcategory": subcategory,
            "price_band": price_band,
            "segment_low_bound": (
                _safe_float(last_vals.get("segment_low_bound"))
                or _safe_float(prev_vals.get("segment_low_bound"))
            ),
            "segment_high_bound": (
                _safe_float(last_vals.get("segment_high_bound"))
                or _safe_float(prev_vals.get("segment_high_bound"))
            ),
            "rev_prev": rev_prev,
            "rev_last": rev_last,
            "qty_prev": qty_prev,
            "qty_last": qty_last,
        }
        subcat_map[(category, subcategory)].append(row_data)

    for (category, subcategory), items in subcat_map.items():
        recommended_band = _calc_recommended_band(items)

        subcat_total_rev_last = sum(r["rev_last"] for r in items)
        subcat_total_qty_last = sum(r["qty_last"] for r in items)
        subcat_total_qty_prev = sum(r["qty_prev"] for r in items)

        w_prev, w_last = _get_year_weights(subcat_total_qty_prev, subcat_total_qty_last)

        for r in items:
            delta_rev = r["rev_last"] - r["rev_prev"]
            delta_qty = r["qty_last"] - r["qty_prev"]

            price_prev = r["rev_prev"] / r["qty_prev"] if r["qty_prev"] else 0
            price_last = r["rev_last"] / r["qty_last"] if r["qty_last"] else 0
            delta_price = price_last - price_prev

            share_qty_last = r["qty_last"] / subcat_total_qty_last if subcat_total_qty_last else 0
            share_rev_last = r["rev_last"] / subcat_total_rev_last if subcat_total_rev_last else 0

            anchor_price = _calc_anchor_price(
                price_prev=price_prev,
                price_last=price_last,
                qty_prev=r["qty_prev"],
                qty_last=r["qty_last"],
                w_prev=w_prev,
                w_last=w_last,
            )

            corridor_text = _build_corridor_text(
                anchor_price=anchor_price,
                band_low=r["segment_low_bound"],
                band_high=r["segment_high_bound"],
                width=0.08,
            )

            if r["price_band"] == recommended_band:
                status = "Ключевой диапазон"
            else:
                status = _trend_status(delta_qty, delta_rev)

            prepared_rows.append(
                [
                    category,                # 1
                    subcategory,             # 2
                    r["price_band"],         # 3
                    r["segment_low_bound"],  # 4
                    r["segment_high_bound"], # 5
                    anchor_price,            # 6
                    corridor_text,           # 7
                    r["rev_prev"],           # 8
                    r["rev_last"],           # 9
                    delta_rev,               # 10
                    r["qty_prev"],           # 11
                    r["qty_last"],           # 12
                    delta_qty,               # 13
                    price_prev,              # 14
                    price_last,              # 15
                    delta_price,             # 16
                    share_qty_last,          # 17
                    share_rev_last,          # 18
                    status,                  # 19
                ]
            )

    prepared_rows.sort(
        key=lambda x: (
            x[0],
            x[1],
            band_order.get(x[2], 999),
        )
    )

    cur_row = data_start_row

    total_rev_prev = 0.0
    total_rev_last = 0.0
    total_qty_prev = 0.0
    total_qty_last = 0.0

    prev_group = None

    for row in prepared_rows:
        current_group = (row[0], row[1])

        if prev_group is not None and current_group != prev_group:
            draw_subcategory_divider(ws, cur_row, col_start=1, col_end=19, height=5)
            cur_row += 1

        style_data_row(
            ws,
            row=cur_row,
            values=row,
            number_formats={
                4: FORMATS["money"],
                5: FORMATS["money"],
                6: FORMATS["money"],
                8: FORMATS["money"],
                9: FORMATS["money"],
                10: FORMATS["money"],
                11: "#,##0",
                12: "#,##0",
                13: "#,##0;[Red]-#,##0",
                14: FORMATS["money"],
                15: FORMATS["money"],
                16: FORMATS["money"],
                17: "0.0%",
                18: "0.0%",
            },
        )

        for col_idx in range(4, 19):
            ws.cell(row=cur_row, column=col_idx).alignment = ALIGNMENTS["center"]

        ws.cell(row=cur_row, column=19).alignment = ALIGNMENTS["left_wrap"]

        if row[18] == "Ключевой диапазон":
            _apply_key_highlight(ws, cur_row)

        total_rev_prev += row[7]
        total_rev_last += row[8]
        total_qty_prev += row[10]
        total_qty_last += row[11]

        prev_group = current_group
        cur_row += 1

    total_delta_rev = total_rev_last - total_rev_prev
    total_delta_qty = total_qty_last - total_qty_prev

    total_price_prev = total_rev_prev / total_qty_prev if total_qty_prev else 0
    total_price_last = total_rev_last / total_qty_last if total_qty_last else 0
    total_delta_price = total_price_last - total_price_prev

    total_anchor_price = _calc_anchor_price(
        price_prev=total_price_prev,
        price_last=total_price_last,
        qty_prev=total_qty_prev,
        qty_last=total_qty_last,
        w_prev=0.40,
        w_last=0.60,
    )

    total_corridor = _build_corridor_text(
        anchor_price=total_anchor_price,
        band_low=0,
        band_high=0,
        width=0.08,
    )

    style_total_row(
        ws,
        row=cur_row,
        values=[
            "ИТОГО",
            "",
            "",
            "",
            "",
            total_anchor_price,
            total_corridor,
            total_rev_prev,
            total_rev_last,
            total_delta_rev,
            total_qty_prev,
            total_qty_last,
            total_delta_qty,
            total_price_prev,
            total_price_last,
            total_delta_price,
            1 if total_qty_last else 0,
            1 if total_rev_last else 0,
            "",
        ],
        number_formats={
            6: FORMATS["money"],
            8: FORMATS["money"],
            9: FORMATS["money"],
            10: FORMATS["money"],
            11: "#,##0",
            12: "#,##0",
            13: "#,##0;[Red]-#,##0",
            14: FORMATS["money"],
            15: FORMATS["money"],
            16: FORMATS["money"],
            17: "0.0%",
            18: "0.0%",
        },
    )

    set_column_widths(
        ws,
        {
            "A": 20,
            "B": 26,
            "C": 18,
            "D": 14,
            "E": 14,
            "F": 14,
            "G": 18,
            "H": 16,
            "I": 16,
            "J": 14,
            "K": 11,
            "L": 11,
            "M": 10,
            "N": 14,
            "O": 14,
            "P": 12,
            "Q": 13,
            "R": 15,
            "S": 22,
        },
    )

    set_row_heights(
        ws,
        {
            2: 24,
            3: 18,
            4: 42,
            8: 42,
        },
    )

    hide_grid_and_freeze(ws, "H9")