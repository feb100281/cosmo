# orders/reports/styles/theme.py
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

COLORS = {
    "dark_green": "2F6656",
    "light_green": "E7F1ED",
    "total_green": "DCECE6",
    "light_gray": "F7F7F7",
    "summary_fill": "FBFBFB",
    "back_fill": "E7F1ED",
    "border_gray": "D9D9D9",
    "text_gray": "666666",
    "text_dark": "666666",
    "white": "FFFFFF",
    "black": "1F1F1F",
    "blue": "2F75B5",
    "back_text_green": "1F5E4E",
    "heat_light": "EEF6F2",
    "divider_fill": "E7F1ED",
    "text_gray": "4F4F4F" ,
    # Новые цвета для акцентов в зелёной гамме
    "accent_green": "1A4D3F",      # для важных чисел
    "warning_green": "6B8E23",     # для умеренных значений
    "warning_red": "7B1F3A",
    "muted_green": "B5C9B5",       # для фона карточек
    "border_light": "C8D9C8",      # для разделителей
     "orange": "E67E22",            # оранжевый для категории B
    "light_blue": "3498DB",        # светло-синий для дата-баров
}

FILLS = {
    "header": PatternFill("solid", fgColor=COLORS["dark_green"]),
    "section": PatternFill("solid", fgColor=COLORS["light_green"]),
    "alt": PatternFill("solid", fgColor=COLORS["light_gray"]),
    "total": PatternFill("solid", fgColor=COLORS["total_green"]),
    "summary": PatternFill("solid", fgColor=COLORS["summary_fill"]),
    "back": PatternFill("solid", fgColor=COLORS["back_fill"]),
    "heat_light": PatternFill("solid", fgColor=COLORS["heat_light"]),
    "divider": PatternFill("solid", fgColor=COLORS["divider_fill"]),
    "none": PatternFill(fill_type=None),
    "kpi_card": PatternFill("solid", fgColor=COLORS["muted_green"]),
    "odd_row": PatternFill("solid", fgColor=COLORS["light_gray"]),     
    "even_row": PatternFill(fill_type=None),  
}

FONTS = {
    "title": Font(name="Roboto", size=16, bold=True, color=COLORS["black"]),
    "subtitle": Font(name="Roboto", size=10, color=COLORS["text_gray"]),
    "section": Font(name="Roboto", size=11, bold=True, color=COLORS["black"]),
    "header_white": Font(name="Roboto", size=10, bold=True, color=COLORS["white"]),
    "bold": Font(name="Roboto", size=10, bold=True, color=COLORS["black"]),
    "normal": Font(name="Roboto", size=10, color=COLORS["black"]),
    "total": Font(name="Roboto", size=11, bold=True, color=COLORS["black"]),
    "back": Font(name="Roboto", size=10, bold=True, color=COLORS["back_text_green"]),
}

thin = Side(style="thin", color=COLORS["border_gray"])
medium_dark = Side(style="medium", color=COLORS["dark_green"])
divider_side = Side(style="thin", color=COLORS["border_gray"])

BORDERS = {
    "thin": Border(left=thin, right=thin, top=thin, bottom=thin),
    "bottom_thin": Border(bottom=thin),
    "bottom_medium": Border(bottom=medium_dark),
    "top_bottom_medium": Border(top=medium_dark, bottom=medium_dark),
    "divider": Border(top=divider_side, bottom=divider_side),
    "none": Border(),
}

ALIGNMENTS = {
    "left": Alignment(horizontal="left", vertical="center"),
    "center": Alignment(horizontal="center", vertical="center"),
    "right": Alignment(horizontal="right", vertical="center"),
    "center_wrap": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "left_wrap": Alignment(horizontal="left", vertical="top", wrap_text=True),
}

FORMATS = {
    "money": '#,##0;(#,##0)',
    "int": '#,##0',
    "pct": '0.0%',
    "date": 'yyyy-mm-dd',
}