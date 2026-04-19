# orders/reports/sheets/payments_analysis_sheet.py
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .base_sheet import BaseSheet
from ..styles.helpers import draw_toc_button
from ..components import create_kpi_cards, create_header, create_table
from ..components.sheet_title import create_sheet_title
from ..components.navigation_link import NavigationLink
from ..queries.payments_analysis_query import PaymentsAnalysisQueries
from ..styles.theme import COLORS, BORDERS


class PaymentsAnalysisSheet(BaseSheet):
    """Анализ оплат - полная аналитика по всем платежам"""

    def __init__(self, workbook, sheet_number, request=None):
        super().__init__(workbook, sheet_number)
        self.sheet_title = create_sheet_title(self.ws)
        self.kpi = create_kpi_cards(self.ws)
        self.header = create_header(self.ws)
        self.table = create_table(self.ws)
        self.queries = PaymentsAnalysisQueries(request)

    def _format_currency(self, value):
        """Форматирование суммы со знаком"""
        if value is None or value == 0:
            return "0 ₽"

        sign = "-" if value < 0 else ""
        formatted = f"{int(round(abs(value))):,} ₽".replace(",", " ")
        return f"{sign}{formatted}"

    def _format_number_str(self, value):
        """Форматирование числа"""
        if value is None or value == 0:
            return "0"
        sign = "-" if value < 0 else ""
        return f"{sign}{int(round(abs(value))):,}".replace(",", " ")

    def _format_dynamics(self, value):
        """Форматирование динамики со стрелкой"""
        if value > 0:
            return f"↑ +{value}%"
        elif value < 0:
            return f"↓ {value}%"
        return "→ 0%"

    def _format_store_name(self, name):
        if not name:
            return ""
        return str(name).upper()

    def _format_register_name(self, name):
        """Форматирование названия регистра/документа"""
        if not name:
            return ""
        name_str = str(name)
        if len(name_str) > 40:
            return name_str[:37] + "..."
        return name_str

    def _format_date(self, date_value):
        if not date_value:
            return ""
        if hasattr(date_value, 'strftime'):
            return date_value.strftime('%d.%m.%Y')
        return str(date_value)

    def _format_month_name(self, month_str):
        """Форматирует месяц из YYYY-MM в МММ YYYY"""
        try:
            from datetime import datetime
            date_obj = datetime.strptime(month_str, '%Y-%m')
            months_ru = {
                1: 'ЯНВ', 2: 'ФЕВ', 3: 'МАР', 4: 'АПР',
                5: 'МАЙ', 6: 'ИЮН', 7: 'ИЮЛ', 8: 'АВГ',
                9: 'СЕН', 10: 'ОКТ', 11: 'НОЯ', 12: 'ДЕК'
            }
            return f"{months_ru[date_obj.month]} {date_obj.year}"
        except Exception:
            return month_str

    def _get_cell_font(self, is_header=False, is_bold=False, is_negative=False):
        """Возвращает стандартный шрифт для ячеек"""
        if is_header:
            return Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
        if is_negative:
            return Font(name="Roboto", size=9, bold=True, color=COLORS["warning_red"])
        if is_bold:
            return Font(name="Roboto", size=9, bold=True, color=COLORS["text_dark"])
        return Font(name="Roboto", size=9, color=COLORS["text_dark"])

    def build(self):
        """Построение листа анализа оплат"""

        metrics = self.queries.get_payments_metrics()
        stores = self.queries.get_payments_by_store(limit=None)
        trends = self.queries.get_payments_trend(months=6)
        monthly_by_store = self.queries.get_payments_by_store_monthly()
        summary_by_type = self.queries.get_payments_summary_by_type()
        large_payments = self.queries.get_large_payments(threshold=300000, limit=20)

        row = 1

        # Шапка
        draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")
        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 8

        self.ws.column_dimensions["A"].width = 2
        row = self.sheet_title.draw(
            row=3,
            title="АНАЛИЗ ОПЛАТ",
            subtitle="Полная аналитика по всем платежам: тренды, магазины, типы операций",
            date_text=f"Сформировано: {self.queries.today.strftime('%d.%m.%Y')}",
            start_col=2,
            end_col=7
        )
        row += 1

        # ============================================================
        # KPI КАРТОЧКИ
        # ============================================================
        mtd = metrics['mtd']
        ytd = metrics['ytd']
        dynamics = metrics['mtd_dynamics']

        outgoing_total_ytd = summary_by_type['outgoing']['total'] if summary_by_type else 0
        outgoing_count_ytd = summary_by_type['outgoing']['count'] if summary_by_type else 0

        row_cards = [
            {
                'title': 'СУММА ОПЛАТ (MTD)',
                'value': self._format_currency(mtd['total_amount']),
                'subtitle': f"{self._format_dynamics(dynamics['amount_vs_last_month'])} к прошлому месяцу",
                'color': COLORS["blue"],
                'width': 2,
            },
            {
                'title': 'СУММА ОПЛАТ (YTD)',
                'value': self._format_currency(ytd['total_amount']),
                'subtitle': f"{self._format_number_str(ytd['total_payments'])} платежей",
                'color': COLORS["blue"],
                'width': 2,
            },
            {
                'title': 'ВОЗВРАТЫ (YTD)',
                'value': self._format_currency(outgoing_total_ytd),
                'subtitle': f"{self._format_number_str(outgoing_count_ytd)} операций возврата",
                'color': COLORS["warning_red"],
                'width': 2,
            },
        ]

        row = self.kpi.draw_row(row, row_cards)
        row += 1

        # ============================================================
        # СВОДКА ПО ТИПАМ ОПЕРАЦИЙ
        # ============================================================
        if summary_by_type:
            row = self.header.draw(row, "СВОДКА ПО ТИПАМ ОПЕРАЦИЙ (YTD)", end_col=5)

            headers = ["ТИП ОПЕРАЦИИ", "СУММА", "КОЛИЧЕСТВО", "ДОЛЯ ОТ ИТОГА"]

            for col_idx, header in enumerate(headers, start=2):
                cell = self.ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
                cell.fill = PatternFill(
                    start_color=COLORS["dark_green"],
                    end_color=COLORS["dark_green"],
                    fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

            row += 1
            start_data_row = row

            total_all = summary_by_type['total_all'] or 0

            data_rows = [
                summary_by_type['incoming'],
                summary_by_type['store_revenue'],
                summary_by_type['outgoing'],
            ]

            for item in data_rows:
                is_negative = (item['total'] or 0) < 0

                cell_type = self.ws.cell(row=row, column=2, value=item['name'])
                cell_type.font = self._get_cell_font(is_negative=is_negative)

                cell_amount = self.ws.cell(row=row, column=3, value=item['total'])
                cell_amount.font = self._get_cell_font(is_negative=is_negative)
                cell_amount.number_format = '#,##0 ₽'
                cell_amount.alignment = Alignment(horizontal="right")

                cell_count = self.ws.cell(row=row, column=4, value=item['count'])
                cell_count.font = self._get_cell_font()
                cell_count.number_format = '#,##0'
                cell_count.alignment = Alignment(horizontal="right")

                pct = ((item['total'] / total_all) * 100) if total_all != 0 else 0
                cell_pct = self.ws.cell(row=row, column=5, value=pct)
                cell_pct.font = self._get_cell_font(is_negative=(pct < 0))
                cell_pct.number_format = '0.0"%"'
                cell_pct.alignment = Alignment(horizontal="right")

                row += 1

            # Итого
            cell_type = self.ws.cell(row=row, column=2, value=summary_by_type['net']['name'])
            cell_type.font = self._get_cell_font(is_bold=True)

            cell_total_amount = self.ws.cell(row=row, column=3, value=total_all)
            cell_total_amount.font = self._get_cell_font(
                is_bold=True,
                is_negative=(total_all < 0)
            )
            cell_total_amount.number_format = '#,##0 ₽'
            cell_total_amount.alignment = Alignment(horizontal="right")
            cell_total_amount.fill = PatternFill(
                start_color=COLORS["light_green"],
                end_color=COLORS["light_green"],
                fill_type="solid"
            )

            cell_total_count = self.ws.cell(row=row, column=4, value="")
            cell_total_count.fill = PatternFill(
                start_color=COLORS["light_green"],
                end_color=COLORS["light_green"],
                fill_type="solid"
            )

            cell_total_pct = self.ws.cell(row=row, column=5, value=100 if total_all != 0 else 0)
            cell_total_pct.font = self._get_cell_font(is_bold=True)
            cell_total_pct.number_format = '0.0"%"'
            cell_total_pct.alignment = Alignment(horizontal="right")
            cell_total_pct.fill = PatternFill(
                start_color=COLORS["light_green"],
                end_color=COLORS["light_green"],
                fill_type="solid"
            )

            end_data_row = row

            for col in range(2, 6):
                for r in range(start_data_row, end_data_row + 1):
                    self.ws.cell(row=r, column=col).border = BORDERS["thin"]

            row += 2

        # ============================================================
        # МЕСЯЧНЫЙ ТРЕНД ОПЛАТ
        # ============================================================
        if trends:
            row = self.header.draw(row, "МЕСЯЧНАЯ ДИНАМИКА ОПЛАТ", end_col=5)

            headers = ["МЕСЯЦ", "СУММА ОПЛАТ", "КОЛ-ВО ПЛАТЕЖЕЙ", "СР. ПЛАТЕЖ"]

            for col_idx, header in enumerate(headers, start=2):
                cell = self.ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
                cell.fill = PatternFill(
                    start_color=COLORS["dark_green"],
                    end_color=COLORS["dark_green"],
                    fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

            row += 1
            start_data_row = row

            for t in trends:
                month_name = t['month'].strftime('%b %Y').upper() if t['month'] else ""
                month_map = {
                    'JAN': 'ЯНВ', 'FEB': 'ФЕВ', 'MAR': 'МАР', 'APR': 'АПР',
                    'MAY': 'МАЙ', 'JUN': 'ИЮН', 'JUL': 'ИЮЛ', 'AUG': 'АВГ',
                    'SEP': 'СЕН', 'OCT': 'ОКТ', 'NOV': 'НОЯ', 'DEC': 'ДЕК'
                }
                for eng, rus in month_map.items():
                    if eng in month_name:
                        month_name = month_name.replace(eng, rus)
                        break

                cell_month = self.ws.cell(row=row, column=2, value=month_name)
                cell_month.font = self._get_cell_font()

                is_negative = (t['total_amount'] or 0) < 0
                cell_amount = self.ws.cell(row=row, column=3, value=t['total_amount'])
                cell_amount.font = self._get_cell_font(is_negative=is_negative)
                cell_amount.number_format = '#,##0 ₽'
                cell_amount.alignment = Alignment(horizontal="right")

                cell_count = self.ws.cell(row=row, column=4, value=t['payments_count'])
                cell_count.font = self._get_cell_font()
                cell_count.number_format = '#,##0'
                cell_count.alignment = Alignment(horizontal="right")

                avg_payment = t['avg_payment'] or 0
                cell_avg = self.ws.cell(row=row, column=5, value=avg_payment)
                cell_avg.font = self._get_cell_font(is_negative=(avg_payment < 0))
                cell_avg.number_format = '#,##0 ₽'
                cell_avg.alignment = Alignment(horizontal="right")

                row += 1

            end_data_row = row - 1

            for col in range(2, 6):
                for r in range(start_data_row, end_data_row + 1):
                    self.ws.cell(row=r, column=col).border = BORDERS["thin"]

            row += 2

        # ============================================================
        # ПОМЕСЯЧНАЯ РАЗБИВКА ПО МАГАЗИНАМ
        # ============================================================
        if monthly_by_store['data'] and monthly_by_store['months']:
            num_months = len(monthly_by_store['months'])
            monthly_by_store['months'].sort()

            row = self.header.draw(
                row,
                f"ПОМЕСЯЧНАЯ ДИНАМИКА ПО МАГАЗИНАМ (за {monthly_by_store['year']} г.)",
                end_col=4 + num_months
            )

            headers = ["МАГАЗИН"] + [self._format_month_name(m) for m in monthly_by_store['months']] + ["ИТОГО"]
            start_col = 2

            for col_idx, header in enumerate(headers, start_col):
                cell = self.ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
                cell.fill = PatternFill(
                    start_color=COLORS["dark_green"],
                    end_color=COLORS["dark_green"],
                    fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
                self.ws.column_dimensions[get_column_letter(col_idx)].width = 14

            row += 1
            start_data_row = row

            # Словарь для хранения итогов по месяцам
            monthly_totals = {month: 0 for month in monthly_by_store['months']}
            grand_total = 0

            for store, months_data in monthly_by_store['data'].items():
                store_display = store[:35] if store and len(store) > 35 else store
                cell_store = self.ws.cell(row=row, column=start_col, value=self._format_store_name(store_display))
                cell_store.font = self._get_cell_font()

                total_by_store = 0
                for idx, month in enumerate(monthly_by_store['months']):
                    amount = months_data.get(month, 0)
                    total_by_store += amount
                    monthly_totals[month] += amount  # Добавляем в общий итог по месяцу

                    is_negative = amount < 0
                    cell_amount = self.ws.cell(row=row, column=start_col + 1 + idx, value=amount)
                    cell_amount.font = self._get_cell_font(is_negative=is_negative)
                    cell_amount.number_format = '#,##0 ₽'
                    cell_amount.alignment = Alignment(horizontal="right")

                is_negative = total_by_store < 0
                cell_total = self.ws.cell(row=row, column=start_col + num_months + 1, value=total_by_store)
                cell_total.font = self._get_cell_font(is_bold=True, is_negative=is_negative)
                cell_total.number_format = '#,##0 ₽'
                cell_total.alignment = Alignment(horizontal="right")
                cell_total.fill = PatternFill(
                    start_color=COLORS["light_green"],
                    end_color=COLORS["light_green"],
                    fill_type="solid"
                )
                
                grand_total += total_by_store
                row += 1

            # ==================== СТРОКА ИТОГО ====================
            # Пустая строка для разделения
            row += 1
            
            # Ячейка "ИТОГО"
            cell_total_label = self.ws.cell(row=row, column=start_col, value="ИТОГО ПО ВСЕМ МАГАЗИНАМ")
            cell_total_label.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
            cell_total_label.fill = PatternFill(
                start_color=COLORS["dark_green"],
                end_color=COLORS["dark_green"],
                fill_type="solid"
            )
            cell_total_label.alignment = Alignment(horizontal="left", vertical="center")
            
            # Итоги по месяцам
            for idx, month in enumerate(monthly_by_store['months']):
                month_total = monthly_totals[month]
                is_negative = month_total < 0
                cell_month_total = self.ws.cell(row=row, column=start_col + 1 + idx, value=month_total)
                cell_month_total.font = self._get_cell_font(is_bold=True, is_negative=is_negative)
                cell_month_total.number_format = '#,##0 ₽'
                cell_month_total.alignment = Alignment(horizontal="right")
                cell_month_total.fill = PatternFill(
                    start_color=COLORS["light_green"],
                    end_color=COLORS["light_green"],
                    fill_type="solid"
                )
            
            # Общий итог
            cell_grand_total = self.ws.cell(row=row, column=start_col + num_months + 1, value=grand_total)
            cell_grand_total.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
            cell_grand_total.fill = PatternFill(
                start_color=COLORS["dark_green"],
                end_color=COLORS["dark_green"],
                fill_type="solid"
            )
            cell_grand_total.number_format = '#,##0 ₽'
            cell_grand_total.alignment = Alignment(horizontal="right")
            
            end_data_row = row
            
            # Рисуем границы для всех строк (включая итоговую)
            for col in range(start_col, start_col + len(headers)):
                for r in range(start_data_row, end_data_row + 1):
                    self.ws.cell(row=r, column=col).border = BORDERS["thin"]

            row += 1
            footnote_cell = self.ws.cell(
                row=row,
                column=start_col,
                value="* Все суммы отражаются как есть в поле amount, включая отрицательные значения"
            )
            footnote_cell.font = Font(name="Roboto", size=8, color=COLORS["text_gray"], italic=True)

            row += 2
            
            # Ссылка 
            nav_link = NavigationLink(self.ws)
            row = nav_link.draw(
                row=row,
                text="💰 Детальную подневную разбивку по магазинам смотрите на отдельном листе «Подневная динамика» →",
                target_sheet="7",
                target_cell="A1",
                start_col=2,
                end_col=6,
                alignment="left",
                with_icon=False
            )

        # ============================================================
        # ВСЕ МАГАЗИНЫ ПО СУММЕ ОПЛАТ
        # ============================================================
        if stores:
            row = self.header.draw(row, "МАГАЗИНЫ ПО СУММЕ ОПЛАТ (YTD)", end_col=6)

            headers = ["МАГАЗИН", "СУММА ОПЛАТ", "КОЛ-ВО ПЛАТЕЖЕЙ", "СР. ПЛАТЕЖ", "КАСС/РЕГИСТРОВ"]

            for col_idx, header in enumerate(headers, start=2):
                cell = self.ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
                cell.fill = PatternFill(
                    start_color=COLORS["dark_green"],
                    end_color=COLORS["dark_green"],
                    fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

            row += 1
            start_data_row = row

            for s in stores:
                cell_store = self.ws.cell(row=row, column=2, value=self._format_store_name(s['store']))
                cell_store.font = self._get_cell_font()

                total_amount = s['total_amount'] or 0
                cell_amount = self.ws.cell(row=row, column=3, value=total_amount)
                cell_amount.font = self._get_cell_font(is_negative=(total_amount < 0))
                cell_amount.number_format = '#,##0 ₽'
                cell_amount.alignment = Alignment(horizontal="right")

                cell_count = self.ws.cell(row=row, column=4, value=s['payments_count'])
                cell_count.font = self._get_cell_font()
                cell_count.number_format = '#,##0'
                cell_count.alignment = Alignment(horizontal="right")

                avg_payment = s['avg_payment'] or 0
                cell_avg = self.ws.cell(row=row, column=5, value=avg_payment)
                cell_avg.font = self._get_cell_font(is_negative=(avg_payment < 0))
                cell_avg.number_format = '#,##0 ₽'
                cell_avg.alignment = Alignment(horizontal="right")

                cell_registers = self.ws.cell(row=row, column=6, value=s['unique_registers'])
                cell_registers.font = self._get_cell_font()
                cell_registers.number_format = '#,##0'
                cell_registers.alignment = Alignment(horizontal="right")

                row += 1

            end_data_row = row - 1

            for col in range(2, 7):
                for r in range(start_data_row, end_data_row + 1):
                    self.ws.cell(row=r, column=col).border = BORDERS["thin"]

            row += 2

        # ============================================================
        # КРУПНЫЕ ПЛАТЕЖИ (≥300 000 ₽)
        # ============================================================
        if large_payments:
            row = self.header.draw(row, "КРУПНЫЕ ПЛАТЕЖИ (≥300 000 ₽)", end_col=5)

            headers = ["ДАТА", "МАГАЗИН", "РЕГИСТР/ДОКУМЕНТ", "СУММА"]

            for col_idx, header in enumerate(headers, start=2):
                cell = self.ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
                cell.fill = PatternFill(
                    start_color=COLORS["dark_green"],
                    end_color=COLORS["dark_green"],
                    fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

            row += 1
            start_data_row = row

            for p in large_payments:
                cell_date = self.ws.cell(row=row, column=2, value=self._format_date(p['date']))
                cell_date.font = self._get_cell_font()

                cell_store = self.ws.cell(row=row, column=3, value=self._format_store_name(p['store']))
                cell_store.font = self._get_cell_font()

                cell_register = self.ws.cell(
                    row=row,
                    column=4,
                    value=self._format_register_name(p.get('register', '—'))
                )
                cell_register.font = self._get_cell_font()

                amount_value = p['amount'] or 0
                cell_amount = self.ws.cell(row=row, column=5, value=amount_value)
                cell_amount.font = self._get_cell_font(is_negative=(amount_value < 0))
                cell_amount.number_format = '#,##0 ₽'
                cell_amount.alignment = Alignment(horizontal="right")

                row += 1

            end_data_row = row - 1

            for col in range(2, 6):
                for r in range(start_data_row, end_data_row + 1):
                    self.ws.cell(row=r, column=col).border = BORDERS["thin"]

            row += 2





        # ============================================================
        # НАСТРОЙКА КОЛОНОК
        # ============================================================
        self.ws.column_dimensions["B"].width = 32
        self.ws.column_dimensions["C"].width = 20
        self.ws.column_dimensions["D"].width = 18
        self.ws.column_dimensions["E"].width = 18
        self.ws.column_dimensions["F"].width = 18

        self.ws.sheet_view.showGridLines = False