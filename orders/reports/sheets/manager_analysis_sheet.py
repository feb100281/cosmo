# # orders/reports/sheets/manager_analysis_sheet.py
# from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
# from .base_sheet import BaseSheet
# from ..components import create_header
# from ..components.sheet_title import create_sheet_title
# from ..styles.helpers import draw_toc_button
# from ..components.footnote import Footnote
# from ..components.navigation_link import NavigationLink
# from ..components.manager_profile import ManagerProfile
# from ..styles.theme import COLORS


# class ManagerAnalysisSheet(BaseSheet):
#     """Анализ по менеджерам - портфолио каждого менеджера"""
    
#     def __init__(self, workbook, sheet_number, request=None):
#         super().__init__(workbook, sheet_number)
#         self.sheet_title = create_sheet_title(self.ws)
#         self.header = create_header(self.ws)
#         self.request = request
    
#     def _format_currency(self, value):
#         if value is None or value == 0:
#             return "0 ₽"
#         return f"{int(round(value)):,} ₽".replace(",", " ")
    
#     def _format_number(self, value):
#         if value is None or value == 0:
#             return "0"
#         return f"{int(round(value)):,}".replace(",", " ")
    
#     def _format_percent(self, value):
#         if value is None:
#             return "0%"
#         return f"{value:.1f}%"
    
#     def _safe_float(self, value, default=0.0):
#         if value is None:
#             return default
#         try:
#             return float(value)
#         except (ValueError, TypeError):
#             return default
    
#     def _safe_int(self, value, default=0):
#         if value is None:
#             return default
#         try:
#             return int(round(float(value)))
#         except (ValueError, TypeError):
#             return default
    
#     def _format_manager_name(self, full_name):
#         """Форматирует имя менеджера: Иванов Иван -> ИВАНОВ И."""
#         if not full_name:
#             return ""
#         parts = str(full_name).strip().split()
#         if len(parts) >= 1:
#             last_name = parts[0].upper()
#             if len(parts) >= 2:
#                 first_initial = parts[1][0].upper() if parts[1] else ""
#                 return f"{last_name} {first_initial}."
#             return last_name
#         return str(full_name).upper()
    
#     def build(self, manager_data):
#         """
#         Построение листа анализа менеджеров
        
#         Args:
#             manager_data: список словарей с данными менеджеров
#                 Каждый словарь должен содержать:
#                 - manager: имя менеджера
#                 - mtd_orders: количество заказов за месяц
#                 - mtd_amount: сумма за месяц
#                 - mtd_paid: оплачено за месяц
#                 - mtd_shipped: отгружено за месяц
#                 - mtd_delivery: доставка за месяц
#                 - ytd_orders: количество заказов с начала года
#                 - ytd_amount: сумма с начала года
#                 - ytd_paid: оплачено с начала года
#                 - ytd_shipped: отгружено с начала года
#                 - ytd_delivery: доставка с начала года
#         """
        
#         row = 1
        
#         # Шапка
        
#         draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

#         self.ws.row_dimensions[1].height = 20
#         self.ws.row_dimensions[2].height = 8
        
#         self.ws.column_dimensions["A"].width = 2
#         row = self.sheet_title.draw(
#             row=3,
#             title="АНАЛИЗ ПО МЕНЕДЖЕРАМ",
#             subtitle="Портфолио менеджеров: ключевые показатели MTD и YTD",
#             date_text=f"Сформировано: {self.ws.parent.properties.created.strftime('%d.%m.%Y') if hasattr(self.ws.parent, 'properties') else ''}",
#             start_col=2,
#             end_col=7
#         )
#         row += 1
        
#         # ============================================================
#         # ВСТУПИТЕЛЬНЫЙ ТЕКСТ
#         # ============================================================
#         intro_row = row
#         self.ws.merge_cells(start_row=intro_row, start_column=2, end_row=intro_row, end_column=7)
#         intro_cell = self.ws.cell(
#             row=intro_row, 
#             column=2, 
#             value="📊 Ниже представлены ключевые показатели эффективности менеджеров за текущий месяц (MTD) и с начала года (YTD)"
#         )
#         intro_cell.font = Font(name="Roboto", size=10, italic=True, color="666666")
#         intro_cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.row_dimensions[intro_row].height = 22
#         row += 1
        
#         # Пустая строка
#         self.ws.row_dimensions[row].height = 6
#         row += 1
        
#         # ============================================================
#         # ПОРТФОЛИО МЕНЕДЖЕРОВ
#         # ============================================================
#         if manager_data:
#             manager_profile = ManagerProfile(self.ws)
            
#             # Сортируем менеджеров по сумме MTD (от большего к меньшему)
#             sorted_managers = sorted(
#                 manager_data, 
#                 key=lambda x: self._safe_float(x.get('mtd_amount', 0)), 
#                 reverse=True
#             )
            
#             for idx, manager in enumerate(sorted_managers):
#                 # Подготовка данных для MTD
#                 mtd_orders = self._safe_int(manager.get('mtd_orders', 0))
#                 mtd_amount = self._safe_float(manager.get('mtd_amount', 0))
#                 mtd_paid = self._safe_float(manager.get('mtd_paid', 0))
#                 mtd_shipped = self._safe_float(manager.get('mtd_shipped', 0))
#                 mtd_delivery = self._safe_float(manager.get('mtd_delivery', 0))
                
#                 mtd_avg_check = mtd_amount / mtd_orders if mtd_orders > 0 else 0
#                 mtd_debt = max(0, mtd_amount - mtd_paid)
#                 mtd_delivery_percent = (mtd_delivery / mtd_amount * 100) if mtd_amount > 0 else 0
#                 mtd_paid_percent = (mtd_paid / mtd_amount * 100) if mtd_amount > 0 else 0
#                 mtd_shipped_percent = (mtd_shipped / mtd_amount * 100) if mtd_amount > 0 else 0
                
#                 mtd_data = {
#                     'manager_name': self._format_manager_name(manager.get('manager', '')),
#                     'orders_count': mtd_orders,
#                     'total_amount': mtd_amount,
#                     'avg_check': mtd_avg_check,
#                     'paid_amount': mtd_paid,
#                     'shipped_amount': mtd_shipped,
#                     'debt_amount': mtd_debt,
#                     'delivery_amount': mtd_delivery,
#                     'delivery_percent': mtd_delivery_percent,
#                     'paid_percent': mtd_paid_percent,
#                     'shipped_percent': mtd_shipped_percent,
#                 }
                
#                 # Подготовка данных для YTD
#                 ytd_orders = self._safe_int(manager.get('ytd_orders', 0))
#                 ytd_amount = self._safe_float(manager.get('ytd_amount', 0))
#                 ytd_paid = self._safe_float(manager.get('ytd_paid', 0))
#                 ytd_shipped = self._safe_float(manager.get('ytd_shipped', 0))
#                 ytd_delivery = self._safe_float(manager.get('ytd_delivery', 0))
                
#                 ytd_avg_check = ytd_amount / ytd_orders if ytd_orders > 0 else 0
#                 ytd_debt = max(0, ytd_amount - ytd_paid)
#                 ytd_delivery_percent = (ytd_delivery / ytd_amount * 100) if ytd_amount > 0 else 0
#                 ytd_paid_percent = (ytd_paid / ytd_amount * 100) if ytd_amount > 0 else 0
#                 ytd_shipped_percent = (ytd_shipped / ytd_amount * 100) if ytd_amount > 0 else 0
                
#                 ytd_data = {
#                     'manager_name': self._format_manager_name(manager.get('manager', '')),
#                     'orders_count': ytd_orders,
#                     'total_amount': ytd_amount,
#                     'avg_check': ytd_avg_check,
#                     'paid_amount': ytd_paid,
#                     'shipped_amount': ytd_shipped,
#                     'debt_amount': ytd_debt,
#                     'delivery_amount': ytd_delivery,
#                     'delivery_percent': ytd_delivery_percent,
#                     'paid_percent': ytd_paid_percent,
#                     'shipped_percent': ytd_shipped_percent,
#                 }
                
#                 # Рисуем профиль с двумя периодами
#                 row = manager_profile.draw_two_periods(
#                     row, 
#                     self._format_manager_name(manager.get('manager', '')),
#                     mtd_data, 
#                     ytd_data
#                 )
                
#                 # Добавляем отступ между менеджерами (кроме последнего)
#                 if idx < len(sorted_managers) - 1:
#                     self.ws.row_dimensions[row].height = 12
#                     row += 1
        
#         else:
#             # Сообщение, если нет данных
#             self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
#             no_data_cell = self.ws.cell(row=row, column=2, value="⚠️ Нет данных по менеджерам за выбранный период")
#             no_data_cell.font = Font(name="Roboto", size=11, color="999999")
#             no_data_cell.alignment = Alignment(horizontal="center", vertical="center")
#             self.ws.row_dimensions[row].height = 40
#             row += 1
        
#         row += 1
        
#         # ============================================================
#         # ПРИМЕЧАНИЕ
#         # ============================================================
#         footnote = Footnote(self.ws)
#         row = footnote.draw(
#             row=row,
#             text="* Показатели рассчитаны без учета полностью отмененных заказов. MTD - с начала месяца, YTD - с начала года."
#         )
#         row += 1
        
#         # ============================================================
#         # ССЫЛКА НА ДЕТАЛИЗАЦИЮ
#         # ============================================================
#         nav_link = NavigationLink(self.ws)
#         row = nav_link.draw(
#             row=row,
#             text="📋 Подробную информацию по заказам смотрите на листе «Детализация заказов» →",
#             target_sheet="2",
#             target_cell="A1",
#             start_col=2,
#             end_col=7,
#             alignment="left",
#             with_icon=False
#         )
        
#         # ============================================================
#         # НАСТРОЙКА КОЛОНОК
#         # ============================================================
#         self.ws.column_dimensions["B"].width = 18
#         self.ws.column_dimensions["C"].width = 18
#         self.ws.column_dimensions["D"].width = 5
#         self.ws.column_dimensions["E"].width = 18
#         self.ws.column_dimensions["F"].width = 18
#         self.ws.column_dimensions["G"].width = 5
        
#         self.ws.sheet_view.showGridLines = False




# orders/reports/sheets/manager_analysis_sheet.py
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .base_sheet import BaseSheet
from ..components import create_header
from ..components.sheet_title import create_sheet_title
from ..styles.helpers import draw_toc_button
from ..components.footnote import Footnote
from ..components.navigation_link import NavigationLink
from ..components.manager_profile import ManagerProfile
from ..styles.theme import COLORS, BORDERS, FILLS


class ManagerAnalysisSheet(BaseSheet):
    """Анализ по менеджерам - портфолио каждого менеджера"""
    
    def __init__(self, workbook, sheet_number, request=None):
        super().__init__(workbook, sheet_number)
        self.sheet_title = create_sheet_title(self.ws)
        self.header = create_header(self.ws)
        self.request = request
    
    def _format_currency(self, value):
        if value is None or value == 0:
            return "0 ₽"
        return f"{int(round(value)):,} ₽".replace(",", " ")
    
    def _format_number(self, value):
        if value is None or value == 0:
            return "0"
        return f"{int(round(value)):,}".replace(",", " ")
    
    def _format_percent(self, value):
        if value is None:
            return "0%"
        return f"{value:.1f}%"
    
    def _safe_float(self, value, default=0.0):
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default
    
    def _safe_int(self, value, default=0):
        if value is None:
            return default
        try:
            return int(round(float(value)))
        except (ValueError, TypeError):
            return default
    
    def _format_manager_name(self, full_name):
        """Форматирует имя менеджера: Иванов Иван -> ИВАНОВ И."""
        if not full_name:
            return ""
        parts = str(full_name).strip().split()
        if len(parts) >= 1:
            last_name = parts[0].upper()
            if len(parts) >= 2:
                first_initial = parts[1][0].upper() if parts[1] else ""
                return f"{last_name} {first_initial}."
            return last_name
        return str(full_name).upper()
    
    def build(self, manager_data):
        """
        Построение листа анализа менеджеров
        
        Args:
            manager_data: список словарей с данными менеджеров
                Каждый словарь должен содержать:
                - manager: имя менеджера
                - mtd_orders: количество заказов за месяц
                - mtd_amount: сумма за месяц
                - mtd_paid: оплачено за месяц
                - mtd_shipped: отгружено за месяц
                - mtd_delivery: доставка за месяц
                - ytd_orders: количество заказов с начала года
                - ytd_amount: сумма с начала года
                - ytd_paid: оплачено с начала года
                - ytd_shipped: отгружено с начала года
                - ytd_delivery: доставка с начала года
                - remaining_to_ship: осталось отгрузить (опционально)
                - top_clients: топ-5 клиентов (опционально)
                - oldest_unpaid_invoice: самый старый счет (опционально)
        """
        
        row = 1
        
        # Шапка
        draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 8
        
        self.ws.column_dimensions["A"].width = 2
        row = self.sheet_title.draw(
            row=3,
            title="АНАЛИЗ ПО МЕНЕДЖЕРАМ",
            subtitle="Портфолио менеджеров: ключевые показатели MTD и YTD",
            date_text=f"Сформировано: {self.ws.parent.properties.created.strftime('%d.%m.%Y') if hasattr(self.ws.parent, 'properties') else ''}",
            start_col=2,
            end_col=7
        )
        row += 1
        
        # ============================================================
        # ВСТУПИТЕЛЬНЫЙ ТЕКСТ
        # ============================================================
        intro_row = row
        self.ws.merge_cells(start_row=intro_row, start_column=2, end_row=intro_row, end_column=7)
        intro_cell = self.ws.cell(
            row=intro_row, 
            column=2, 
            value="Ниже представлены ключевые показатели эффективности менеджеров за текущий месяц (MTD) и с начала года (YTD)"
        )
        intro_cell.font = Font(name="Roboto", size=10, italic=True, color="666666")
        intro_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[intro_row].height = 22
        row += 1
        
        # Пустая строка
        self.ws.row_dimensions[row].height = 6
        row += 1
        
        # ============================================================
        # ПОРТФОЛИО МЕНЕДЖЕРОВ
        # ============================================================
        if manager_data:
            manager_profile = ManagerProfile(self.ws)
            
            # Сортируем менеджеров по сумме MTD (от большего к меньшему)
            sorted_managers = sorted(
                manager_data, 
                key=lambda x: self._safe_float(x.get('mtd_amount', 0)), 
                reverse=True
            )
            
            for idx, manager in enumerate(sorted_managers):
                # Подготовка данных для MTD
                mtd_orders = self._safe_int(manager.get('mtd_orders', 0))
                mtd_amount = self._safe_float(manager.get('mtd_amount', 0))
                mtd_paid = self._safe_float(manager.get('mtd_paid', 0))
                mtd_shipped = self._safe_float(manager.get('mtd_shipped', 0))
                mtd_delivery = self._safe_float(manager.get('mtd_delivery', 0))
                
                mtd_avg_check = mtd_amount / mtd_orders if mtd_orders > 0 else 0
                mtd_debt = max(0, mtd_amount - mtd_paid)
                mtd_delivery_percent = (mtd_delivery / mtd_amount * 100) if mtd_amount > 0 else 0
                mtd_paid_percent = (mtd_paid / mtd_amount * 100) if mtd_amount > 0 else 0
                mtd_shipped_percent = (mtd_shipped / mtd_amount * 100) if mtd_amount > 0 else 0
                
                # НОВЫЕ ПОЛЯ ДЛЯ MTD
                remaining_to_ship = self._safe_float(manager.get('remaining_to_ship', 0))
                # top_clients = manager.get('top_clients', [])
                # oldest_invoice = manager.get('oldest_unpaid_invoice')
                
                mtd_data = {
                    'manager_name': self._format_manager_name(manager.get('manager', '')),
                    'orders_count': mtd_orders,
                    'total_amount': mtd_amount,
                    'avg_check': mtd_avg_check,
                    'paid_amount': mtd_paid,
                    'shipped_amount': mtd_shipped,
                    'debt_amount': mtd_debt,
                    'delivery_amount': mtd_delivery,
                    'delivery_percent': mtd_delivery_percent,
                    'paid_percent': mtd_paid_percent,
                    'shipped_percent': mtd_shipped_percent,
                    'remaining_to_ship': remaining_to_ship,
                    # 'top_clients': top_clients,
                    # 'oldest_unpaid_invoice': oldest_invoice,
                }
                
                # Подготовка данных для YTD
                ytd_orders = self._safe_int(manager.get('ytd_orders', 0))
                ytd_amount = self._safe_float(manager.get('ytd_amount', 0))
                ytd_paid = self._safe_float(manager.get('ytd_paid', 0))
                ytd_shipped = self._safe_float(manager.get('ytd_shipped', 0))
                ytd_delivery = self._safe_float(manager.get('ytd_delivery', 0))
                
                ytd_avg_check = ytd_amount / ytd_orders if ytd_orders > 0 else 0
                ytd_debt = max(0, ytd_amount - ytd_paid)
                ytd_delivery_percent = (ytd_delivery / ytd_amount * 100) if ytd_amount > 0 else 0
                ytd_paid_percent = (ytd_paid / ytd_amount * 100) if ytd_amount > 0 else 0
                ytd_shipped_percent = (ytd_shipped / ytd_amount * 100) if ytd_amount > 0 else 0
                
                ytd_data = {
                    'manager_name': self._format_manager_name(manager.get('manager', '')),
                    'orders_count': ytd_orders,
                    'total_amount': ytd_amount,
                    'avg_check': ytd_avg_check,
                    'paid_amount': ytd_paid,
                    'shipped_amount': ytd_shipped,
                    'debt_amount': ytd_debt,
                    'delivery_amount': ytd_delivery,
                    'delivery_percent': ytd_delivery_percent,
                    'paid_percent': ytd_paid_percent,
                    'shipped_percent': ytd_shipped_percent,
                }
                
                # Рисуем профиль с двумя периодами
                row = manager_profile.draw_two_periods(
                    row, 
                    self._format_manager_name(manager.get('manager', '')),
                    mtd_data, 
                    ytd_data
                )
                
                # Добавляем отступ между менеджерами (кроме последнего)
                if idx < len(sorted_managers) - 1:
                    self.ws.row_dimensions[row].height = 12
                    row += 1
        
        else:
            # Сообщение, если нет данных
            self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
            no_data_cell = self.ws.cell(row=row, column=2, value="⚠️ Нет данных по менеджерам за выбранный период")
            no_data_cell.font = Font(name="Roboto", size=11, color="999999")
            no_data_cell.alignment = Alignment(horizontal="center", vertical="center")
            self.ws.row_dimensions[row].height = 40
            row += 1
        
        row += 1
        
        # ============================================================
        # ПРИМЕЧАНИЕ
        # ============================================================
        footnote = Footnote(self.ws)
        row = footnote.draw(
            row=row,
            text="* Показатели рассчитаны без учета полностью отмененных заказов. MTD - с начала месяца, YTD - с начала года. Топ-5 клиентов и информация о старых счетах отображаются по активным заказам."
        )
        row += 1
        
        # ============================================================
        # ССЫЛКА НА ДЕТАЛИЗАЦИЮ
        # ============================================================
        nav_link = NavigationLink(self.ws)
        row = nav_link.draw(
            row=row,
            text="📋 Подробную информацию по заказам смотрите на листе «Детализация заказов» →",
            target_sheet="2",
            target_cell="A1",
            start_col=2,
            end_col=7,
            alignment="left",
            with_icon=False
        )
        
        # ============================================================
        # НАСТРОЙКА КОЛОНОК
        # ============================================================
        self.ws.column_dimensions["B"].width = 18
        self.ws.column_dimensions["C"].width = 18
        self.ws.column_dimensions["D"].width = 5
        self.ws.column_dimensions["E"].width = 18
        self.ws.column_dimensions["F"].width = 18
        self.ws.column_dimensions["G"].width = 5
        
        self.ws.sheet_view.showGridLines = False