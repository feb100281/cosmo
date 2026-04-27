# # orders/reports/sheets/completed_orders_wrong_status_sheet.py
# from datetime import datetime
# from openpyxl.styles import Font, PatternFill, Border, Side
# from .base_sheet import BaseSheet
# from ..styles.theme import COLORS, BORDERS, FILLS, FONTS, ALIGNMENTS
# from ..styles.helpers import draw_toc_button
# from ..components.footnote import Footnote
# from ..components.sheet_title import create_sheet_title
# from ..components.kpi_cards import KPICards


# class CompletedOrdersWrongStatusSheet(BaseSheet):
#     """Лист с полностью выполненными заказами, у которых статус не Закрыт"""

#     def __init__(self, workbook, sheet_number):
#         super().__init__(workbook, sheet_number)
#         self.sheet_title = create_sheet_title(self.ws)
#         self.kpi_cards = KPICards(self.ws)

#     def _format_currency(self, value):
#         if value is None or value == 0:
#             return "0 ₽"
#         return f"{int(round(abs(value))):,} ₽".replace(",", " ")

#     def _format_number(self, value):
#         if value is None or value == 0:
#             return "0"
#         return f"{int(round(abs(value))):,}".replace(",", " ")

#     def _safe_float(self, value, default=0.0):
#         if value is None:
#             return default
#         try:
#             return float(value)
#         except (ValueError, TypeError):
#             return default

#     def _format_manager_name(self, full_name):
#         if not full_name:
#             return ""
#         parts = str(full_name).strip().split()
#         if len(parts) >= 1:
#             last_name = parts[0].upper()
#             if len(parts) >= 2:
#                 first_initial = parts[1][0].upper() if parts[1] else ""
#                 return f"{last_name} {first_initial}."
#             return last_name
#         return str(full_name).upper()

#     def _format_client_name(self, name):
#         if not name:
#             return ""
#         return str(name).upper()

#     def build(self, orders_data, summary_data):
#         """Построение листа"""
#         row = 1

#         # Кнопка назад
#         draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

#         # Заголовок
#         row = self.sheet_title.draw(
#             row=3,
#             title="ЗАКАЗЫ ГОТОВЫЕ К ЗАКРЫТИЮ",
#             subtitle=f"Заказы полностью отгружены и оплачены (отгрузка = сумма заказа = оплата), но статус не Закрыт | Найдено: {summary_data['total_orders']} заказов на сумму {self._format_currency(summary_data['total_amount'])}",
#             date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y')}",
#             start_col=2,
#             end_col=11
#         )

#         row += 2

#         # KPI карточки
#         row1_cards = [
#             {'title': 'ЗАКАЗОВ К ЗАКРЫТИЮ', 'value': self._format_number(summary_data['total_orders']), 'color': COLORS["warning_red"], 'width': 2},
#             {'title': 'ОБЩАЯ СУММА', 'value': self._format_currency(summary_data['total_amount']), 'color': COLORS["warning_red"], 'width': 2},
#             {'title': 'КЛИЕНТОВ', 'value': self._format_number(summary_data['unique_clients']), 'color': COLORS["dark_green"], 'width': 3},
#             {'title': 'МЕНЕДЖЕРОВ', 'value': self._format_number(summary_data['unique_managers']), 'color': COLORS["dark_green"], 'width': 3},
#         ]
        
#         row = self.kpi_cards.draw_row(row, row1_cards)
#         row += 1

#         # Заголовок таблицы
#         headers = [
#             "ЗАКАЗ",
#             "ДАТА\nСОЗДАНИЯ",
#             "СТАТУС",
#             "КЛИЕНТ",
#             "МЕНЕДЖЕР",
#             "МАГАЗИН",
#             "СУММА\nЗАКАЗА",
#             "ОТГРУЖЕНО",
#             "ОПЛАЧЕНО",
#             "ДНЕЙ В\nСТАТУСЕ",
#         ]

#         table_start_row = row

#         # Рисуем заголовки
#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = FONTS["header_white"]
#             cell.fill = FILLS["header"]
#             cell.alignment = ALIGNMENTS["center_wrap"]
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 40
#         row += 1

#         # Данные
#         for idx, order in enumerate(orders_data):
#             # Чередование фона
#             if idx % 2 == 0:
#                 row_fill = FILLS.get("alt", PatternFill(fill_type=None))
#             else:
#                 row_fill = PatternFill(fill_type=None)

#             values = [
#                 order.get('order_name', ''),
#                 order.get('date_from', ''),
#                 order.get('status', ''),
#                 self._format_client_name(order.get('client', '')),
#                 self._format_manager_name(order.get('manager', '')),
#                 (order.get('store', '') or '').upper()[:30],
#                 self._safe_float(order.get('amount_active')),
#                 self._safe_float(order.get('total_shiped_amount')),
#                 self._safe_float(order.get('cash_pmts')),
#                 f"{order.get('days_since_update', 0)}",
#             ]

#             for col_idx, value in enumerate(values):
#                 col_num = col_idx + 2
#                 cell = self.ws.cell(row=row, column=col_num, value=value)
#                 cell.fill = row_fill
#                 cell.border = BORDERS["thin"]
#                 cell.font = FONTS["normal"]

#                 # Настройка выравнивания
#                 if col_idx in [0, 3, 4, 5]:  # ЗАКАЗ, КЛИЕНТ, МЕНЕДЖЕР, МАГАЗИН
#                     cell.alignment = ALIGNMENTS["left"]
#                 elif col_idx in [1, 2, 9]:  # ДАТА, СТАТУС, ДНЕЙ
#                     cell.alignment = ALIGNMENTS["center"]
#                     if col_idx == 9 and order.get('days_since_update', 0) > 30:
#                         cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["warning_red"])
#                 elif col_idx in [6, 7, 8]:  # Суммовые колонки
#                     cell.alignment = ALIGNMENTS["right"]
#                     if isinstance(value, (int, float)):
#                         cell.number_format = '#,##0.00'

#             self.ws.row_dimensions[row].height = 25
#             row += 1

#         # ИТОГОВАЯ СТРОКА
#         if orders_data:
#             total_amount = summary_data['total_amount']

#             total_fill = PatternFill("solid", fgColor=COLORS["total_green"])
#             thin_border = Border(
#                 left=Side(style='thin'),
#                 right=Side(style='thin'),
#                 top=Side(style='thin'),
#                 bottom=Side(style='thin')
#             )

#             # Пишем "ИТОГО"
#             cell = self.ws.cell(row=row, column=2, value="ИТОГО:")
#             cell.fill = total_fill
#             cell.border = thin_border
#             cell.font = FONTS["bold"]
#             cell.alignment = ALIGNMENTS["center"]

#             # Сумма заказа
#             cell = self.ws.cell(row=row, column=7, value=total_amount)
#             cell.fill = total_fill
#             cell.border = thin_border
#             cell.font = FONTS["bold"]
#             cell.alignment = ALIGNMENTS["right"]
#             cell.number_format = '#,##0.00'

#             # Отгружено
#             cell = self.ws.cell(row=row, column=8, value=total_amount)
#             cell.fill = total_fill
#             cell.border = thin_border
#             cell.font = FONTS["bold"]
#             cell.alignment = ALIGNMENTS["right"]
#             cell.number_format = '#,##0.00'

#             # Оплачено
#             cell = self.ws.cell(row=row, column=9, value=total_amount)
#             cell.fill = total_fill
#             cell.border = thin_border
#             cell.font = FONTS["bold"]
#             cell.alignment = ALIGNMENTS["right"]
#             cell.number_format = '#,##0.00'

#             # Заполняем остальные колонки
#             for col in [3, 4, 5, 6, 10, 11]:
#                 cell = self.ws.cell(row=row, column=col, value="")
#                 cell.fill = total_fill
#                 cell.border = thin_border
#                 cell.font = FONTS["bold"]
#                 cell.alignment = ALIGNMENTS["center"]

#             row += 2

#         # Примечания
#         footnote = Footnote(self.ws)
#         footnote.draw(row, text="⚠️ Эти заказы полностью отгружены и оплачены, но не закрыты в системе")
#         row += 1
#         footnote.draw(row, text="💡 Рекомендуется сменить статус на 'Закрыт'")
#         row += 1
#         footnote.draw(row, text="📦 Дни в статусе считаются от даты последнего изменения заказа (update_at)")

#         # Настройка колонок
#         col_widths = {
#             "B": 35,  # ЗАКАЗ
#             "C": 14,  # ДАТА СОЗДАНИЯ
#             "D": 25,  # СТАТУС
#             "E": 30,  # КЛИЕНТ
#             "F": 20,  # МЕНЕДЖЕР
#             "G": 25,  # МАГАЗИН
#             "H": 16,  # СУММА ЗАКАЗА
#             "I": 16,  # ОТГРУЖЕНО
#             "J": 16,  # ОПЛАЧЕНО
#             "K": 16,  # ДНЕЙ В СТАТУСЕ
#         }
#         for col, width in col_widths.items():
#             self.ws.column_dimensions[col].width = width

#         self.ws.freeze_panes = f'G{table_start_row + 1}'
#         if orders_data:
#             self.ws.auto_filter.ref = f'B{table_start_row}:K{row - 5}'
#         self.ws.sheet_view.showGridLines = False




# orders/reports/sheets/completed_orders_wrong_status_sheet.py
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Border, Side
from .base_sheet import BaseSheet
from ..styles.theme import COLORS, BORDERS, FILLS, FONTS, ALIGNMENTS
from ..styles.helpers import draw_toc_button
from ..components.footnote import Footnote
from ..components.sheet_title import create_sheet_title
from ..components.kpi_cards import KPICards


class CompletedOrdersWrongStatusSheet(BaseSheet):
    """Лист с полностью выполненными заказами, у которых статус не Закрыт"""

    def __init__(self, workbook, sheet_number):
        super().__init__(workbook, sheet_number)
        self.sheet_title = create_sheet_title(self.ws)
        self.kpi_cards = KPICards(self.ws)

    def _format_currency(self, value):
        if value is None or value == 0:
            return "0 ₽"
        return f"{int(round(abs(value))):,} ₽".replace(",", " ")

    def _format_number(self, value):
        if value is None or value == 0:
            return "0"
        return f"{int(round(abs(value))):,}".replace(",", " ")

    def _safe_float(self, value, default=0.0):
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default

    def _format_manager_name(self, full_name):
        if not full_name:
            return ""
        parts = str(full_name).strip().split()
        if len(parts) >= 1:
            last_name = parts[0].upper()
            if len(parts) >= 2:
                first_initial = parts[1][0].upper() if parts[1] else ""
                return f"{last_name} {first_initial}."
            return last_name
        return str(full_name).upper()

    def _format_client_name(self, name):
        if not name:
            return ""
        return str(name).upper()

    def _format_datetime(self, date_value):
        """Форматирует дату без времени: YYYY-MM-DD"""
        if not date_value:
            return ""
        if hasattr(date_value, 'strftime'):
            return date_value.strftime('%Y-%m-%d')
        if isinstance(date_value, str):
            # Пробуем распарсить строку с временем
            for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d.%m.%Y']:
                try:
                    dt = datetime.strptime(date_value, fmt)
                    return dt.strftime('%Y-%m-%d')
                except ValueError:
                    continue
        return str(date_value)[:10] if len(str(date_value)) > 10 else str(date_value)

    def build(self, orders_data, summary_data):
        """Построение листа"""
        row = 1

        # Кнопка назад
        draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 8

        # Заголовок
        row = self.sheet_title.draw(
            row=3,
            title="ЗАКАЗЫ ГОТОВЫЕ К ЗАКРЫТИЮ",
            subtitle=f"Заказы полностью отгружены и оплачены (отгрузка = сумма заказа = оплата), но статус не Закрыт | Найдено: {summary_data['total_orders']} заказов на сумму {self._format_currency(summary_data['total_amount'])}",
            date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            start_col=2,
            end_col=11
        )

        row += 1

        # KPI карточки
        row1_cards = [
            {'title': 'ЗАКАЗОВ К ЗАКРЫТИЮ', 'value': self._format_number(summary_data['total_orders']), 'color': COLORS["warning_red"], 'width': 2},
            {'title': 'ОБЩАЯ СУММА', 'value': self._format_currency(summary_data['total_amount']), 'color': COLORS["warning_red"], 'width': 2},
            {'title': 'КЛИЕНТОВ', 'value': self._format_number(summary_data['unique_clients']), 'color': COLORS["dark_green"], 'width': 3},
            {'title': 'МЕНЕДЖЕРОВ', 'value': self._format_number(summary_data['unique_managers']), 'color': COLORS["dark_green"], 'width': 3},
        ]
        
        row = self.kpi_cards.draw_row(row, row1_cards)
        row += 2

        # Заголовок таблицы
        headers = [
            "ЗАКАЗ",
            "ДАТА\nСОЗДАНИЯ",
            "СТАТУС",
            "КЛИЕНТ",
            "МЕНЕДЖЕР",
            "МАГАЗИН",
            "СУММА\nЗАКАЗА",
            "ОТГРУЖЕНО",
            "ОПЛАЧЕНО",
            "ДНЕЙ В\nСТАТУСЕ",
        ]

        table_start_row = row

        # Рисуем заголовки
        for col_idx, header in enumerate(headers, start=2):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
            cell.fill = FILLS["header_blue"] if 'header_blue' in FILLS else PatternFill("solid", fgColor=COLORS["dark_green"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDERS["thin"]

        self.ws.row_dimensions[row].height = 40
        row += 1

        # Данные
        for idx, order in enumerate(orders_data):
            # Чередование фона
            if idx % 2 == 0:
                fill = FILLS.get("odd_row", PatternFill(fill_type=None))
            else:
                fill = FILLS.get("even_row", PatternFill(fill_type=None))

            values = [
                order.get('order_name', ''),
                self._format_datetime(order.get('date_from')),  # ДАТА СОЗДАНИЯ
                order.get('status', ''),
                self._format_client_name(order.get('client', '')),
                self._format_manager_name(order.get('manager', '')),
                (order.get('store', '') or '').upper()[:30],
                self._safe_float(order.get('amount_active')),
                self._safe_float(order.get('total_shiped_amount')),
                self._safe_float(order.get('cash_pmts')),
                order.get('days_since_update', 0),
            ]

            for col_idx, value in enumerate(values):
                col_num = col_idx + 2
                cell = self.ws.cell(row=row, column=col_num, value=value)
                cell.fill = fill
                cell.border = BORDERS["thin"]
                # ЯВНО ЗАДАЁМ ЧЁРНЫЙ ЦВЕТ ШРИФТА (как в orders_detail_sheet)
                cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))

                # Настройка выравнивания
                if col_idx in [0, 3, 4, 5]:  # ЗАКАЗ, КЛИЕНТ, МЕНЕДЖЕР, МАГАЗИН
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif col_idx in [1, 2, 9]:  # ДАТА, СТАТУС, ДНЕЙ
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if col_idx == 9 and value > 30:
                        cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["warning_red"])
                elif col_idx in [6, 7, 8]:  # Суммовые колонки
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0'

                # Специальная обработка статуса
                if col_idx == 2:  # СТАТУС
                    status = value
                    if status and "Закрыт" not in status:
                        cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS['dark_green'])

            self.ws.row_dimensions[row].height = 25
            row += 1

        # ИТОГОВАЯ СТРОКА
        if orders_data:
            total_amount = summary_data['total_amount']

            fill = FILLS["total"]
            thin_border = BORDERS["thin"]

            # Пишем "ИТОГО"
            cell = self.ws.cell(row=row, column=2, value="ИТОГО:")
            cell.fill = fill
            cell.border = thin_border
            cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
            cell.alignment = Alignment(horizontal="right", vertical="center")

            # Заполняем пустые колонки
            for col in [3, 4, 5, 6]:
                cell = self.ws.cell(row=row, column=col, value="")
                cell.fill = fill
                cell.border = thin_border

            # Сумма заказа
            cell = self.ws.cell(row=row, column=7, value=total_amount)
            cell.fill = fill
            cell.border = thin_border
            cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0'

            # Отгружено
            cell = self.ws.cell(row=row, column=8, value=total_amount)
            cell.fill = fill
            cell.border = thin_border
            cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0'

            # Оплачено
            cell = self.ws.cell(row=row, column=9, value=total_amount)
            cell.fill = fill
            cell.border = thin_border
            cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0'

            # Дней в статусе - пусто
            cell = self.ws.cell(row=row, column=10, value="")
            cell.fill = fill
            cell.border = thin_border

            # Магазин - пусто (колонка 11 у нас нет, но если есть - заполняем)
            if self.ws.max_column >= 11:
                cell = self.ws.cell(row=row, column=11, value="")
                cell.fill = fill
                cell.border = thin_border

            self.ws.row_dimensions[row].height = 30
            row += 2

        # Примечания
        footnote = Footnote(self.ws)
        footnote.draw(row, text="⚠️ Эти заказы полностью отгружены и оплачены, но не закрыты в системе")
        row += 1
        footnote.draw(row, text="💡 Рекомендуется сменить статус на 'Закрыт'")
        row += 1
        footnote.draw(row, text="📦 Дни в статусе считаются от даты последнего изменения заказа (update_at)")

        # Настройка колонок
        col_widths = {
            "B": 30,  # ЗАКАЗ
            "C": 14,  # ДАТА СОЗДАНИЯ
            "D": 25,  # СТАТУС
            "E": 30,  # КЛИЕНТ
            "F": 20,  # МЕНЕДЖЕР
            "G": 25,  # МАГАЗИН
            "H": 16,  # СУММА ЗАКАЗА
            "I": 16,  # ОТГРУЖЕНО
            "J": 16,  # ОПЛАЧЕНО
            "K": 16,  # ДНЕЙ В СТАТУСЕ
        }
        for col, width in col_widths.items():
            self.ws.column_dimensions[col].width = width

        self.ws.freeze_panes = f'G{table_start_row + 1}'
        if orders_data:
            self.ws.auto_filter.ref = f'B{table_start_row}:K{row - 5}'
        self.ws.sheet_view.showGridLines = False


# Добавляем импорт Alignment если нужен
from openpyxl.styles import Alignment