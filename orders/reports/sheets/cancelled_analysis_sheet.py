# # orders/reports/sheets/cancelled_analysis_sheet.py
# from datetime import datetime
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from .base_sheet import BaseSheet
# from ..styles.theme import COLORS, BORDERS, FILLS
# from ..styles.helpers import draw_toc_button
# from ..components.footnote import Footnote
# from ..components.sheet_title import create_sheet_title
# from ..components.kpi_cards import KPICards


# class CancelledAnalysisSheet(BaseSheet):
#     """Лист с анализом отменённых заказов"""

#     def __init__(self, workbook, sheet_number):
#         super().__init__(workbook, sheet_number)
#         self.sheet_title = create_sheet_title(self.ws)
#         self.kpi_cards = KPICards(self.ws)

#     def _format_currency(self, value):
#         if value is None or value == 0:
#             return "0 ₽"
#         return f"{int(round(abs(value))):,} ₽".replace(",", " ")

#     def _format_number(self, value):
#         if value is None or value == 0:
#             return "0"
#         return f"{int(round(abs(value))):,}".replace(",", " ")

#     def _safe_float(self, value, default=0.0):
#         if value is None:
#             return default
#         try:
#             return float(value)
#         except (ValueError, TypeError):
#             return default

#     def _format_manager_name(self, full_name):
#         if not full_name:
#             return ""
#         parts = str(full_name).strip().split()
#         if len(parts) >= 1:
#             last_name = parts[0].upper()
#             if len(parts) >= 2:
#                 first_initial = parts[1][0].upper() if parts[1] else ""
#                 return f"{last_name} {first_initial}."
#             return last_name
#         return str(full_name).upper()

#     def _format_client_name(self, name):
#         if not name:
#             return ""
#         return str(name).upper()

#     def _format_datetime(self, date_value):
#         """Форматирует дату без времени: YYYY-MM-DD"""
#         if not date_value:
#             return ""
#         if hasattr(date_value, 'strftime'):
#             return date_value.strftime('%Y-%m-%d')
#         if isinstance(date_value, str):
#             for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d.%m.%Y']:
#                 try:
#                     dt = datetime.strptime(date_value, fmt)
#                     return dt.strftime('%Y-%m-%d')
#                 except ValueError:
#                     continue
#         return str(date_value)[:10] if len(str(date_value)) > 10 else str(date_value)

#     def draw_reason_breakdown(self, start_row, reasons_data):
#         """Рисует breakdown по причинам отмен"""
        
#         if not reasons_data:
#             return start_row
        
#         # Заголовок
#         row = start_row
#         cell = self.ws.cell(row=row, column=2, value="ОТМЕНЫ ПО ПРИЧИНАМ")
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
#         row += 1
        
#         # Заголовки таблицы причин
#         reason_headers = ["ПРИЧИНА ОТМЕНЫ", "КОЛ-ВО ЗАКАЗОВ", "СУММА ОТМЕН", "СРЕДНЯЯ СУММА"]
#         for col_idx, header in enumerate(reason_headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
#             cell.fill = FILLS.get("header_blue", PatternFill("solid", fgColor=COLORS["dark_green"]))
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]
        
#         self.ws.row_dimensions[row].height = 30
#         row += 1
        
#         # Данные по причинам
#         for idx, reason in enumerate(reasons_data):
#             fill = FILLS.get("odd_row", PatternFill(fill_type=None)) if idx % 2 == 0 else FILLS.get("even_row", PatternFill(fill_type=None))
            
#             values = [
#                 reason.get('reason', ''),
#                 reason.get('count', 0),
#                 reason.get('amount', 0),
#                 reason.get('avg_amount', 0),
#             ]
            
#             for col_idx, value in enumerate(values, start=2):
#                 cell = self.ws.cell(row=row, column=col_idx, value=value)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]
#                 cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))
                
#                 if col_idx == 2:  # Причина
#                     cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
#                 elif col_idx == 3:  # Кол-во
#                     cell.alignment = Alignment(horizontal="center", vertical="center")
#                 else:  # Суммы
#                     cell.alignment = Alignment(horizontal="right", vertical="center")
#                     if isinstance(value, (int, float)):
#                         cell.number_format = '#,##0'
            
#             self.ws.row_dimensions[row].height = 25
#             row += 1
        
#         row += 1
#         return row

#     def draw_manager_breakdown(self, start_row, managers_data):
#         """Рисует breakdown по менеджерам"""
        
#         if not managers_data:
#             return start_row
        
#         # Заголовок
#         row = start_row
#         cell = self.ws.cell(row=row, column=2, value="ОТМЕНЫ ПО МЕНЕДЖЕРАМ")
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
#         row += 1
        
#         # Заголовки таблицы
#         manager_headers = ["МЕНЕДЖЕР", "КОЛ-ВО ОТМЕН", "СУММА ОТМЕН", "СРЕДНЯЯ СУММА"]
#         for col_idx, header in enumerate(manager_headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
#             cell.fill = FILLS.get("header_blue", PatternFill("solid", fgColor=COLORS["dark_green"]))
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]
        
#         self.ws.row_dimensions[row].height = 30
#         row += 1
        
#         # Данные по менеджерам
#         for idx, manager in enumerate(managers_data[:10]):  # Топ 10 менеджеров
#             fill = FILLS.get("odd_row", PatternFill(fill_type=None)) if idx % 2 == 0 else FILLS.get("even_row", PatternFill(fill_type=None))
            
#             values = [
#                 self._format_manager_name(manager.get('manager', '')),
#                 manager.get('orders_count', 0),
#                 manager.get('total_amount', 0),
#                 manager.get('avg_amount', 0),
#             ]
            
#             for col_idx, value in enumerate(values, start=2):
#                 cell = self.ws.cell(row=row, column=col_idx, value=value)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]
#                 cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))
                
#                 if col_idx == 2:
#                     cell.alignment = Alignment(horizontal="left", vertical="center")
#                 elif col_idx == 3:
#                     cell.alignment = Alignment(horizontal="center", vertical="center")
#                 else:
#                     cell.alignment = Alignment(horizontal="right", vertical="center")
#                     if isinstance(value, (int, float)):
#                         cell.number_format = '#,##0'
            
#             self.ws.row_dimensions[row].height = 25
#             row += 1
        
#         row += 1
#         return row

#     def build(self, orders_data, summary_data, reasons_data, managers_data, top_orders):
#         """Построение листа с анализом отмен"""
#         row = 1

#         # Кнопка назад
#         draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

#         self.ws.row_dimensions[1].height = 20
#         self.ws.row_dimensions[2].height = 8

#         # Заголовок
#         period_text = ""
#         if summary_data.get('period_start') and summary_data.get('period_end'):
#             period_text = f"Период: {summary_data['period_start']} — {summary_data['period_end']}"
        
#         row = self.sheet_title.draw(
#             row=3,
#             title="АНАЛИЗ ОТМЕНЁННЫХ ЗАКАЗОВ",
#             subtitle=f"Детальный анализ всех отменённых заказов | Всего отмен: {summary_data['total_orders']} на сумму {self._format_currency(summary_data['total_cancelled_amount'])}",
#             date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')} | {period_text}",
#             start_col=2,
#             end_col=14
#         )

#         row += 1

#         # KPI карточки
#         row1_cards = [
#             {'title': 'ОТМЕНЁННЫХ ЗАКАЗОВ', 'value': self._format_number(summary_data['total_orders']), 'color': COLORS["warning_red"], 'width': 2},
#             {'title': 'СУММА ОТМЕН', 'value': self._format_currency(summary_data['total_cancelled_amount']), 'color': COLORS["warning_red"], 'width': 2},
#         ]
        
#         row = self.kpi_cards.draw_row(row, row1_cards)
#         row += 1
        
#         row2_cards = [
#             {'title': 'УНИКАЛЬНЫХ КЛИЕНТОВ', 'value': self._format_number(summary_data['unique_clients']), 'color': COLORS["blue"], 'width': 2},
#             {'title': 'МЕНЕДЖЕРОВ', 'value': self._format_number(summary_data['unique_managers']), 'color': COLORS["blue"], 'width': 2},
#         ]
        
#         row = self.kpi_cards.draw_row(row, row2_cards)
#         row += 2

#         # ============================================================
#         # АНАЛИТИКА ПО ПРИЧИНАМ И МЕНЕДЖЕРАМ
#         # ============================================================
        
#         # Breakdown по причинам
#         row = self.draw_reason_breakdown(row, reasons_data)
        
#         # Breakdown по менеджерам
#         row = self.draw_manager_breakdown(row, managers_data)

#         # ============================================================
#         # ЗАГОЛОВОК ТАБЛИЦЫ
#         # ============================================================
#         cell = self.ws.cell(row=row, column=2, value="ДЕТАЛЬНАЯ ИНФОРМАЦИЯ ПО ОТМЕНЁННЫМ ЗАКАЗАМ")
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)
#         row += 2

#         # Заголовки таблицы
#         headers = [
#             "ЗАКАЗ",
#             "ДАТА\nСОЗДАНИЯ",
#             "КЛИЕНТ",
#             "МЕНЕДЖЕР",
#             "МАГАЗИН",
#             "КОЛ-ВО\nТОВАРОВ",
#             "СУММА\nЗАКАЗА",
#             "ОПЛАЧЕНО",
#             "СУММА\nОТМЕНЫ",
#             "СТАТУС",
#             "ПРИЧИНА ОТМЕНЫ",
#         ]

#         table_start_row = row

#         # Рисуем заголовки
#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
#             cell.fill = FILLS.get("header_blue", PatternFill("solid", fgColor=COLORS["dark_green"]))
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 40
#         row += 1

#         # Данные (только топ отмен)
#         for idx, order in enumerate(top_orders):
#             # Чередование фона
#             if idx % 2 == 0:
#                 fill = FILLS.get("odd_row", PatternFill(fill_type=None))
#             else:
#                 fill = FILLS.get("even_row", PatternFill(fill_type=None))

#             values = [
#                 order.get('order_name', '') or order.get('number', ''),
#                 self._format_datetime(order.get('date_from')),
#                 self._format_client_name(order.get('client', '')),
#                 self._format_manager_name(order.get('manager', '')),
#                 (order.get('store', '') or '').upper()[:30],
#                 self._safe_float(order.get('order_qty', 0)),
#                 self._safe_float(order.get('cancelled_amount', 0)) + self._safe_float(order.get('paid_amount', 0)),  # Полная сумма заказа
#                 self._safe_float(order.get('paid_amount', 0)),
#                 self._safe_float(order.get('cancelled_amount', 0)),
#                 order.get('status', ''),
#                 order.get('cancellation_reason', 'Не указана'),
#             ]

#             for col_idx, value in enumerate(values):
#                 col_num = col_idx + 2
#                 cell = self.ws.cell(row=row, column=col_num, value=value)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]
#                 cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))

#                 # Настройка выравнивания
#                 if col_idx in [0, 2, 3, 4, 10]:  # ЗАКАЗ, КЛИЕНТ, МЕНЕДЖЕР, МАГАЗИН, ПРИЧИНА
#                     cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
#                 elif col_idx in [1, 5, 9]:  # ДАТА, КОЛ-ВО, СТАТУС
#                     cell.alignment = Alignment(horizontal="center", vertical="center")
#                 else:  # Числовые колонки
#                     cell.alignment = Alignment(horizontal="right", vertical="center")
#                     if isinstance(value, (int, float)):
#                         cell.number_format = '#,##0'

#                 # Подсветка крупных отмен
#                 if col_idx == 8 and value > 1000000:  # СУММА ОТМЕНЫ > 1 млн
#                     cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["warning_red"])

#             self.ws.row_dimensions[row].height = 25
#             row += 1

#         # ИТОГОВАЯ СТРОКА
#         if orders_data:
#             total_fill = FILLS["total"]
            
#             # Пишем "ИТОГО:"
#             cell = self.ws.cell(row=row, column=2, value="ИТОГО:")
#             cell.fill = total_fill
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
#             cell.alignment = Alignment(horizontal="right", vertical="center")
            
#             # Количество товаров
#             cell = self.ws.cell(row=row, column=6, value=summary_data['total_qty'])
#             cell.fill = total_fill
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
#             cell.alignment = Alignment(horizontal="center", vertical="center")
#             cell.number_format = '#,##0'
            
#             # Сумма заказа (пропускаем)
#             cell = self.ws.cell(row=row, column=7, value="")
#             cell.fill = total_fill
#             cell.border = BORDERS["thin"]
            
#             # Оплачено
#             cell = self.ws.cell(row=row, column=8, value=summary_data['total_paid_amount'])
#             cell.fill = total_fill
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
#             cell.alignment = Alignment(horizontal="right", vertical="center")
#             cell.number_format = '#,##0'
            
#             # Сумма отмен
#             cell = self.ws.cell(row=row, column=9, value=summary_data['total_cancelled_amount'])
#             cell.fill = total_fill
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
#             cell.alignment = Alignment(horizontal="right", vertical="center")
#             cell.number_format = '#,##0'
            
#             # Заполняем остальные колонки
#             for col in [3, 4, 5, 10, 11, 12]:
#                 cell = self.ws.cell(row=row, column=col, value="")
#                 cell.fill = total_fill
#                 cell.border = BORDERS["thin"]

#             self.ws.row_dimensions[row].height = 30
#             row += 2

#         # Примечания
#         footnote = Footnote(self.ws)
#         footnote.draw(row, text="В таблице показаны TOP-50 отмен по сумме (остальные отмены не отображаются для читаемости)")
#         row += 1
#         footnote.draw(row, text="Сумма заказа = Сумма отмены + Оплачено (если были частичные оплаты)")
#         row += 1
#         footnote.draw(row, text="Аналитика по причинам и менеджерам включает ВСЕ отмены за период")

#         # Настройка колонок
#         col_widths = {
#             "B": 35,  # ЗАКАЗ
#             "C": 14,  # ДАТА СОЗДАНИЯ
#             "D": 30,  # КЛИЕНТ
#             "E": 22,  # МЕНЕДЖЕР
#             "F": 22,  # МАГАЗИН
#             "G": 12,  # КОЛ-ВО ТОВАРОВ
#             "H": 16,  # СУММА ЗАКАЗА
#             "I": 16,  # ОПЛАЧЕНО
#             "J": 16,  # СУММА ОТМЕНЫ
#             "K": 25,  # СТАТУС
#             "L": 35,  # ПРИЧИНА ОТМЕНЫ
#         }
#         for col, width in col_widths.items():
#             self.ws.column_dimensions[col].width = width


        
#         # Автофильтр
#         if orders_data:
#             self.ws.auto_filter.ref = f'B{table_start_row}:L{row - 5}'
        
#         self.ws.sheet_view.showGridLines = False




# # orders/reports/sheets/cancelled_analysis_sheet.py
# from datetime import datetime
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from .base_sheet import BaseSheet
# from ..styles.theme import COLORS, BORDERS, FILLS
# from ..styles.helpers import draw_toc_button
# from ..components.footnote import Footnote
# from ..components.sheet_title import create_sheet_title
# from ..components.kpi_cards import KPICards


# class CancelledAnalysisSheet(BaseSheet):
#     """Лист с анализом отменённых заказов"""

#     def __init__(self, workbook, sheet_number):
#         super().__init__(workbook, sheet_number)
#         self.sheet_title = create_sheet_title(self.ws)
#         self.kpi_cards = KPICards(self.ws)

#     def _format_currency(self, value):
#         if value is None or value == 0:
#             return "0 ₽"
#         return f"{int(round(abs(value))):,} ₽".replace(",", " ")

#     def _format_number(self, value):
#         if value is None or value == 0:
#             return "0"
#         return f"{int(round(abs(value))):,}".replace(",", " ")

#     def _safe_float(self, value, default=0.0):
#         if value is None:
#             return default
#         try:
#             return float(value)
#         except (ValueError, TypeError):
#             return default

#     def _format_manager_name(self, full_name):
#         if not full_name:
#             return ""
#         parts = str(full_name).strip().split()
#         if len(parts) >= 1:
#             last_name = parts[0].upper()
#             if len(parts) >= 2:
#                 first_initial = parts[1][0].upper() if parts[1] else ""
#                 return f"{last_name} {first_initial}."
#             return last_name
#         return str(full_name).upper()

#     def _format_client_name(self, name):
#         if not name:
#             return ""
#         return str(name).upper()

#     def _format_datetime(self, date_value):
#         """Форматирует дату без времени: YYYY-MM-DD"""
#         if not date_value:
#             return ""
#         if hasattr(date_value, 'strftime'):
#             return date_value.strftime('%Y-%m-%d')
#         if isinstance(date_value, str):
#             for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d.%m.%Y']:
#                 try:
#                     dt = datetime.strptime(date_value, fmt)
#                     return dt.strftime('%Y-%m-%d')
#                 except ValueError:
#                     continue
#         return str(date_value)[:10] if len(str(date_value)) > 10 else str(date_value)

#     def draw_reason_breakdown(self, start_row, reasons_data):
#         """Рисует breakdown по причинам отмен"""
        
#         if not reasons_data:
#             return start_row
        
#         # Заголовок
#         row = start_row
#         cell = self.ws.cell(row=row, column=2, value="ОТМЕНЫ ПО ПРИЧИНАМ")
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
#         row += 1
        
#         # Заголовки таблицы причин
#         reason_headers = ["ПРИЧИНА ОТМЕНЫ", "КОЛ-ВО ЗАКАЗОВ", "СУММА ОТМЕН", "СРЕДНЯЯ СУММА"]
#         for col_idx, header in enumerate(reason_headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
#             cell.fill = FILLS.get("header_blue", PatternFill("solid", fgColor=COLORS["dark_green"]))
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]
        
#         self.ws.row_dimensions[row].height = 30
#         row += 1
        
#         # Данные по причинам
#         for idx, reason in enumerate(reasons_data):
#             fill = FILLS.get("odd_row", PatternFill(fill_type=None)) if idx % 2 == 0 else FILLS.get("even_row", PatternFill(fill_type=None))
            
#             values = [
#                 reason.get('reason', ''),
#                 reason.get('count', 0),
#                 reason.get('amount', 0),
#                 reason.get('avg_amount', 0),
#             ]
            
#             for col_idx, value in enumerate(values, start=2):
#                 cell = self.ws.cell(row=row, column=col_idx, value=value)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]
#                 cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))
                
#                 if col_idx == 2:  # Причина
#                     cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
#                 elif col_idx == 3:  # Кол-во
#                     cell.alignment = Alignment(horizontal="center", vertical="center")
#                 else:  # Суммы
#                     cell.alignment = Alignment(horizontal="right", vertical="center")
#                     if isinstance(value, (int, float)):
#                         cell.number_format = '#,##0'
            
#             self.ws.row_dimensions[row].height = 25
#             row += 1
        
#         row += 1
#         return row

#     def draw_manager_breakdown(self, start_row, managers_data):
#         """Рисует breakdown по менеджерам"""
        
#         if not managers_data:
#             return start_row
        
#         # Заголовок
#         row = start_row
#         cell = self.ws.cell(row=row, column=2, value="ОТМЕНЫ ПО МЕНЕДЖЕРАМ")
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
#         row += 1
        
#         # Заголовки таблицы
#         manager_headers = ["МЕНЕДЖЕР", "КОЛ-ВО ОТМЕН", "СУММА ОТМЕН", "СРЕДНЯЯ СУММА"]
#         for col_idx, header in enumerate(manager_headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
#             cell.fill = FILLS.get("header_blue", PatternFill("solid", fgColor=COLORS["dark_green"]))
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]
        
#         self.ws.row_dimensions[row].height = 30
#         row += 1
        
#         # Данные по менеджерам
#         for idx, manager in enumerate(managers_data[:10]):  # Топ 10 менеджеров
#             fill = FILLS.get("odd_row", PatternFill(fill_type=None)) if idx % 2 == 0 else FILLS.get("even_row", PatternFill(fill_type=None))
            
#             values = [
#                 self._format_manager_name(manager.get('manager', '')),
#                 manager.get('orders_count', 0),
#                 manager.get('total_amount', 0),
#                 manager.get('avg_amount', 0),
#             ]
            
#             for col_idx, value in enumerate(values, start=2):
#                 cell = self.ws.cell(row=row, column=col_idx, value=value)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]
#                 cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))
                
#                 if col_idx == 2:
#                     cell.alignment = Alignment(horizontal="left", vertical="center")
#                 elif col_idx == 3:
#                     cell.alignment = Alignment(horizontal="center", vertical="center")
#                 else:
#                     cell.alignment = Alignment(horizontal="right", vertical="center")
#                     if isinstance(value, (int, float)):
#                         cell.number_format = '#,##0'
            
#             self.ws.row_dimensions[row].height = 25
#             row += 1
        
#         row += 1
#         return row

#     def build(self, orders_data, summary_data, reasons_data, managers_data, top_orders):
#         """Построение листа с анализом отмен"""
#         row = 1

#         # Кнопка назад
#         draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

#         self.ws.row_dimensions[1].height = 20
#         self.ws.row_dimensions[2].height = 8

#         # Заголовок
#         period_text = ""
#         if summary_data.get('period_start') and summary_data.get('period_end'):
#             period_text = f"Период: {summary_data['period_start']} — {summary_data['period_end']}"
        
#         row = self.sheet_title.draw(
#             row=3,
#             title="АНАЛИЗ ОТМЕНЁННЫХ ЗАКАЗОВ",
#             subtitle=f"Детальный анализ всех отменённых заказов за период с 01.04.2025 | Всего отмен: {summary_data['total_orders']} на сумму {self._format_currency(summary_data['total_cancelled_amount'])}",
#             date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y')} | {period_text}",
#             start_col=2,
#             end_col=5
#         )

#         row += 1

#         # KPI карточки (убрали лишние)
#         row1_cards = [
#             {'title': 'ОТМЕНЁННЫХ ЗАКАЗОВ', 'value': self._format_number(summary_data['total_orders']), 'color': COLORS["warning_red"], 'width': 2},
#             {'title': 'СУММА ОТМЕН', 'value': self._format_currency(summary_data['total_cancelled_amount']), 'color': COLORS["warning_red"], 'width': 2},
#         ]
        
#         row = self.kpi_cards.draw_row(row, row1_cards)
#         row += 1
        
#         row2_cards = [
#             {'title': 'УНИКАЛЬНЫХ КЛИЕНТОВ', 'value': self._format_number(summary_data['unique_clients']), 'color': COLORS["blue"], 'width': 2},
#             {'title': 'МЕНЕДЖЕРОВ', 'value': self._format_number(summary_data['unique_managers']), 'color': COLORS["blue"], 'width': 2},
#         ]
        
#         row = self.kpi_cards.draw_row(row, row2_cards)
#         row += 2

#         # ============================================================
#         # АНАЛИТИКА ПО ПРИЧИНАМ И МЕНЕДЖЕРАМ
#         # ============================================================
        
#         # Breakdown по причинам
#         row = self.draw_reason_breakdown(row, reasons_data)
        
#         # Breakdown по менеджерам
#         row = self.draw_manager_breakdown(row, managers_data)

#         # ============================================================
#         # ЗАГОЛОВОК ТАБЛИЦЫ
#         # ============================================================
#         cell = self.ws.cell(row=row, column=2, value="ДЕТАЛЬНАЯ ИНФОРМАЦИЯ ПО ОТМЕНЁННЫМ ЗАКАЗАМ (ТОП-50 ПО СУММЕ ОТМЕНЫ)")
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
#         row += 2

#         # Заголовки таблицы (убран столбец ОПЛАЧЕНО)
#         headers = [
#             "ЗАКАЗ",
#             "ДАТА\nСОЗДАНИЯ",
#             "КЛИЕНТ",
#             "МЕНЕДЖЕР",
#             "МАГАЗИН",
#             "СУММА\nЗАКАЗА",
#             "СУММА\nОТМЕНЫ",
#             "СТАТУС",
#             "ПРИЧИНА ОТМЕНЫ",
#         ]

#         table_start_row = row

#         # Рисуем заголовки
#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
#             cell.fill = FILLS.get("header_blue", PatternFill("solid", fgColor=COLORS["dark_green"]))
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 40
#         row += 1

#         # Данные (только топ отмен)
#         for idx, order in enumerate(top_orders):
#             # Чередование фона
#             if idx % 2 == 0:
#                 fill = FILLS.get("odd_row", PatternFill(fill_type=None))
#             else:
#                 fill = FILLS.get("even_row", PatternFill(fill_type=None))

#             values = [
#                 order.get('order_name', '') or order.get('number', ''),
#                 self._format_datetime(order.get('date_from')),
#                 self._format_client_name(order.get('client', '')),
#                 self._format_manager_name(order.get('manager', '')),
#                 (order.get('store', '') or '').upper()[:30],
#                 self._safe_float(order.get('cancelled_amount', 0)) + self._safe_float(order.get('paid_amount', 0)),  # Полная сумма заказа
#                 self._safe_float(order.get('cancelled_amount', 0)),
#                 order.get('status', ''),
#                 order.get('cancellation_reason', 'Не указана'),
#             ]

#             for col_idx, value in enumerate(values):
#                 col_num = col_idx + 2
#                 cell = self.ws.cell(row=row, column=col_num, value=value)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]
#                 cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))

#                 # Настройка выравнивания
#                 if col_idx in [0, 2, 3, 4, 9]:  # ЗАКАЗ, КЛИЕНТ, МЕНЕДЖЕР, МАГАЗИН, ПРИЧИНА
#                     cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
#                 elif col_idx in [1, 5, 8]:  # ДАТА, КОЛ-ВО, СТАТУС
#                     cell.alignment = Alignment(horizontal="center", vertical="center")
#                 else:  # Числовые колонки
#                     cell.alignment = Alignment(horizontal="right", vertical="center")
#                     if isinstance(value, (int, float)):
#                         cell.number_format = '#,##0'

#                 # Подсветка крупных отмен
#                 if col_idx == 7 and value > 1000000:  # СУММА ОТМЕНЫ > 1 млн
#                     cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["warning_red"])

#             self.ws.row_dimensions[row].height = 25
#             row += 1

#         # ИТОГОВАЯ СТРОКА - теперь понятная
#         if top_orders:
#             total_fill = FILLS["total"]
            
#             # Рассчитываем итоги ТОЛЬКО по отображаемым заказам (топ-50)
#             total_displayed_qty = sum(self._safe_float(o.get('order_qty', 0)) for o in top_orders)
#             total_displayed_amount = sum(
#                 self._safe_float(o.get('cancelled_amount', 0)) + self._safe_float(o.get('paid_amount', 0)) 
#                 for o in top_orders
#             )
#             total_displayed_cancelled = sum(self._safe_float(o.get('cancelled_amount', 0)) for o in top_orders)
            
#             # Пишем "ИТОГО ПО ТОП-50:"
#             cell = self.ws.cell(row=row, column=2, value="ИТОГО ПО ТОП-50:")
#             cell.fill = total_fill
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
#             cell.alignment = Alignment(horizontal="right", vertical="center")
            
#             # Количество товаров
#             cell = self.ws.cell(row=row, column=6, value=total_displayed_qty)
#             cell.fill = total_fill
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
#             cell.alignment = Alignment(horizontal="center", vertical="center")
#             cell.number_format = '#,##0'
            
#             # Сумма заказа
#             cell = self.ws.cell(row=row, column=7, value=total_displayed_amount)
#             cell.fill = total_fill
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
#             cell.alignment = Alignment(horizontal="right", vertical="center")
#             cell.number_format = '#,##0'
            
#             # Сумма отмен
#             cell = self.ws.cell(row=row, column=8, value=total_displayed_cancelled)
#             cell.fill = total_fill
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
#             cell.alignment = Alignment(horizontal="right", vertical="center")
#             cell.number_format = '#,##0'
            
#             # Заполняем остальные колонки
#             for col in [3, 4, 5, 9, 10, 11]:
#                 cell = self.ws.cell(row=row, column=col, value="")
#                 cell.fill = total_fill
#                 cell.border = BORDERS["thin"]

#             self.ws.row_dimensions[row].height = 30
#             row += 2
            
#             # Добавляем строку с общими итогами по ВСЕМ отменам
#             total_fill_light = PatternFill("solid", fgColor="E8F5E9")  # Светло-зелёный
            
#             cell = self.ws.cell(row=row, column=2, value="ВСЕГО ПО ПЕРИОДУ:")
#             cell.fill = total_fill_light
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="right", vertical="center")
            
#             # Количество товаров
#             cell = self.ws.cell(row=row, column=6, value=summary_data.get('total_qty', 0))
#             cell.fill = total_fill_light
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="center", vertical="center")
#             cell.number_format = '#,##0'
            
#             # Сумма заказа (для всех отмен)
#             total_all_amount = summary_data.get('total_cancelled_amount', 0) + summary_data.get('total_paid_amount', 0)
#             cell = self.ws.cell(row=row, column=7, value=total_all_amount)
#             cell.fill = total_fill_light
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="right", vertical="center")
#             cell.number_format = '#,##0'
            
#             # Сумма отмен (всего по периоду)
#             cell = self.ws.cell(row=row, column=8, value=summary_data.get('total_cancelled_amount', 0))
#             cell.fill = total_fill_light
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="right", vertical="center")
#             cell.number_format = '#,##0'
            
#             # Заполняем остальные колонки
#             for col in [3, 4, 5, 9, 10, 11]:
#                 cell = self.ws.cell(row=row, column=col, value="")
#                 cell.fill = total_fill_light
#                 cell.border = BORDERS["thin"]
            
#             self.ws.row_dimensions[row].height = 30
#             row += 2

#         # Примечания
#         footnote = Footnote(self.ws)
#         footnote.draw(row, text="В таблице показаны TOP-50 отмен по сумме (остальные отмены не отображаются для читаемости)")
#         row += 1
#         footnote.draw(row, text="Сумма заказа = Сумма отмены + Оплачено (если были частичные оплаты)")
#         row += 1
#         footnote.draw(row, text="Аналитика по причинам и менеджерам включает ВСЕ отмены за период с 01 апреля 2025 г.")

#         # Настройка колонок (убран столбец Оплачено)
#         col_widths = {
#             "B": 35,  # ЗАКАЗ
#             "C": 14,  # ДАТА СОЗДАНИЯ
#             "D": 30,  # КЛИЕНТ
#             "E": 22,  # МЕНЕДЖЕР
#             "F": 22,  # МАГАЗИН
#             "G": 12,  # КОЛ-ВО ТОВАРОВ
#             "H": 16,  # СУММА ЗАКАЗА
#             "I": 16,  # СУММА ОТМЕНЫ
#             "J": 25,  # СТАТУС
#             "K": 35,  # ПРИЧИНА ОТМЕНЫ
#         }
#         for col, width in col_widths.items():
#             self.ws.column_dimensions[col].width = width

#         # Автофильтр
#         if top_orders:
#             self.ws.auto_filter.ref = f'B{table_start_row}:K{row - 5}'
        
#         self.ws.sheet_view.showGridLines = False




# orders/reports/sheets/cancelled_analysis_sheet.py
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .base_sheet import BaseSheet
from ..styles.theme import COLORS, BORDERS, FILLS
from ..styles.helpers import draw_toc_button
from ..components.footnote import Footnote
from ..components.sheet_title import create_sheet_title
from ..components.kpi_cards import KPICards


class CancelledAnalysisSheet(BaseSheet):
    """Лист с анализом отменённых заказов"""

    def __init__(self, workbook, sheet_number):
        super().__init__(workbook, sheet_number)
        self.sheet_title = create_sheet_title(self.ws)
        self.kpi_cards = KPICards(self.ws)

    def _format_currency(self, value):
        if value is None or value == 0:
            return "0 ₽"
        return f"{int(round(abs(value))):,} ₽".replace(",", " ")

    def _format_number(self, value):
        if value is None or value == 0:
            return "0"
        return f"{int(round(abs(value))):,}".replace(",", " ")

    def _safe_float(self, value, default=0.0):
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default

    def _format_manager_name(self, full_name):
        if not full_name:
            return ""
        parts = str(full_name).strip().split()
        if len(parts) >= 1:
            last_name = parts[0].upper()
            if len(parts) >= 2:
                first_initial = parts[1][0].upper() if parts[1] else ""
                return f"{last_name} {first_initial}."
            return last_name
        return str(full_name).upper()

    def _format_client_name(self, name):
        if not name:
            return ""
        return str(name).upper()

    def _format_datetime(self, date_value):
        """Форматирует дату без времени: YYYY-MM-DD"""
        if not date_value:
            return ""
        if hasattr(date_value, 'strftime'):
            return date_value.strftime('%Y-%m-%d')
        if isinstance(date_value, str):
            for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d.%m.%Y']:
                try:
                    dt = datetime.strptime(date_value, fmt)
                    return dt.strftime('%Y-%m-%d')
                except ValueError:
                    continue
        return str(date_value)[:10] if len(str(date_value)) > 10 else str(date_value)

    def draw_reason_breakdown(self, start_row, reasons_data):
        """Рисует breakdown по причинам отмен"""
        
        if not reasons_data:
            return start_row
        
        # Заголовок
        row = start_row
        cell = self.ws.cell(row=row, column=2, value="ОТМЕНЫ ПО ПРИЧИНАМ")
        cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
        cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1
        
        # Заголовки таблицы причин
        reason_headers = ["ПРИЧИНА ОТМЕНЫ", "КОЛ-ВО ЗАКАЗОВ", "СУММА ОТМЕН", "СРЕДНЯЯ СУММА"]
        for col_idx, header in enumerate(reason_headers, start=2):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
            cell.fill = FILLS.get("header_blue", PatternFill("solid", fgColor=COLORS["dark_green"]))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDERS["thin"]
        
        self.ws.row_dimensions[row].height = 30
        row += 1
        
        # Данные по причинам
        for idx, reason in enumerate(reasons_data):
            fill = FILLS.get("odd_row", PatternFill(fill_type=None)) if idx % 2 == 0 else FILLS.get("even_row", PatternFill(fill_type=None))
            
            values = [
                reason.get('reason', ''),
                reason.get('count', 0),
                reason.get('amount', 0),
                reason.get('avg_amount', 0),
            ]
            
            for col_idx, value in enumerate(values, start=2):
                cell = self.ws.cell(row=row, column=col_idx, value=value)
                cell.fill = fill
                cell.border = BORDERS["thin"]
                cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))
                
                if col_idx == 2:  # Причина
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif col_idx == 3:  # Кол-во
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:  # Суммы
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0'
            
            self.ws.row_dimensions[row].height = 25
            row += 1
        
        row += 1
        return row

    def draw_manager_breakdown(self, start_row, managers_data):
        """Рисует breakdown по менеджерам"""
        
        if not managers_data:
            return start_row
        
        # Заголовок
        row = start_row
        cell = self.ws.cell(row=row, column=2, value="ОТМЕНЫ ПО МЕНЕДЖЕРАМ")
        cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
        cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1
        
        # Заголовки таблицы
        manager_headers = ["МЕНЕДЖЕР", "КОЛ-ВО ОТМЕН", "СУММА ОТМЕН", "СРЕДНЯЯ СУММА"]
        for col_idx, header in enumerate(manager_headers, start=2):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
            cell.fill = FILLS.get("header_blue", PatternFill("solid", fgColor=COLORS["dark_green"]))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDERS["thin"]
        
        self.ws.row_dimensions[row].height = 30
        row += 1
        
        # Данные по менеджерам
        for idx, manager in enumerate(managers_data[:10]):  # Топ 10 менеджеров
            fill = FILLS.get("odd_row", PatternFill(fill_type=None)) if idx % 2 == 0 else FILLS.get("even_row", PatternFill(fill_type=None))
            
            values = [
                self._format_manager_name(manager.get('manager', '')),
                manager.get('orders_count', 0),
                manager.get('total_amount', 0),
                manager.get('avg_amount', 0),
            ]
            
            for col_idx, value in enumerate(values, start=2):
                cell = self.ws.cell(row=row, column=col_idx, value=value)
                cell.fill = fill
                cell.border = BORDERS["thin"]
                cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))
                
                if col_idx == 2:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif col_idx == 3:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0'
            
            self.ws.row_dimensions[row].height = 25
            row += 1
        
        row += 1
        return row

    def build(self, orders_data, summary_data, reasons_data, managers_data, top_orders):
        """Построение листа с анализом отмен"""
        row = 1

        # Кнопка назад
        draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 8

        # Заголовок
        row = self.sheet_title.draw(
            row=3,
            title="АНАЛИЗ ОТМЕНЁННЫХ ЗАКАЗОВ",
            subtitle=f"Детальный анализ всех отменённых заказов за период с 01.04.2025 | Всего отмен: {summary_data['total_orders']} на сумму {self._format_currency(summary_data['total_cancelled_amount'])}",
            date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            start_col=2,
            end_col=10
        )

        row += 1

        # KPI карточки
        row1_cards = [
            {'title': 'ОТМЕНЁННЫХ ЗАКАЗОВ', 'value': self._format_number(summary_data['total_orders']), 'color': COLORS["warning_red"], 'width': 2},
            {'title': 'СУММА ОТМЕН', 'value': self._format_currency(summary_data['total_cancelled_amount']), 'color': COLORS["warning_red"], 'width': 2},
        ]
        
        row = self.kpi_cards.draw_row(row, row1_cards)
        row += 1
        
        row2_cards = [
            {'title': 'УНИКАЛЬНЫХ КЛИЕНТОВ', 'value': self._format_number(summary_data['unique_clients']), 'color': COLORS["blue"], 'width': 2},
            {'title': 'МЕНЕДЖЕРОВ', 'value': self._format_number(summary_data['unique_managers']), 'color': COLORS["blue"], 'width': 2},
        ]
        
        row = self.kpi_cards.draw_row(row, row2_cards)
        row += 2

        # ============================================================
        # АНАЛИТИКА ПО ПРИЧИНАМ И МЕНЕДЖЕРАМ
        # ============================================================
        
        # Breakdown по причинам
        row = self.draw_reason_breakdown(row, reasons_data)
        
        # Breakdown по менеджерам
        row = self.draw_manager_breakdown(row, managers_data)

        # ============================================================
        # ЗАГОЛОВОК ТАБЛИЦЫ
        # ============================================================
        cell = self.ws.cell(row=row, column=2, value="ДЕТАЛЬНАЯ ИНФОРМАЦИЯ ПО ОТМЕНЁННЫМ ЗАКАЗАМ (ТОП-50 ПО СУММЕ ОТМЕНЫ)")
        cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
        cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        row += 2

        # Заголовки таблицы (без колонки КОЛ-ВО ТОВАРОВ)
        headers = [
            "ЗАКАЗ",
            "ДАТА\nСОЗДАНИЯ",
            "КЛИЕНТ",
            "МЕНЕДЖЕР",
            "МАГАЗИН",
            "СУММА\nЗАКАЗА",
            "СУММА\nОТМЕНЫ",
            "СТАТУС",
            "ПРИЧИНА ОТМЕНЫ",
        ]

        table_start_row = row

        # Рисуем заголовки
        for col_idx, header in enumerate(headers, start=2):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
            cell.fill = FILLS.get("header_blue", PatternFill("solid", fgColor=COLORS["dark_green"]))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDERS["thin"]

        self.ws.row_dimensions[row].height = 40
        row += 1

        # Данные (только топ отмен)
        for idx, order in enumerate(top_orders):
            # Чередование фона
            if idx % 2 == 0:
                fill = FILLS.get("odd_row", PatternFill(fill_type=None))
            else:
                fill = FILLS.get("even_row", PatternFill(fill_type=None))

            values = [
                order.get('order_name', '') or order.get('number', ''),
                self._format_datetime(order.get('date_from')),
                self._format_client_name(order.get('client', '')),
                self._format_manager_name(order.get('manager', '')),
                (order.get('store', '') or '').upper()[:30],
                self._safe_float(order.get('cancelled_amount', 0)) + self._safe_float(order.get('paid_amount', 0)),
                self._safe_float(order.get('cancelled_amount', 0)),
                order.get('status', ''),
                order.get('cancellation_reason', 'Не указана'),
            ]

            for col_idx, value in enumerate(values):
                col_num = col_idx + 2
                cell = self.ws.cell(row=row, column=col_num, value=value)
                cell.fill = fill
                cell.border = BORDERS["thin"]
                cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))

                # Настройка выравнивания
                if col_idx in [0, 2, 3, 4, 8]:  # ЗАКАЗ, КЛИЕНТ, МЕНЕДЖЕР, МАГАЗИН, ПРИЧИНА
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif col_idx in [1, 7]:  # ДАТА, СТАТУС
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:  # Числовые колонки
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0'

                # Подсветка крупных отмен
                if col_idx == 6 and value > 1000000:  # СУММА ОТМЕНЫ > 1 млн
                    cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["warning_red"])

            self.ws.row_dimensions[row].height = 25
            row += 1

        # ИТОГОВАЯ СТРОКА - только по топ-50
        if top_orders:
            total_fill = FILLS["total"]
            
            # Рассчитываем итоги по отображаемым заказам (топ-50)
            total_displayed_amount = sum(
                self._safe_float(o.get('cancelled_amount', 0)) + self._safe_float(o.get('paid_amount', 0)) 
                for o in top_orders
            )
            total_displayed_cancelled = sum(self._safe_float(o.get('cancelled_amount', 0)) for o in top_orders)
            
            # Пишем "ИТОГО ПО ТОП-50:"
            cell = self.ws.cell(row=row, column=2, value="ИТОГО ПО ТОП-50:")
            cell.fill = total_fill
            cell.border = BORDERS["thin"]
            cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
            cell.alignment = Alignment(horizontal="right", vertical="center")
            
            # Сумма заказа
            cell = self.ws.cell(row=row, column=6, value=total_displayed_amount)
            cell.fill = total_fill
            cell.border = BORDERS["thin"]
            cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0'
            
            # Сумма отмен
            cell = self.ws.cell(row=row, column=7, value=total_displayed_cancelled)
            cell.fill = total_fill
            cell.border = BORDERS["thin"]
            cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0'
            
            # Заполняем остальные колонки
            for col in [3, 4, 5, 8, 9, 10]:
                cell = self.ws.cell(row=row, column=col, value="")
                cell.fill = total_fill
                cell.border = BORDERS["thin"]

            self.ws.row_dimensions[row].height = 30
            row += 2

        # Примечания
        footnote = Footnote(self.ws)
        footnote.draw(row, text="Период анализа: с 01 апреля 2025 года по текущую дату")
        row += 1
        footnote.draw(row, text="В таблице показаны TOP-50 отмен по сумме (остальные отмены не отображаются для читаемости)")
        row += 1
        footnote.draw(row, text="Аналитика по причинам и менеджерам включает ВСЕ отмены за период")

        # Настройка колонок (без колонки КОЛ-ВО ТОВАРОВ)
        col_widths = {
            "B": 35,  # ЗАКАЗ
            "C": 14,  # ДАТА СОЗДАНИЯ
            "D": 30,  # КЛИЕНТ
            "E": 22,  # МЕНЕДЖЕР
            "F": 22,  # МАГАЗИН
            "G": 16,  # СУММА ЗАКАЗА
            "H": 16,  # СУММА ОТМЕНЫ
            "I": 25,  # СТАТУС
            "J": 35,  # ПРИЧИНА ОТМЕНЫ
        }
        for col, width in col_widths.items():
            self.ws.column_dimensions[col].width = width

        # Автофильтр
        if top_orders:
            self.ws.auto_filter.ref = f'B{table_start_row}:J{row - 5}'
        
        self.ws.sheet_view.showGridLines = False