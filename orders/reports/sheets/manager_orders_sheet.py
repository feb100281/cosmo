# orders/reports/sheets/manager_orders_sheet.py
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from .base_sheet import BaseSheet
from ..styles.theme import COLORS, BORDERS, FILLS
from ..styles.helpers import draw_toc_button


class ManagerOrdersSheet(BaseSheet):
    """Лист с детализацией заказов для конкретного менеджера"""
    
    def __init__(self, workbook, sheet_name, manager_name):
        safe_name = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
        super().__init__(workbook, safe_name)
        self.manager_name = manager_name
        self.formatted_manager_name = self._format_manager_name(manager_name)
    
    def _format_manager_name(self, manager_name):
        """Форматирует имя менеджера: первое слово полностью, остальные - первая буква и точка"""
        if not manager_name:
            return manager_name
        words = manager_name.split()
        if not words:
            return manager_name
        formatted_words = [words[0]]
        for word in words[1:]:
            if word and len(word) > 0:
                formatted_words.append(f"{word[0]}.")
            else:
                formatted_words.append(word)
        return " ".join(formatted_words)
    
    def _set_number(self, cell, value, fmt="# ##0"):
        cell.value = value or 0
        cell.number_format = fmt
    
    def _set_money(self, cell, value):
        cell.value = value or 0
        cell.number_format = '# ##0 ₽'
    
    def _set_date(self, cell, value):
        if value:
            cell.value = value
            cell.number_format = "DD.MM.YYYY"
        else:
            cell.value = ""
    
    def _get_status_color(self, status):
        status_colors = {
            "Закрыт": COLORS["dark_green"],
            "Закрыто": COLORS["dark_green"],
            "Отгружен": COLORS["blue"],
            "Оплачен": COLORS["orange"],
            "Новый": COLORS["text_gray"],
            "В обработке": COLORS["text_gray"],
        }
        for key, color in status_colors.items():
            if key in status:
                return color
        return COLORS["text_dark"]
    
    def build(self, orders_data, summary_data):
        """Строит лист с заказами менеджера"""
        row = 1
        
        draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")
        
        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 8
        row = 3
        
        # ЗАГОЛОВОК
        title_cell = self.ws.cell(row=row, column=2, value=f"МЕНЕДЖЕР: {self.formatted_manager_name}")
        title_cell.font = Font(name="Roboto", size=16, bold=True, color=COLORS["dark_green"])
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        row += 1
        
        date_cell = self.ws.cell(row=row, column=2, value=f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        date_cell.font = Font(name="Roboto", size=9, italic=True, color=COLORS["text_gray"])
        date_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        row += 2
        
        # СВОДНАЯ СТАТИСТИКА
        stats_row = row
        stats_headers = ["Всего заказов", "Сумма заказов", "Отменено", "Без оплаты", "Частичная оплата"]
        
        for col_idx, header in enumerate(stats_headers, start=2):
            cell = self.ws.cell(row=stats_row, column=col_idx, value=header)
            cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
            cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDERS["thin"]
        
        stats_row += 1
        
        stats_values = [
            summary_data.get("total_orders", 0),
            summary_data.get("total_amount", 0),
            summary_data.get("canceled_orders", 0),
            summary_data.get("no_payment_orders", 0),
            summary_data.get("partial_payment_orders", 0),
        ]
        
        for col_idx, value in enumerate(stats_values, start=2):
            cell = self.ws.cell(row=stats_row, column=col_idx)
            cell.fill = FILLS["odd_row"]
            cell.border = BORDERS["thin"]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx == 3:
                self._set_money(cell, value)
            else:
                self._set_number(cell, value)
        
        row = stats_row + 2
        
        # ТАБЛИЦА ЗАКАЗОВ
        headers = [
            "№ заказа", "Дата", "Клиент", "Статус", 
            "Сумма заказа", "Оплачено", "Сумма отгрузки", 
            "Статус оплаты", "Комментарий"
        ]
        
        for col_idx, header in enumerate(headers, start=2):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["white"])
            cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDERS["thin"]
        
        self.ws.row_dimensions[row].height = 35
        row += 1
        
        # Данные заказов
        if orders_data:
            for idx, order in enumerate(orders_data):
                fill = FILLS["odd_row"] if idx % 2 == 0 else FILLS["even_row"]
                
                payment_status = "Не оплачен"
                cash_pmts = order.get("cash_pmts", 0)
                amount_active = order.get("amount_active", 0)
                if cash_pmts > 0:
                    if cash_pmts >= amount_active:
                        payment_status = "Полностью оплачен"
                    else:
                        payment_status = "Частично оплачен"
                
                # Строка 2: № заказа
                cell = self.ws.cell(row=row, column=2, value=order.get("order_id", ""))
                cell.fill = fill
                cell.border = BORDERS["thin"]
                cell.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Строка 3: Дата
                cell = self.ws.cell(row=row, column=3)
                cell.fill = fill
                cell.border = BORDERS["thin"]
                self._set_date(cell, order.get("date_from"))
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                
                # Строка 4: Клиент
                client_name = order.get("client_name") or order.get("store_name") or ""
                cell = self.ws.cell(row=row, column=4, value=client_name)
                cell.fill = fill
                cell.border = BORDERS["thin"]
                cell.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Строка 5: Статус
                status = order.get("status", "")
                cell = self.ws.cell(row=row, column=5, value=status)
                cell.fill = fill
                cell.border = BORDERS["thin"]
                cell.font = Font(name="Roboto", size=9, color=self._get_status_color(status))
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Строка 6: Сумма заказа
                cell = self.ws.cell(row=row, column=6)
                cell.fill = fill
                cell.border = BORDERS["thin"]
                self._set_money(cell, amount_active)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                
                # Строка 7: Оплачено
                cell = self.ws.cell(row=row, column=7)
                cell.fill = fill
                cell.border = BORDERS["thin"]
                self._set_money(cell, cash_pmts)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                
                # Строка 8: Сумма отгрузки
                cell = self.ws.cell(row=row, column=8)
                cell.fill = fill
                cell.border = BORDERS["thin"]
                self._set_money(cell, order.get("total_shiped_amount", 0))
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                
                # Строка 9: Статус оплаты
                cell = self.ws.cell(row=row, column=9, value=payment_status)
                cell.fill = fill
                cell.border = BORDERS["thin"]
                color = COLORS["dark_green"] if "Полностью" in payment_status else (COLORS["warning_red"] if "Не" in payment_status else COLORS["orange"])
                cell.font = Font(name="Roboto", size=9, bold=True, color=color)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Строка 10: Комментарий
                comment = order.get("comment") or order.get("order_info") or ""
                if len(str(comment)) > 80:
                    comment = str(comment)[:80] + "..."
                cell = self.ws.cell(row=row, column=10, value=comment)
                cell.fill = fill
                cell.border = BORDERS["thin"]
                cell.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                self.ws.row_dimensions[row].height = 28
                row += 1
        else:
            # Если нет данных
            cell = self.ws.cell(row=row, column=2, value="Нет активных заказов")
            cell.fill = FILLS["odd_row"]
            cell.border = BORDERS["thin"]
            cell.font = Font(name="Roboto", size=10, italic=True, color=COLORS["text_gray"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        
        # Ширина колонок
        col_widths = {
            "B": 14, "C": 12, "D": 30, "E": 18, "F": 15,
            "G": 15, "H": 15, "I": 18, "J": 50,  # J увеличен для комментария
        }
        for col, width in col_widths.items():
            self.ws.column_dimensions[col].width = width
        
        self.ws.sheet_view.showGridLines = False