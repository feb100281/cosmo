# # orders/reports/sheets/unpaid_shipments_sheet.py
# from datetime import datetime
# from openpyxl.styles import Font, PatternFill, Border, Side
# from .base_sheet import BaseSheet
# from ..styles.theme import COLORS, BORDERS, FILLS, Alignment
# from ..styles.helpers import draw_toc_button
# from ..components.footnote import Footnote
# from ..components.sheet_title import create_sheet_title


# class UnpaidShipmentsSheet(BaseSheet):
#     """Лист с заказами, где есть отгрузка, но нет/неполная оплата"""

#     def __init__(self, workbook, sheet_number):
#         super().__init__(workbook, sheet_number)
#         self.sheet_title = create_sheet_title(self.ws)
        
#         self.CRITICAL_COLOR = "D32F2F"
#         self.HIGH_RISK_COLOR = "F57C00"
#         self.MEDIUM_RISK_COLOR = "FBC02D"
#         self.LOW_RISK_COLOR = "7CB342"

#     def _format_currency(self, value):
#         if value is None or value == 0:
#             return "0 ₽"
#         return f"{int(round(abs(value))):,} ₽".replace(",", " ")

#     def _format_number(self, value):
#         if value is None or value == 0:
#             return "0"
#         return f"{int(round(abs(value))):,}".replace(",", " ")

#     def _format_percent(self, value):
#         if value is None:
#             return "0%"
#         return f"{value:.1f}%"

#     def _safe_float(self, value, default=0.0):
#         if value is None:
#             return default
#         try:
#             return float(value)
#         except (ValueError, TypeError):
#             return default

#     def _format_datetime(self, date_value):
#         if not date_value:
#             return ""
#         if hasattr(date_value, 'strftime'):
#             return date_value.strftime('%Y-%m-%d')
#         return str(date_value)[:10]

#     def _format_manager_name(self, full_name):
#         if not full_name:
#             return ""
#         parts = str(full_name).strip().split()
#         if len(parts) >= 1:
#             last_name = parts[0].upper()
#             if len(parts) >= 2:
#                 first_initial = parts[1][0].upper() if parts[1] else ""
#                 return f"{last_name} {first_initial}."
#             return last_name
#         return str(full_name).upper()

#     def _format_client_name(self, name):
#         if not name:
#             return ""
#         return str(name).upper()

#     def _format_payment_summary(self, payment_details):
#         """Форматирует краткую сводку по оплатам"""
#         if not payment_details:
#             return "Нет оплат"
        
#         payments_str = []
#         for p in payment_details[:3]:
#             sign = "-" if p['amount'] < 0 else ""
#             payments_str.append(f"{p['date']}: {sign}{self._format_currency(abs(p['amount']))}")
        
#         result = "\n".join(payments_str)
#         if len(payment_details) > 3:
#             result += f"\n+ еще {len(payment_details) - 3}"
#         return result

#     def build(self, orders_data, summary_data):
#         """Построение листа"""
#         row = 1

#         # Кнопка назад
#         draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

#         # Заголовок
#         row = self.sheet_title.draw(
#             row=3,
#             title="ОТГРУЗКИ БЕЗ ОПЛАТЫ / НЕПОЛНАЯ ОПЛАТА",
#             subtitle=f"Заказы с отгрузкой, но оплата меньше суммы отгрузки | Найдено: {summary_data['total_orders']} заказов | Общая сумма недоплаты: {self._format_currency(summary_data['total_underpayment'])}",
#             date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
#             start_col=2,
#             end_col=16
#         )

#         row += 2

#         # # KPI карточки
#         # row = self._draw_kpi_cards(row, summary_data)
#         # row += 2

#         # Заголовок таблицы
#         headers = [
#             "ЗАКАЗ",
#             "ДАТА\nСОЗДАНИЯ",
#             "ПОСЛЕДНЯЯ\nОТГРУЗКА",
#             "КОЛ-ВО\nОТГРУЗОК",
#             "ДНЕЙ БЕЗ\nОПЛАТЫ",
#             "СТАТУС",
#             "КЛИЕНТ",
#             "МЕНЕДЖЕР",
#             "МАГАЗИН",
#             "СУММА\nЗАКАЗА",
#             "ОТГРУЖЕНО\nВСЕГО",
#             "ОПЛАЧЕНО",
#             "НЕДОПЛАТА",
#             "%\nОПЛАТЫ",
#             "РИСК",
#             "ОПЛАТЫ\n(сводка)",
#         ]

#         table_start_row = row

#         # Рисуем заголовки
#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
#             cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 40
#         row += 1

#         # Данные
#         for idx, order in enumerate(orders_data):
#             # Чередование фона
#             fill = FILLS.get("odd_row" if idx % 2 == 0 else "even_row", PatternFill(fill_type=None))

#             # Определяем цвет для строки в зависимости от риска
#             risk = order.get('risk_category', '')
#             if "КРИТИЧНО" in risk:
#                 row_fill = PatternFill("solid", fgColor="FFEBEE")
#             elif "ВЫСОКИЙ" in risk:
#                 row_fill = PatternFill("solid", fgColor="FFF3E0")
#             else:
#                 row_fill = fill

#             values = [
#                 order.get('order_name', '') or order.get('number', ''),
#                 self._format_datetime(order.get('date_from')),
#                 order.get('last_shipment_date', ''),  # НОВАЯ КОЛОНКА
#                 order.get('shipments_count', 0),       # НОВАЯ КОЛОНКА
#                 f"{order.get('days_without_payment', 0)} дн.",  # НОВАЯ КОЛОНКА
#                 order.get('status', ''),
#                 self._format_client_name(order.get('client', '')),
#                 self._format_manager_name(order.get('manager', '')),
#                 (order.get('store', '') or '').upper()[:30],
#                 self._safe_float(order.get('amount_active')),
#                 self._safe_float(order.get('total_shiped_amount')),
#                 self._safe_float(order.get('cash_pmts')),
#                 order.get('underpayment', 0),
#                 order.get('payment_percent', 0),
#                 order.get('risk_category', ''),
#                 self._format_payment_summary(order.get('payment_details', [])),
#             ]

#             for col_idx, value in enumerate(values):
#                 col_num = col_idx + 2
#                 cell = self.ws.cell(row=row, column=col_num, value=value)
#                 cell.fill = row_fill
#                 cell.border = BORDERS["thin"]

#                 # Настройка выравнивания и цветов
#                 if col_idx in [0, 6, 7, 8]:  # ЗАКАЗ, КЛИЕНТ, МЕНЕДЖЕР, МАГАЗИН
#                     cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
#                     cell.font = Font(name="Roboto", size=9)
#                 elif col_idx in [1, 2, 4, 5]:  # ДАТА СОЗДАНИЯ, ПОСЛЕДНЯЯ ОТГРУЗКА, ДНЕЙ БЕЗ ОПЛАТЫ, СТАТУС
#                     cell.alignment = Alignment(horizontal="center", vertical="center")
#                     cell.font = Font(name="Roboto", size=9)
#                 elif col_idx == 3:  # КОЛ-ВО ОТГРУЗОК
#                     cell.alignment = Alignment(horizontal="center", vertical="center")
#                     cell.font = Font(name="Roboto", size=9, bold=True)
#                     if value > 1:
#                         cell.font = Font(name="Roboto", size=9, bold=True, color=self.HIGH_RISK_COLOR)
#                 elif col_idx == 12:  # НЕДОПЛАТА
#                     cell.alignment = Alignment(horizontal="right", vertical="center")
#                     cell.font = Font(name="Roboto", size=9, bold=True, color=self.CRITICAL_COLOR)
#                     cell.number_format = '#,##0.00'
#                 elif col_idx == 14:  # РИСК
#                     cell.alignment = Alignment(horizontal="left", vertical="center")
#                     if "КРИТИЧНО" in str(value):
#                         cell.font = Font(name="Roboto", size=9, bold=True, color=self.CRITICAL_COLOR)
#                     elif "ВЫСОКИЙ" in str(value):
#                         cell.font = Font(name="Roboto", size=9, bold=True, color=self.HIGH_RISK_COLOR)
#                     elif "СРЕДНИЙ" in str(value):
#                         cell.font = Font(name="Roboto", size=9, color=self.MEDIUM_RISK_COLOR)
#                     else:
#                         cell.font = Font(name="Roboto", size=9, color=self.LOW_RISK_COLOR)
#                 elif col_idx == 15:  # ОПЛАТЫ (сводка)
#                     cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
#                     cell.font = Font(name="Roboto", size=8)
#                 else:  # Числовые колонки
#                     cell.alignment = Alignment(horizontal="right", vertical="center")
#                     cell.font = Font(name="Roboto", size=9)
#                     if isinstance(value, (int, float)) and col_idx not in [13]:
#                         cell.number_format = '#,##0.00'

#             self.ws.row_dimensions[row].height = 28
#             row += 1

#         # ИТОГОВАЯ СТРОКА
#         total_shipped = summary_data['total_shipped_amount']
#         total_underpayment = summary_data['total_underpayment']
#         total_amount_active = sum(o.get('amount_active', 0) for o in orders_data)
#         total_cash_pmts = sum(o.get('cash_pmts', 0) for o in orders_data)
#         total_shipments_count = summary_data.get('total_shipments_count', 0)

#         total_fill = PatternFill("solid", fgColor="E8F0FE")
#         thin_border = Border(
#             left=Side(style='thin'),
#             right=Side(style='thin'),
#             top=Side(style='thin'),
#             bottom=Side(style='thin')
#         )

#         # Пишем "ИТОГО" в первой колонке
#         cell = self.ws.cell(row=row, column=2, value="ИТОГО:")
#         cell.fill = total_fill
#         cell.border = thin_border
#         cell.font = Font(name="Roboto", bold=True, size=10)
#         cell.alignment = Alignment(horizontal="center", vertical="center")

#         # Колонка 4 (D) - КОЛ-ВО ОТГРУЗОК (всего)
#         cell = self.ws.cell(row=row, column=4, value=total_shipments_count)
#         cell.fill = total_fill
#         cell.border = thin_border
#         cell.font = Font(name="Roboto", bold=True, size=10)
#         cell.alignment = Alignment(horizontal="center", vertical="center")
#         cell.number_format = '#,##0'

#         # Колонка 10 (J) - Сумма заказа
#         cell = self.ws.cell(row=row, column=10, value=total_amount_active)
#         cell.fill = total_fill
#         cell.border = thin_border
#         cell.font = Font(name="Roboto", bold=True, size=10)
#         cell.alignment = Alignment(horizontal="right", vertical="center")
#         cell.number_format = '#,##0.00'

#         # Колонка 11 (K) - Отгружено всего
#         cell = self.ws.cell(row=row, column=11, value=total_shipped)
#         cell.fill = total_fill
#         cell.border = thin_border
#         cell.font = Font(name="Roboto", bold=True, size=10)
#         cell.alignment = Alignment(horizontal="right", vertical="center")
#         cell.number_format = '#,##0.00'

#         # Колонка 12 (L) - Оплачено
#         cell = self.ws.cell(row=row, column=12, value=total_cash_pmts)
#         cell.fill = total_fill
#         cell.border = thin_border
#         cell.font = Font(name="Roboto", bold=True, size=10)
#         cell.alignment = Alignment(horizontal="right", vertical="center")
#         cell.number_format = '#,##0.00'

#         # Колонка 13 (M) - Недоплата
#         cell = self.ws.cell(row=row, column=13, value=total_underpayment)
#         cell.fill = total_fill
#         cell.border = thin_border
#         cell.font = Font(name="Roboto", bold=True, size=10, color=self.CRITICAL_COLOR)
#         cell.alignment = Alignment(horizontal="right", vertical="center")
#         cell.number_format = '#,##0.00'

#         # Заполняем остальные колонки пустыми с заливкой
#         for col in [3, 5, 6, 7, 8, 9, 14, 15, 16, 17]:
#             cell = self.ws.cell(row=row, column=col, value="")
#             cell.fill = total_fill
#             cell.border = thin_border
#             cell.alignment = Alignment(horizontal="center", vertical="center")
#             cell.font = Font(name="Roboto", bold=True, size=10)

#         row += 2

#         # Примечания
#         footnote = Footnote(self.ws)
#         footnote.draw(row, text="⚠️ КРИТИЧНО: полное отсутствие оплаты при наличии отгрузки")
#         row += 1
#         footnote.draw(row, text="🟠 ВЫСОКИЙ РИСК: оплачено менее 50% от суммы отгрузки")
#         row += 1
#         footnote.draw(row, text="🟡 СРЕДНИЙ РИСК: оплачено 50-80% от суммы отгрузки")
#         row += 1
#         footnote.draw(row, text="🟢 НЕЗНАЧИТЕЛЬНО: оплачено более 80% от суммы отгрузки")
#         row += 1
#         footnote.draw(row, text="📦 Дни без оплаты считаются от ПОСЛЕДНЕЙ даты отгрузки (из sales_salesdata)")

#         # Настройка колонок
#         col_widths = {
#             "B": 28,  # ЗАКАЗ
#             "C": 12,  # ДАТА СОЗДАНИЯ
#             "D": 14,  # ПОСЛЕДНЯЯ ОТГРУЗКА
#             "E": 12,  # КОЛ-ВО ОТГРУЗОК
#             "F": 14,  # ДНЕЙ БЕЗ ОПЛАТЫ
#             "G": 22,  # СТАТУС
#             "H": 28,  # КЛИЕНТ
#             "I": 20,  # МЕНЕДЖЕР
#             "J": 22,  # МАГАЗИН
#             "K": 16,  # СУММА ЗАКАЗА
#             "L": 16,  # ОТГРУЖЕНО ВСЕГО
#             "M": 14,  # ОПЛАЧЕНО
#             "N": 14,  # НЕДОПЛАТА
#             "O": 12,  # % ОПЛАТЫ
#             "P": 25,  # РИСК
#             "Q": 35,  # ОПЛАТЫ (сводка)
#         }
#         for col, width in col_widths.items():
#             self.ws.column_dimensions[col].width = width

#         self.ws.freeze_panes = f'G{table_start_row + 1}'
#         self.ws.auto_filter.ref = f'B{table_start_row}:Q{row - 5}'
#         self.ws.sheet_view.showGridLines = False

#     def _draw_kpi_cards(self, start_row, summary):
#         """Рисует KPI карточки"""
#         cards = [
#             {'title': 'ПРОБЛЕМНЫХ ЗАКАЗОВ', 'value': self._format_number(summary['total_orders']), 'color': self.HIGH_RISK_COLOR},
#             {'title': 'НЕДОПЛАТА ВСЕГО', 'value': self._format_currency(summary['total_underpayment']), 'color': self.CRITICAL_COLOR},
#             {'title': 'КРИТИЧНЫХ (0% оплаты)', 'value': self._format_number(summary['critical_orders']), 'color': self.CRITICAL_COLOR},
#             {'title': 'ВЫСОКИЙ РИСК', 'value': self._format_number(summary['high_risk_orders']), 'color': self.HIGH_RISK_COLOR},
#             {'title': 'ВСЕГО ОТГРУЗОК', 'value': self._format_number(summary.get('total_shipments_count', 0)), 'color': self.MEDIUM_RISK_COLOR},
#             {'title': 'СР. ДНЕЙ БЕЗ ОПЛАТЫ', 'value': f"{summary.get('avg_days_without_payment', 0):.0f}", 'color': self.LOW_RISK_COLOR},
#         ]

#         # Разбиваем на 2 ряда по 3 карточки
#         for idx, card in enumerate(cards[:3]):
#             col = 2 + idx * 3
#             # Заголовок
#             title_cell = self.ws.cell(row=start_row, column=col, value=card['title'])
#             title_cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#             title_cell.alignment = Alignment(horizontal="center")
            
#             # Значение
#             value_cell = self.ws.cell(row=start_row + 1, column=col, value=card['value'])
#             value_cell.font = Font(name="Roboto", size=20, bold=True, color=card['color'])
#             value_cell.alignment = Alignment(horizontal="center")
            
#             # Объединяем ячейки
#             self.ws.merge_cells(start_row=start_row, start_column=col, end_row=start_row + 1, end_column=col + 2)
            
#             # Границы
#             thin_border = Border(
#                 left=Side(style='thin', color='E0E0E0'),
#                 right=Side(style='thin', color='E0E0E0'),
#                 top=Side(style='thin', color='E0E0E0'),
#                 bottom=Side(style='thin', color='E0E0E0')
#             )
#             for r in range(start_row, start_row + 2):
#                 for c in range(col, col + 3):
#                     border_cell = self.ws.cell(row=r, column=c)
#                     border_cell.border = thin_border

#         # Второй ряд карточек
#         for idx, card in enumerate(cards[3:]):
#             col = 2 + idx * 3
#             # Заголовок
#             title_cell = self.ws.cell(row=start_row + 2, column=col, value=card['title'])
#             title_cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#             title_cell.alignment = Alignment(horizontal="center")
            
#             # Значение
#             value_cell = self.ws.cell(row=start_row + 3, column=col, value=card['value'])
#             value_cell.font = Font(name="Roboto", size=20, bold=True, color=card['color'])
#             value_cell.alignment = Alignment(horizontal="center")
            
#             # Объединяем ячейки
#             self.ws.merge_cells(start_row=start_row + 2, start_column=col, end_row=start_row + 3, end_column=col + 2)
            
#             # Границы
#             thin_border = Border(
#                 left=Side(style='thin', color='E0E0E0'),
#                 right=Side(style='thin', color='E0E0E0'),
#                 top=Side(style='thin', color='E0E0E0'),
#                 bottom=Side(style='thin', color='E0E0E0')
#             )
#             for r in range(start_row + 2, start_row + 4):
#                 for c in range(col, col + 3):
#                     border_cell = self.ws.cell(row=r, column=c)
#                     border_cell.border = thin_border

#         return start_row + 3





# orders/reports/sheets/unpaid_shipments_sheet.py
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Border, Side
from .base_sheet import BaseSheet
from ..styles.theme import COLORS, BORDERS, FILLS, Alignment
from ..styles.helpers import draw_toc_button
from ..components.footnote import Footnote
from ..components.sheet_title import create_sheet_title


class UnpaidShipmentsSheet(BaseSheet):
    """Лист с заказами, где есть отгрузка, но нет/неполная оплата"""

    def __init__(self, workbook, sheet_number):
        super().__init__(workbook, sheet_number)
        self.sheet_title = create_sheet_title(self.ws)
        
        self.CRITICAL_COLOR = "D32F2F"
        self.HIGH_RISK_COLOR = "F57C00"
        self.MEDIUM_RISK_COLOR = "FBC02D"
        self.LOW_RISK_COLOR = "7CB342"

    def _format_currency(self, value):
        if value is None or value == 0:
            return "0 ₽"
        return f"{int(round(abs(value))):,} ₽".replace(",", " ")

    def _format_number(self, value):
        if value is None or value == 0:
            return "0"
        return f"{int(round(abs(value))):,}".replace(",", " ")

    def _format_percent(self, value):
        if value is None:
            return "0%"
        return f"{value:.1f}%"

    def _safe_float(self, value, default=0.0):
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default

    def _format_datetime(self, date_value):
        if not date_value:
            return ""
        if hasattr(date_value, 'strftime'):
            return date_value.strftime('%Y-%m-%d')
        return str(date_value)[:10]

    def _format_manager_name(self, full_name):
        if not full_name:
            return ""
        parts = str(full_name).strip().split()
        if len(parts) >= 1:
            last_name = parts[0].upper()
            if len(parts) >= 2:
                first_initial = parts[1][0].upper() if parts[1] else ""
                return f"{last_name} {first_initial}."
            return last_name
        return str(full_name).upper()

    def _format_client_name(self, name):
        if not name:
            return ""
        return str(name).upper()

    def _format_payment_summary(self, payment_details):
        """Форматирует краткую сводку по оплатам"""
        if not payment_details:
            return "Нет оплат"
        
        payments_str = []
        for p in payment_details[:3]:
            sign = "-" if p['amount'] < 0 else ""
            payments_str.append(f"{p['date']}: {sign}{self._format_currency(abs(p['amount']))}")
        
        result = "\n".join(payments_str)
        if len(payment_details) > 3:
            result += f"\n+ еще {len(payment_details) - 3}"
        return result

    def build(self, orders_data, summary_data):
        """Построение листа"""
        row = 1

        # Кнопка назад
        draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 8

        # Заголовок
        row = self.sheet_title.draw(
            row=3,
            title="ОТГРУЗКИ БЕЗ ОПЛАТЫ / НЕПОЛНАЯ ОПЛАТА",
            subtitle=f"Заказы с отгрузкой, но оплата меньше суммы отгрузки | Найдено: {summary_data['total_orders']} заказов | Общая сумма недоплаты: {self._format_currency(summary_data['total_underpayment'])}",
            date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            start_col=2,
            end_col=16
        )

        row += 2

        # Заголовок таблицы
        headers = [
            "ЗАКАЗ",
            "ДАТА\nСОЗДАНИЯ",
            "ПОСЛЕДНЯЯ\nОТГРУЗКА",
            "КОЛ-ВО\nОТГРУЗОК",
            "ДНЕЙ БЕЗ\nОПЛАТЫ",
            "СТАТУС",
            "КЛИЕНТ",
            "МЕНЕДЖЕР",
            "МАГАЗИН",
            "СУММА\nЗАКАЗА",
            "ОТГРУЖЕНО\nВСЕГО",
            "ОПЛАЧЕНО",
            "НЕДОПЛАТА",
            "%\nОПЛАТЫ",
            "РИСК",
            "ОПЛАТЫ\n(сводка)",
        ]

        table_start_row = row

        # Рисуем заголовки
        for col_idx, header in enumerate(headers, start=2):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDERS["thin"]

        self.ws.row_dimensions[row].height = 40
        row += 1

        # Данные
        for idx, order in enumerate(orders_data):
            # Чередование фона
            if idx % 2 == 0:
                fill = FILLS.get("odd_row", PatternFill(fill_type=None))
            else:
                fill = FILLS.get("even_row", PatternFill(fill_type=None))

            # Определяем цвет для строки в зависимости от риска
            risk = order.get('risk_category', '')
            if "КРИТИЧНО" in risk:
                row_fill = PatternFill("solid", fgColor="FFEBEE")  # очень светлый красный
            elif "ВЫСОКИЙ" in risk:
                row_fill = PatternFill("solid", fgColor="FFF3E0")  # очень светлый оранжевый
            else:
                row_fill = fill

            values = [
                order.get('order_name', '') or order.get('number', ''),
                self._format_datetime(order.get('date_from')),
                order.get('last_shipment_date', ''),
                order.get('shipments_count', 0),
                f"{order.get('days_without_payment', 0)} дн.",
                order.get('status', ''),
                self._format_client_name(order.get('client', '')),
                self._format_manager_name(order.get('manager', '')),
                (order.get('store', '') or '').upper()[:30],
                self._safe_float(order.get('amount_active')),
                self._safe_float(order.get('total_shiped_amount')),
                self._safe_float(order.get('cash_pmts')),
                order.get('underpayment', 0),
                order.get('payment_percent', 0),
                order.get('risk_category', ''),
                self._format_payment_summary(order.get('payment_details', [])),
            ]

            for col_idx, value in enumerate(values):
                col_num = col_idx + 2
                cell = self.ws.cell(row=row, column=col_num, value=value)
                cell.fill = row_fill
                cell.border = BORDERS["thin"]
                cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))

                # Настройка выравнивания
                if col_idx in [0, 6, 7, 8]:  # ЗАКАЗ, КЛИЕНТ, МЕНЕДЖЕР, МАГАЗИН
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif col_idx in [1, 2, 4, 5]:  # ДАТА СОЗДАНИЯ, ПОСЛЕДНЯЯ ОТГРУЗКА, ДНЕЙ БЕЗ ОПЛАТЫ, СТАТУС
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx == 3:  # КОЛ-ВО ОТГРУЗОК
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if value > 1:
                        cell.font = Font(name="Roboto", size=9, bold=True, color=self.HIGH_RISK_COLOR)
                elif col_idx == 12:  # НЕДОПЛАТА
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.font = Font(name="Roboto", size=9, bold=True, color=self.CRITICAL_COLOR)
                    cell.number_format = '#,##0'
                elif col_idx == 14:  # РИСК
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    if "КРИТИЧНО" in str(value):
                        cell.font = Font(name="Roboto", size=9, bold=True, color=self.CRITICAL_COLOR)
                    elif "ВЫСОКИЙ" in str(value):
                        cell.font = Font(name="Roboto", size=9, bold=True, color=self.HIGH_RISK_COLOR)
                    elif "СРЕДНИЙ" in str(value):
                        cell.font = Font(name="Roboto", size=9, color=self.MEDIUM_RISK_COLOR)
                    else:
                        cell.font = Font(name="Roboto", size=9, color=self.LOW_RISK_COLOR)
                elif col_idx == 15:  # ОПЛАТЫ (сводка)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.font = Font(name="Roboto", size=8, color=COLORS.get("text_dark", "1F1F1F"))
                else:  # Числовые колонки
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(value, (int, float)) and col_idx not in [13]:
                        cell.number_format = '#,##0'

            self.ws.row_dimensions[row].height = 28
            row += 1

        # ИТОГОВАЯ СТРОКА - теперь без голубой заливки
        if orders_data:
            total_shipped = summary_data.get('total_shipped_amount', 0)
            total_underpayment = summary_data.get('total_underpayment', 0)
            total_amount_active = summary_data.get('total_amount_active', 0)
            total_cash_pmts = summary_data.get('total_cash_pmts', 0)
            total_shipments_count = summary_data.get('total_shipments_count', 0)

            # Используем стандартную заливку для итогов
            total_fill = FILLS["total"]
            thin_border = BORDERS["thin"]

            # Пишем "ИТОГО:" в первой колонке
            cell = self.ws.cell(row=row, column=2, value="ИТОГО:")
            cell.fill = total_fill
            cell.border = thin_border
            cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
            cell.alignment = Alignment(horizontal="right", vertical="center")

            # Колонка 4 (D) - КОЛ-ВО ОТГРУЗОК (всего)
            cell = self.ws.cell(row=row, column=4, value=total_shipments_count)
            cell.fill = total_fill
            cell.border = thin_border
            cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = '#,##0'

            # Колонка 10 (J) - Сумма заказа
            cell = self.ws.cell(row=row, column=10, value=total_amount_active)
            cell.fill = total_fill
            cell.border = thin_border
            cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0'

            # Колонка 11 (K) - Отгружено всего
            cell = self.ws.cell(row=row, column=11, value=total_shipped)
            cell.fill = total_fill
            cell.border = thin_border
            cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0'

            # Колонка 12 (L) - Оплачено
            cell = self.ws.cell(row=row, column=12, value=total_cash_pmts)
            cell.fill = total_fill
            cell.border = thin_border
            cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("text_dark", "1F1F1F"))
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0'

            # Колонка 13 (M) - Недоплата (оставляем красной, это важно)
            cell = self.ws.cell(row=row, column=13, value=total_underpayment)
            cell.fill = total_fill
            cell.border = thin_border
            cell.font = Font(name="Roboto", size=10, bold=True, color=self.CRITICAL_COLOR)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0'

            # Заполняем остальные колонки пустыми с заливкой
            for col in [3, 5, 6, 7, 8, 9, 14, 15, 16, 17]:
                cell = self.ws.cell(row=row, column=col, value="")
                cell.fill = total_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            self.ws.row_dimensions[row].height = 30
            row += 2

        # Примечания
        footnote = Footnote(self.ws)
        footnote.draw(row, text="⚠️ КРИТИЧНО: полное отсутствие оплаты при наличии отгрузки")
        row += 1
        footnote.draw(row, text="🟠 ВЫСОКИЙ РИСК: оплачено менее 50% от суммы отгрузки")
        row += 1
        footnote.draw(row, text="🟡 СРЕДНИЙ РИСК: оплачено 50-80% от суммы отгрузки")
        row += 1
        footnote.draw(row, text="🟢 НЕЗНАЧИТЕЛЬНО: оплачено более 80% от суммы отгрузки")
        row += 1
        footnote.draw(row, text="📦 Дни без оплаты считаются от ПОСЛЕДНЕЙ даты отгрузки (из sales_salesdata)")

        # Настройка колонок
        col_widths = {
            "B": 28,  # ЗАКАЗ
            "C": 12,  # ДАТА СОЗДАНИЯ
            "D": 14,  # ПОСЛЕДНЯЯ ОТГРУЗКА
            "E": 12,  # КОЛ-ВО ОТГРУЗОК
            "F": 14,  # ДНЕЙ БЕЗ ОПЛАТЫ
            "G": 22,  # СТАТУС
            "H": 28,  # КЛИЕНТ
            "I": 20,  # МЕНЕДЖЕР
            "J": 22,  # МАГАЗИН
            "K": 16,  # СУММА ЗАКАЗА
            "L": 16,  # ОТГРУЖЕНО ВСЕГО
            "M": 14,  # ОПЛАЧЕНО
            "N": 14,  # НЕДОПЛАТА
            "O": 12,  # % ОПЛАТЫ
            "P": 25,  # РИСК
            "Q": 35,  # ОПЛАТЫ (сводка)
        }
        for col, width in col_widths.items():
            self.ws.column_dimensions[col].width = width

        self.ws.freeze_panes = f'G{table_start_row + 1}'
        if orders_data:
            self.ws.auto_filter.ref = f'B{table_start_row}:Q{row - 5}'
        self.ws.sheet_view.showGridLines = False