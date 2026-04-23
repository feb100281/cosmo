# # orders/reports/sheets/base_sheet.py
# from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
# from ..styles.theme import FILLS, FONTS, BORDERS, ALIGNMENTS, COLORS


# class BaseSheet:
#     def __init__(self, workbook, sheet_number):
#         """sheet_number - это цифра (1, 2, 3...) или "TOC" для оглавления"""
#         self.wb = workbook
#         self.sheet_number = sheet_number
#         # Создаем лист с названием
#         self.ws = self.wb.create_sheet(title=str(sheet_number))
#         self.current_row = 1
    
#     def add_header(self, title, subtitle=None, note=None):
#         """Добавляет стандартный заголовок"""
#         # Ссылка в оглавление (всегда на лист TOC)
#         toc_link = self.ws.cell(row=self.current_row, column=2, value="← ВЕРНУТЬСЯ В ОГЛАВЛЕНИЕ")
#         toc_link.font = Font(name="Roboto", size=9, bold=True, color=COLORS["blue"], underline="single")
#         toc_link.alignment = ALIGNMENTS["left"]
#         toc_link.hyperlink = "#'TOC'!A1"
        
#         self.current_row += 1
        
#         # Заголовок
#         title_cell = self.ws.cell(row=self.current_row, column=2, value=title)
#         title_cell.font = Font(name="Roboto", size=16, bold=True, color=COLORS["dark_green"])
#         self.current_row += 1
        
#         if subtitle:
#             subtitle_cell = self.ws.cell(row=self.current_row, column=2, value=subtitle)
#             subtitle_cell.font = Font(name="Roboto", size=10, color=COLORS["text_gray"])
#             self.current_row += 1
        
#         if note:
#             note_cell = self.ws.cell(row=self.current_row, column=2, value=note)
#             note_cell.font = Font(name="Roboto", size=9, italic=True, color=COLORS["text_gray"])
#             self.current_row += 1
        
#         self.current_row += 1
    
#     def add_toc_button(self):
#         """Добавляет кнопку возврата в оглавление"""
#         toc_link = self.ws.cell(row=self.current_row, column=2, value="← ВЕРНУТЬСЯ В ОГЛАВЛЕНИЕ")
#         toc_link.font = Font(name="Roboto", size=9, bold=True, color=COLORS["blue"], underline="single")
#         toc_link.alignment = ALIGNMENTS["left"]
#         toc_link.hyperlink = "#'TOC'!A1"
#         self.current_row += 2
    
#     def add_section_title(self, title, col_start=1, col_end=5):
#         """Добавляет заголовок секции"""
#         for col in range(col_start, col_end + 1):
#             cell = self.ws.cell(row=self.current_row, column=col)
#             cell.fill = FILLS["section"]
#             cell.border = BORDERS["bottom_thin"]
#             if col == col_start:
#                 cell.value = title
#                 cell.font = FONTS["section"]
#                 cell.alignment = ALIGNMENTS["left"]
#         self.current_row += 1
    
#     def add_table_header(self, headers, start_col=1):
#         """Добавляет заголовки таблицы"""
#         for i, header in enumerate(headers, start=start_col):
#             cell = self.ws.cell(row=self.current_row, column=i)
#             cell.value = header
#             cell.fill = FILLS["header"]
#             cell.font = FONTS["header_white"]
#             cell.alignment = ALIGNMENTS["center"]
#             cell.border = BORDERS["thin"]
#         self.current_row += 1
    
#     def add_data_row(self, values, start_col=1, number_formats=None):
#         """Добавляет строку данных"""
#         for i, value in enumerate(values, start=start_col):
#             cell = self.ws.cell(row=self.current_row, column=i, value=value)
#             cell.font = FONTS["normal"]
#             cell.alignment = ALIGNMENTS["left"] if i == start_col else ALIGNMENTS["right"]
#             cell.border = BORDERS["thin"]
#             cell.fill = FILLS["alt"] if self.current_row % 2 == 0 else FILLS["none"]
            
#             if number_formats and i in number_formats:
#                 cell.number_format = number_formats[i]
#         self.current_row += 1
    
#     def add_blank_row(self, height=6):
#         """Добавляет пустую строку"""
#         self.ws.row_dimensions[self.current_row].height = height
#         self.current_row += 1
    
#     def set_column_widths(self, widths):
#         """Устанавливает ширину колонок"""
#         for col, width in widths.items():
#             self.ws.column_dimensions[col].width = width
    
#     def freeze_and_hide_grid(self, freeze_cell="A8"):
#         """Замораживает панель и скрывает сетку"""
#         self.ws.freeze_panes = freeze_cell
#         self.ws.sheet_view.showGridLines = False




# orders/reports/sheets/base_sheet.py
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from ..styles.theme import FILLS, FONTS, BORDERS, ALIGNMENTS, COLORS, FORMATS
from datetime import datetime


class BaseSheet:
    def __init__(self, workbook, sheet_number):
        self.wb = workbook
        self.sheet_number = sheet_number
        self.ws = self.wb.create_sheet(title=str(sheet_number))
        self.current_row = 1
    
    def add_header(self, title, subtitle=None, note=None):
        """Добавляет заголовок"""
        self.current_row += 1
        title_cell = self.ws.cell(row=self.current_row, column=2, value=title)
        title_cell.font = Font(name="Roboto", size=16, bold=True, color=COLORS["dark_green"])
        
        if subtitle:
            self.current_row += 1
            sub_cell = self.ws.cell(row=self.current_row, column=2, value=subtitle)
            sub_cell.font = Font(name="Roboto", size=10, color=COLORS["text_gray"])
        
        if note:
            self.current_row += 1
            note_cell = self.ws.cell(row=self.current_row, column=2, value=note)
            note_cell.font = Font(name="Roboto", size=9, italic=True, color=COLORS["text_gray"])
        
        self.current_row += 2
    
    def add_toc_button(self):
        """Добавляет кнопку возврата в оглавление"""
        toc_link = self.ws.cell(row=self.current_row, column=2, value="← ВЕРНУТЬСЯ В ОГЛАВЛЕНИЕ")
        toc_link.font = Font(name="Roboto", size=9, bold=True, color=COLORS["blue"], underline="single")
        toc_link.alignment = ALIGNMENTS["left"]
        toc_link.hyperlink = "#'TOC'!A1"
        self.current_row += 2
    
    def add_table_header(self, headers, start_col=2):
        """Добавляет заголовки таблицы"""
        for i, header in enumerate(headers, start=start_col):
            cell = self.ws.cell(row=self.current_row, column=i, value=header)
            cell.font = FONTS["header_white"]
            cell.fill = FILLS["header"]
            cell.alignment = ALIGNMENTS["center"]
            cell.border = BORDERS["thin"]
        self.current_row += 1
    
    def add_data_row(self, values, start_col=2, number_formats=None, highlight_negative=False):
        """Добавляет строку данных"""
        for i, value in enumerate(values, start=start_col):
            cell = self.ws.cell(row=self.current_row, column=i, value=value)
            cell.font = FONTS["normal"]
            cell.border = BORDERS["thin"]
            cell.fill = FILLS["alt"] if self.current_row % 2 == 0 else FILLS["none"]
            
            # Выравнивание
            if i == start_col:
                cell.alignment = ALIGNMENTS["left"]
            else:
                cell.alignment = ALIGNMENTS["right"]
            
            # Форматы чисел
            if number_formats and i in number_formats:
                cell.number_format = number_formats[i]
            
            # Подсветка отрицательных значений
            if highlight_negative and isinstance(value, (int, float)) and value < 0:
                cell.font = Font(color="DC3545")
        
        self.current_row += 1
    
    def add_total_row(self, values, start_col=2):
        """Добавляет итоговую строку"""
        for i, value in enumerate(values, start=start_col):
            cell = self.ws.cell(row=self.current_row, column=i, value=value)
            cell.font = FONTS["total"]
            cell.fill = FILLS["total"]
            cell.border = BORDERS["top_bottom_medium"]
            
            if i == start_col:
                cell.alignment = ALIGNMENTS["left"]
            else:
                cell.alignment = ALIGNMENTS["right"]
        
        self.current_row += 2
    
    def add_blank_row(self, height=6):
        """Добавляет пустую строку"""
        self.ws.row_dimensions[self.current_row].height = height
        self.current_row += 1
    
    def set_column_widths(self, widths):
        """Устанавливает ширину колонок"""
        for col, width in widths.items():
            self.ws.column_dimensions[col].width = width
    
    def freeze_and_hide_grid(self, freeze_cell="A8"):
        """Замораживает панель и скрывает сетку"""
        self.ws.freeze_panes = freeze_cell
        self.ws.sheet_view.showGridLines = False