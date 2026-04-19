# orders/reports/sheets/orders_detail_sheet.py
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from .base_sheet import BaseSheet
from ..styles.theme import COLORS, ALIGNMENTS, BORDERS, FILLS, Alignment
from ..components import create_kpi_cards, create_header, create_table
from ..styles.helpers import draw_toc_button
from ..components.footnote import Footnote
from ..components.sheet_title import create_sheet_title


class OrdersDetailSheet(BaseSheet):
    """Лист с детализацией активных заказов"""

    def __init__(self, workbook, sheet_number):
        super().__init__(workbook, sheet_number)
        self.sheet_title = create_sheet_title(self.ws)
        self.kpi = create_kpi_cards(self.ws)
        self.header = create_header(self.ws)
        self.table = create_table(self.ws)

    def _format_currency(self, value):
        if value is None or value == 0:
            return "0 ₽"
        return f"{int(round(value)):,} ₽".replace(",", " ")

    def _format_number(self, value):
        if value is None or value == 0:
            return "0"
        return f"{int(round(value)):,}".replace(",", " ")

    def _safe_float(self, value, default=0.0):
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default

    def _safe_int(self, value, default=0):
        if value is None:
            return default
        try:
            return int(round(float(value)))
        except (ValueError, TypeError):
            return default

    def _format_datetime(self, date_value):
        """Форматирует дату без времени: YYYY-MM-DD"""
        if not date_value:
            return ""
        if hasattr(date_value, 'strftime'):
            return date_value.strftime('%Y-%m-%d')
        if isinstance(date_value, str):
            # Пробуем распарсить строку с временем
            for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d.%m.%Y']:
                try:
                    dt = datetime.strptime(date_value, fmt)
                    return dt.strftime('%Y-%m-%d')
                except ValueError:
                    continue
        return str(date_value)[:10] if len(str(date_value)) > 10 else str(date_value)

    def _format_manager_name(self, full_name):
        """Форматирует имя менеджера: Иванов Иван -> ИВАНОВ И."""
        if not full_name:
            return ""
        parts = str(full_name).strip().split()
        if len(parts) >= 1:
            last_name = parts[0].upper()
            if len(parts) >= 2:
                first_initial = parts[1][0].upper() if parts[1] else ""
                return f"{last_name} {first_initial}."
            return last_name
        return str(full_name).upper()

    def _format_client_name(self, name):
        """Форматирует имя клиента: ООО Ромашка -> ООО РОМАШКА"""
        if not name:
            return ""
        return str(name).upper()

    def _format_payment_dates(self, payment_dates):
        """Форматирует даты оплат для отображения"""
        if not payment_dates:
            return "—"
        
        # Если это строка
        if isinstance(payment_dates, str):
            import re
            
            # Пробуем найти все вхождения "YYYY-MM-DD: сумма руб"
            pattern = r'(\d{4}-\d{2}-\d{2}):\s*([\d\.]+)\s*руб'
            matches = re.findall(pattern, payment_dates, re.IGNORECASE)
            
            if matches:
                formatted = []
                for date, amount_str in matches[:5]:  # Показываем максимум 5 оплат
                    try:
                        amount = float(amount_str)
                        formatted.append(f"{date}: {self._format_currency(amount)}")
                    except ValueError:
                        formatted.append(f"{date}: {amount_str} руб")
                
                result = "\n".join(formatted)
                if len(matches) > 5:
                    result += f"\n+ еще {len(matches) - 5}"
                return result
            
            # Если не нашли по паттерну, пробуем просто разбить по запятым
            if ',' in payment_dates:
                parts = payment_dates.split(',')
                formatted = []
                for part in parts[:5]:
                    part = part.strip().strip("'\"")
                    if ':' in part:
                        date, amount = part.split(':', 1)
                        amount = amount.replace('руб', '').strip()
                        try:
                            amount_float = float(amount)
                            formatted.append(f"{date.strip()}: {self._format_currency(amount_float)}")
                        except:
                            formatted.append(part)
                    else:
                        formatted.append(part)
                
                result = "\n".join(formatted)
                if len(parts) > 5:
                    result += f"\n+ еще {len(parts) - 5}"
                return result
            
            # Если ничего не подошло, возвращаем как есть
            return payment_dates[:50]
        
        # Если это список
        if isinstance(payment_dates, (list, tuple)):
            formatted = []
            for item in payment_dates[:5]:
                if isinstance(item, str):
                    # Обрабатываем строку из списка
                    if ':' in item:
                        date, amount_part = item.split(':', 1)
                        amount = amount_part.replace('руб', '').strip()
                        try:
                            amount_float = float(amount)
                            formatted.append(f"{date.strip()}: {self._format_currency(amount_float)}")
                        except:
                            formatted.append(item[:30])
                    else:
                        formatted.append(item[:30])
                else:
                    formatted.append(str(item)[:30])
            
            result = "\n".join(formatted)
            if len(payment_dates) > 5:
                result += f"\n+ еще {len(payment_dates) - 5}"
            return result
        
        return str(payment_dates)[:50]

    def build(self, orders_data):
        """Построение листа с детализацией активных заказов"""
        row = 1

        # ============================================================
        # КНОПКА НАЗАД
        # ============================================================
        draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 8

        # ============================================================
        # ШАПКА
        # ============================================================
        row = self.sheet_title.draw(
            row=3,
            title="ДЕТАЛИЗАЦИЯ АКТИВНЫХ ЗАКАЗОВ",
            subtitle="Операционная таблица по заказам в работе: К выполнению / В резерве / На согласовании",
            date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            start_col=2,
            end_col=16
        )

        row += 1


        
        
        
        # ============================================================
        # KPI КАРТОЧКИ
        # ============================================================
        total_orders = len(orders_data)
        total_amount = sum(self._safe_float(o.get('amount_active')) for o in orders_data)
        total_paid = sum(self._safe_float(o.get('cash_pmts')) for o in orders_data)
        total_shipped = sum(self._safe_float(o.get('total_shiped_amount')) for o in orders_data)
        total_delivery = sum(self._safe_float(o.get('amount_delivery')) for o in orders_data)
        avg_check = total_amount / total_orders if total_orders > 0 else 0

        unique_clients = len(set(
            self._format_client_name(o.get('client', '')) 
            for o in orders_data 
            if o.get('client')
        ))

        # Первый ряд KPI
        row1_cards = [
            {
                'title': 'АКТИВНЫЕ ЗАКАЗЫ',
                'value': self._format_number(total_orders),
                'subtitle': 'в работе',
                'color': COLORS["dark_green"],
                'width': 1
            },
            {
                'title': 'СУММА АКТИВНЫХ',
                'value': self._format_currency(total_amount),
                'subtitle': 'по всем активным',
                'color': COLORS["blue"]
            },
            {
                'title': 'СРЕДНИЙ ЧЕК',
                'value': self._format_currency(avg_check),
                'subtitle': 'по активным заказам',
                'color': COLORS.get("warning_red", "C62828"),  
                 'width': 1
            },
        ]

        row = self.kpi.draw_two_rows(row, row1_cards, [])
        row -= 2

        # Второй ряд KPI
        row2_cards = [
            {
                'title': 'ОПЛАЧЕНО',
                'value': self._format_currency(total_paid),
                'subtitle': 'получено оплат',
                'color': COLORS["dark_green"],
                'width': 1
            },
            {
                'title': 'ОТГРУЖЕНО',
                'value': self._format_currency(total_shipped),
                'subtitle': 'на сумму',
                'color': COLORS["blue"]
            },
            {
                'title': 'УСЛУГИ (ДОСТАВКА)',
                'value': self._format_currency(total_delivery),
                'subtitle': 'в том числе',
                'color': COLORS["dark_green"],
                 'width': 1
            },
        ]

        row = self.kpi.draw_two_rows(row, row2_cards, [])
        row -= 2
        row += 1

        # ============================================================
        # ДОПОЛНИТЕЛЬНАЯ ИНФОРМАЦИЯ
        # ============================================================
        # Аккуратно выводим количество уникальных клиентов
        row = self.header.draw(row, "ДОПОЛНИТЕЛЬНАЯ ИНФОРМАЦИЯ", start_col=2, end_col=16)

        # Создаем аккуратную строку с уникальными клиентами
        clients_text = f"Уникальных клиентов в активных заказах: {self._format_number(unique_clients)}"
        cell = self.ws.cell(row=row, column=2, value=clients_text)
        cell.font = Font(name="Roboto", size=10, color=COLORS.get("text_gray", "404040"))
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDERS["thin"]
        cell.fill = PatternFill("solid", fgColor=COLORS.get("light_gray", "F8F9FA"))

        # Объединяем ячейки для красоты
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=16)
        self.ws.row_dimensions[row].height = 30

        row += 1  # Один отступ после дополнительной информации вместо 2

        # ============================================================
        # ЗАГОЛОВОК ТАБЛИЦЫ
        # ============================================================
        row = self.header.draw(row, "ДЕТАЛЬНАЯ ИНФОРМАЦИЯ ПО ЗАКАЗАМ", start_col=2, end_col=16)
        row += 1

        # ============================================================
        # ЗАГОЛОВКИ ТАБЛИЦЫ
        # ============================================================
        headers = [
                "ЗАКАЗ",
                "ДАТА\nСОЗДАНИЯ",
                "ДАТА\nИЗМЕНЕНИЯ",
                "СТАТУС",
                "ИЗМЕНЕНИЯ",
                "КЛИЕНТ",
                "МЕНЕДЖЕР",
                "МАГАЗИН",
                "КОЛ-ВО\nЗАКАЗАНО",
                "СУММА\nЗАКАЗА",
                "В Т.Ч.\nУСЛУГИ",
                "ОТГРУЖЕНО\nРУБ",
                "ОПЛАЧЕНО",
                "К ОПЛАТЕ",
                "ДАТЫ ОПЛАТ",
            ]

        # Рисуем таблицу
        table_start_row = row
        
        # Рисуем заголовки вручную для лучшего контроля
        for col_idx, header in enumerate(headers, start=2):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
            cell.fill = FILLS["header_blue"] if 'header_blue' in FILLS else PatternFill("solid", fgColor=COLORS["dark_green"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDERS["thin"]
        
        self.ws.row_dimensions[row].height = 40
        row += 1

        # ============================================================
        # ДАННЫЕ
        # ============================================================
        for idx, order in enumerate(orders_data):
            # Чередование фона
            if idx % 2 == 0:
                fill = FILLS.get("odd_row", PatternFill(fill_type=None))
            else:
                fill = FILLS.get("even_row", PatternFill(fill_type=None))

            # Получаем данные
            order_display = order.get('order_name', '') or order.get('number', '')
            
            values = [
                    order.get('order_name', '') or order.get('number', ''),  # ЗАКАЗ
                    self._format_datetime(order.get('date_from')),  # ДАТА СОЗДАНИЯ
                    self._format_datetime(order.get('update_at')),  # ДАТА ИЗМЕНЕНИЯ
                    order.get('status', ''),  # СТАТУС
                    order.get('change_status', 'Без изменений'),  # ИЗМЕНЕНИЯ
                    self._format_client_name(order.get('client', '')),  # КЛИЕНТ
                    self._format_manager_name(order.get('manager', '')),  # МЕНЕДЖЕР
                    (order.get('store', '') or '').upper()[:30],  # МАГАЗИН
                    self._safe_int(order.get('order_qty')),  # КОЛ-ВО ЗАКАЗАНО (итого к заказу)
                    self._safe_float(order.get('amount_active')),  # СУММА ЗАКАЗА (итого сумма)
                    self._safe_float(order.get('amount_delivery')),  # В Т.Ч. УСЛУГИ (доставка)
                    self._safe_float(order.get('total_shiped_amount')),  # ОТГРУЖЕНО РУБ
                    self._safe_float(order.get('cash_pmts')),  # ОПЛАЧЕНО
                    self._safe_float(order.get('remaining_amount', 0)),  # К ОПЛАТЕ
                    self._format_payment_dates(order.get('payment_dates')),  # ДАТЫ ОПЛАТ
                ]
            # Записываем значения в ячейки
            for col_idx, value in enumerate(values):
                col_num = col_idx + 2
                cell = self.ws.cell(row=row, column=col_num, value=value)
                cell.fill = fill
                cell.border = BORDERS["thin"]
                cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))

                # Настройка выравнивания
                if col_idx in [0, 5, 6, 7]:  # ЗАКАЗ, КЛИЕНТ, МЕНЕДЖЕР, МАГАЗИН
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif col_idx in [1, 2, 3, 4]:  # ДАТЫ и СТАТУСЫ
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx == 14:  # ДАТЫ ОПЛАТ - включаем перенос текста
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.font = Font(name="Roboto", size=8)  # Можно чуть меньше шрифт
                else:  # Числовые колонки
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0'

                # Специальная обработка статуса
                if col_idx == 3:  # СТАТУС
                    status = value
                    if "К выполнению" in status or "В резерве" in status:
                        cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
                    elif "На согласовании" in status:
                        cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS.get("warning_orange", "F57C00"))

            self.ws.row_dimensions[row].height = 28
            row += 1

        # ============================================================
        # ИТОГОВАЯ СТРОКА
        # ============================================================
        total_order_qty = sum(self._safe_int(o.get('order_qty')) for o in orders_data)
        total_amount_sum = sum(self._safe_float(o.get('amount_active')) for o in orders_data)
        total_delivery_sum = sum(self._safe_float(o.get('amount_delivery')) for o in orders_data)
        total_shipped_sum = sum(self._safe_float(o.get('total_shiped_amount')) for o in orders_data)
        total_paid_sum = sum(self._safe_float(o.get('cash_pmts')) for o in orders_data)
        total_remaining_sum = sum(self._safe_float(o.get('remaining_amount', 0)) for o in orders_data)

        total_row = [
            "ИТОГО:",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            total_order_qty,
            total_amount_sum,
            total_delivery_sum,
            total_shipped_sum,
            total_paid_sum,
            total_remaining_sum,
            "",
        ]

        fill = FILLS["total"]
        for col_idx, value in enumerate(total_row):
            cell = self.ws.cell(row=row, column=col_idx + 2, value=value)
            cell.fill = fill
            cell.border = BORDERS["thin"]
            cell.font = Font(name="Roboto", bold=True, size=10)

            if col_idx == 0:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx >= 8 and col_idx <= 13:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        self.ws.row_dimensions[row].height = 30
        row += 2

        # ============================================================
        # ПРИМЕЧАНИЕ
        # ============================================================
        footnote = Footnote(self.ws)
        footnote.draw(
            row=row,
            text="* Статусы: К выполнению / В резерве - зеленым, На согласовании - оранжевым"
        )

        # ============================================================
        # НАСТРОЙКА КОЛОНОК
        # ============================================================
        self.ws.column_dimensions["A"].width = 2
        self.ws.column_dimensions["B"].width = 28  # ЗАКАЗ
        self.ws.column_dimensions["C"].width = 12  # ДАТА СОЗДАНИЯ
        self.ws.column_dimensions["D"].width = 12  # ДАТА ИЗМЕНЕНИЯ
        self.ws.column_dimensions["E"].width = 22  # СТАТУС
        self.ws.column_dimensions["F"].width = 14  # ИЗМЕНЕНИЯ
        self.ws.column_dimensions["G"].width = 28  # КЛИЕНТ
        self.ws.column_dimensions["H"].width = 20  # МЕНЕДЖЕР
        self.ws.column_dimensions["I"].width = 22  # МАГАЗИН
        self.ws.column_dimensions["J"].width = 12  # КОЛ-ВО ЗАКАЗАНО
        self.ws.column_dimensions["K"].width = 16  # СУММА ЗАКАЗА
        self.ws.column_dimensions["L"].width = 14  # В Т.Ч. УСЛУГИ
        self.ws.column_dimensions["M"].width = 14  # ОТГРУЖЕНО РУБ
        self.ws.column_dimensions["N"].width = 14  # ОПЛАЧЕНО
        self.ws.column_dimensions["O"].width = 14  # К ОПЛАТЕ
        self.ws.column_dimensions["P"].width = 22  # ДАТЫ ОПЛАТ

        # Заморозка панелей: замораживаем до колонки СТАТУС (колонка E)
        # Замораживаем строку заголовка + 1, и колонки до E
        self.ws.freeze_panes = f'E{table_start_row + 1}'
        
        # Автофильтр
        self.ws.auto_filter.ref = f'B{table_start_row}:P{row - 2}'
        
        # Скрываем сетку
        self.ws.sheet_view.showGridLines = False


# Добавляем импорт PatternFill если нужен
from openpyxl.styles import PatternFill