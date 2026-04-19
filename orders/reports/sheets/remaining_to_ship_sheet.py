# # orders/reports/sheets/remaining_to_ship_sheet.py
# from openpyxl.styles import Alignment, Font, PatternFill
# from openpyxl.utils import get_column_letter

# from .base_sheet import BaseSheet
# from ..styles.helpers import draw_toc_button
# from ..components import create_kpi_cards, create_header, create_table
# from ..components.sheet_title import create_sheet_title
# from ..components.navigation_link import NavigationLink
# from ..styles.theme import COLORS, BORDERS


# class RemainingToShipSheet(BaseSheet):
#     """Остатки к отгрузке по категориям"""

#     def __init__(self, workbook, sheet_number, request=None):
#         super().__init__(workbook, sheet_number)
#         self.sheet_title = create_sheet_title(self.ws)
#         self.kpi = create_kpi_cards(self.ws)
#         self.header = create_header(self.ws)
#         self.table = create_table(self.ws)
#         self.request = request

#     def _format_currency(self, value):
#         """Форматирование суммы"""
#         if value is None or value == 0:
#             return "0 ₽"
#         return f"{int(round(abs(value))):,} ₽".replace(",", " ")

#     def _format_number_str(self, value):
#         """Форматирование числа"""
#         if value is None or value == 0:
#             return "0"
#         return f"{int(round(abs(value))):,}".replace(",", " ")

#     def _get_cell_font(self, is_header=False, is_bold=False):
#         if is_header:
#             return Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#         if is_bold:
#             return Font(name="Roboto", size=9, bold=True, color=COLORS["text_dark"])
#         return Font(name="Roboto", size=9, color=COLORS["text_dark"])

#     def build(self, remaining_data, summary, parent_cat_summary):
#         """Построение листа остатков к отгрузке"""
        
#         row = 1
        
#         # Шапка
#         draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")
#         self.ws.row_dimensions[1].height = 20
#         self.ws.row_dimensions[2].height = 8
#         self.ws.column_dimensions["A"].width = 2
        
#         from datetime import datetime
#         current_date = datetime.now()
        
#         # Заголовок
#         row = self.sheet_title.draw(
#             row=3,
#             title="ОСТАТКИ К ОТГРУЗКЕ ПО КАТЕГОРИЯМ",
#             subtitle="Товары в незакрытых заказах (штуки и рубли)",
#             date_text=f"Сформировано: {current_date.strftime('%d.%m.%Y в %H:%M')}",
#             start_col=2,
#             end_col=8
#         )
#         row += 2
        
#         # ============================================================
#         # KPI КАРТОЧКИ
#         # ============================================================
#         row_cards = [
#             {
#                 'title': 'ВСЕГО К ОТГРУЗКЕ (шт)',
#                 'value': self._format_number_str(summary['total_qty']),
#                 'subtitle': f"в {self._format_number_str(summary['total_orders'])} заказах",
#                 'color': COLORS["blue"],
#                 'width': 3,
#             },
#             {
#                 'title': 'ВСЕГО К ОТГРУЗКЕ (₽)',
#                 'value': self._format_currency(summary['total_amount']),
#                 'subtitle': f"{self._format_number_str(summary['total_items'])} позиций",
#                 'color': COLORS["blue"],
#                 'width': 4,
#             },
#         ]
        
#         row = self.kpi.draw_row(row, row_cards)
#         row += 2
        
#         # ============================================================
#         # СВОДКА ПО РОДИТЕЛЬСКИМ КАТЕГОРИЯМ
#         # ============================================================
#         if parent_cat_summary:
#             row = self.header.draw(row, "СВОДКА ПО РОДИТЕЛЬСКИМ КАТЕГОРИЯМ", end_col=5)
            
#             headers = ["РОДИТЕЛЬСКАЯ КАТЕГОРИЯ", "КОЛИЧЕСТВО (шт)", "СУММА (₽)", "ЗАКАЗОВ"]
            
#             for col_idx, header in enumerate(headers, start=2):
#                 cell = self.ws.cell(row=row, column=col_idx, value=header)
#                 cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#                 cell.fill = PatternFill(
#                     start_color=COLORS["dark_green"],
#                     end_color=COLORS["dark_green"],
#                     fill_type="solid"
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
            
#             row += 1
#             start_data_row = row
            
#             for cat in parent_cat_summary:
#                 if not cat['parent_cat']:
#                     continue
                    
#                 cell_name = self.ws.cell(row=row, column=2, value=cat['parent_cat'])
#                 cell_name.font = self._get_cell_font()
                
#                 cell_qty = self.ws.cell(row=row, column=3, value=cat['total_qty'])
#                 cell_qty.font = self._get_cell_font()
#                 cell_qty.number_format = '#,##0'
#                 cell_qty.alignment = Alignment(horizontal="right")
                
#                 cell_amount = self.ws.cell(row=row, column=4, value=cat['total_amount'])
#                 cell_amount.font = self._get_cell_font()
#                 cell_amount.number_format = '#,##0 ₽'
#                 cell_amount.alignment = Alignment(horizontal="right")
                
#                 cell_orders = self.ws.cell(row=row, column=5, value=cat['order_count'])
#                 cell_orders.font = self._get_cell_font()
#                 cell_orders.number_format = '#,##0'
#                 cell_orders.alignment = Alignment(horizontal="right")
                
#                 row += 1
            
#             end_data_row = row - 1
            
#             for col in range(2, 6):
#                 for r in range(start_data_row, end_data_row + 1):
#                     self.ws.cell(row=r, column=col).border = BORDERS["thin"]
            
#             row += 2
        
#         # ============================================================
#         # ДЕТАЛЬНАЯ РАЗБИВКА ПО КАТЕГОРИЯМ
#         # ============================================================
#         if remaining_data:
#             row = self.header.draw(row, "ДЕТАЛЬНАЯ РАЗБИВКА ПО КАТЕГОРИЯМ", end_col=6)
            
#             headers = ["РОДИТЕЛЬСКАЯ", "КАТЕГОРИЯ", "ПОДКАТЕГОРИЯ", "КОЛИЧЕСТВО (шт)", "СУММА (₽)", "ЗАКАЗОВ"]
            
#             for col_idx, header in enumerate(headers, start=2):
#                 cell = self.ws.cell(row=row, column=col_idx, value=header)
#                 cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#                 cell.fill = PatternFill(
#                     start_color=COLORS["dark_green"],
#                     end_color=COLORS["dark_green"],
#                     fill_type="solid"
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
            
#             row += 1
#             start_data_row = row
            
#             for item in remaining_data:
#                 # Родительская категория
#                 cell_parent = self.ws.cell(row=row, column=2, value=item['parent_cat'] or '—')
#                 cell_parent.font = self._get_cell_font()
                
#                 # Категория
#                 cell_cat = self.ws.cell(row=row, column=3, value=item['cat'] or '—')
#                 cell_cat.font = self._get_cell_font()
                
#                 # Подкатегория
#                 cell_subcat = self.ws.cell(row=row, column=4, value=item['subcat'] or '—')
#                 cell_subcat.font = self._get_cell_font()
                
#                 # Количество
#                 cell_qty = self.ws.cell(row=row, column=5, value=item['total_qty'])
#                 cell_qty.font = self._get_cell_font(is_bold=(item['total_qty'] > 100))
#                 cell_qty.number_format = '#,##0'
#                 cell_qty.alignment = Alignment(horizontal="right")
                
#                 # Сумма
#                 cell_amount = self.ws.cell(row=row, column=6, value=item['total_amount'])
#                 cell_amount.font = self._get_cell_font(is_bold=(item['total_amount'] > 1000000))
#                 cell_amount.number_format = '#,##0 ₽'
#                 cell_amount.alignment = Alignment(horizontal="right")
                
#                 # Количество заказов
#                 cell_orders = self.ws.cell(row=row, column=7, value=item['order_count'])
#                 cell_orders.font = self._get_cell_font()
#                 cell_orders.number_format = '#,##0'
#                 cell_orders.alignment = Alignment(horizontal="right")
                
#                 row += 1
            
#             end_data_row = row - 1
            
#             for col in range(2, 8):
#                 for r in range(start_data_row, end_data_row + 1):
#                     self.ws.cell(row=r, column=col).border = BORDERS["thin"]
            
#             row += 2
        
#         # ============================================================
#         # МЕТОДОЛОГИЯ
#         # ============================================================
#         row = self.header.draw(row, "МЕТОДОЛОГИЯ", end_col=6)
#         row += 1
        
#         methodology_texts = [
#             "• Учитываются только НЕЗАКРЫТЫЕ заказы",
#             "• Показываются остатки к отгрузке в штуках и рублях",
#             "• Сумма рассчитывается как amount из mv_orders_items",
#         ]
        
#         for text in methodology_texts:
#             cell = self.ws.cell(row=row, column=2, value=text)
#             cell.font = Font(name="Roboto", size=8, color=COLORS["text_gray"])
#             row += 1
        
#         row += 1
        
#         # ============================================================
#         # ССЫЛКА НА ОГЛАВЛЕНИЕ
#         # ============================================================
#         nav_link = NavigationLink(self.ws)
#         row = nav_link.draw(
#             row=row,
#             text="← Вернуться к оглавлению",
#             target_sheet="TOC",
#             target_cell="A1",
#             start_col=2,
#             end_col=6,
#             alignment="left",
#             with_icon=True,
#             icon="📊"
#         )
        
#         # Настройка колонок
#         self.ws.column_dimensions["B"].width = 25
#         self.ws.column_dimensions["C"].width = 25
#         self.ws.column_dimensions["D"].width = 25
#         self.ws.column_dimensions["E"].width = 18
#         self.ws.column_dimensions["F"].width = 18
#         self.ws.column_dimensions["G"].width = 12
        
#         self.ws.sheet_view.showGridLines = False




# orders/reports/sheets/remaining_to_ship_sheet.py
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

from .base_sheet import BaseSheet
from ..styles.helpers import draw_toc_button
from ..components import create_kpi_cards, create_header, create_table
from ..components.sheet_title import create_sheet_title
from ..components.navigation_link import NavigationLink
from ..styles.theme import COLORS, BORDERS


class RemainingToShipSheet(BaseSheet):
    """Остатки к отгрузке по категориям"""

    def __init__(self, workbook, sheet_number, request=None):
        super().__init__(workbook, sheet_number)
        self.sheet_title = create_sheet_title(self.ws)
        self.kpi = create_kpi_cards(self.ws)
        self.header = create_header(self.ws)
        self.table = create_table(self.ws)
        self.request = request

    def _format_currency(self, value):
        """Форматирование суммы"""
        if value is None or value == 0:
            return "0 ₽"
        return f"{int(round(abs(float(value)))):,} ₽".replace(",", " ")

    def _format_number_str(self, value):
        """Форматирование числа"""
        if value is None or value == 0:
            return "0"
        return f"{int(round(abs(float(value)))):,}".replace(",", " ")

    def _get_cell_font(self, is_header=False, is_bold=False):
        if is_header:
            return Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
        if is_bold:
            return Font(name="Roboto", size=9, bold=True, color=COLORS["text_dark"])
        return Font(name="Roboto", size=9, color=COLORS["text_dark"])

    def build(self, remaining_data, summary, parent_cat_summary):
        """Построение листа остатков к отгрузке"""
        
        row = 1
        
        # Шапка
        draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")
        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 8
        self.ws.column_dimensions["A"].width = 2
        
        current_date = datetime.now()
        
        # Заголовок
        row = self.sheet_title.draw(
            row=3,
            title="ОСТАТКИ К ОТГРУЗКЕ ПО КАТЕГОРИЯМ",
            subtitle="Товары в незакрытых заказах (штуки и рубли)",
            date_text=f"Сформировано: {current_date.strftime('%d.%m.%Y в %H:%M')}",
            start_col=2,
            end_col=8
        )
        row += 2
        
        # ============================================================
        # KPI КАРТОЧКИ
        # ============================================================
        row_cards = [
            {
                'title': 'ВСЕГО К ОТГРУЗКЕ (шт)',
                'value': self._format_number_str(summary.get('total_qty', 0)),
                'subtitle': f"в {self._format_number_str(summary.get('total_orders', 0))} заказах",
                'color': COLORS["blue"],
                'width': 3,
            },
            {
                'title': 'ВСЕГО К ОТГРУЗКЕ (₽)',
                'value': self._format_currency(summary.get('total_amount', 0)),
                'subtitle': f"{self._format_number_str(summary.get('total_items', 0))} позиций",
                'color': COLORS["blue"],
                'width': 4,
            },
        ]
        
        row = self.kpi.draw_row(row, row_cards)
        row += 2
        
        # ============================================================
        # СВОДКА ПО РОДИТЕЛЬСКИМ КАТЕГОРИЯМ
        # ============================================================
        if parent_cat_summary:
            row = self.header.draw(row, "СВОДКА ПО РОДИТЕЛЬСКИМ КАТЕГОРИЯМ", end_col=5)
            
            headers = ["РОДИТЕЛЬСКАЯ КАТЕГОРИЯ", "КОЛИЧЕСТВО (шт)", "СУММА (₽)", "ЗАКАЗОВ"]
            
            for col_idx, header in enumerate(headers, start=2):
                cell = self.ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
                cell.fill = PatternFill(
                    start_color=COLORS["dark_green"],
                    end_color=COLORS["dark_green"],
                    fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Ширина колонок
                if col_idx == 2:
                    self.ws.column_dimensions[get_column_letter(col_idx)].width = 30
                else:
                    self.ws.column_dimensions[get_column_letter(col_idx)].width = 18
            
            row += 1
            start_data_row = row
            
            # Фильтруем пустые категории и сортируем по сумме
            filtered_cats = [cat for cat in parent_cat_summary if cat.get('parent_cat')]
            
            for cat in filtered_cats:
                cell_name = self.ws.cell(row=row, column=2, value=cat['parent_cat'])
                cell_name.font = self._get_cell_font()
                
                cell_qty = self.ws.cell(row=row, column=3, value=float(cat.get('total_qty') or 0))
                cell_qty.font = self._get_cell_font()
                cell_qty.number_format = '#,##0'
                cell_qty.alignment = Alignment(horizontal="right")
                
                cell_amount = self.ws.cell(row=row, column=4, value=float(cat.get('total_amount') or 0))
                cell_amount.font = self._get_cell_font(is_bold=(float(cat.get('total_amount') or 0) > 1000000))
                cell_amount.number_format = '#,##0 ₽'
                cell_amount.alignment = Alignment(horizontal="right")
                
                cell_orders = self.ws.cell(row=row, column=5, value=cat.get('order_count') or 0)
                cell_orders.font = self._get_cell_font()
                cell_orders.number_format = '#,##0'
                cell_orders.alignment = Alignment(horizontal="right")
                
                row += 1
            
            end_data_row = row - 1
            
            # Рисуем границы
            for col in range(2, 6):
                for r in range(start_data_row, end_data_row + 1):
                    self.ws.cell(row=r, column=col).border = BORDERS["thin"]
            
            row += 2
        
        # ============================================================
        # ДЕТАЛЬНАЯ РАЗБИВКА ПО КАТЕГОРИЯМ
        # ============================================================
        if remaining_data:
            row = self.header.draw(row, "ДЕТАЛЬНАЯ РАЗБИВКА ПО КАТЕГОРИЯМ", end_col=7)
            
            headers = ["РОДИТЕЛЬСКАЯ", "КАТЕГОРИЯ", "ПОДКАТЕГОРИЯ", "КОЛИЧЕСТВО (шт)", "СУММА (₽)", "ЗАКАЗОВ"]
            
            for col_idx, header in enumerate(headers, start=2):
                cell = self.ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
                cell.fill = PatternFill(
                    start_color=COLORS["dark_green"],
                    end_color=COLORS["dark_green"],
                    fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Ширина колонок
                if col_idx in [2, 3, 4]:
                    self.ws.column_dimensions[get_column_letter(col_idx)].width = 25
                else:
                    self.ws.column_dimensions[get_column_letter(col_idx)].width = 18
            
            row += 1
            start_data_row = row
            
            for item in remaining_data:
                # Родительская категория
                cell_parent = self.ws.cell(row=row, column=2, value=item.get('parent_cat') or '—')
                cell_parent.font = self._get_cell_font()
                
                # Категория
                cell_cat = self.ws.cell(row=row, column=3, value=item.get('cat') or '—')
                cell_cat.font = self._get_cell_font()
                
                # Подкатегория
                cell_subcat = self.ws.cell(row=row, column=4, value=item.get('subcat') or '—')
                cell_subcat.font = self._get_cell_font()
                
                # Количество
                qty = float(item.get('total_qty') or 0)
                cell_qty = self.ws.cell(row=row, column=5, value=qty)
                cell_qty.font = self._get_cell_font(is_bold=(qty > 100))
                cell_qty.number_format = '#,##0'
                cell_qty.alignment = Alignment(horizontal="right")
                
                # Сумма
                amount = float(item.get('total_amount') or 0)
                cell_amount = self.ws.cell(row=row, column=6, value=amount)
                cell_amount.font = self._get_cell_font(is_bold=(amount > 1000000))
                cell_amount.number_format = '#,##0 ₽'
                cell_amount.alignment = Alignment(horizontal="right")
                
                # Количество заказов
                cell_orders = self.ws.cell(row=row, column=7, value=item.get('order_count') or 0)
                cell_orders.font = self._get_cell_font()
                cell_orders.number_format = '#,##0'
                cell_orders.alignment = Alignment(horizontal="right")
                
                row += 1
            
            end_data_row = row - 1
            
            # Рисуем границы
            for col in range(2, 8):
                for r in range(start_data_row, end_data_row + 1):
                    self.ws.cell(row=r, column=col).border = BORDERS["thin"]
            
            row += 2
        
        # ============================================================
        # МЕТОДОЛОГИЯ
        # ============================================================
        row = self.header.draw(row, "МЕТОДОЛОГИЯ", end_col=6)
        row += 1
        
        methodology_texts = [
            "• Учитываются только НЕЗАКРЫТЫЕ заказы",
            "• Показываются остатки к отгрузке в штуках и рублях",
            "• Сумма рассчитывается как amount из mv_orders_items",
            "• В сумму включены только активные позиции (без отмененных)",
        ]
        
        for text in methodology_texts:
            cell = self.ws.cell(row=row, column=2, value=text)
            cell.font = Font(name="Roboto", size=8, color=COLORS["text_gray"])
            cell.alignment = Alignment(horizontal="left", vertical="center")
            row += 1
        
        row += 1
        
        # ============================================================
        # ССЫЛКА НА ОГЛАВЛЕНИЕ
        # ============================================================
        nav_link = NavigationLink(self.ws)
        row = nav_link.draw(
            row=row,
            text="← Вернуться к оглавлению",
            target_sheet="TOC",
            target_cell="A1",
            start_col=2,
            end_col=6,
            alignment="left",
            with_icon=True,
            icon="📊"
        )
        
        # Настройка колонок (если еще не настроены)
        if self.ws.column_dimensions["B"].width != 25:
            self.ws.column_dimensions["B"].width = 25
        if self.ws.column_dimensions["C"].width != 25:
            self.ws.column_dimensions["C"].width = 25
        if self.ws.column_dimensions["D"].width != 25:
            self.ws.column_dimensions["D"].width = 25
        if self.ws.column_dimensions["E"].width != 18:
            self.ws.column_dimensions["E"].width = 18
        if self.ws.column_dimensions["F"].width != 18:
            self.ws.column_dimensions["F"].width = 18
        if self.ws.column_dimensions["G"].width != 12:
            self.ws.column_dimensions["G"].width = 12
        
        self.ws.sheet_view.showGridLines = False