# # orders/reports/sheets/daily_payments_sheet.py
# from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
# from openpyxl.utils import get_column_letter

# from .base_sheet import BaseSheet
# from ..styles.helpers import draw_toc_button
# from ..components import create_kpi_cards, create_header, create_table
# from ..components.sheet_title import create_sheet_title
# from ..components.navigation_link import NavigationLink
# from ..styles.theme import COLORS, BORDERS


# class DailyPaymentsSheet(BaseSheet):
#     """Подневная аналитика оплат по магазинам"""

#     def __init__(self, workbook, sheet_number, request=None):
#         super().__init__(workbook, sheet_number)
#         self.sheet_title = create_sheet_title(self.ws)
#         self.kpi = create_kpi_cards(self.ws)
#         self.header = create_header(self.ws)
#         self.table = create_table(self.ws)
#         self.request = request

#     def _format_currency(self, value):
#         """Форматирование суммы"""
#         if value is None or value == 0:
#             return "0 ₽"
#         sign = "-" if value < 0 else ""
#         formatted = f"{int(round(abs(value))):,} ₽".replace(",", " ")
#         return f"{sign}{formatted}"

#     def _format_number_str(self, value):
#         """Форматирование числа"""
#         if value is None or value == 0:
#             return "0"
#         sign = "-" if value < 0 else ""
#         return f"{sign}{int(round(abs(value))):,}".replace(",", " ")

#     def _format_date(self, date_value):
#         """Форматирование даты"""
#         if not date_value:
#             return ""
#         if hasattr(date_value, 'strftime'):
#             return date_value.strftime('%d.%m.%Y')
#         return str(date_value)

#     def _get_cell_font(self, is_header=False, is_bold=False, is_negative=False, is_warning=False):
#         """Возвращает стандартный шрифт для ячеек"""
#         if is_header:
#             return Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#         if is_negative:
#             return Font(name="Roboto", size=9, bold=True, color=COLORS["warning_red"])
#         if is_warning:
#             return Font(name="Roboto", size=9, bold=True, color=COLORS["warning_orange"])
#         if is_bold:
#             return Font(name="Roboto", size=9, bold=True, color=COLORS["text_dark"])
#         return Font(name="Roboto", size=9, color=COLORS["text_dark"])

#     def _get_conditional_fill(self, amount):
#         """Условное форматирование для сумм"""
#         if amount < 0:
#             return PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
#         elif amount > 100000:
#             return PatternFill(start_color="E5F3E5", end_color="E5F3E5", fill_type="solid")
#         return None

#     def build(self, daily_data):
#         """Построение листа подневной аналитики - даты в строках, магазины в колонках"""
        
#         if not daily_data['data']:
#             self._build_empty_state()
#             return
        
#         row = 1
        
#         # Шапка
#         draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")
#         self.ws.row_dimensions[1].height = 20
#         self.ws.row_dimensions[2].height = 8
#         self.ws.column_dimensions["A"].width = 2
        
#         # Заголовок
#         row = self.sheet_title.draw(
#             row=3,
#             title="ПОДНЕВНАЯ ДИНАМИКА ОПЛАТ ПО МАГАЗИНАМ",
#             subtitle=f"Детальная разбивка по дням с {daily_data['start_date'].strftime('%d.%m.%Y')} по {daily_data['end_date'].strftime('%d.%m.%Y')}",
#             date_text=f"Сформировано: {self.request.user.username if self.request and hasattr(self.request, 'user') else 'System'} | {daily_data['end_date'].strftime('%d.%m.%Y')}",
#             start_col=2,
#             end_col=10
#         )
#         row += 2
        
#         # ============================================================
#         # KPI КАРТОЧКИ
#         # ============================================================
#         total_stores = len(daily_data['data'])
#         total_days = len(daily_data['dates'])
        
#         # Считаем общую сумму и кол-во операций
#         total_amount = 0
#         total_operations = 0
#         store_totals = {}  # Суммы по магазинам для итоговой строки
        
#         for store, days in daily_data['data'].items():
#             store_total = 0
#             for date, metrics in days.items():
#                 store_total += metrics['total_amount']
#                 total_amount += metrics['total_amount']
#                 total_operations += metrics['total_count']
#             store_totals[store] = store_total
        
#         row_cards = [
#             {
#                 'title': 'ВСЕГО МАГАЗИНОВ',
#                 'value': self._format_number_str(total_stores),
#                 'subtitle': f"активных магазинов",
#                 'color': COLORS["blue"],
#                 'width': 3,
#             },
#             {
#                 'title': 'ОБЩАЯ СУММА ОПЛАТ',
#                 'value': self._format_currency(total_amount),
#                 'subtitle': f"{self._format_number_str(total_operations)} операций всего",
#                 'color': COLORS["blue"],
#                 'width': 3,
#             },
#             {
#                 'title': 'ПЕРИОД АНАЛИЗА',
#                 'value': f"{total_days} дней",
#                 'subtitle': f"с {daily_data['start_date'].strftime('%d.%m')} по {daily_data['end_date'].strftime('%d.%m.%Y')}",
#                 'color': COLORS["blue"],
#                 'width': 3,
#             },
#         ]
        
#         row = self.kpi.draw_row(row, row_cards)
#         row += 2
        
#         # ============================================================
#         # ТАБЛИЦА: ДАТЫ (строки) → МАГАЗИНЫ (колонки)
#         # ============================================================
        
#         # Получаем отсортированный список магазинов
#         stores_sorted = sorted(daily_data['data'].keys())
#         num_stores = len(stores_sorted)
        
#         # Заголовок таблицы
#         row = self.header.draw(
#             row,
#             "ДЕТАЛЬНАЯ РАЗБИВКА ПО ДНЯМ",
#             end_col=3 + num_stores
#         )
        
#         # Формируем заголовки: ДАТА | МАГАЗИН 1 | МАГАЗИН 2 | ... | ИТОГО ЗА ДЕНЬ
#         headers = ["ДАТА"] + [self._format_store_name(s) for s in stores_sorted] + ["ИТОГО ЗА ДЕНЬ"]
#         start_col = 2
        
#         # Рисуем заголовки
#         for col_idx, header in enumerate(headers, start_col):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill(
#                 start_color=COLORS["dark_green"],
#                 end_color=COLORS["dark_green"],
#                 fill_type="solid"
#             )
#             cell.alignment = Alignment(horizontal="center", vertical="center")
            
#             # Ширина колонок
#             if col_idx == start_col:
#                 self.ws.column_dimensions[get_column_letter(col_idx)].width = 14
#             elif col_idx == start_col + num_stores:
#                 self.ws.column_dimensions[get_column_letter(col_idx)].width = 18
#             else:
#                 self.ws.column_dimensions[get_column_letter(col_idx)].width = 16
        
#         row += 1
#         start_data_row = row
        
#         # Собираем данные по дням
#         daily_totals = {}
        
#         # Выводим данные по каждой дате
#         for date in daily_data['dates']:
#             date_str = self._format_date(date)
#             cell_date = self.ws.cell(row=row, column=start_col, value=date_str)
#             cell_date.font = self._get_cell_font(is_bold=True)
#             cell_date.alignment = Alignment(horizontal="left", vertical="center")
            
#             daily_total = 0
            
#             # Проходим по всем магазинам
#             for idx, store in enumerate(stores_sorted):
#                 amount = daily_data['data'][store].get(date, {}).get('total_amount', 0)
#                 daily_total += amount
                
#                 cell_amount = self.ws.cell(row=row, column=start_col + 1 + idx, value=amount)
#                 cell_amount.font = self._get_cell_font(is_negative=(amount < 0))
#                 cell_amount.number_format = '#,##0 ₽'
#                 cell_amount.alignment = Alignment(horizontal="right")
                
#                 # Условное форматирование
#                 fill = self._get_conditional_fill(amount)
#                 if fill:
#                     cell_amount.fill = fill
            
#             # Итого за день
#             daily_totals[date] = daily_total
#             cell_daily_total = self.ws.cell(row=row, column=start_col + num_stores + 1, value=daily_total)
#             cell_daily_total.font = self._get_cell_font(is_bold=True, is_negative=(daily_total < 0))
#             cell_daily_total.number_format = '#,##0 ₽'
#             cell_daily_total.alignment = Alignment(horizontal="right")
#             cell_daily_total.fill = PatternFill(
#                 start_color=COLORS["light_green"],
#                 end_color=COLORS["light_green"],
#                 fill_type="solid"
#             )
            
#             row += 1
        
#         # ==================== СТРОКА ИТОГО ПО МАГАЗИНАМ ====================
#         row += 1
        
#         # Ячейка "ИТОГО ПО МАГАЗИНУ"
#         cell_total_label = self.ws.cell(row=row, column=start_col, value="ИТОГО ПО МАГАЗИНУ")
#         cell_total_label.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#         cell_total_label.fill = PatternFill(
#             start_color=COLORS["dark_green"],
#             end_color=COLORS["dark_green"],
#             fill_type="solid"
#         )
#         cell_total_label.alignment = Alignment(horizontal="left", vertical="center")
        
#         # Итоги по каждому магазину
#         for idx, store in enumerate(stores_sorted):
#             store_total = store_totals.get(store, 0)
#             cell_store_total = self.ws.cell(row=row, column=start_col + 1 + idx, value=store_total)
#             cell_store_total.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#             cell_store_total.fill = PatternFill(
#                 start_color=COLORS["dark_green"],
#                 end_color=COLORS["dark_green"],
#                 fill_type="solid"
#             )
#             cell_store_total.number_format = '#,##0 ₽'
#             cell_store_total.alignment = Alignment(horizontal="right")
        
#         # Общий итог за все дни (должен совпасть с total_amount из KPI)
#         cell_grand_total = self.ws.cell(row=row, column=start_col + num_stores + 1, value=total_amount)
#         cell_grand_total.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#         cell_grand_total.fill = PatternFill(
#             start_color=COLORS["dark_green"],
#             end_color=COLORS["dark_green"],
#             fill_type="solid"
#         )
#         cell_grand_total.number_format = '#,##0 ₽'
#         cell_grand_total.alignment = Alignment(horizontal="right")
        
#         end_data_row = row
        
#         # Рисуем границы
#         for col in range(start_col, start_col + len(headers)):
#             for r in range(start_data_row, end_data_row + 1):
#                 self.ws.cell(row=r, column=col).border = BORDERS["thin"]
        
#         # ==================== ЗАМОРОЗКА ОБЛАСТИ ====================
#         # Замораживаем первую строку с заголовками и первую колонку с датами
#         self.ws.freeze_panes = self.ws.cell(row=start_data_row, column=start_col + 1)
        
#         row += 2
        
#         # ============================================================
#         # ЛЕГЕНДА
#         # ============================================================
#         row = self.header.draw(row, "ЛЕГЕНДА", end_col=6)
#         row += 1
        
#         legend_items = [
#             ("🟢 Зеленый фон", "Сумма > 100 000 ₽"),
#             ("🔴 Красный фон", "Отрицательная сумма (возвраты)"),

#         ]
        
#         for text, desc in legend_items:
#             cell_legend = self.ws.cell(row=row, column=2, value=f"{text}: {desc}")
#             cell_legend.font = Font(name="Roboto", size=8, color=COLORS["text_gray"])
#             row += 1
        
#         row += 1
        
#         # ============================================================
#         # ССЫЛКА НА ОГЛАВЛЕНИЕ
#         # ============================================================
#         nav_link = NavigationLink(self.ws)
#         row = nav_link.draw(
#             row=row,
#             text="← Вернуться к оглавлению",
#             target_sheet="TOC",
#             target_cell="A1",
#             start_col=2,
#             end_col=6,
#             alignment="left",
#             with_icon=True,
#             icon="📊"
#         )
        
#         # Настройка отображения
#         self.ws.sheet_view.showGridLines = False
    
#     def _format_store_name(self, name):
#         """Форматирование названия магазина для заголовка"""
#         if not name:
#             return "БЕЗ МАГАЗИНА"
#         name_str = str(name).upper()
#         if len(name_str) > 25:
#             return name_str[:22] + "..."
#         return name_str
    
#     def _build_empty_state(self):
#         """Если нет данных"""
#         row = self.sheet_title.draw(
#             row=3,
#             title="ПОДНЕВНАЯ ДИНАМИКА ОПЛАТ ПО МАГАЗИНАМ",
#             subtitle="Нет данных за выбранный период",
#             date_text=f"Сформировано: {self.request.user.username if self.request and hasattr(self.request, 'user') else 'System'}",
#             start_col=2,
#             end_col=7
#         )
        
#         row += 2
#         cell_empty = self.ws.cell(row=row, column=2, value="⚠️ Нет данных для отображения")
#         cell_empty.font = Font(name="Roboto", size=12, bold=True, color=COLORS["warning_red"])
#         cell_empty.alignment = Alignment(horizontal="center", vertical="center")
        
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)




# orders/reports/sheets/daily_payments_sheet.py
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .base_sheet import BaseSheet
from ..styles.helpers import draw_toc_button
from ..components import create_kpi_cards, create_header, create_table
from ..components.sheet_title import create_sheet_title
from ..components.navigation_link import NavigationLink
from ..styles.theme import COLORS, BORDERS


class DailyPaymentsSheet(BaseSheet):
    """Подневная аналитика оплат по магазинам"""

    def __init__(self, workbook, sheet_number, request=None):
        super().__init__(workbook, sheet_number)
        self.sheet_title = create_sheet_title(self.ws)
        self.kpi = create_kpi_cards(self.ws)
        self.header = create_header(self.ws)
        self.table = create_table(self.ws)
        self.request = request

    def _format_currency(self, value):
        """Форматирование суммы"""
        if value is None or value == 0:
            return "0 ₽"
        sign = "-" if value < 0 else ""
        formatted = f"{int(round(abs(value))):,} ₽".replace(",", " ")
        return f"{sign}{formatted}"

    def _format_number_str(self, value):
        """Форматирование числа"""
        if value is None or value == 0:
            return "0"
        sign = "-" if value < 0 else ""
        return f"{sign}{int(round(abs(value))):,}".replace(",", " ")

    def _format_date(self, date_value):
        """Форматирование даты"""
        if not date_value:
            return ""
        if hasattr(date_value, 'strftime'):
            return date_value.strftime('%d.%m.%Y')
        return str(date_value)

    def _get_cell_font(self, is_header=False, is_bold=False, is_negative=False, is_warning=False):
        """Возвращает стандартный шрифт для ячеек"""
        if is_header:
            return Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
        if is_negative:
            return Font(name="Roboto", size=9, bold=True, color=COLORS["warning_red"])
        if is_warning:
            return Font(name="Roboto", size=9, bold=True, color=COLORS["warning_orange"])
        if is_bold:
            return Font(name="Roboto", size=9, bold=True, color=COLORS["text_dark"])
        return Font(name="Roboto", size=9, color=COLORS["text_dark"])

    def _get_conditional_fill(self, amount):
        """Условное форматирование для сумм"""
        if amount < 0:
            return PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
        elif amount > 100000:
            return PatternFill(start_color="E5F3E5", end_color="E5F3E5", fill_type="solid")
        return None

    def build(self, daily_data):
        """Построение листа подневной аналитики - даты в строках, магазины в колонках"""
        
        if not daily_data['data']:
            self._build_empty_state()
            return
        
        row = 1
        
        # Получаем реальные даты из данных
        if daily_data['dates']:
            first_date = min(daily_data['dates'])
            last_date = max(daily_data['dates'])
            period_start = first_date
            period_end = last_date
        else:
            period_start = daily_data['start_date']
            period_end = daily_data['end_date']
        
        from datetime import datetime
        current_date = datetime.now()
        
        # Шапка
        draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")
        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 8
        self.ws.column_dimensions["A"].width = 2
        
        # Заголовок с реальными датами
        row = self.sheet_title.draw(
            row=3,
            title="ПОДНЕВНАЯ ДИНАМИКА ОПЛАТ ПО МАГАЗИНАМ",
            subtitle=f"Детальная разбивка по дням с {period_start.strftime('%d.%m.%Y')} по {period_end.strftime('%d.%m.%Y')}",
            date_text=f"Сформировано: {current_date.strftime('%d.%m.%Y в %H:%M')}",
            start_col=2,
            end_col=10
        )
        row += 2
        
        # ============================================================
        # KPI КАРТОЧКИ
        # ============================================================
        total_stores = len(daily_data['data'])
        total_days = len(daily_data['dates'])
        
        # Считаем общую сумму и кол-во операций
        total_amount = 0
        total_operations = 0
        store_totals = {}  # Суммы по магазинам для итоговой строки
        
        for store, days in daily_data['data'].items():
            store_total = 0
            for date, metrics in days.items():
                store_total += metrics['total_amount']
                total_amount += metrics['total_amount']
                total_operations += metrics['total_count']
            store_totals[store] = store_total
        
        row_cards = [
            {
                'title': 'ВСЕГО МАГАЗИНОВ',
                'value': self._format_number_str(total_stores),
                'subtitle': f"активных магазинов",
                'color': COLORS["blue"],
                'width': 3,
            },
            {
                'title': 'ОБЩАЯ СУММА ОПЛАТ',
                'value': self._format_currency(total_amount),
                'subtitle': f"{self._format_number_str(total_operations)} операций всего",
                'color': COLORS["blue"],
                'width': 3,
            },
            {
                'title': 'ПЕРИОД АНАЛИЗА',
                'value': f"{total_days} дней",
                'subtitle': f"с {period_start.strftime('%d.%m.%Y')} по {period_end.strftime('%d.%m.%Y')}",
                'color': COLORS["blue"],
                'width': 3,
            },
        ]
        
        row = self.kpi.draw_row(row, row_cards)
        row += 2
        
        # ============================================================
        # ТАБЛИЦА: ДАТЫ (строки) → МАГАЗИНЫ (колонки)
        # ============================================================
        
        # Получаем отсортированный список магазинов
        stores_sorted = sorted(daily_data['data'].keys())
        num_stores = len(stores_sorted)
        
        # Заголовок таблицы
        row = self.header.draw(
            row,
            "ДЕТАЛЬНАЯ РАЗБИВКА ПО ДНЯМ",
            end_col=3 + num_stores
        )
        
        # Формируем заголовки: ДАТА | МАГАЗИН 1 | МАГАЗИН 2 | ... | ИТОГО ЗА ДЕНЬ
        headers = ["ДАТА"] + [self._format_store_name(s) for s in stores_sorted] + ["ИТОГО ЗА ДЕНЬ"]
        start_col = 2
        
        # Рисуем заголовки
        for col_idx, header in enumerate(headers, start_col):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
            cell.fill = PatternFill(
                start_color=COLORS["dark_green"],
                end_color=COLORS["dark_green"],
                fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Ширина колонок
            if col_idx == start_col:
                self.ws.column_dimensions[get_column_letter(col_idx)].width = 14
            elif col_idx == start_col + num_stores:
                self.ws.column_dimensions[get_column_letter(col_idx)].width = 18
            else:
                self.ws.column_dimensions[get_column_letter(col_idx)].width = 16
        
        row += 1
        start_data_row = row
        
        # Собираем данные по дням
        daily_totals = {}
        
        # Выводим данные по каждой дате
        for date in daily_data['dates']:
            date_str = self._format_date(date)
            cell_date = self.ws.cell(row=row, column=start_col, value=date_str)
            cell_date.font = self._get_cell_font(is_bold=True)
            cell_date.alignment = Alignment(horizontal="left", vertical="center")
            
            daily_total = 0
            
            # Проходим по всем магазинам
            for idx, store in enumerate(stores_sorted):
                amount = daily_data['data'][store].get(date, {}).get('total_amount', 0)
                daily_total += amount
                
                cell_amount = self.ws.cell(row=row, column=start_col + 1 + idx, value=amount)
                cell_amount.font = self._get_cell_font(is_negative=(amount < 0))
                cell_amount.number_format = '#,##0 ₽'
                cell_amount.alignment = Alignment(horizontal="right")
                
                # Условное форматирование
                fill = self._get_conditional_fill(amount)
                if fill:
                    cell_amount.fill = fill
            
            # Итого за день
            daily_totals[date] = daily_total
            cell_daily_total = self.ws.cell(row=row, column=start_col + num_stores + 1, value=daily_total)
            cell_daily_total.font = self._get_cell_font(is_bold=True, is_negative=(daily_total < 0))
            cell_daily_total.number_format = '#,##0 ₽'
            cell_daily_total.alignment = Alignment(horizontal="right")
            cell_daily_total.fill = PatternFill(
                start_color=COLORS["light_green"],
                end_color=COLORS["light_green"],
                fill_type="solid"
            )
            
            row += 1
        
        # ==================== СТРОКА ИТОГО ПО МАГАЗИНАМ ====================
        row += 1
        
        # Ячейка "ИТОГО ПО МАГАЗИНУ"
        cell_total_label = self.ws.cell(row=row, column=start_col, value="ИТОГО ПО МАГАЗИНУ")
        cell_total_label.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
        cell_total_label.fill = PatternFill(
            start_color=COLORS["dark_green"],
            end_color=COLORS["dark_green"],
            fill_type="solid"
        )
        cell_total_label.alignment = Alignment(horizontal="left", vertical="center")
        
        # Итоги по каждому магазину
        for idx, store in enumerate(stores_sorted):
            store_total = store_totals.get(store, 0)
            cell_store_total = self.ws.cell(row=row, column=start_col + 1 + idx, value=store_total)
            cell_store_total.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
            cell_store_total.fill = PatternFill(
                start_color=COLORS["dark_green"],
                end_color=COLORS["dark_green"],
                fill_type="solid"
            )
            cell_store_total.number_format = '#,##0 ₽'
            cell_store_total.alignment = Alignment(horizontal="right")
        
        # Общий итог за все дни (должен совпасть с total_amount из KPI)
        cell_grand_total = self.ws.cell(row=row, column=start_col + num_stores + 1, value=total_amount)
        cell_grand_total.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
        cell_grand_total.fill = PatternFill(
            start_color=COLORS["dark_green"],
            end_color=COLORS["dark_green"],
            fill_type="solid"
        )
        cell_grand_total.number_format = '#,##0 ₽'
        cell_grand_total.alignment = Alignment(horizontal="right")
        
        end_data_row = row
        
        # Рисуем границы
        for col in range(start_col, start_col + len(headers)):
            for r in range(start_data_row, end_data_row + 1):
                self.ws.cell(row=r, column=col).border = BORDERS["thin"]
        
        # ==================== ЗАМОРОЗКА ОБЛАСТИ ====================
        # Замораживаем первую строку с заголовками и первую колонку с датами
        self.ws.freeze_panes = self.ws.cell(row=start_data_row, column=start_col + 1)
        
        row += 2
        
        # ============================================================
        # ЛЕГЕНДА
        # ============================================================
        row = self.header.draw(row, "ЛЕГЕНДА", end_col=6)
        row += 1
        
        legend_items = [
            ("🟢 Зеленый фон", "Сумма > 100 000 ₽"),
            ("🔴 Красный фон", "Отрицательная сумма (возвраты)"),

        ]
        
        for text, desc in legend_items:
            cell_legend = self.ws.cell(row=row, column=2, value=f"{text}: {desc}")
            cell_legend.font = Font(name="Roboto", size=8, color=COLORS["text_gray"])
            row += 1
        
        row += 1
        

        
        # Настройка отображения
        self.ws.sheet_view.showGridLines = False
    
    def _format_store_name(self, name):
        """Форматирование названия магазина для заголовка"""
        if not name:
            return "БЕЗ МАГАЗИНА"
        name_str = str(name).upper()
        if len(name_str) > 25:
            return name_str[:22] + "..."
        return name_str
    
    def _build_empty_state(self):
        """Если нет данных"""
        row = self.sheet_title.draw(
            row=3,
            title="ПОДНЕВНАЯ ДИНАМИКА ОПЛАТ ПО МАГАЗИНАМ",
            subtitle="Нет данных за выбранный период",
            date_text=f"Сформировано: {self.request.user.username if self.request and hasattr(self.request, 'user') else 'System'}",
            start_col=2,
            end_col=7
        )
        
        row += 2
        cell_empty = self.ws.cell(row=row, column=2, value="⚠️ Нет данных для отображения")
        cell_empty.font = Font(name="Roboto", size=12, bold=True, color=COLORS["warning_red"])
        cell_empty.alignment = Alignment(horizontal="center", vertical="center")
        
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)