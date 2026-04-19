# orders/reports/sheets/client_analysis_sheet.py
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from .base_sheet import BaseSheet
from ..styles.helpers import draw_toc_button
from ..components import create_kpi_cards, create_header, create_table
from ..components.sheet_title import create_sheet_title
from ..components.footnote import Footnote
from ..components.navigation_link import NavigationLink
from ..components.methodology import Methodology
from ..queries.client_analysis_query import ClientAnalysisQueries
from ..styles.theme import COLORS, FILLS, BORDERS


class ClientAnalysisSheet(BaseSheet):
    """Анализ клиентской базы - полная аналитика по клиентам"""
    
    def __init__(self, workbook, sheet_number, request=None):
        super().__init__(workbook, sheet_number)
        self.sheet_title = create_sheet_title(self.ws)
        self.kpi = create_kpi_cards(self.ws)
        self.header = create_header(self.ws)
        self.table = create_table(self.ws)
        self.queries = ClientAnalysisQueries(request)
    
    def _format_currency(self, value):
        if value is None or value == 0:
            return "0 ₽"
        return f"{int(round(value)):,} ₽".replace(",", " ")
    
    def _format_number(self, value):
        if value is None or value == 0:
            return 0
        return int(round(value))
    
    def _format_number_str(self, value):
        if value is None or value == 0:
            return "0"
        return f"{int(round(value)):,}".replace(",", " ")
    
    def _format_percent(self, value):
        if value is None:
            return "0%"
        return f"{value:.1f}%"
    
    def _safe_float(self, value, default=0.0):
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default
    
    def _format_client_name(self, name):
        if not name:
            return ""
        return str(name).upper()
    
    def _format_manager_name(self, full_name):
        if not full_name:
            return ""
        parts = str(full_name).strip().split()
        if len(parts) >= 1:
            last_name = parts[0].upper()
            if len(parts) >= 2:
                first_initial = parts[1][0].upper() if parts[1] else ""
                return f"{last_name} {first_initial}."
            return last_name
        return str(full_name).upper()
    
    def _apply_data_font(self, row, start_col, end_col):
        """Применяет темно-серый шрифт к ячейкам с данными"""
        for col in range(start_col, end_col + 1):
            cell = self.ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
    
    def _draw_abc_pie_chart_text(self, row, abc_data):
        """Рисует текстовую имитацию круговой диаграммы ABC-анализа"""
        total_clients = abc_data['total_clients']
        if total_clients == 0:
            return row
        
        # Заголовок
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        title_cell = self.ws.cell(
            row=row, 
            column=2, 
            value="ABC-АНАЛИЗ КЛИЕНТОВ (Правило Парето 80/20) - данные за YTD"
        )
        title_cell.font = Font(name="Roboto", size=11, bold=True, color=COLORS["dark_green"])
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 1
        
        # Статистика по категориям (только реальные цифры, без противоречий)
        categories = [
            {
                'name': 'A (основные клиенты)',
                'count': abc_data['a_count'],
                'percent_clients': abc_data['a_percent_of_clients'],
                'percent_amount': abc_data['a_percent_of_amount'],
                'color': COLORS["dark_green"],
                'icon': '▰',
            },
            {
                'name': 'B (средние клиенты)',
                'count': abc_data['b_count'],
                'percent_clients': abc_data['b_percent_of_clients'],
                'percent_amount': abc_data['b_percent_of_amount'],
                'color': COLORS["warning_green"],
                'icon': '▱',
            },
            {
                'name': 'C (мелкие клиенты)',
                'count': abc_data['c_count'],
                'percent_clients': abc_data['c_percent_of_clients'],
                'percent_amount': abc_data['c_percent_of_amount'],
                'color': COLORS["warning_red"],
                'icon': '▰',
            },
        ]
        
        for i, cat in enumerate(categories):
            col = 2 + (i * 2)
            
            # Название категории
            cat_cell = self.ws.cell(row=row, column=col, value=f"{cat['icon']} {cat['name']}")
            cat_cell.font = Font(name="Roboto", size=9, bold=True, color=cat['color'])
            cat_cell.alignment = Alignment(horizontal="left")
            
            # Количество клиентов и процент
            stats_cell = self.ws.cell(
                row=row + 1, 
                column=col, 
                value=f"{self._format_number_str(cat['count'])} клиентов ({self._format_percent(cat['percent_clients'])})"
            )
            stats_cell.font = Font(name="Roboto", size=8, color=COLORS["text_gray"])
            stats_cell.alignment = Alignment(horizontal="left")
            
            # Реальный процент выручки (без ложных обещаний)
            amount_cell = self.ws.cell(
                row=row + 2, 
                column=col, 
                value=f"→ {self._format_percent(cat['percent_amount'])} выручки"
            )
            amount_cell.font = Font(name="Roboto", size=9, bold=True, color=cat['color'])
            amount_cell.alignment = Alignment(horizontal="left")
        
        row += 4
        return row
    
    def build(self):
        """Построение листа анализа клиентов"""
        
        # Получаем данные
        metrics = self.queries.get_client_metrics()
        top_by_amount_ytd = self.queries.get_top_clients_by_amount(limit=15, period='ytd')
        top_by_amount_mtd = self.queries.get_top_clients_by_amount(limit=10, period='mtd')
        top_by_orders = self.queries.get_top_clients_by_orders(limit=10, period='ytd')
        abc_analysis = self.queries.get_abc_analysis(period='ytd')
        high_risk_clients = self.queries.get_high_risk_clients(limit=10, debt_threshold_percent=30)
        clients_by_manager = self.queries.get_clients_by_manager(limit=10)
        
        row = 1
        
        # Шапка
        draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")
        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 8
        
        self.ws.column_dimensions["A"].width = 2
        row = self.sheet_title.draw(
            row=3,
            title="АНАЛИЗ КЛИЕНТСКОЙ БАЗЫ",
            subtitle="Полная аналитика по клиентам: ключевые метрики, ABC-анализ, топ-клиенты, риски",
            date_text=f"Сформировано: {self.queries.today.strftime('%d.%m.%Y')}",
            start_col=2,
            end_col=7
        )
        row += 1
        
        # ============================================================
        # KPI КАРТОЧКИ - MTD и YTD
        # ============================================================
        mtd = metrics['mtd']
        ytd = metrics['ytd']
        
        # Первый ряд: MTD
        row1_cards = [
            {
                'title': 'АКТИВНЫЕ КЛИЕНТЫ (MTD)',
                'value': self._format_number_str(mtd['total_clients']),
                'subtitle': f"ср.чек: {self._format_currency(mtd['avg_check'])}",
                'color': COLORS["blue"],
            },
            {
                'title': 'КЛИЕНТЫ С 2+ ЗАКАЗАМИ',
                'value': f"{self._format_percent(mtd['repeat_rate'])}",
                'subtitle': f"{self._format_number_str(mtd['repeat_clients'])} клиентов повторно",
                'color': COLORS["dark_green"],
            },
            {
                'title': 'СУММА ЗАКАЗА ОТ КЛИЕНТОВ',
                'value': self._format_currency(mtd['total_amount']),
                'subtitle': f"{self._format_number_str(mtd['total_orders'])} заказов",
                'color': COLORS["dark_green"],
            },
        ]
        
        # Рисуем первый ряд карточек (без start_col)
        row = self.kpi.draw_two_rows(row, row1_cards, [])
        row -= 2
        
        # Второй ряд: YTD
        row2_cards = [
            {
                'title': 'АКТИВНЫЕ КЛИЕНТЫ (YTD)',
                'value': self._format_number_str(ytd['total_clients']),
                'subtitle': f"ср.чек: {self._format_currency(ytd['avg_check'])}",
                'color': COLORS["blue"],
            },
            {
                'title': 'НОВЫЕ / ПОВТОРНЫЕ',
                'value': f"{self._format_percent(ytd['repeat_rate'])}",
                'subtitle': f"{self._format_number_str(ytd['repeat_clients'])} клиентов повторно",
                'color': COLORS["dark_green"],
            },
            {
                'title': 'СУММА ЗАКАЗА ОТ КЛИЕНТОВ',
                'value': self._format_currency(ytd['total_amount']),
                'subtitle': f"{self._format_number_str(ytd['total_orders'])} заказов",
                'color': COLORS["dark_green"],
            },
        ]
        
        # Рисуем второй ряд карточек (без start_col)
        row = self.kpi.draw_two_rows(row, row2_cards, [])
        row -= 2
        row += 2
        
        # ============================================================
        # ABC-АНАЛИЗ
        # ============================================================
        row = self._draw_abc_pie_chart_text(row, abc_analysis)
        row += 1
        
        # ============================================================
        # ТОП КЛИЕНТОВ ПО СУММЕ ЗАКАЗОВ (YTD)
        # ============================================================
        if top_by_amount_ytd:
            row = self.header.draw(row, "ТОП-15 КЛИЕНТОВ ПО СУММЕ ЗАКАЗОВ (YTD)", end_col=7)
            
            headers = ["КЛИЕНТ", "СУММА ЗАКАЗОВ", "ЗАКАЗОВ", "СР. ЧЕК", "ОПЛАЧЕНО", "ДОЛГ"]
            
            # Рисуем заголовки таблицы
            for col_idx, header in enumerate(headers, start=2):
                cell = self.ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
                cell.fill = PatternFill(start_color=COLORS["dark_green"], end_color=COLORS["dark_green"], fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            row += 1
            
            # Записываем данные
            start_data_row = row
            for c in top_by_amount_ytd:
                # Клиент (текст)
                cell_client = self.ws.cell(row=row, column=2, value=self._format_client_name(c['client']))
                cell_client.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                
                # СУММА ЗАКАЗОВ (число + формат валюты)
                cell_amount = self.ws.cell(row=row, column=3, value=c['total_amount'])
                cell_amount.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell_amount.number_format = '#,##0 ₽'
                cell_amount.alignment = Alignment(horizontal="right")
                
                # ЗАКАЗОВ (число)
                cell_orders = self.ws.cell(row=row, column=4, value=c['orders_count'])
                cell_orders.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell_orders.number_format = '#,##0'
                cell_orders.alignment = Alignment(horizontal="right")
                
                # СР. ЧЕК (число + формат валюты)
                cell_avg = self.ws.cell(row=row, column=5, value=c['avg_check'])
                cell_avg.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell_avg.number_format = '#,##0 ₽'
                cell_avg.alignment = Alignment(horizontal="right")
                
                # ОПЛАЧЕНО (число + формат валюты)
                cell_paid = self.ws.cell(row=row, column=6, value=c['paid_amount'])
                cell_paid.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell_paid.number_format = '#,##0 ₽'
                cell_paid.alignment = Alignment(horizontal="right")
                
                # ДОЛГ (число + формат валюты)
                cell_debt = self.ws.cell(row=row, column=7, value=c['debt'])
                cell_debt.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell_debt.number_format = '#,##0 ₽'
                cell_debt.alignment = Alignment(horizontal="right")
                
                row += 1
            
            end_data_row = row - 1
            
            # Дата-бар на колонке СУММА ЗАКАЗОВ (колонка C)
            self.ws.conditional_formatting.add(
                f"C{start_data_row}:C{end_data_row}",
                DataBarRule(
                    start_type="min",
                    end_type="max",
                    color=COLORS["light_green"],
                    showValue=True
                )
            )
            
            # Подсветка долга (колонка G)
            self.ws.conditional_formatting.add(
                f"G{start_data_row}:G{end_data_row}",
                ColorScaleRule(
                    start_type='min', start_color='FFFFFF',
                    mid_type='percentile', mid_value=50, mid_color='FFF4E5',
                    end_type='max', end_color='FFE5E5'
                )
            )
            
            # Применяем границы таблицы и шрифт для всех данных
            for col in range(2, 8):
                for r in range(start_data_row, end_data_row + 1):
                    self.ws.cell(row=r, column=col).border = BORDERS["thin"]
            
            row += 2
        
        # ============================================================
        # ТОП КЛИЕНТОВ ПО КОЛИЧЕСТВУ ЗАКАЗОВ
        # ============================================================
        if top_by_orders:
            row = self.header.draw(row, "ТОП-10 КЛИЕНТОВ ПО КОЛИЧЕСТВУ ЗАКАЗОВ (YTD)", end_col=5)
            
            headers = ["КЛИЕНТ", "ЗАКАЗОВ", "СУММА ЗАКАЗОВ", "СР. ЧЕК"]
            
            # Рисуем заголовки таблицы
            for col_idx, header in enumerate(headers, start=2):
                cell = self.ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
                cell.fill = PatternFill(start_color=COLORS["dark_green"], end_color=COLORS["dark_green"], fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            row += 1
            
            # Записываем данные
            start_data_row = row
            for c in top_by_orders:
                # Клиент (текст)
                cell_client = self.ws.cell(row=row, column=2, value=self._format_client_name(c['client']))
                cell_client.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                
                # ЗАКАЗОВ (число для дата-бара)
                cell_orders = self.ws.cell(row=row, column=3, value=c['orders_count'])
                cell_orders.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell_orders.number_format = '#,##0'
                cell_orders.alignment = Alignment(horizontal="right")
                
                # СУММА ЗАКАЗОВ (число)
                cell_amount = self.ws.cell(row=row, column=4, value=c['total_amount'])
                cell_amount.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell_amount.number_format = '#,##0 ₽'
                cell_amount.alignment = Alignment(horizontal="right")
                
                # СР. ЧЕК (число)
                cell_avg = self.ws.cell(row=row, column=5, value=c['avg_check'])
                cell_avg.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell_avg.number_format = '#,##0 ₽'
                cell_avg.alignment = Alignment(horizontal="right")
                
                row += 1
            
            end_data_row = row - 1
            
            # Дата-бар на колонке ЗАКАЗОВ (колонка C)
            self.ws.conditional_formatting.add(
                f"C{start_data_row}:C{end_data_row}",
                DataBarRule(
                    start_type="min",
                    end_type="max",
                    color=COLORS["light_green"],
                    showValue=True
                )
            )
            
            # Применяем границы таблицы
            for col in range(2, 6):
                for r in range(start_data_row, end_data_row + 1):
                    self.ws.cell(row=r, column=col).border = BORDERS["thin"]
            
            row += 2
        
        # ============================================================
        # ТОП КЛИЕНТОВ ПО СУММЕ (MTD) - быстрый взгляд на текущий месяц
        # ============================================================
        if top_by_amount_mtd:
            row = self.header.draw(row, "ТОП-10 КЛИЕНТОВ ЗА ТЕКУЩИЙ МЕСЯЦ (MTD)", end_col=5)
            
            headers = ["КЛИЕНТ", "СУММА ЗАКАЗОВ", "ЗАКАЗОВ", "СР. ЧЕК"]
            
            # Рисуем заголовки таблицы
            for col_idx, header in enumerate(headers, start=2):
                cell = self.ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
                cell.fill = PatternFill(start_color=COLORS["dark_green"], end_color=COLORS["dark_green"], fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            row += 1
            
            # Записываем данные
            start_data_row = row
            for c in top_by_amount_mtd:
                # Клиент (текст)
                cell_client = self.ws.cell(row=row, column=2, value=self._format_client_name(c['client']))
                cell_client.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                
                # СУММА ЗАКАЗОВ (число для дата-бара)
                cell_amount = self.ws.cell(row=row, column=3, value=c['total_amount'])
                cell_amount.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell_amount.number_format = '#,##0 ₽'
                cell_amount.alignment = Alignment(horizontal="right")
                
                # ЗАКАЗОВ (число)
                cell_orders = self.ws.cell(row=row, column=4, value=c['orders_count'])
                cell_orders.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell_orders.number_format = '#,##0'
                cell_orders.alignment = Alignment(horizontal="right")
                
                # СР. ЧЕК (число)
                cell_avg = self.ws.cell(row=row, column=5, value=c['avg_check'])
                cell_avg.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell_avg.number_format = '#,##0 ₽'
                cell_avg.alignment = Alignment(horizontal="right")
                
                row += 1
            
            end_data_row = row - 1
            
            # Дата-бар на колонке СУММА ЗАКАЗОВ (колонка C)
            self.ws.conditional_formatting.add(
                f"C{start_data_row}:C{end_data_row}",
                DataBarRule(
                    start_type="min",
                    end_type="max",
                    color=COLORS["light_green"],
                    showValue=True
                )
            )
            
            # Применяем границы таблицы
            for col in range(2, 6):
                for r in range(start_data_row, end_data_row + 1):
                    self.ws.cell(row=r, column=col).border = BORDERS["thin"]
            
            row += 2
        
        # ============================================================
        # КЛИЕНТЫ С ВЫСОКИМ РИСКОМ (БОЛЬШОЙ ДОЛГ)
        # ============================================================
        if high_risk_clients:
            row = self.header.draw(row, "КЛИЕНТЫ С ВЫСОКИМ РИСКОМ (ДОЛГ >30% ОТ СУММЫ) - данные за YTD", end_col=7)
            
            headers = ["КЛИЕНТ", "МЕНЕДЖЕР", "СУММА ЗАКАЗОВ", "ДОЛГ", "% ДОЛГА", "ЗАКАЗОВ"]
            
            # Рисуем заголовки таблицы
            for col_idx, header in enumerate(headers, start=2):
                cell = self.ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
                cell.fill = PatternFill(start_color=COLORS["dark_green"], end_color=COLORS["dark_green"], fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            row += 1
            
            # Записываем данные
            start_data_row = row
            for c in high_risk_clients:
                # Клиент (текст)
                cell_client = self.ws.cell(row=row, column=2, value=self._format_client_name(c['client']))
                cell_client.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                
                # Менеджер (текст)
                cell_manager = self.ws.cell(row=row, column=3, value=self._format_manager_name(c['manager']))
                cell_manager.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                
                # СУММА ЗАКАЗОВ (число)
                cell_amount = self.ws.cell(row=row, column=4, value=c['total_amount'])
                cell_amount.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell_amount.number_format = '#,##0 ₽'
                cell_amount.alignment = Alignment(horizontal="right")
                
                # ДОЛГ (число)
                cell_debt = self.ws.cell(row=row, column=5, value=c['debt'])
                cell_debt.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell_debt.number_format = '#,##0 ₽'
                cell_debt.alignment = Alignment(horizontal="right")
                
                # % ДОЛГА (число)
                cell_percent = self.ws.cell(row=row, column=6, value=c['debt_percent'])
                cell_percent.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell_percent.number_format = '0.0"%"'
                cell_percent.alignment = Alignment(horizontal="right")
                
                # ЗАКАЗОВ (число)
                cell_orders = self.ws.cell(row=row, column=7, value=c['orders_count'])
                cell_orders.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell_orders.number_format = '#,##0'
                cell_orders.alignment = Alignment(horizontal="right")
                
                row += 1
            
            end_data_row = row - 1
            
            # Подсветка долга (колонка E)
            self.ws.conditional_formatting.add(
                f"E{start_data_row}:E{end_data_row}",
                ColorScaleRule(
                    start_type='min', start_color='FFFFFF',
                    mid_type='percentile', mid_value=50, mid_color='FFF4E5',
                    end_type='max', end_color='FFE5E5'
                )
            )
            
            # Применяем границы таблицы
            for col in range(2, 8):
                for r in range(start_data_row, end_data_row + 1):
                    self.ws.cell(row=r, column=col).border = BORDERS["thin"]
            
            row += 2
        
        # ============================================================
        # РАСПРЕДЕЛЕНИЕ КЛИЕНТОВ ПО МЕНЕДЖЕРАМ
        # ============================================================
        if clients_by_manager:
            row = self.header.draw(row, "РАСПРЕДЕЛЕНИЕ КЛИЕНТОВ ПО МЕНЕДЖЕРАМ", end_col=6)
            
            headers = ["МЕНЕДЖЕР", "КЛИЕНТОВ", "ЗАКАЗОВ", "СУММА ЗАКАЗОВ", "СР. НА КЛИЕНТА"]
            
            # Рисуем заголовки таблицы
            for col_idx, header in enumerate(headers, start=2):
                cell = self.ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
                cell.fill = PatternFill(start_color=COLORS["dark_green"], end_color=COLORS["dark_green"], fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            row += 1
            
            # Записываем данные
            start_data_row = row
            for m in clients_by_manager:
                # Менеджер (текст)
                cell_manager = self.ws.cell(row=row, column=2, value=self._format_manager_name(m['manager']))
                cell_manager.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                
                # КЛИЕНТОВ (число)
                cell_clients = self.ws.cell(row=row, column=3, value=m['clients_count'])
                cell_clients.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell_clients.number_format = '#,##0'
                cell_clients.alignment = Alignment(horizontal="right")
                
                # ЗАКАЗОВ (число)
                cell_orders = self.ws.cell(row=row, column=4, value=m['orders_count'])
                cell_orders.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell_orders.number_format = '#,##0'
                cell_orders.alignment = Alignment(horizontal="right")
                
                # СУММА ЗАКАЗОВ (число)
                cell_amount = self.ws.cell(row=row, column=5, value=m['total_amount'])
                cell_amount.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell_amount.number_format = '#,##0 ₽'
                cell_amount.alignment = Alignment(horizontal="right")
                
                # СР. НА КЛИЕНТА (число)
                cell_avg = self.ws.cell(row=row, column=6, value=m['avg_per_client'])
                cell_avg.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
                cell_avg.number_format = '#,##0 ₽'
                cell_avg.alignment = Alignment(horizontal="right")
                
                row += 1
            
            end_data_row = row - 1
            
            # Применяем границы таблицы
            for col in range(2, 7):
                for r in range(start_data_row, end_data_row + 1):
                    self.ws.cell(row=r, column=col).border = BORDERS["thin"]
            
            # Дата-бар на количестве клиентов (колонка C)
            clients_col_letter = get_column_letter(3)
            self.ws.conditional_formatting.add(
                f"{clients_col_letter}{start_data_row}:{clients_col_letter}{end_data_row}",
                DataBarRule(
                    start_type="min",
                    end_type="max",
                    color=COLORS["light_green"],
                    showValue=True
                )
            )
            
            row += 2
        
        # ============================================================
        # МЕТОДОЛОГИЯ РАСЧЕТОВ
        # ============================================================
        row = self.header.draw(row, "МЕТОДОЛОГИЯ РАСЧЕТОВ", end_col=6)

        methodology = Methodology(self.ws)
        row = methodology.draw_metrics_methodology(row, start_col=2, end_col=8)

        # ============================================================
        # ПРИМЕЧАНИЕ (обновленное)
        # ============================================================
        footnote = Footnote(self.ws)
        row = footnote.draw(
            row=row,
            text="* MTD — с начала месяца, YTD — с начала года. Данные исключают полностью отмененные заказы. ABC-анализ основан на правиле Парето 80/20."
        )
        row += 1
        
        # ============================================================
        # ССЫЛКИ
        # ============================================================
        nav_link = NavigationLink(self.ws)
        row = nav_link.draw(
            row=row,
            text="Подробную информацию по активным заказам смотрите на листе «Детализация заказов» →",
            target_sheet="2",
            target_cell="A1",
            start_col=2,
            end_col=8,
            alignment="left",
            with_icon=True,
            icon="📋"
        )
        
        # ============================================================
        # НАСТРОЙКА КОЛОНОК
        # ============================================================
        self.ws.column_dimensions["B"].width = 28
        self.ws.column_dimensions["C"].width = 18
        self.ws.column_dimensions["D"].width = 18
        self.ws.column_dimensions["E"].width = 18
        self.ws.column_dimensions["F"].width = 18
        self.ws.column_dimensions["G"].width = 18
        self.ws.column_dimensions["H"].width = 18
        self.ws.column_dimensions["I"].width = 18
        self.ws.column_dimensions["J"].width = 16
        
        self.ws.sheet_view.showGridLines = False