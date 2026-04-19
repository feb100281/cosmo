# orders/reports/sheets/delivery_analysis_sheet.py
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter
from .base_sheet import BaseSheet
from ..styles.helpers import draw_toc_button
from ..components import create_kpi_cards, create_header, create_table
from ..components.sheet_title import create_sheet_title
from ..components.footnote import Footnote
from ..components.navigation_link import NavigationLink
from ..queries.delivery_analysis_query import DeliveryAnalysisQueries
from ..styles.theme import COLORS, FILLS


class DeliveryAnalysisSheet(BaseSheet):
    """Анализ доставки - отдельный лист с полной аналитикой"""
    
    def __init__(self, workbook, sheet_number, request=None):
        super().__init__(workbook, sheet_number)
        self.sheet_title = create_sheet_title(self.ws)
        self.kpi = create_kpi_cards(self.ws)
        self.header = create_header(self.ws)
        self.table = create_table(self.ws)
        self.queries = DeliveryAnalysisQueries(request)
    
    def _format_currency(self, value):
        if value is None or value == 0:
            return "0 ₽"
        return f"{int(round(value)):,} ₽".replace(",", " ")
    
    def _format_number(self, value):
        if value is None or value == 0:
            return 0
        return int(round(value))
    
    def _format_number_str(self, value):
        """Для отображения в ячейках с форматированием"""
        if value is None or value == 0:
            return "0"
        return f"{int(round(value)):,}".replace(",", " ")
    
    def _format_percent(self, value):
        if value is None:
            return "0%"
        return f"{value:.1f}%"
    
    def _safe_float(self, value, default=0.0):
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default
    
    def _format_manager_name(self, full_name):
        if not full_name:
            return ""
        parts = str(full_name).strip().split()
        if len(parts) >= 1:
            last_name = parts[0].upper()
            if len(parts) >= 2:
                first_initial = parts[1][0].upper() if parts[1] else ""
                return f"{last_name} {first_initial}."
            return last_name
        return str(full_name).upper()
    
    def _format_client_name(self, name):
        if not name:
            return ""
        return str(name).upper()
    
    def _format_month_name(self, month_date):
        """Возвращает название месяца в формате: ЯНВ 2026, ФЕВ 2026"""
        if not month_date:
            return ""
        
        month_names = {
            1: "ЯНВ", 2: "ФЕВ", 3: "МАР", 4: "АПР",
            5: "МАЙ", 6: "ИЮН", 7: "ИЮЛ", 8: "АВГ",
            9: "СЕН", 10: "ОКТ", 11: "НОЯ", 12: "ДЕК"
        }
        
        if hasattr(month_date, 'month'):
            return f"{month_names[month_date.month]} {month_date.year}"
        return str(month_date)
    
    def build(self):
        """Построение листа анализа доставки"""
        
        # Получаем данные
        metrics = self.queries.get_delivery_metrics()
        managers_mtd = self.queries.get_delivery_by_manager(limit=12)
        managers_ytd = self.queries.get_delivery_by_manager_ytd(limit=12)  
        trends = self.queries.get_delivery_trend(months=6)
        top_clients = self.queries.get_top_clients_by_delivery(limit=5)
        
        row = 1
        
        # Шапка
        draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")
        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 8
        
        self.ws.column_dimensions["A"].width = 2
        row = self.sheet_title.draw(
            row=3,
            title="АНАЛИЗ ДОСТАВКИ",
            subtitle="Аналитика по услугам доставки в заказах",
            date_text=f"Сформировано: {self.queries.today.strftime('%d.%m.%Y')}",
            start_col=2,
            end_col=7
        )
        row += 1
        
        # ============================================================
        # KPI КАРТОЧКИ - MTD и YTD
        # ============================================================
        mtd = metrics['mtd']
        ytd = metrics['ytd']
        
        # Первый ряд: MTD
        row1_cards = [
            {
                'title': 'ДОСТАВКА (MTD)',
                'value': self._format_currency(mtd['delivery_sum']),
                'subtitle': f"{self._format_percent(mtd['delivery_percent_of_amount'])} от суммы заказов",
                'color': COLORS["blue"],
          
            },
            {
                'title': 'ЗАКАЗОВ С ДОСТАВКОЙ',
                'value': self._format_percent(mtd['orders_with_delivery_pct']),
                'subtitle': 'доля от всех заказов',
                'color': COLORS["dark_green"]
            },
            {
                'title': 'СР. СТОИМОСТЬ ДОСТАВКИ',
                'value': self._format_currency(mtd['avg_delivery_per_order']),
                'subtitle': 'на 1 заказ с доставкой',
                'color': COLORS["dark_green"],

            },
        ]
        
        # Рисуем первый ряд карточек
        row = self.kpi.draw_two_rows(row, row1_cards, [])
        row -= 2
        
        # Второй ряд: YTD
        row2_cards = [
            {
                'title': 'ДОСТАВКА (YTD)',
                'value': self._format_currency(ytd['delivery_sum']),
                'subtitle': f"{self._format_percent(ytd['delivery_percent_of_amount'])} от суммы заказов",
                'color': COLORS["blue"],
         
            },
            {
                'title': 'ЗАКАЗОВ С ДОСТАВКОЙ',
                'value': self._format_percent(ytd['orders_with_delivery_pct']),
                'subtitle': 'доля от всех заказов',
                'color': COLORS["dark_green"]
            },
            {
                'title': 'СР. СТОИМОСТЬ ДОСТАВКИ',
                'value': self._format_currency(ytd['avg_delivery_per_order']),
                'subtitle': 'на 1 заказ с доставкой',
                'color': COLORS["dark_green"],
  
            },
        ]
        
        # Рисуем второй ряд карточек
        row = self.kpi.draw_two_rows(row, row2_cards, [])
        row -= 2
        row += 2
        
        # ============================================================
        # ТОП МЕНЕДЖЕРОВ ПО СУММЕ ДОСТАВКИ (MTD)
        # ============================================================
        if managers_mtd:
            row = self.header.draw(row, "ТОП МЕНЕДЖЕРОВ ПО СУММЕ ДОСТАВКИ (MTD)", end_col=7)
            
            headers = ["МЕНЕДЖЕР", "ЗАКАЗОВ", "С ДОСТАВКОЙ", "СУММА ДОСТАВКИ", "% ОТ СУММЫ", "СР. ДОСТАВКА"]
            
            data_rows = []
            for m in managers_mtd:
                data_rows.append([
                    self._format_manager_name(m['manager']),
                    self._format_number(m['orders_count']),
                    f"{self._format_number_str(m['orders_with_delivery'])} ({self._format_percent(m['coverage_pct'])})",
                    m['total_delivery'],
                    m['delivery_percent'],
                    m['avg_delivery'],
                ])
            
            row = self.table.draw(row, headers, data_rows)
            
            # Дата-бар на колонке с суммой доставки
            start_col = 2
            amount_col_idx = headers.index("СУММА ДОСТАВКИ")
            amount_col_num = start_col + amount_col_idx
            amount_col_letter = get_column_letter(amount_col_num)
            
            first_data_row = row - len(data_rows)
            last_data_row = row - 1
            
            self.ws.conditional_formatting.add(
                f"{amount_col_letter}{first_data_row}:{amount_col_letter}{last_data_row}",
                DataBarRule(
                    start_type="min",
                    end_type="max",
                    color=COLORS["light_green"],
                    showValue=True
                )
            )
            
            row += 2
        
        # ============================================================
        # ТОП МЕНЕДЖЕРОВ ПО СУММЕ ДОСТАВКИ (YTD)
        # ============================================================
        if managers_ytd:
            row = self.header.draw(row, "ТОП МЕНЕДЖЕРОВ ПО СУММЕ ДОСТАВКИ (YTD)", end_col=7)
            
            headers = ["МЕНЕДЖЕР", "ЗАКАЗОВ", "С ДОСТАВКОЙ", "СУММА ДОСТАВКИ", "% ОТ СУММЫ", "СР. ДОСТАВКА"]
            
            data_rows = []
            for m in managers_ytd:
                data_rows.append([
                    self._format_manager_name(m['manager']),
                    self._format_number(m['orders_count']),
                    f"{self._format_number_str(m['orders_with_delivery'])} ({self._format_percent(m['coverage_pct'])})",
                    m['total_delivery'],
                    m['delivery_percent'],
                    m['avg_delivery'],
                ])
            
            row = self.table.draw(row, headers, data_rows)
            
            # Дата-бар на колонке с суммой доставки
            start_col = 2
            amount_col_idx = headers.index("СУММА ДОСТАВКИ")
            amount_col_num = start_col + amount_col_idx
            amount_col_letter = get_column_letter(amount_col_num)
            
            first_data_row = row - len(data_rows)
            last_data_row = row - 1
            
            self.ws.conditional_formatting.add(
                f"{amount_col_letter}{first_data_row}:{amount_col_letter}{last_data_row}",
                DataBarRule(
                    start_type="min",
                    end_type="max",
                    color=COLORS["light_green"],
                    showValue=True
                )
            )
            
            row += 2
        
        # ============================================================
        # МЕСЯЧНЫЙ ТРЕНД
        # ============================================================
        if trends:
            row = self.header.draw(row, "МЕСЯЧНАЯ ДИНАМИКА ДОСТАВКИ", end_col=6)
            
            headers = ["МЕСЯЦ", "СУММА ДОСТАВКИ", "% ОТ СУММЫ ЗАКАЗА", "ЗАКАЗОВ С ДОСТАВКОЙ", "ДОЛЯ ЗАКАЗОВ"]
            
            data_rows = []
            for t in trends:
                month_name = self._format_month_name(t['month'])
                data_rows.append([
                    month_name,
                    t['delivery_sum'],
                    t['delivery_percent'],
                    self._format_number(t['orders_with_delivery']),
                    t['coverage_pct'],
                ])
            
            row = self.table.draw(row, headers, data_rows)
            
            # Дата-бар на сумме доставки
            start_col = 2
            amount_col_idx = headers.index("СУММА ДОСТАВКИ")
            amount_col_num = start_col + amount_col_idx
            amount_col_letter = get_column_letter(amount_col_num)
            
            first_data_row = row - len(data_rows)
            last_data_row = row - 1
            
            self.ws.conditional_formatting.add(
                f"{amount_col_letter}{first_data_row}:{amount_col_letter}{last_data_row}",
                DataBarRule(
                    start_type="min",
                    end_type="max",
                    color=COLORS["dark_green"],
                    showValue=True
                )
            )
            
            row += 2
        
        # ============================================================
        # ТОП КЛИЕНТОВ ПО ДОСТАВКЕ (MTD) - используем ту же таблицу, что и другие
        # ============================================================
        if top_clients:
            row = self.header.draw(row, "ТОП КЛИЕНТОВ ПО ДОСТАВКЕ (MTD)", end_col=5)
            
            headers = ["КЛИЕНТ", "ЗАКАЗОВ", "СУММА ДОСТАВКИ", "СР. ДОСТАВКА"]
            
            data_rows = []
            for c in top_clients:
                data_rows.append([
                    self._format_client_name(c['client']),
                    self._format_number(c['orders_count']),
                    c['total_delivery'],
                    c['avg_delivery'],
                ])
            
            row = self.table.draw(row, headers, data_rows)
            row += 2
        
        # ============================================================
        # ССЫЛКА НА ДЕТАЛИЗАЦИЮ
        # ============================================================
        nav_link = NavigationLink(self.ws)
        row = nav_link.draw(
            row=row,
            text="Подробную информацию по заказам смотрите на листе «Детализация заказов» →",
            target_sheet="2",
            target_cell="A1",
            start_col=2,
            end_col=6,
            alignment="left",
            with_icon=True,
            icon="📋"
        )
        
        # ============================================================
        # НАСТРОЙКА КОЛОНОК
        # ============================================================
        self.ws.column_dimensions["B"].width = 28
        self.ws.column_dimensions["C"].width = 18
        self.ws.column_dimensions["D"].width = 18
        self.ws.column_dimensions["E"].width = 18
        self.ws.column_dimensions["F"].width = 18
        self.ws.column_dimensions["G"].width = 20
        self.ws.column_dimensions["H"].width = 16
        
        self.ws.sheet_view.showGridLines = False