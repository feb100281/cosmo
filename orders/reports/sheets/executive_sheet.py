# orders/reports/sheets/executive_sheet.py
from openpyxl.styles import Font
from .base_sheet import BaseSheet
from ..styles.theme import COLORS, ALIGNMENTS, BORDERS, FILLS, Alignment
from ..components import create_kpi_cards, create_header, create_table
from ..styles.helpers import draw_toc_button
from ..components.alerts import create_alert
from ..components.sheet_title import create_sheet_title
from ..components.footnote import Footnote
from ..components.navigation_link import NavigationLink
from ..queries import ExecutiveQueries
from datetime import datetime
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter


class ExecutiveSheet(BaseSheet):
    """Executive Summary - профессиональный дашборд"""

    def __init__(self, workbook, sheet_number, request=None):
        super().__init__(workbook, sheet_number)
        self.sheet_title = create_sheet_title(self.ws)
        self.kpi = create_kpi_cards(self.ws)
        self.header = create_header(self.ws)
        self.table = create_table(self.ws)
        self.alert = create_alert(self.ws)
        self.queries = ExecutiveQueries(request)

    def _format_currency(self, value):
        if value is None or value == 0:
            return "0 ₽"
        return f"{int(round(value)):,} ₽".replace(",", " ")

    def _format_number(self, value):
        if value is None or value == 0:
            return "0"
        return f"{int(round(value)):,}".replace(",", " ")

    def _safe_float(self, value, default=0.0):
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default

    def _safe_int(self, value, default=0):
        if value is None:
            return default
        try:
            return int(round(float(value)))
        except (ValueError, TypeError):
            return default

    def _format_date(self, date_value):
        if not date_value:
            return ""
        if hasattr(date_value, 'strftime'):
            return date_value.strftime('%d.%m.%Y')
        return str(date_value)

    def _format_manager_name(self, full_name):
        """Форматирует имя менеджера: Иванов Иван -> ИВАНОВ И."""
        if not full_name:
            return ""
        parts = str(full_name).strip().split()
        if len(parts) >= 1:
            last_name = parts[0].upper()
            if len(parts) >= 2:
                first_initial = parts[1][0].upper() if parts[1] else ""
                return f"{last_name} {first_initial}."
            return last_name
        return str(full_name).upper()

    def _format_client_name(self, name):
        """Форматирует имя клиента: ООО Ромашка -> ООО РОМАШКА"""
        if not name:
            return ""
        return str(name).upper()

    def _format_month_name(self, month_date):
        """Возвращает название месяца в именительном падеже (короткий формат)"""
        if not month_date:
            return ""

        month_names = {
            1: "Янв", 2: "Фев", 3: "Мар", 4: "Апр",
            5: "Май", 6: "Июн", 7: "Июл", 8: "Авг",
            9: "Сен", 10: "Окт", 11: "Ноя", 12: "Дек"
        }

        if hasattr(month_date, 'month'):
            return f"{month_names[month_date.month]} {month_date.year}"
        return str(month_date)

    def _calculate_forecast(self, mtd_amount, monthly_trends):
        """
        Рассчитывает прогноз на месяц и динамику по среднедневному темпу.

        Логика:
        - текущий темп = MTD / число прошедших дней текущего месяца
        - прогноз = текущий темп * число дней в текущем месяце
        - динамика = сравнение текущего среднедневного темпа
          с последним завершенным месяцем
        """
        import calendar

        today = self.queries.today
        current_year = today.year
        current_month = today.month
        days_passed = today.day
        days_in_month = calendar.monthrange(current_year, current_month)[1]

        daily_rate = mtd_amount / days_passed if days_passed > 0 else 0
        forecast = daily_rate * days_in_month

        result = {
            'daily_rate': daily_rate,
            'forecast': forecast,
            'trend': None,
            'trend_pct': None,
            'trend_arrow': '→'
        }

        if not monthly_trends:
            return result

        # Надежно сортируем месяцы по дате
        sorted_trends = sorted(
            monthly_trends,
            key=lambda x: x.get('month') or datetime.min.date()
        )

        # Ищем последний завершенный месяц до текущего
        prev_month_data = None
        for item in reversed(sorted_trends):
            month_value = item.get('month')
            if not month_value:
                continue

            if month_value.year < current_year or (
                month_value.year == current_year and month_value.month < current_month
            ):
                prev_month_data = item
                break

        if not prev_month_data:
            return result

        prev_amount = self._safe_float(prev_month_data.get('total_amount'))
        prev_date = prev_month_data.get('month')

        if prev_amount <= 0 or not prev_date:
            return result

        prev_days = calendar.monthrange(prev_date.year, prev_date.month)[1]
        prev_daily = prev_amount / prev_days if prev_days > 0 else 0

        if prev_daily <= 0:
            return result

        pct = ((daily_rate - prev_daily) / prev_daily) * 100

        result['trend'] = pct
        result['trend_pct'] = abs(pct)
        result['trend_arrow'] = "▲" if pct > 0 else "▼" if pct < 0 else "◆"

        return result

    def build(self):
        """Построение дашборда - все данные берутся из ExecutiveQueries"""

        # ============================================================
        # ПОЛУЧЕНИЕ ДАННЫХ ИЗ ЗАПРОСОВ
        # ============================================================
        financial_metrics = self.queries.get_financial_metrics()
        monthly_trends = self.queries.get_monthly_trends()
        manager_summary = self.queries.get_manager_summary()
        top_clients = self.queries.get_top_clients(limit=5)

        active_stats = financial_metrics.get('active_stats', {})
        cancelled_stats = financial_metrics.get('cancelled_stats', {})
        completed_stats = financial_metrics.get('completed_stats', {})
        large_debts = financial_metrics.get('large_debts', [])
        last_days_stats = financial_metrics.get('last_days_stats', {})

        ALERT_PERIOD_DAYS = 1
        period_stats = last_days_stats.get(f'last_{ALERT_PERIOD_DAYS}_days', {})

        row = 1

        # ============================================================
        # ШАПКА
        # ============================================================
        draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

        # Немного воздуха между кнопкой и шапкой
        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 8

        row = self.sheet_title.draw(
            row=3,
            title="EXECUTIVE SUMMARY",
            subtitle="Аналитика активных заказов и ключевые показатели",
            date_text=f"Сформировано: {self.queries.today.strftime('%d.%m.%Y')}",
            start_col=2,
            end_col=6
        )

        row += 1

        # ============================================================
        # KPI КАРТОЧКИ
        # ============================================================
        active_amount = self._safe_float(active_stats.get('total_active_amount'))
        active_paid = self._safe_float(active_stats.get('total_active_paid'))
        paid_pct = (active_paid / active_amount * 100) if active_amount > 0 else 0
        active_remaining = max(0, active_amount - active_paid)

        row1_cards = [
            {
                'title': 'СУММА АКТИВНЫХ',
                'value': self._format_currency(active_amount),
                'subtitle': 'по всем активным',
                'color': COLORS["blue"],
                'width': 1
            },
            {
                'title': 'ОПЛАЧЕНО',
                'value': self._format_currency(active_paid),
                'subtitle': f"{paid_pct:.0f}% от суммы",
                'color': COLORS["dark_green"]
            },
            {
                'title': 'К ОПЛАТЕ',
                'value': self._format_currency(active_remaining),
                'subtitle': 'остаток по активным',
                'color': COLORS.get("warning_red", "C62828")
            },
        ]

        row2_cards = [
            {
                'title': 'АКТИВНЫЕ ЗАКАЗЫ',
                'value': f"{self._format_number(active_stats.get('total_active_count'))} заказов",
                'subtitle': 'в работе',
                'color': COLORS["dark_green"], 
                'width': 1
            },
            {
                'title': 'ОТМЕНЕНО YTD',
                'value': f"{self._format_number(cancelled_stats.get('ytd_count'))} заказов",
                'subtitle': f"на {self._format_currency(cancelled_stats.get('ytd_amount'))}",
                'color': COLORS.get("warning_red", "C62828")
            },
            {
                'title': 'ЗАКРЫТО YTD',
                'value': f"{self._format_number(completed_stats.get('ytd_count'))} заказов",
                'subtitle': f"на {self._format_currency(completed_stats.get('ytd_amount'))}",
                'color': COLORS["dark_green"]
            },
        ]

        row = self.kpi.draw_two_rows(row, row1_cards, row2_cards)

        # ============================================================
        # АЛЕРТЫ
        # ============================================================
        row += 1

        cancelled_period = period_stats.get('cancelled', {})
        if cancelled_period.get('count', 0) > 0:
            row = self.alert.draw_cancelled_alert(
                row,
                cancelled_period,
                period_days=ALERT_PERIOD_DAYS
            )

        new_orders_period = period_stats.get('new_orders', {})
        if new_orders_period.get('count', 0) > 0:
            row = self.alert.draw_new_orders_alert(
                row,
                new_orders_period,
                period_days=ALERT_PERIOD_DAYS
            )

        row += 1

        # ============================================================
        # КЛЮЧЕВЫЕ МЕТРИКИ
        # ============================================================
        row = self.header.draw(row, "КЛЮЧЕВЫЕ МЕТРИКИ")

        mtd_metrics = financial_metrics.get('mtd', {})
        ytd_metrics = financial_metrics.get('ytd', {})

        mtd_amount = self._safe_float(mtd_metrics.get('amount'))
        mtd_orders = self._safe_int(mtd_metrics.get('orders'))
        mtd_avg = mtd_amount / mtd_orders if mtd_orders > 0 else 0

        ytd_amount = self._safe_float(ytd_metrics.get('amount'))
        ytd_orders = self._safe_int(ytd_metrics.get('orders'))
        ytd_avg = ytd_amount / ytd_orders if ytd_orders > 0 else 0

        forecast = self._calculate_forecast(mtd_amount, monthly_trends)

        metrics_headers = ["МЕТРИКА", "MTD", "YTD", "ДИНАМИКА", "ПРОГНОЗ (МЕС.)"]
        metrics_data = [
            [
                "Сумма заказов",
                self._format_currency(mtd_amount),
                self._format_currency(ytd_amount),
                f"{forecast['trend_arrow']} {forecast['trend_pct']:.0f}%"
                if forecast['trend'] is not None else "—",
                self._format_currency(forecast['forecast'])
            ],
            [
                "Количество заказов",
                self._format_number(mtd_orders),
                self._format_number(ytd_orders),
                "—",
                "—"
            ],
            [
                "Средний чек",
                self._format_currency(mtd_avg),
                self._format_currency(ytd_avg),
                "—",
                "—"
            ],
        ]

        row = self.table.draw(row, metrics_headers, metrics_data)
        row += 1  
        footnote = Footnote(self.ws)
        row = footnote.draw(
            row=row,
            text="* Показаны результаты без учета полностью отмененных заказов"
        )
        row += 2

        # ============================================================
        # СТРУКТУРА АКТИВНЫХ ЗАКАЗОВ
        # ============================================================
        if active_stats:
            row = self.header.draw(
                        row, "СТРУКТУРА АКТИВНЫХ ЗАКАЗОВ", start_col=2, end_col=6)

            headers = ["СТАТУС", "ЗАКАЗОВ", "СУММА", "ОПЛАЧЕНО", "ОСТАТОК"]

            to_do_count = self._safe_int(active_stats.get('to_do_count'))
            to_do_qty = self._safe_int(active_stats.get('to_do_qty'))
            to_do_amount = self._safe_float(active_stats.get('to_do_amount'))
            to_do_paid = self._safe_float(active_stats.get('to_do_paid'))

            pending_count = self._safe_int(active_stats.get('pending_count'))
            pending_qty = self._safe_int(active_stats.get('pending_qty'))
            pending_amount = self._safe_float(active_stats.get('pending_amount'))
            pending_paid = self._safe_float(active_stats.get('pending_paid'))

            data_rows = [
                [
                    "К выполнению / В резерве",
                    f"{self._format_number(to_do_count)} шт\n{self._format_number(to_do_qty)} тов.",
                    to_do_amount,
                    to_do_paid,
                    max(0, to_do_amount - to_do_paid)
                ],
                [
                    "На согласовании",
                    f"{self._format_number(pending_count)} шт\n{self._format_number(pending_qty)} тов.",
                    pending_amount,
                    pending_paid,
                    max(0, pending_amount - pending_paid)
                ],
            ]

            table_start_row = row
            row = self.table.draw(row, headers, data_rows, highlight_cols=[4])

            # Немного увеличиваем высоту строк данных, потому что теперь в колонке "ЗАКАЗОВ" две строки
            for data_row_idx in range(table_start_row + 1, row):
                self.ws.row_dimensions[data_row_idx].height = 28

            total_count = self._safe_int(active_stats.get('total_active_count'))
            total_qty = self._safe_int(active_stats.get('total_active_qty'))
            total_amount = self._safe_float(active_stats.get('total_active_amount'))
            total_paid = self._safe_float(active_stats.get('total_active_paid'))
            total_remaining = max(0, total_amount - total_paid)

            total_row = [
                "ИТОГО",
                f"{self._format_number(total_count)} шт\n{self._format_number(total_qty)} тов.",
                total_amount,
                total_paid,
                total_remaining
            ]

            fill = FILLS["total"]
            for col_idx, value in enumerate(total_row):
                cell = self.ws.cell(row=row, column=col_idx + 2, value=value)
                cell.fill = fill
                cell.border = BORDERS["thin"]
                cell.font = Font(name="Roboto", bold=True, size=10)

                if col_idx == 0:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif col_idx == 1:
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0'

            self.ws.row_dimensions[row].height = 30
            row += 2
            
            
            # ============================================================
            # ССЫЛКА НА ДЕТАЛИЗАЦИЮ ЗАКАЗОВ
            # ============================================================
            nav_link = NavigationLink(self.ws)
            

            row = nav_link.draw(
                row=row,
                text="Подробно посмотреть активные заказы вы можете на листе «Детализация заказов» →",
                target_sheet="2",  # или какой номер у листа с детализацией
                target_cell="A1",
                start_col=2,
                end_col=6,
                alignment="left",
                with_icon=True,
                icon="📋"
            )
            
            # Ссылка на анализ доставки
            row = nav_link.draw(
                row=row,
                text="🚚 Анализ доставки смотрите на отдельном листе «Анализ доставки» →",
                target_sheet="3",
                target_cell="A1",
                start_col=2,
                end_col=6,
                alignment="left",
                with_icon=False
            )
            
            # ИЛИ Вариант 2: Ссылка в виде кнопки (раскомментируйте если нужен такой стиль)
            # row = nav_link.draw_as_button(
            #     row=row,
            #     text="Перейти к детализации заказов",
            #     target_sheet="3",
            #     target_cell="A1",
            #     start_col=2,
            #     end_col=4,
            #     bg_color="E8F0FE"
            # )
            
            row += 1  # дополнительный отступ после ссылки
            
        # ============================================================
        # МЕСЯЧНАЯ ДИНАМИКА
        # ============================================================
        if monthly_trends:
            row = self.header.draw(row, "МЕСЯЧНАЯ ДИНАМИКА")

            headers = ["МЕСЯЦ", "ЗАКАЗОВ", "СУММА", "СР.ЧЕК", "% ОПЛАТЫ"]

            sorted_trends = sorted(
                monthly_trends,
                key=lambda x: x.get('month') or datetime.min.date()
            )

            data_rows = []
            for month_data in sorted_trends[-6:]:
                month_amount = self._safe_float(month_data.get('total_amount'))
                month_paid = self._safe_float(month_data.get('total_paid'))
                payment_pct = (month_paid / month_amount * 100) if month_amount > 0 else 0

                month_name = self._format_month_name(month_data.get('month')) if month_data.get('month') else ""

                data_rows.append([
                    month_name,
                    self._safe_int(month_data.get('orders_count')),
                    month_amount,
                    self._safe_float(month_data.get('avg_check')),
                    f"{payment_pct:.0f}%"
                ])

            table_start_row = row
            row = self.table.draw(row, headers, data_rows)
            if data_rows:
                start_col = 2
                amount_col_idx = headers.index("СУММА")  # 2
                amount_col_num = start_col + amount_col_idx
                amount_col_letter = get_column_letter(amount_col_num)

                first_data_row = table_start_row + 1
                last_data_row = row - 1

                amount_range = f"{amount_col_letter}{first_data_row}:{amount_col_letter}{last_data_row}"

                self.ws.conditional_formatting.add(
                    amount_range,
                    DataBarRule(
                        start_type="min",
                        end_type="max",
                        color=COLORS["dark_green"],
                        showValue=True
                    )
                )
                
                row += 1  
                footnote = Footnote(self.ws)
                row = footnote.draw(
                    row=row,
                    text="* Показаны результаты без учета полностью отмененных заказов"
                )

            row += 2

        # ============================================================
        # КРУПНЫЕ ОТМЕНЫ
        # ============================================================
        top_cancelled = cancelled_stats.get('top_cancelled', [])
        if top_cancelled:
            row = self.header.draw(row, "КРУПНЫЕ ОТМЕНЫ")

            headers = ["ЗАКАЗ", "МАГАЗИН", "МЕНЕДЖЕР", "ПОКУПАТЕЛЬ", "СУММА"]

            data_rows = []
            for canc in top_cancelled[:8]:
                order_display = canc.get('order_name', '')
                if not order_display and canc.get('number'):
                    date_part = self._format_date(canc.get('date_from'))
                    order_display = f"{canc.get('number')} ({date_part})" if date_part else canc.get('number')

                reason = (canc.get('cancellation_reason') or '').strip()
                amount_value = self._safe_float(canc.get('cancelled_amount'))

                amount_display = self._format_currency(amount_value)
                if reason:
                    amount_display = f"{amount_display}\n{reason}"

                data_rows.append([
                    order_display,
                    (canc.get('store', '') or '')[:25],
                    self._format_manager_name(canc.get('manager', '')),
                    self._format_client_name(canc.get('client', '')),
                    amount_display
                ])

            table_start_row = row
            row = self.table.draw(row, headers, data_rows)

            for data_row_idx in range(table_start_row + 1, row):
                self.ws.row_dimensions[data_row_idx].height = 32
                
            row += 1  
            footnote = Footnote(self.ws)
            row = footnote.draw(
                    row=row,
                    text="* С начала текущего года"
                )

            row += 2
            
            
            

        # # ============================================================
        # # КРУПНЫЕ ДОЛГИ
        # # ============================================================
        # if large_debts:
        #     row = self.header.draw(row, "КРУПНЫЕ ДОЛГИ (БОЛЕЕ 1 МЛН ₽)")

        #     headers = ["ЗАКАЗ", "КЛИЕНТ", "МЕНЕДЖЕР", "СУММА", "К ОПЛАТЕ"]

        #     data_rows = []
        #     for debt in large_debts[:6]:
        #         data_rows.append([
        #             debt.get('number', ''),
        #             self._format_client_name(debt.get('client', '')),
        #             self._format_manager_name(debt.get('manager', '')),
        #             self._safe_float(debt.get('amount_active')),
        #             self._safe_float(debt.get('debt'))
        #         ])

        #     row = self.table.draw(row, headers, data_rows, highlight_cols=[4])
        #     row += 2

        # # ============================================================
        # # ТОП КЛИЕНТОВ
        # # ============================================================
        # if top_clients:
        #     row = self.header.draw(row, "ТОП КЛИЕНТОВ ПО СУММЕ ДОЛГА")

        #     headers = ["КЛИЕНТ", "ЗАКАЗОВ", "СУММА", "ДОЛГ"]

        #     data_rows = []
        #     for client in top_clients[:5]:
        #         total = self._safe_float(client.get('total_amount'))
        #         paid = self._safe_float(client.get('paid_amount', 0))
        #         debt = max(0, total - paid)

        #         data_rows.append([
        #             self._format_client_name(client.get('client', '')),
        #             self._safe_int(client.get('orders_count')),
        #             total,
        #             debt
        #         ])

        #     row = self.table.draw(row, headers, data_rows, highlight_cols=[3])
        #     row += 2

        
        # ============================================================
        # ЭФФЕКТИВНОСТЬ МЕНЕДЖЕРОВ - МЕСЯЦ
        # ============================================================
        if manager_summary:
            row = self.header.draw(row, "ПОКАЗАТЕЛИ МЕНЕДЖЕРОВ (MTD)")
            
            # 5 колонок: Менеджер | Заказов | Сумма | Средний чек | Отгрузка / Остаток
            headers = ["МЕНЕДЖЕР", "ЗАКАЗОВ", "СУММА", "СР.ЧЕК", "ОТГР. / ОСТАТОК"]
            
            data_rows = []
            for manager in manager_summary[:10]:  # топ-10 за месяц
                orders_count = manager['mtd_orders']
                order_amount = manager['mtd_amount']
                avg_check = order_amount / orders_count if orders_count > 0 else 0
                
                shipped_amount = manager['mtd_shipped_amount']
                remaining_amount = max(0, order_amount - shipped_amount)
                
                data_rows.append([
                    self._format_manager_name(manager['manager']),
                    orders_count,
                    order_amount,
                    avg_check,
                    f"📦 {self._format_currency(shipped_amount)}\n⚡ {self._format_currency(remaining_amount)}"
                ])
            
    
            row = self.table.draw(row, headers, data_rows)
            row += 1  
            footnote = Footnote(self.ws)
            row = footnote.draw(
                    row=row,
                    text="* Показаны ТОП-10 менеджеров"
                )
            row += 2
            
            # ============================================================
            # ЭФФЕКТИВНОСТЬ МЕНЕДЖЕРОВ - ГОД
            # ============================================================
            row = self.header.draw(row, "ПОКАЗАТЕЛИ МЕНЕДЖЕРОВ (YTD)")
            
            data_rows = []
            for manager in manager_summary[:10]:  # топ-10 за год
                orders_count = manager['ytd_orders']
                order_amount = manager['ytd_amount']
                avg_check = order_amount / orders_count if orders_count > 0 else 0
                
                shipped_amount = manager['ytd_shipped_amount']
                remaining_amount = max(0, order_amount - shipped_amount)
                
                data_rows.append([
                    self._format_manager_name(manager['manager']),
                    orders_count,
                    order_amount,
                    avg_check,
                    f"📦 {self._format_currency(shipped_amount)}\n⚡ {self._format_currency(remaining_amount)}"
                ])
            
        
            row = self.table.draw(row, headers, data_rows)
            row += 1  
            footnote = Footnote(self.ws)
            row = footnote.draw(
                    row=row,
                    text="* Показаны ТОП-10 менеджеров"
                )
            row += 2
            
            
        row = nav_link.draw(
        row=row,
        text="Подробная информация по каждому менеджеру доступна на листе «Детализация по менеджерам» →",
        target_sheet="4", 
        target_cell="A1",
        start_col=2,
        end_col=6,  
        alignment="left",
        with_icon=True,
        icon="📋"
    )
    
    
        # ============================================================
        # НАСТРОЙКА КОЛОНОК
        # ============================================================
        self.ws.column_dimensions["A"].width = 2
        self.ws.column_dimensions["B"].width = 28
        self.ws.column_dimensions["C"].width = 14
        self.ws.column_dimensions["D"].width = 16
        self.ws.column_dimensions["E"].width = 16
        self.ws.column_dimensions["F"].width = 16
        self.ws.column_dimensions["G"].width = 16
        self.ws.column_dimensions["H"].width = 20
        self.ws.column_dimensions["I"].width = 16
        self.ws.column_dimensions["J"].width = 16
        self.ws.column_dimensions["K"].width = 14
        self.ws.column_dimensions["L"].width = 14
        self.ws.column_dimensions["M"].width = 14
        self.ws.column_dimensions["N"].width = 16

        self.ws.sheet_view.showGridLines = False




