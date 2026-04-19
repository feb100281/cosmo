# orders/reports/sheets/debt_analysis_sheet.py
from datetime import datetime
from django.utils import timezone
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .base_sheet import BaseSheet
from ..styles.theme import COLORS, ALIGNMENTS, BORDERS, FILLS
from ..components import create_kpi_cards, create_header, create_table
from ..styles.helpers import draw_toc_button
from ..components.footnote import Footnote
from ..components.sheet_title import create_sheet_title


class DebtAnalysisSheet(BaseSheet):
    """Лист с анализом дебиторской задолженности"""

    def __init__(self, workbook, sheet_number, request=None):
        super().__init__(workbook, sheet_number)
        self.request = request
        self.sheet_title = create_sheet_title(self.ws)
        self.kpi = create_kpi_cards(self.ws)
        self.header = create_header(self.ws)
        self.table = create_table(self.ws)

    def _format_currency(self, value):
        if value is None or value == 0:
            return "0 ₽"
        return f"{int(round(value)):,} ₽".replace(",", " ")

    def _format_number(self, value):
        if value is None or value == 0:
            return "0"
        return f"{int(round(value)):,}".replace(",", " ")

    def _safe_float(self, value, default=0.0):
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default

    def _safe_int(self, value, default=0):
        if value is None:
            return default
        try:
            return int(round(float(value)))
        except (ValueError, TypeError):
            return default

    def _format_datetime(self, date_value):
        """Форматирует дату без времени: YYYY-MM-DD"""
        if not date_value:
            return ""
        if hasattr(date_value, 'strftime'):
            return date_value.strftime('%Y-%m-%d')
        if isinstance(date_value, str):
            for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d.%m.%Y']:
                try:
                    dt = datetime.strptime(date_value, fmt)
                    return dt.strftime('%Y-%m-%d')
                except ValueError:
                    continue
        return str(date_value)[:10] if len(str(date_value)) > 10 else str(date_value)

    def _format_manager_name(self, full_name):
        """Форматирует имя менеджера: Иванов Иван -> ИВАНОВ И."""
        if not full_name:
            return ""
        parts = str(full_name).strip().split()
        if len(parts) >= 1:
            last_name = parts[0].upper()
            if len(parts) >= 2:
                first_initial = parts[1][0].upper() if parts[1] else ""
                return f"{last_name} {first_initial}."
            return last_name
        return str(full_name).upper()

    def _format_client_name(self, name):
        """Форматирует имя клиента"""
        if not name:
            return ""
        return str(name).upper()

    def _get_debt_status(self, debt_percentage, days_old=None):
        """Текстовый статус долга"""
        if debt_percentage >= 90:
            return "КРИТИЧЕСКИЙ"
        elif debt_percentage >= 50:
            return "ВЫСОКИЙ"
        elif debt_percentage >= 25:
            return "СРЕДНИЙ"
        elif debt_percentage > 0:
            return "НОРМАЛЬНЫЙ"
        return "ОПЛАЧЕНО"

    def _get_days_old(self, date_value):
        """Рассчитывает количество дней с даты создания заказа"""
        if not date_value:
            return 0
        
        try:
            # Если это datetime объект
            if hasattr(date_value, 'date'):
                date_obj = date_value.date()
            else:
                date_obj = date_value
            
            today = timezone.now().date()
            return (today - date_obj).days
        except Exception:
            return 0

    def _format_payment_dates(self, payment_dates):
        """Форматирует даты оплат для отображения"""
        if not payment_dates:
            return "—"
        
        if isinstance(payment_dates, str):
            import re
            pattern = r'(\d{4}-\d{2}-\d{2}):\s*([\d\.]+)\s*руб'
            matches = re.findall(pattern, payment_dates, re.IGNORECASE)
            
            if matches:
                formatted = []
                for date, amount_str in matches[:3]:
                    try:
                        amount = float(amount_str)
                        formatted.append(f"{date}: {self._format_currency(amount)}")
                    except ValueError:
                        formatted.append(f"{date}: {amount_str} руб")
                
                result = "\n".join(formatted)
                if len(matches) > 3:
                    result += f"\n+ еще {len(matches) - 3}"
                return result
        
        return str(payment_dates)[:50]

    def build(self, orders_data, summary_data=None):
        """Построение листа с анализом задолженности"""
        
        # Защита от пустых данных
        if orders_data is None:
            orders_data = []
        
        row = 1

        # ============================================================
        # КНОПКА НАЗАД
        # ============================================================
        draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 8

        # ============================================================
        # ШАПКА
        # ============================================================
        row = self.sheet_title.draw(
            row=3,
            title="ПРОСРОЧЕННАЯ ДЕБИТОРСКАЯ ЗАДОЛЖЕННОСТЬ",
            subtitle="Все заказы с остатком к оплате > 0 (независимо от статуса)",
            date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            start_col=2,
            end_col=16
        )

        row += 1

        # ============================================================
        # KPI КАРТОЧКИ
        # ============================================================
        if summary_data:
            total_debt_orders = summary_data.get('total_debt_orders', 0)
            total_debt_amount = summary_data.get('total_debt_amount', 0)
            avg_debt = summary_data.get('avg_debt', 0)
            max_debt = summary_data.get('max_debt', 0)
            unique_clients = summary_data.get('unique_clients', 0)
            old_debts_count = summary_data.get('old_debts', {}).get('count', 0)
            old_debts_amount = summary_data.get('old_debts', {}).get('amount', 0)
        else:
            total_debt_orders = len(orders_data)
            total_debt_amount = sum(self._safe_float(o.get('remaining_debt', 0)) for o in orders_data)
            avg_debt = total_debt_amount / total_debt_orders if total_debt_orders > 0 else 0
            max_debt = max([self._safe_float(o.get('remaining_debt', 0)) for o in orders_data]) if orders_data else 0
            unique_clients = len(set(o.get('client', '') for o in orders_data if o.get('client')))
            old_debts = [o for o in orders_data if self._get_days_old(o.get('date_from')) > 90]
            old_debts_count = len(old_debts)
            old_debts_amount = sum(self._safe_float(o.get('remaining_debt', 0)) for o in old_debts)

        # Первый ряд KPI
        row1_cards = [
            {
                'title': 'ДОЛЖНИКИ',
                'value': self._format_number(total_debt_orders),
                'subtitle': 'заказов с долгом',
                'color': COLORS.get("danger_red", "C62828"),
                'width': 1
            },
            {
                'title': 'СУММА ДОЛГА',
                'value': self._format_currency(total_debt_amount),
                'subtitle': 'общая задолженность',
                'color': COLORS.get("danger_red", "C62828")
            },
            {
                'title': 'СРЕДНИЙ ДОЛГ',
                'value': self._format_currency(avg_debt),
                'subtitle': 'на один заказ',
                'color': COLORS.get("warning_orange", "F57C00"),
                'width': 1
            },
        ]

        row = self.kpi.draw_two_rows(row, row1_cards, [])
        row -= 2

        # Второй ряд KPI
        row2_cards = [
            {
                'title': 'МАКС. ДОЛГ',
                'value': self._format_currency(max_debt),
                'subtitle': 'максимальная сумма',
                'color': COLORS.get("danger_red", "C62828"),
                'width': 1
            },
            {
                'title': 'КЛИЕНТЫ-ДОЛЖНИКИ',
                'value': self._format_number(unique_clients),
                'subtitle': 'уникальных клиентов',
                'color': COLORS.get("warning_orange", "F57C00")
            },
            {
                'title': 'СТАРЫЕ ДОЛГИ (>90 дн)',
                'value': f"{self._format_number(old_debts_count)} / {self._format_currency(old_debts_amount)}",
                'subtitle': 'заказов / сумма',
                'color': COLORS.get("danger_red", "C62828"),
                'width': 1
            },
        ]

        row = self.kpi.draw_two_rows(row, row2_cards, [])
        row -= 2
        row += 1

        # ============================================================
        # ВОЗРАСТНАЯ СТРУКТУРА
        # ============================================================
        if summary_data and summary_data.get('aging_buckets') and total_debt_orders > 0:
            row = self.header.draw(row, "ВОЗРАСТНАЯ СТРУКТУРА ЗАДОЛЖЕННОСТИ", start_col=2, end_col=16)
            row += 1
            
            aging_data = summary_data['aging_buckets']
            
            headers = ["Категория", "Кол-во заказов", "Сумма долга", "Доля"]
            for col_idx, header in enumerate(headers, start=2):
                cell = self.ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=COLORS.get("dark_green", "1B5E20"))
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = BORDERS["thin"]
            
            row += 1
            
            total_aging_amount = sum(b['amount'] for b in aging_data.values())
            
            for category, data in aging_data.items():
                if data['count'] > 0 or data['amount'] > 0:
                    values = [
                        category,
                        data['count'],
                        data['amount'],
                        (data['amount'] / total_aging_amount * 100) if total_aging_amount > 0 else 0
                    ]
                    
                    for col_idx, value in enumerate(values, start=2):
                        cell = self.ws.cell(row=row, column=col_idx, value=value)
                        cell.border = BORDERS["thin"]
                        cell.font = Font(name="Roboto", size=9)
                        
                        if col_idx == 2:
                            cell.alignment = Alignment(horizontal="center")
                            if isinstance(value, (int, float)):
                                cell.number_format = '#,##0'
                        elif col_idx == 3:
                            cell.alignment = Alignment(horizontal="right")
                            if isinstance(value, (int, float)):
                                cell.number_format = '#,##0'
                        elif col_idx == 4:
                            cell.alignment = Alignment(horizontal="right")
                            cell.number_format = '0.00"%"'
                        
                        if category in ['91-180 дней', 'более 180 дней'] and data['amount'] > 0:
                            cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS.get("danger_red", "C62828"))
                    
                    row += 1
            
            # Итого
            total_row = ["ИТОГО:", total_debt_orders, total_debt_amount, 100]
            for col_idx, value in enumerate(total_row, start=2):
                cell = self.ws.cell(row=row, column=col_idx, value=value)
                cell.fill = PatternFill("solid", fgColor=COLORS.get("light_gray", "F5F5F5"))
                cell.font = Font(name="Roboto", bold=True, size=9)
                cell.border = BORDERS["thin"]
                if col_idx in [3, 4]:
                    cell.alignment = Alignment(horizontal="right")
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = Alignment(horizontal="center")
            
            row += 2

        # ============================================================
        # ТОП-10 ДОЛЖНИКОВ
        # ============================================================
        if summary_data and summary_data.get('top_debtors') and len(summary_data['top_debtors']) > 0:
            row = self.header.draw(row, "ТОП-10 КЛИЕНТОВ-ДОЛЖНИКОВ", start_col=2, end_col=16)
            row += 1
            
            headers = ["Клиент", "Сумма долга", "Кол-во заказов"]
            for col_idx, header in enumerate(headers, start=2):
                cell = self.ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=COLORS.get("dark_green", "1B5E20"))
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = BORDERS["thin"]
            
            row += 1
            
            for debtor in summary_data['top_debtors'][:10]:
                values = [
                    self._format_client_name(debtor.get('client', '')),
                    debtor.get('total_debt', 0),
                    debtor.get('orders_count', 0)
                ]
                
                for col_idx, value in enumerate(values, start=2):
                    cell = self.ws.cell(row=row, column=col_idx, value=value)
                    cell.border = BORDERS["thin"]
                    cell.font = Font(name="Roboto", size=9)
                    
                    if col_idx == 2:
                        cell.alignment = Alignment(horizontal="left")
                    elif col_idx == 3:
                        cell.alignment = Alignment(horizontal="right")
                        if isinstance(value, (int, float)):
                            cell.number_format = '#,##0'
                    else:
                        cell.alignment = Alignment(horizontal="center")
                
                row += 1
            
            row += 1

        # ============================================================
        # ТАБЛИЦА С ЗАКАЗАМИ
        # ============================================================
        row = self.header.draw(row, "ДЕТАЛЬНЫЙ ПЕРЕЧЕНЬ ЗАКАЗОВ С ЗАДОЛЖЕННОСТЬЮ", start_col=2, end_col=16)
        row += 1

        if len(orders_data) == 0:
            # Нет данных
            cell = self.ws.cell(row=row, column=2, value="Нет заказов с задолженностью")
            cell.font = Font(name="Roboto", size=12, bold=True, color="666666")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
            self.ws.row_dimensions[row].height = 50
        else:
            # Заголовки таблицы
            headers = [
                "ЗАКАЗ",
                "ДАТА\nСОЗДАНИЯ",
                "СТАТУС",
                "КЛИЕНТ",
                "МЕНЕДЖЕР",
                "СУММА\nЗАКАЗА",
                "ОПЛАЧЕНО",
                "СУММА\nДОЛГА",
                "ДОЛЯ\nДОЛГА",
                "СТАТУС\nДОЛГА",
                "ВОЗРАСТ\n(ДНЕЙ)",
                "ДАТЫ ОПЛАТ",
            ]

            table_start_row = row

            for col_idx, header in enumerate(headers, start=2):
                cell = self.ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=COLORS.get("danger_red", "C62828"))
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = BORDERS["thin"]

            self.ws.row_dimensions[row].height = 40
            row += 1

            # Данные
            for idx, order in enumerate(orders_data):
                if idx % 2 == 0:
                    fill = FILLS.get("odd_row", PatternFill(fill_type=None))
                else:
                    fill = FILLS.get("even_row", PatternFill(fill_type=None))

                debt_amount = self._safe_float(order.get('remaining_debt', 0))
                debt_percentage = self._safe_float(order.get('debt_percentage', 0))
                order_amount = self._safe_float(order.get('amount_active', 0))
                paid_amount = self._safe_float(order.get('cash_pmts', 0))
                days_old = self._get_days_old(order.get('date_from'))
                debt_status = self._get_debt_status(debt_percentage, days_old)
                
                values = [
                    order.get('order_name', '') or order.get('number', ''),
                    self._format_datetime(order.get('date_from')),
                    order.get('status', ''),
                    self._format_client_name(order.get('client', '')),
                    self._format_manager_name(order.get('manager', '')),
                    order_amount,
                    paid_amount,
                    debt_amount,
                    debt_percentage,
                    debt_status,
                    days_old if days_old > 0 else 0,
                    self._format_payment_dates(order.get('payment_dates')),
                ]

                for col_idx, value in enumerate(values):
                    col_num = col_idx + 2
                    cell = self.ws.cell(row=row, column=col_num, value=value)
                    cell.fill = fill
                    cell.border = BORDERS["thin"]
                    
                    base_color = COLORS.get("text_dark", "1F1F1F")
                    
                    if debt_percentage >= 50 or (days_old > 90 and debt_amount > 0):
                        base_color = COLORS.get("danger_red", "C62828")
                        cell.font = Font(name="Roboto", size=9, bold=(debt_percentage >= 90), color=base_color)
                    else:
                        cell.font = Font(name="Roboto", size=9, color=base_color)

                    if col_idx in [0, 3, 4]:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    elif col_idx in [1, 2, 9, 10]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_idx == 11:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        cell.font = Font(name="Roboto", size=8)
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        if isinstance(value, (int, float)):
                            if col_idx in [5, 6, 7]:
                                cell.number_format = '#,##0'
                            elif col_idx == 8:
                                cell.number_format = '0.00"%"'

                    if col_idx == 9:
                        if debt_status == "КРИТИЧЕСКИЙ":
                            cell.fill = PatternFill("solid", fgColor="FFCDD2")
                            cell.font = Font(name="Roboto", size=9, bold=True, color="C62828")
                        elif debt_status == "ВЫСОКИЙ":
                            cell.fill = PatternFill("solid", fgColor="FFE0B2")
                            cell.font = Font(name="Roboto", size=9, bold=True, color="E65100")

                self.ws.row_dimensions[row].height = 28
                row += 1

            # Итоговая строка
            total_order_amount = sum(self._safe_float(o.get('amount_active', 0)) for o in orders_data)
            total_paid = sum(self._safe_float(o.get('cash_pmts', 0)) for o in orders_data)
            total_debt = sum(self._safe_float(o.get('remaining_debt', 0)) for o in orders_data)

            total_row = ["ИТОГО:", "", "", "", "", total_order_amount, total_paid, total_debt, "", "", "", ""]

            fill_total = PatternFill("solid", fgColor=COLORS.get("light_gray", "F5F5F5"))
            for col_idx, value in enumerate(total_row):
                cell = self.ws.cell(row=row, column=col_idx + 2, value=value)
                cell.fill = fill_total
                cell.border = BORDERS["thin"]
                cell.font = Font(name="Roboto", bold=True, size=10)

                if col_idx == 0:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_idx in [5, 6, 7]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            self.ws.row_dimensions[row].height = 30
            row += 2

            # Настройка колонок
            self.ws.column_dimensions["A"].width = 2
            self.ws.column_dimensions["B"].width = 28
            self.ws.column_dimensions["C"].width = 12
            self.ws.column_dimensions["D"].width = 22
            self.ws.column_dimensions["E"].width = 30
            self.ws.column_dimensions["F"].width = 20
            self.ws.column_dimensions["G"].width = 16
            self.ws.column_dimensions["H"].width = 14
            self.ws.column_dimensions["I"].width = 14
            self.ws.column_dimensions["J"].width = 12
            self.ws.column_dimensions["K"].width = 16
            self.ws.column_dimensions["L"].width = 12
            self.ws.column_dimensions["M"].width = 22

            # Заморозка
            self.ws.freeze_panes = f'G{table_start_row + 1}'
            
            # Автофильтр
            self.ws.auto_filter.ref = f'B{table_start_row}:M{row - 2}'

        # Примечание
        footnote = Footnote(self.ws)
        footnote.draw(
            row=row,
            text="* Долг считается как разница между суммой заказа и полученными оплатами. "
                 "Красным выделены критически просроченные долги (>50% от суммы или >90 дней)."
        )
        
        # Скрываем сетку
        self.ws.sheet_view.showGridLines = False