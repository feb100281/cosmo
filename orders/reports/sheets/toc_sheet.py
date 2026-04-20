# orders/reports/sheets/toc_sheet.py
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .base_sheet import BaseSheet
from ..styles.theme import FILLS, FONTS, BORDERS, ALIGNMENTS, COLORS


class TOCSheet(BaseSheet):
    def __init__(self, workbook):
        super().__init__(workbook, "TOC")  # Лист называется "TOC"
        # Удаляем дефолтный лист, если он есть
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

    def _draw_separator(self, row, start_col, end_col, color=COLORS["border_gray"]):
        """Рисует разделительную линию"""
        for col in range(start_col, end_col + 1):
            cell = self.ws.cell(row=row, column=col)
            cell.border = Border(bottom=Side(style="thin", color=color))

    def build(self, sheets_info, request=None, filters=None):
        row = 1
        start_col = 2  # B
        end_col = 4    # D

        # -------------------------------------------------
        # ЗАГОЛОВОК
        # -------------------------------------------------
        # Декоративная полоса сверху
        self._draw_separator(row, start_col, end_col, COLORS["dark_green"])
        top_separator_row = row
        row += 1

        # Главный заголовок
        title_row = row
        title_cell = self.ws.cell(row=row, column=start_col, value="ОТЧЕТ ПО ЗАКАЗАМ")
        title_cell.font = Font(name="Roboto", size=20, bold=True, color=COLORS["dark_green"])
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        # Подзаголовок
        subtitle_row = row
        subtitle_cell = self.ws.cell(row=row, column=start_col, value="Аналитика и детализация заказов")
        subtitle_cell.font = Font(name="Roboto", size=11, color=COLORS["text_gray"])
        subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        # Дата и время генерации
        generated_row = row
        generated_text = f"Сформировано: {datetime.now().strftime('%d.%m.%Y в %H:%M')}"
        date_cell = self.ws.cell(row=row, column=start_col, value=generated_text)
        date_cell.font = Font(name="Roboto", size=9, italic=True, color=COLORS["text_gray"])
        date_cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 2

        # Декоративная полоса
        middle_separator_row = row
        self._draw_separator(row, start_col, end_col, COLORS["light_green"])
        row += 1

        # -------------------------------------------------
        # НАВИГАЦИЯ ПО ОТЧЕТУ
        # -------------------------------------------------
        nav_title_row = row
        nav_cell = self.ws.cell(row=row, column=start_col, value="НАВИГАЦИЯ ПО ОТЧЕТУ")
        nav_cell.font = Font(name="Roboto", size=13, bold=True, color=COLORS["dark_green"])
        nav_cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        nav_desc_row = row
        desc_cell = self.ws.cell(row=row, column=start_col, value="Кликните на название раздела, чтобы перейти к нему")
        desc_cell.font = Font(name="Roboto", size=9, italic=True, color=COLORS["text_gray"])
        desc_cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 2

        # -------------------------------------------------
        # ТАБЛИЦА ОГЛАВЛЕНИЯ
        # -------------------------------------------------
        headers = ["№", "РАЗДЕЛ", "ОПИСАНИЕ"]
        table_header_row = row

        # Заголовки таблицы
        for col_idx, header in enumerate(headers, start_col):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["white"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(
                start_color=COLORS["dark_green"],
                end_color=COLORS["dark_green"],
                fill_type="solid"
            )
            cell.border = BORDERS["thin"]

        row += 1
        table_data_start_row = row

        # Данные таблицы
        for sheet in sheets_info:
            # Номер раздела
            cell_num = self.ws.cell(
                row=row,
                column=start_col,
                value=f"0{sheet['number']}" if sheet["number"] < 10 else sheet["number"]
            )
            cell_num.font = Font(name="Roboto", size=10, bold=True, color=COLORS["dark_green"])
            cell_num.alignment = Alignment(horizontal="center", vertical="center")
            cell_num.border = BORDERS["thin"]
            cell_num.fill = PatternFill(
                start_color=COLORS["light_green"],
                end_color=COLORS["light_green"],
                fill_type="solid"
            )

            # Название раздела с гиперссылкой
            cell_name = self.ws.cell(row=row, column=start_col + 1, value=sheet["name"])
            cell_name.font = Font(name="Roboto", size=10, bold=True, color=COLORS["blue"])
            cell_name.alignment = Alignment(horizontal="left", vertical="center")
            cell_name.border = BORDERS["thin"]
            cell_name.hyperlink = f"#'{sheet['number']}'!A1"

            # Описание раздела
            cell_desc = self.ws.cell(row=row, column=start_col + 2, value=sheet.get("description", ""))
            cell_desc.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
            cell_desc.alignment = Alignment(horizontal="left", vertical="center")
            cell_desc.border = BORDERS["thin"]

            # Чередование цвета строки для столбцов C и D
            row_fill_color = COLORS["light_gray"] if sheet["number"] % 2 == 0 else COLORS["white"]

            cell_name.fill = PatternFill(
                start_color=row_fill_color,
                end_color=row_fill_color,
                fill_type="solid"
            )
            cell_desc.fill = PatternFill(
                start_color=row_fill_color,
                end_color=row_fill_color,
                fill_type="solid"
            )

            row += 1

        table_data_end_row = row - 1
        table_bottom_border_row = row

        # Нижняя граница таблицы
        for col in range(start_col, start_col + 3):
            cell = self.ws.cell(row=table_bottom_border_row, column=col)
            cell.border = Border(bottom=Side(style="medium", color=COLORS["dark_green"]))

        row += 1

        # -------------------------------------------------
        # ИТОГОВАЯ СТАТИСТИКА
        # -------------------------------------------------
        total_row = row
        total_text = f"Всего разделов в отчете: {len(sheets_info)}"
        total_cell = self.ws.cell(row=total_row, column=start_col, value=total_text)
        total_cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["white"])
        total_cell.fill = PatternFill(
            start_color=COLORS["dark_green"],
            end_color=COLORS["dark_green"],
            fill_type="solid"
        )
        total_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Объединяем ячейки для итоговой строки
        self.ws.merge_cells(
            start_row=total_row,
            start_column=start_col,
            end_row=total_row,
            end_column=start_col + 2
        )

        row += 2

        # -------------------------------------------------
        # ДОПОЛНИТЕЛЬНАЯ ИНФОРМАЦИЯ
        # -------------------------------------------------
        info_title_row = row
        info_cell = self.ws.cell(row=row, column=start_col, value="ИНФОРМАЦИЯ")
        info_cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["dark_green"])
        info_cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        info_items_start_row = row
        info_items = [
            "• Данные актуальны на момент формирования отчета",
            "• Для перехода к разделу кликните на его название",
            "• Отчет включает анализ заказов, доставки, оплат и клиентов",
        ]

        for item in info_items:
            item_cell = self.ws.cell(row=row, column=start_col + 1, value=item)
            item_cell.font = Font(name="Roboto", size=8, color=COLORS["text_gray"])
            item_cell.alignment = Alignment(horizontal="left", vertical="center")
            row += 1

        info_items_end_row = row - 1

        row += 1

        # -------------------------------------------------
        # ДЕКОРАТИВНЫЙ БЛОК
        # -------------------------------------------------
        bottom_separator_row = row
        self._draw_separator(row, start_col, end_col, COLORS["light_green"])
        row += 1

        thanks_row = row
        thanks_cell = self.ws.cell(row=row, column=start_col, value="✨ Спасибо за использование отчета ✨")
        thanks_cell.font = Font(name="Roboto", size=9, italic=True, color=COLORS["text_gray"])
        thanks_cell.alignment = Alignment(horizontal="left", vertical="center")

        # -------------------------------------------------
        # ПЕРЕНОС ТЕКСТА
        # -------------------------------------------------
        for r in range(table_data_start_row, table_data_end_row + 1):
            self.ws.cell(r, start_col + 1).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True
            )
            self.ws.cell(r, start_col + 2).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True
            )

        # -------------------------------------------------
        # НАСТРОЙКА ШИРИНЫ КОЛОНОК
        # -------------------------------------------------
        self.ws.column_dimensions["A"].width = 3
        self.ws.column_dimensions["B"].width = 10   # №
        self.ws.column_dimensions["C"].width = 35   # РАЗДЕЛ
        self.ws.column_dimensions["D"].width = 55   # ОПИСАНИЕ
        self.ws.column_dimensions["E"].width = 20

        # -------------------------------------------------
        # НАСТРОЙКА ВЫСОТЫ СТРОК
        # -------------------------------------------------
        self.ws.row_dimensions[top_separator_row].height = 10
        self.ws.row_dimensions[title_row].height = 35
        self.ws.row_dimensions[subtitle_row].height = 20
        self.ws.row_dimensions[generated_row].height = 20
        self.ws.row_dimensions[middle_separator_row].height = 10
        self.ws.row_dimensions[nav_title_row].height = 25
        self.ws.row_dimensions[nav_desc_row].height = 18

        # Заголовок таблицы
        self.ws.row_dimensions[table_header_row].height = 28

        # Все строки таблицы оглавления
        for r in range(table_data_start_row, table_data_end_row + 1):
            self.ws.row_dimensions[r].height = 24

        # Строка с нижней границей таблицы
        self.ws.row_dimensions[table_bottom_border_row].height = 6

        # Итоговая строка
        self.ws.row_dimensions[total_row].height = 28

        # Блок информации
        self.ws.row_dimensions[info_title_row].height = 20
        for r in range(info_items_start_row, info_items_end_row + 1):
            self.ws.row_dimensions[r].height = 18

        # Нижний декоративный блок
        self.ws.row_dimensions[bottom_separator_row].height = 10
        self.ws.row_dimensions[thanks_row].height = 20

        # -------------------------------------------------
        # СКРЫВАЕМ СЕТКУ
        # -------------------------------------------------
        self.ws.sheet_view.showGridLines = False
        
