# # orders/reports/sheets/funnel_analysis_sheet.py
# from datetime import datetime
# from openpyxl.styles import Font, PatternFill, Alignment
# from .base_sheet import BaseSheet
# from ..styles.theme import COLORS, BORDERS, FILLS
# from ..styles.helpers import draw_toc_button
# from ..components.footnote import Footnote
# from ..components.sheet_title import create_sheet_title


# class FunnelAnalysisSheet(BaseSheet):
#     def __init__(self, workbook, sheet_number):
#         super().__init__(workbook, sheet_number)
#         self.ws.title = str(sheet_number)
#         self.sheet_title = create_sheet_title(self.ws)

#     def _get_conv_color(self, percent):
#         if percent >= 70:
#             return COLORS["dark_green"]
#         if percent >= 40:
#             return COLORS["orange"]
#         return COLORS["warning_red"]

#     def _format_manager_name(self, manager_name):
#         """
#         Формат:
#         Иванов Иван Иванович -> ИВАНОВ И.
#         Иванов Иван -> ИВАНОВ И.
#         Иванов -> ИВАНОВ
#         """
#         if not manager_name:
#             return ""

#         words = str(manager_name).strip().split()
#         if not words:
#             return ""

#         first = words[0].upper()

#         if len(words) >= 2 and words[1]:
#             return f"{first} {words[1][0].upper()}."

#         return first

#     def _number_font(self, bold=False, color=None):
#         return Font(
#             name="Roboto",
#             size=9,
#             bold=bold,
#             color=color or COLORS["text_gray"],
#         )

#     def _set_number(self, cell, value, bold=False):
#         cell.value = int(round(value or 0))
#         cell.number_format = '#,##0'
#         cell.font = self._number_font(bold=bold)

#     def _set_money(self, cell, value, bold=False):
#         cell.value = int(round(value or 0))
#         cell.number_format = '#,##0 ₽'
#         cell.font = self._number_font(bold=bold)

#     def _set_percent(self, cell, value):
#         cell.value = int(round(value or 0))
#         cell.number_format = '0"%"'

#     def _draw_period_section(self, start_row, funnel_data):
#         row = start_row

#         cell = self.ws.cell(row=row, column=2, value=funnel_data["period_name"])
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
#         row += 1

#         stages = funnel_data["stages"]
#         first = stages[0]
#         last = stages[-1]

#         summary = (
#             f"Создано заказов без учета отмен: {int(first['orders']):,}".replace(",", " ")
#             + f" | Сумма: {int(first['amount']):,} ₽".replace(",", " ")
#             + f" | Закрыто: {int(last['orders']):,}".replace(",", " ")
#             + f" | Отменено: {int(funnel_data['canceled_orders']):,}".replace(",", " ")
#         )

#         cell = self.ws.cell(row=row, column=2, value=summary)
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
#         row += 2

#         headers = [
#             "ЭТАП",
#             "ЗАКАЗОВ",
#             "СУММА",
#             "КОНВ. ОТ ЭТАПА",
#             "КОНВ. ОТ НАЧАЛА",
#             "ЗАВИСЛО ЗАКАЗОВ",
#             "ЗАВИСЛО СУММА",
#             "ЗАВИСЛО %",
#         ]

#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 34
#         row += 1

#         for idx, stage in enumerate(stages):
#             prev = stages[idx - 1] if idx > 0 else None

#             stuck_orders = prev["orders"] - stage["orders"] if prev else 0
#             stuck_amount = prev["amount"] - stage["amount"] if prev else 0
#             stuck_percent = (stuck_orders / prev["orders"] * 100) if prev and prev["orders"] else 0

#             fill = FILLS["odd_row"] if idx % 2 == 0 else FILLS["even_row"]

#             cell = self.ws.cell(row=row, column=2, value=stage["name"])
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
#             cell.alignment = Alignment(horizontal="left", vertical="center")

#             cell = self.ws.cell(row=row, column=3)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_number(cell, stage["orders"], bold=True)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=4)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_money(cell, stage["amount"], bold=True)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=5)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_percent(cell, stage["conversion_from_prev"])
#             cell.font = Font(
#                 name="Roboto",
#                 size=9,
#                 bold=True,
#                 color=self._get_conv_color(stage["conversion_from_prev"]),
#             )
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#             cell = self.ws.cell(row=row, column=6)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_percent(cell, stage["conversion_from_start"])
#             cell.font = Font(
#                 name="Roboto",
#                 size=9,
#                 bold=True,
#                 color=self._get_conv_color(stage["conversion_from_start"]),
#             )
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#             cell = self.ws.cell(row=row, column=7)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_number(cell, stuck_orders)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=8)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_money(cell, stuck_amount)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=9)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_percent(cell, stuck_percent)
#             cell.font = self._number_font()
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#             self.ws.row_dimensions[row].height = 24
#             row += 1

#         row += 1

#         cell = self.ws.cell(row=row, column=2, value="ОТМЕНЕННЫЕ ЗАКАЗЫ")
#         cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["warning_red"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         row += 1

#         cancel_headers = ["ПЕРИОД", "ОТМЕНЕНО ЗАКАЗОВ", "СУММА ОТМЕН", "КОММЕНТАРИЙ"]

#         for col_idx, header in enumerate(cancel_headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["warning_red"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         row += 1

#         cell = self.ws.cell(row=row, column=2, value=funnel_data["period_key"])
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")

#         cell = self.ws.cell(row=row, column=3)
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         self._set_number(cell, funnel_data["canceled_orders"])
#         cell.alignment = Alignment(horizontal="right", vertical="center")

#         cell = self.ws.cell(row=row, column=4)
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         self._set_money(cell, funnel_data["canceled_amount"])
#         cell.alignment = Alignment(horizontal="right", vertical="center")

#         cell = self.ws.cell(row=row, column=5, value="Исключены из основной воронки.")
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

#         self.ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)

#         row += 2
#         return row

#     def _draw_manager_funnel(self, start_row, managers_data):
#         row = start_row

#         cell = self.ws.cell(row=row, column=2, value="ВОРОНКА ПО МЕНЕДЖЕРАМ: YTD")
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
#         row += 2

#         headers = [
#             "МЕНЕДЖЕР",
#             "ЗАКАЗОВ",
#             "СУММА",
#             "ОПЛАТА",
#             "КОНВ.%",
#             "ОТГРУЗКА",
#             "КОНВ.%",
#             "ЗАКРЫТО",
#             "КОНВ.%",
#             "БЕЗ ОПЛАТЫ",
#             "БЕЗ ОТГРУЗКИ",
#             "НЕ ЗАКРЫТО",
#         ]

#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 35
#         row += 1

#         for idx, mgr in enumerate(managers_data[:15]):
#             ytd = mgr["ytd"]
#             fill = FILLS["odd_row"] if idx % 2 == 0 else FILLS["even_row"]
#             formatted_name = self._format_manager_name(mgr["manager"])

#             cell = self.ws.cell(row=row, column=2, value=formatted_name)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="left", vertical="center")

#             numeric_columns = [
#                 (3, ytd["total_orders"], "number"),
#                 (4, ytd["total_amount"], "money"),
#                 (5, ytd["paid_orders"], "number"),
#                 (7, ytd["shipped_orders"], "number"),
#                 (9, ytd["closed_orders"], "number"),
#                 (11, ytd["stuck_without_payment"], "number_red"),
#                 (12, ytd["stuck_without_shipment"], "number_red"),
#                 (13, ytd["stuck_not_closed"], "number_red"),
#             ]

#             for col, value, kind in numeric_columns:
#                 cell = self.ws.cell(row=row, column=col)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]

#                 if kind == "money":
#                     self._set_money(cell, value)
#                 else:
#                     self._set_number(cell, value)

#                 if kind == "number_red":
#                     cell.font = Font(name="Roboto", size=9, color=COLORS["warning_red"])

#                 cell.alignment = Alignment(horizontal="right", vertical="center")

#             percent_columns = [
#                 (6, ytd["conv_paid"]),
#                 (8, ytd["conv_shipped"]),
#                 (10, ytd["conv_closed"]),
#             ]

#             for col, value in percent_columns:
#                 cell = self.ws.cell(row=row, column=col)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]
#                 self._set_percent(cell, value)
#                 cell.font = Font(
#                     name="Roboto",
#                     size=9,
#                     bold=True,
#                     color=self._get_conv_color(value),
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")

#             self.ws.row_dimensions[row].height = 22
#             row += 1

#         return row

#     def build(self, funnel_overall, funnel_by_manager, loss_analysis):
#         row = 1

#         draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

#         self.ws.row_dimensions[1].height = 20
#         self.ws.row_dimensions[2].height = 8

#         row = self.sheet_title.draw(
#             row=3,
#             title="ОПЕРАЦИОННАЯ ВОРОНКА ЗАКАЗОВ",
#             subtitle="Заказ → Оплата → Отгрузка → Полная оплата → Закрытие",
#             date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y')}",
#             start_col=2,
#             end_col=9,
#         )

#         row += 2

#         row = self._draw_period_section(row, funnel_overall["ytd"])
#         row += 1

#         row = self._draw_period_section(row, funnel_overall["mtd"])
#         row += 2

#         row = self._draw_manager_funnel(row, funnel_by_manager)

#         footnote = Footnote(self.ws)
#         footnote.draw(row + 2, text="Отмененные заказы исключены из основной воронки.")

#         col_widths = {
#             "B": 36,
#             "C": 14,
#             "D": 20,
#             "E": 15,
#             "F": 14,
#             "G": 16,
#             "H": 16,
#             "I": 14,
#             "J": 14,
#             "K": 14,
#             "L": 15,
#             "M": 15,
#         }

#         for col, width in col_widths.items():
#             self.ws.column_dimensions[col].width = width

#         self.ws.sheet_view.showGridLines = False




# # orders/reports/sheets/funnel_analysis_sheet.py

# from datetime import datetime
# from openpyxl.styles import Font, PatternFill, Alignment
# from openpyxl.worksheet.table import Table, TableStyleInfo

# from .base_sheet import BaseSheet
# from ..styles.theme import COLORS, BORDERS, FILLS
# from ..styles.helpers import draw_toc_button
# from ..components.footnote import Footnote
# from ..components.sheet_title import create_sheet_title


# class FunnelAnalysisSheet(BaseSheet):
#     def __init__(self, workbook, sheet_number):
#         super().__init__(workbook, sheet_number)
#         self.ws.title = str(sheet_number)
#         self.sheet_title = create_sheet_title(self.ws)

#     def _get_conv_color(self, percent):
#         if percent >= 70:
#             return COLORS["dark_green"]
#         if percent >= 40:
#             return COLORS["orange"]
#         return COLORS["warning_red"]

#     def _format_manager_name(self, manager_name):
#         """
#         Иванов Иван Иванович -> ИВАНОВ И.
#         Иванов Иван -> ИВАНОВ И.
#         Иванов -> ИВАНОВ
#         """
#         if not manager_name:
#             return ""

#         words = str(manager_name).strip().split()
#         if not words:
#             return ""

#         first = words[0].upper()

#         if len(words) >= 2 and words[1]:
#             return f"{first} {words[1][0].upper()}."

#         return first

#     def _number_font(self, bold=False, color=None, size=9):
#         return Font(
#             name="Roboto",
#             size=size,
#             bold=bold,
#             color=color or COLORS["text_gray"],
#         )

#     def _set_number(self, cell, value, bold=False, size=9):
#         cell.value = int(round(value or 0))
#         cell.number_format = '#,##0'
#         cell.font = self._number_font(bold=bold, size=size)

#     def _set_money(self, cell, value, bold=False, size=9):
#         cell.value = int(round(value or 0))
#         cell.number_format = '#,##0 ₽'
#         cell.font = self._number_font(bold=bold, size=size)

#     def _set_percent(self, cell, value):
#         cell.value = int(round(value or 0))
#         cell.number_format = '0"%"'

#     def _format_date(self, value):
#         if not value:
#             return None
#         if hasattr(value, "date"):
#             return value.date()
#         return value

#     def _draw_period_section(self, start_row, funnel_data):
#         row = start_row

#         cell = self.ws.cell(row=row, column=2, value=funnel_data["period_name"])
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
#         row += 1

#         stages = funnel_data["stages"]
#         first = stages[0]
#         last = stages[-1]

#         summary = (
#             f"Создано заказов без учета отмен: {int(first['orders']):,}".replace(",", " ")
#             + f" | Сумма: {int(first['amount']):,} ₽".replace(",", " ")
#             + f" | Закрыто: {int(last['orders']):,}".replace(",", " ")
#             + f" | Отменено: {int(funnel_data['canceled_orders']):,}".replace(",", " ")
#         )

#         cell = self.ws.cell(row=row, column=2, value=summary)
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
#         row += 2

#         headers = [
#             "ЭТАП",
#             "ЗАКАЗОВ",
#             "СУММА",
#             "КОНВ. ОТ ЭТАПА",
#             "КОНВ. ОТ НАЧАЛА",
#             "ЗАВИСЛО ЗАКАЗОВ",
#             "ЗАВИСЛО СУММА",
#             "ЗАВИСЛО %",
#         ]

#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 34
#         row += 1

#         for idx, stage in enumerate(stages):
#             prev = stages[idx - 1] if idx > 0 else None

#             stuck_orders = prev["orders"] - stage["orders"] if prev else 0
#             stuck_amount = prev["amount"] - stage["amount"] if prev else 0
#             stuck_percent = (stuck_orders / prev["orders"] * 100) if prev and prev["orders"] else 0

#             fill = FILLS["odd_row"] if idx % 2 == 0 else FILLS["even_row"]

#             cell = self.ws.cell(row=row, column=2, value=stage["name"])
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
#             cell.alignment = Alignment(horizontal="left", vertical="center")

#             cell = self.ws.cell(row=row, column=3)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_number(cell, stage["orders"], bold=True)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=4)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_money(cell, stage["amount"], bold=True)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=5)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_percent(cell, stage["conversion_from_prev"])
#             cell.font = Font(
#                 name="Roboto",
#                 size=9,
#                 bold=True,
#                 color=self._get_conv_color(stage["conversion_from_prev"]),
#             )
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#             cell = self.ws.cell(row=row, column=6)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_percent(cell, stage["conversion_from_start"])
#             cell.font = Font(
#                 name="Roboto",
#                 size=9,
#                 bold=True,
#                 color=self._get_conv_color(stage["conversion_from_start"]),
#             )
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#             cell = self.ws.cell(row=row, column=7)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_number(cell, stuck_orders)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=8)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_money(cell, stuck_amount)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=9)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_percent(cell, stuck_percent)
#             cell.font = self._number_font()
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#             self.ws.row_dimensions[row].height = 24
#             row += 1

#         row += 1

#         cell = self.ws.cell(row=row, column=2, value="ОТМЕНЕННЫЕ ЗАКАЗЫ")
#         cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["warning_red"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         row += 1

#         cancel_headers = ["ПЕРИОД", "ОТМЕНЕНО ЗАКАЗОВ", "СУММА ОТМЕН", "КОММЕНТАРИЙ"]

#         for col_idx, header in enumerate(cancel_headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["warning_red"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         row += 1

#         cell = self.ws.cell(row=row, column=2, value=funnel_data["period_key"])
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")

#         cell = self.ws.cell(row=row, column=3)
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         self._set_number(cell, funnel_data["canceled_orders"])
#         cell.alignment = Alignment(horizontal="right", vertical="center")

#         cell = self.ws.cell(row=row, column=4)
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         self._set_money(cell, funnel_data["canceled_amount"])
#         cell.alignment = Alignment(horizontal="right", vertical="center")

#         cell = self.ws.cell(row=row, column=5, value="Исключены из основной воронки.")
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

#         self.ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)

#         row += 2
#         return row

#     def _draw_manager_funnel(self, start_row, managers_data):
#         row = start_row

#         cell = self.ws.cell(row=row, column=2, value="ВОРОНКА ПО МЕНЕДЖЕРАМ: YTD")
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
#         row += 2

#         headers = [
#             "МЕНЕДЖЕР",
#             "ЗАКАЗОВ",
#             "СУММА",
#             "ОПЛАТА",
#             "КОНВ.%",
#             "ОТГРУЗКА",
#             "КОНВ.%",
#             "ЗАКРЫТО",
#             "КОНВ.%",
#             "БЕЗ ОПЛАТЫ",
#             "БЕЗ ОТГРУЗКИ",
#             "НЕ ЗАКРЫТО",
#         ]

#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 35
#         row += 1

#         for idx, mgr in enumerate(managers_data[:15]):
#             ytd = mgr["ytd"]
#             fill = FILLS["odd_row"] if idx % 2 == 0 else FILLS["even_row"]
#             formatted_name = self._format_manager_name(mgr["manager"])

#             cell = self.ws.cell(row=row, column=2, value=formatted_name)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="left", vertical="center")

#             numeric_columns = [
#                 (3, ytd["total_orders"], "number"),
#                 (4, ytd["total_amount"], "money"),
#                 (5, ytd["paid_orders"], "number"),
#                 (7, ytd["shipped_orders"], "number"),
#                 (9, ytd["closed_orders"], "number"),
#                 (11, ytd["stuck_without_payment"], "number_red"),
#                 (12, ytd["stuck_without_shipment"], "number_red"),
#                 (13, ytd["stuck_not_closed"], "number_red"),
#             ]

#             for col, value, kind in numeric_columns:
#                 cell = self.ws.cell(row=row, column=col)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]

#                 if kind == "money":
#                     self._set_money(cell, value)
#                 else:
#                     self._set_number(cell, value)

#                 if kind == "number_red":
#                     cell.font = Font(name="Roboto", size=9, color=COLORS["warning_red"])

#                 cell.alignment = Alignment(horizontal="right", vertical="center")

#             percent_columns = [
#                 (6, ytd["conv_paid"]),
#                 (8, ytd["conv_shipped"]),
#                 (10, ytd["conv_closed"]),
#             ]

#             for col, value in percent_columns:
#                 cell = self.ws.cell(row=row, column=col)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]
#                 self._set_percent(cell, value)
#                 cell.font = Font(
#                     name="Roboto",
#                     size=9,
#                     bold=True,
#                     color=self._get_conv_color(value),
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")

#             self.ws.row_dimensions[row].height = 22
#             row += 1

#         return row

#     def _draw_orders_detail_table(self, start_row, orders_detail):
#         row = start_row

#         cell = self.ws.cell(row=row, column=2, value="ДЕТАЛИЗАЦИЯ ЗАКАЗОВ: YTD")
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=19)
#         row += 1

#         cell = self.ws.cell(
#             row=row,
#             column=2,
#             value="Таблица для фильтрации заказов и поиска проблем: нет оплаты, нет отгрузки, частичная оплата, не закрыт в системе, отменен.",
#         )
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=19)
#         row += 2

#         headers = [
#             "ПРОБЛЕМА",
#             "КОММЕНТАРИЙ",
#             "НОМЕР",
#             "ДАТА",
#             "ИЗМЕНЕН",
#             "СТАТУС",
#             "ИЗМ. ЗАКАЗА",
#             "КЛИЕНТ",
#             "МЕНЕДЖЕР",
#             "МАГАЗИН",
#             "КОЛ-ВО",
#             "ОТМЕНЕНО КОЛ-ВО",
#             "СУММА ЗАКАЗА",
#             "СУММА АКТИВНАЯ",
#             "ОПЛАЧЕНО",
#             "ОТГРУЖЕНО",
#             "ВОЗВРАТ",
#             "ДАТЫ ОПЛАТ",
#         ]

#         fields = [
#             "problem_status",
#             "problem_comment",
#             "number",
#             "date_from",
#             "update_at",
#             "status",
#             "change_status",
#             "client",
#             "manager",
#             "store",
#             "order_qty",
#             "qty_cancelled",
#             "order_amount",
#             "amount_active",
#             "cash_pmts",
#             "total_shiped_amount",
#             "returned_amount",
#             "payment_dates",
#         ]

#         money_fields = {
#             "order_amount",
#             "amount_active",
#             "cash_pmts",
#             "total_shiped_amount",
#             "returned_amount",
#         }

#         numeric_fields = {
#             "order_qty",
#             "qty_cancelled",
#         }

#         header_row = row

#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 34
#         row += 1

#         for idx, order in enumerate(orders_detail):
#             fill = FILLS["odd_row"] if idx % 2 == 0 else FILLS["even_row"]

#             for col_idx, field in enumerate(fields, start=2):
#                 value = order.get(field)

#                 cell = self.ws.cell(row=row, column=col_idx)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]
#                 cell.font = Font(name="Roboto", size=8, color=COLORS["text_gray"])

#                 if field in ["date_from", "update_at"]:
#                     cell.value = self._format_date(value)
#                     cell.number_format = "DD.MM.YYYY"
#                     cell.alignment = Alignment(horizontal="center", vertical="center")

#                 elif field in money_fields:
#                     self._set_money(cell, value, size=8)
#                     cell.alignment = Alignment(horizontal="right", vertical="center")

#                 elif field in numeric_fields:
#                     self._set_number(cell, value, size=8)
#                     cell.alignment = Alignment(horizontal="right", vertical="center")

#                 else:
#                     cell.value = value
#                     cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

#                 if field == "problem_status":
#                     if value == "ОК":
#                         cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["dark_green"])
#                     elif value == "Отменен":
#                         cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["warning_red"])
#                     else:
#                         cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["orange"])

#             self.ws.row_dimensions[row].height = 22
#             row += 1

#         last_row = row - 1

#         if last_row > header_row:
#             table_ref = f"B{header_row}:S{last_row}"
#             table = Table(displayName="OrdersDetailTable", ref=table_ref)

#             style = TableStyleInfo(
#                 name="TableStyleMedium2",
#                 showFirstColumn=False,
#                 showLastColumn=False,
#                 showRowStripes=False,
#                 showColumnStripes=False,
#             )

#             table.tableStyleInfo = style
#             self.ws.add_table(table)

#         return row + 2

#     def build(self, funnel_overall, funnel_by_manager, loss_analysis, orders_detail=None):
#         row = 1

#         if orders_detail is None:
#             orders_detail = []

#         draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

#         self.ws.row_dimensions[1].height = 20
#         self.ws.row_dimensions[2].height = 8

#         row = self.sheet_title.draw(
#             row=3,
#             title="ОПЕРАЦИОННАЯ ВОРОНКА ЗАКАЗОВ",
#             subtitle="Заказ → Оплата → Отгрузка → Полная оплата → Закрытие",
#             date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y')}",
#             start_col=2,
#             end_col=9,
#         )

#         row += 2

#         row = self._draw_period_section(row, funnel_overall["ytd"])
#         row += 1

#         row = self._draw_period_section(row, funnel_overall["mtd"])
#         row += 2

#         row = self._draw_manager_funnel(row, funnel_by_manager)
#         row += 3

#         row = self._draw_orders_detail_table(row, orders_detail)

#         footnote = Footnote(self.ws)
#         footnote.draw(row + 2, text="Отмененные заказы исключены из основной воронки.")

#         col_widths = {
#             "B": 22,
#             "C": 38,
#             "D": 18,
#             "E": 12,
#             "F": 12,
#             "G": 24,
#             "H": 16,
#             "I": 30,
#             "J": 20,
#             "K": 18,
#             "L": 12,
#             "M": 17,
#             "N": 18,
#             "O": 18,
#             "P": 18,
#             "Q": 18,
#             "R": 18,
#             "S": 40,
#         }

#         for col, width in col_widths.items():
#             self.ws.column_dimensions[col].width = width

#         self.ws.sheet_view.showGridLines = False




# # orders/reports/sheets/funnel_analysis_sheet.py

# from datetime import datetime
# from openpyxl.styles import Font, PatternFill, Alignment
# from openpyxl.worksheet.table import Table, TableStyleInfo

# from .base_sheet import BaseSheet
# from ..styles.theme import COLORS, BORDERS, FILLS
# from ..styles.helpers import draw_toc_button
# from ..components.footnote import Footnote
# from ..components.sheet_title import create_sheet_title


# class FunnelAnalysisSheet(BaseSheet):
#     def __init__(self, workbook, sheet_number):
#         super().__init__(workbook, sheet_number)
#         self.ws.title = str(sheet_number)
#         self.sheet_title = create_sheet_title(self.ws)

#     def _get_conv_color(self, percent):
#         if percent >= 70:
#             return COLORS["dark_green"]
#         if percent >= 40:
#             return COLORS["orange"]
#         return COLORS["warning_red"]

#     def _number_font(self, bold=False, color=None, size=9):
#         return Font(
#             name="Roboto",
#             size=size,
#             bold=bold,
#             color=color or COLORS["text_gray"],
#         )

#     def _set_number(self, cell, value, bold=False, size=9):
#         cell.value = int(round(value or 0))
#         cell.number_format = '#,##0'
#         cell.font = self._number_font(bold=bold, size=size)

#     def _set_money(self, cell, value, bold=False, size=9):
#         cell.value = int(round(value or 0))
#         cell.number_format = '#,##0 ₽'
#         cell.font = self._number_font(bold=bold, size=size)

#     def _set_percent(self, cell, value):
#         cell.value = int(round(value or 0))
#         cell.number_format = '0"%"'

#     def _format_date(self, value):
#         if not value:
#             return None
#         if hasattr(value, "date"):
#             value = value.date()
#         # Формат YYYY-MM-DD
#         if value:
#             return value.strftime("%Y-%m-%d")
#         return None

#     def _draw_period_section(self, start_row, funnel_data):
#         row = start_row

#         cell = self.ws.cell(row=row, column=2, value=funnel_data["period_name"])
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
#         row += 1

#         stages = funnel_data["stages"]
#         first = stages[0]
#         last = stages[-1]

#         summary = (
#             f"Создано заказов без учета отмен: {int(first['orders']):,}".replace(",", " ")
#             + f" | Сумма: {int(first['amount']):,} ₽".replace(",", " ")
#             + f" | Закрыто: {int(last['orders']):,}".replace(",", " ")
#             + f" | Отменено: {int(funnel_data['canceled_orders']):,}".replace(",", " ")
#         )

#         cell = self.ws.cell(row=row, column=2, value=summary)
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
#         row += 2

#         headers = [
#             "ЭТАП",
#             "ЗАКАЗОВ",
#             "СУММА",
#             "КОНВ. ОТ ЭТАПА",
#             "КОНВ. ОТ НАЧАЛА",
#             "ЗАВИСЛО ЗАКАЗОВ",
#             "ЗАВИСЛО СУММА",
#             "ЗАВИСЛО %",
#         ]

#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 34
#         row += 1

#         for idx, stage in enumerate(stages):
#             prev = stages[idx - 1] if idx > 0 else None

#             stuck_orders = prev["orders"] - stage["orders"] if prev else 0
#             stuck_amount = prev["amount"] - stage["amount"] if prev else 0
#             stuck_percent = (stuck_orders / prev["orders"] * 100) if prev and prev["orders"] else 0

#             fill = FILLS["odd_row"] if idx % 2 == 0 else FILLS["even_row"]

#             cell = self.ws.cell(row=row, column=2, value=stage["name"])
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
#             cell.alignment = Alignment(horizontal="left", vertical="center")

#             cell = self.ws.cell(row=row, column=3)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_number(cell, stage["orders"], bold=True)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=4)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_money(cell, stage["amount"], bold=True)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=5)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_percent(cell, stage["conversion_from_prev"])
#             cell.font = Font(
#                 name="Roboto",
#                 size=9,
#                 bold=True,
#                 color=self._get_conv_color(stage["conversion_from_prev"]),
#             )
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#             cell = self.ws.cell(row=row, column=6)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_percent(cell, stage["conversion_from_start"])
#             cell.font = Font(
#                 name="Roboto",
#                 size=9,
#                 bold=True,
#                 color=self._get_conv_color(stage["conversion_from_start"]),
#             )
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#             cell = self.ws.cell(row=row, column=7)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_number(cell, stuck_orders)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=8)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_money(cell, stuck_amount)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=9)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_percent(cell, stuck_percent)
#             cell.font = self._number_font()
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#             self.ws.row_dimensions[row].height = 24
#             row += 1

#         row += 1

#         cell = self.ws.cell(row=row, column=2, value="ОТМЕНЕННЫЕ ЗАКАЗЫ")
#         cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["warning_red"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         row += 1

#         cancel_headers = ["ПЕРИОД", "ОТМЕНЕНО ЗАКАЗОВ", "СУММА ОТМЕН", "КОММЕНТАРИЙ"]

#         for col_idx, header in enumerate(cancel_headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["warning_red"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         row += 1

#         cell = self.ws.cell(row=row, column=2, value=funnel_data["period_key"])
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")

#         cell = self.ws.cell(row=row, column=3)
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         self._set_number(cell, funnel_data["canceled_orders"])
#         cell.alignment = Alignment(horizontal="right", vertical="center")

#         cell = self.ws.cell(row=row, column=4)
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         self._set_money(cell, funnel_data["canceled_amount"])
#         cell.alignment = Alignment(horizontal="right", vertical="center")

#         cell = self.ws.cell(row=row, column=5, value="Исключены из основной воронки.")
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

     

#         row += 2
#         return row

#     def _draw_manager_funnel(self, start_row, managers_data):
#         row = start_row

#         cell = self.ws.cell(row=row, column=2, value="ВОРОНКА ПО МЕНЕДЖЕРАМ: YTD")
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
#         row += 2

#         headers = [
#             "МЕНЕДЖЕР",
#             "ЗАКАЗОВ",
#             "СУММА",
#             "ОПЛАТА",
#             "КОНВ.%",
#             "ОТГРУЗКА",
#             "КОНВ.%",
#             "ЗАКРЫТО",
#             "КОНВ.%",
#             "БЕЗ ОПЛАТЫ",
#             "БЕЗ ОТГРУЗКИ",
#             "НЕ ЗАКРЫТО",
#         ]

#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 35
#         row += 1

#         def format_manager_name(manager_name):
#             if not manager_name:
#                 return ""
#             words = str(manager_name).strip().split()
#             if not words:
#                 return ""
#             first = words[0].upper()
#             if len(words) >= 2 and words[1]:
#                 return f"{first} {words[1][0].upper()}."
#             return first

#         for idx, mgr in enumerate(managers_data[:15]):
#             ytd = mgr["ytd"]
#             fill = FILLS["odd_row"] if idx % 2 == 0 else FILLS["even_row"]
#             formatted_name = format_manager_name(mgr["manager"])

#             cell = self.ws.cell(row=row, column=2, value=formatted_name)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="left", vertical="center")

#             numeric_columns = [
#                 (3, ytd["total_orders"], "number"),
#                 (4, ytd["total_amount"], "money"),
#                 (5, ytd["paid_orders"], "number"),
#                 (7, ytd["shipped_orders"], "number"),
#                 (9, ytd["closed_orders"], "number"),
#                 (11, ytd["stuck_without_payment"], "number_red"),
#                 (12, ytd["stuck_without_shipment"], "number_red"),
#                 (13, ytd["stuck_not_closed"], "number_red"),
#             ]

#             for col, value, kind in numeric_columns:
#                 cell = self.ws.cell(row=row, column=col)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]

#                 if kind == "money":
#                     self._set_money(cell, value)
#                 else:
#                     self._set_number(cell, value)

#                 if kind == "number_red":
#                     cell.font = Font(name="Roboto", size=9, color=COLORS["warning_red"])

#                 cell.alignment = Alignment(horizontal="right", vertical="center")

#             percent_columns = [
#                 (6, ytd["conv_paid"]),
#                 (8, ytd["conv_shipped"]),
#                 (10, ytd["conv_closed"]),
#             ]

#             for col, value in percent_columns:
#                 cell = self.ws.cell(row=row, column=col)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]
#                 self._set_percent(cell, value)
#                 cell.font = Font(
#                     name="Roboto",
#                     size=9,
#                     bold=True,
#                     color=self._get_conv_color(value),
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")

#             self.ws.row_dimensions[row].height = 22
#             row += 1

#         return row

#     def _draw_orders_detail_table(self, start_row, orders_detail):
#         row = start_row

#         cell = self.ws.cell(row=row, column=2, value="ДЕТАЛИЗАЦИЯ ПРОБЛЕМНЫХ ЗАКАЗОВ: YTD")
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
#         row += 1

#         cell = self.ws.cell(
#             row=row,
#             column=2,
#             value="Таблица для фильтрации заказов и поиска проблем: нет оплаты, нет отгрузки, частичная оплата, не закрыт в системе.",
#         )
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
#         row += 2

#         headers = [
#             "ПРОБЛЕМА",
#             "КОММЕНТАРИЙ",
#             "НОМЕР",
#             "ДАТА",
#             "СТАТУС",
#             "КЛИЕНТ",
#             "МЕНЕДЖЕР",
#             "МАГАЗИН",
#             "СУММА ЗАКАЗА",
#             "ОПЛАЧЕНО",
#             "ОТГРУЖЕНО",
#         ]

#         fields = [
#             "problem_status",
#             "problem_comment",
#             "number",
#             "date_from",
#             "status",
#             "client",
#             "manager",
#             "store",
#             "amount_active",
#             "cash_pmts",
#             "total_shiped_amount",
#         ]

#         header_row = row

#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 34
#         row += 1

#         for idx, order in enumerate(orders_detail):
#             fill = FILLS["odd_row"] if idx % 2 == 0 else FILLS["even_row"]

#             for col_idx, field in enumerate(fields, start=2):
#                 value = order.get(field)

#                 cell = self.ws.cell(row=row, column=col_idx)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]
#                 cell.font = Font(name="Roboto", size=8, color=COLORS["text_gray"])

#                 if field == "date_from":
#                     cell.value = self._format_date(value)
#                     cell.alignment = Alignment(horizontal="center", vertical="center")

#                 elif field in ["amount_active", "cash_pmts", "total_shiped_amount"]:
#                     if value:
#                         cell.value = int(round(value))
#                         cell.number_format = '#,##0 ₽'
#                     else:
#                         cell.value = 0
#                         cell.number_format = '#,##0 ₽'
#                     cell.alignment = Alignment(horizontal="right", vertical="center")

#                 else:
#                     cell.value = value
#                     cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

#                 if field == "problem_status":
#                     if value == "Нет оплаты":
#                         cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["warning_red"])
#                     elif value == "Нет отгрузки":
#                         cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["orange"])
#                     elif value == "Частичная оплата":
#                         cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["orange"])
#                     elif value == "Не закрыт в системе":
#                         cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["dark_green"])

#             self.ws.row_dimensions[row].height = 22
#             row += 1

#         last_row = row - 1

#         if last_row > header_row:
#             table_ref = f"B{header_row}:L{last_row}"
#             table = Table(displayName="OrdersDetailTable", ref=table_ref)

#             style = TableStyleInfo(
#                 name="TableStyleMedium2",
#                 showFirstColumn=False,
#                 showLastColumn=False,
#                 showRowStripes=False,
#                 showColumnStripes=False,
#             )

#             table.tableStyleInfo = style
#             self.ws.add_table(table)

#         return row + 2

#     def build(self, funnel_overall, funnel_by_manager, loss_analysis, orders_detail=None):
#         row = 1

#         if orders_detail is None:
#             orders_detail = []

#         draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

#         self.ws.row_dimensions[1].height = 20
#         self.ws.row_dimensions[2].height = 8

#         row = self.sheet_title.draw(
#             row=3,
#             title="ОПЕРАЦИОННАЯ ВОРОНКА ЗАКАЗОВ",
#             subtitle="Заказ → Оплата → Отгрузка → Полная оплата → Закрытие",
#             date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y')}",
#             start_col=2,
#             end_col=9,
#         )

#         row += 2

#         row = self._draw_period_section(row, funnel_overall["ytd"])
#         row += 1

#         row = self._draw_period_section(row, funnel_overall["mtd"])
#         row += 2

#         row = self._draw_manager_funnel(row, funnel_by_manager)
#         row += 3

#         row = self._draw_orders_detail_table(row, orders_detail)



#         col_widths = {
#             "B": 22,
#             "C": 15,
#             "D": 15,
#             "E": 12,
#             "F": 12,
#             "G": 15,
#             "H": 15,
#             "I": 15,
#             "J": 15,
#             "K": 15,
#             "L": 15,
#             "M": 15,
#         }

#         for col, width in col_widths.items():
#             self.ws.column_dimensions[col].width = width

#         self.ws.sheet_view.showGridLines = False



# # orders/reports/sheets/funnel_analysis_sheet.py
# from datetime import datetime
# from openpyxl.styles import Font, PatternFill, Alignment
# from openpyxl.worksheet.table import Table, TableStyleInfo

# from .base_sheet import BaseSheet
# from ..styles.theme import COLORS, BORDERS, FILLS
# from ..styles.helpers import draw_toc_button
# from ..components.footnote import Footnote
# from ..components.sheet_title import create_sheet_title


# class FunnelAnalysisSheet(BaseSheet):
#     def __init__(self, workbook, sheet_number):
#         super().__init__(workbook, sheet_number)
#         self.ws.title = str(sheet_number)
#         self.sheet_title = create_sheet_title(self.ws)

#     def _get_conv_color(self, percent):
#         if percent >= 70:
#             return COLORS["dark_green"]
#         if percent >= 40:
#             return COLORS["orange"]
#         return COLORS["warning_red"]

#     def _number_font(self, bold=False, color=None, size=9):
#         return Font(
#             name="Roboto",
#             size=size,
#             bold=bold,
#             color=color or COLORS["text_gray"],
#         )

#     def _set_number(self, cell, value, bold=False, size=9):
#         cell.value = int(round(value or 0))
#         cell.number_format = '#,##0'
#         cell.font = self._number_font(bold=bold, size=size)

#     def _set_money(self, cell, value, bold=False, size=9):
#         cell.value = int(round(value or 0))
#         cell.number_format = '#,##0 ₽'
#         cell.font = self._number_font(bold=bold, size=size)

#     def _set_percent(self, cell, value):
#         cell.value = int(round(value or 0))
#         cell.number_format = '0"%"'

#     def _format_date(self, value):
#         if not value:
#             return None
#         if hasattr(value, "date"):
#             value = value.date()
#         if value:
#             return value.strftime("%Y-%m-%d")
#         return None

#     def _draw_period_section(self, start_row, funnel_data):
#         row = start_row

#         cell = self.ws.cell(row=row, column=2, value=funnel_data["period_name"])
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
#         row += 1

#         stages = funnel_data["stages"]
#         first = stages[0]
#         last = stages[-1]

#         summary = (
#             f"Создано заказов без учета отмен и возвратов: {int(first['orders']):,}".replace(",", " ")
#             + f" | Сумма: {int(first['amount']):,} ₽".replace(",", " ")
#             + f" | Закрыто: {int(last['orders']):,}".replace(",", " ")
#             + f" | Отменено: {int(funnel_data['canceled_orders']):,}".replace(",", " ")
#             + f" | Возвращено: {int(funnel_data['returned_orders']):,}".replace(",", " ")
#         )

#         cell = self.ws.cell(row=row, column=2, value=summary)
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
#         row += 2

#         headers = [
#             "ЭТАП",
#             "ЗАКАЗОВ",
#             "СУММА",
#             "КОНВ. ОТ ЭТАПА",
#             "КОНВ. ОТ НАЧАЛА",
#             "ЗАВИСЛО ЗАКАЗОВ",
#             "ЗАВИСЛО СУММА",
#             "ЗАВИСЛО %",
#         ]

#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 34
#         row += 1

#         for idx, stage in enumerate(stages):
#             prev = stages[idx - 1] if idx > 0 else None

#             stuck_orders = prev["orders"] - stage["orders"] if prev else 0
#             stuck_amount = prev["amount"] - stage["amount"] if prev else 0
#             stuck_percent = (stuck_orders / prev["orders"] * 100) if prev and prev["orders"] else 0

#             fill = FILLS["odd_row"] if idx % 2 == 0 else FILLS["even_row"]

#             cell = self.ws.cell(row=row, column=2, value=stage["name"])
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
#             cell.alignment = Alignment(horizontal="left", vertical="center")

#             cell = self.ws.cell(row=row, column=3)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_number(cell, stage["orders"], bold=True)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=4)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_money(cell, stage["amount"], bold=True)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=5)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_percent(cell, stage["conversion_from_prev"])
#             cell.font = Font(
#                 name="Roboto",
#                 size=9,
#                 bold=True,
#                 color=self._get_conv_color(stage["conversion_from_prev"]),
#             )
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#             cell = self.ws.cell(row=row, column=6)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_percent(cell, stage["conversion_from_start"])
#             cell.font = Font(
#                 name="Roboto",
#                 size=9,
#                 bold=True,
#                 color=self._get_conv_color(stage["conversion_from_start"]),
#             )
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#             cell = self.ws.cell(row=row, column=7)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_number(cell, stuck_orders)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=8)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_money(cell, stuck_amount)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=9)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_percent(cell, stuck_percent)
#             cell.font = self._number_font()
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#             self.ws.row_dimensions[row].height = 24
#             row += 1

#         row += 1

#         # Секция отмененных заказов
#         cell = self.ws.cell(row=row, column=2, value="ОТМЕНЕННЫЕ ЗАКАЗЫ")
#         cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["warning_red"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         row += 1

#         cancel_headers = ["ПЕРИОД", "ОТМЕНЕНО ЗАКАЗОВ", "СУММА ОТМЕН", "КОММЕНТАРИЙ"]

#         for col_idx, header in enumerate(cancel_headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["warning_red"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         row += 1

#         # ПЕРИОД
#         cell = self.ws.cell(row=row, column=2, value=funnel_data["period_key"])
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")

#         # ОТМЕНЕНО ЗАКАЗОВ
#         cell = self.ws.cell(row=row, column=3)
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         self._set_number(cell, funnel_data["canceled_orders"])
#         cell.alignment = Alignment(horizontal="right", vertical="center")

#         # СУММА ОТМЕН
#         cell = self.ws.cell(row=row, column=4)
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         self._set_money(cell, funnel_data["canceled_amount"])
#         cell.alignment = Alignment(horizontal="right", vertical="center")

#         # КОММЕНТАРИЙ - объединяем колонки 5 и 6 (E и F)
#         comment_text = "Исключены из основной воронки."
#         cell = self.ws.cell(row=row, column=5, value=comment_text)
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
#         # Объединяем колонки 5 и 6
#         self.ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)

#         row += 2

#         # Секция возвращенных заказов
#         cell = self.ws.cell(row=row, column=2, value="ВОЗВРАЩЕННЫЕ ЗАКАЗЫ")
#         cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["orange"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         row += 1

#         returned_headers = ["ПЕРИОД", "ВОЗВРАЩЕНО ЗАКАЗОВ", "СУММА ВОЗВРАТОВ", "КОММЕНТАРИЙ"]

#         for col_idx, header in enumerate(returned_headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["orange"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         row += 1

#         # ПЕРИОД
#         cell = self.ws.cell(row=row, column=2, value=funnel_data["period_key"])
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")

#         # ВОЗВРАЩЕНО ЗАКАЗОВ
#         cell = self.ws.cell(row=row, column=3)
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         self._set_number(cell, funnel_data["returned_orders"])
#         cell.alignment = Alignment(horizontal="right", vertical="center")

#         # СУММА ВОЗВРАТОВ
#         cell = self.ws.cell(row=row, column=4)
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         self._set_money(cell, funnel_data["returned_amount"])
#         cell.alignment = Alignment(horizontal="right", vertical="center")

#         # КОММЕНТАРИЙ - объединяем колонки 5 и 6 (E и F)
#         comment_text = "Была оплата/отгрузка, но клиент отказался (полный возврат)."
#         cell = self.ws.cell(row=row, column=5, value=comment_text)
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
#         # Объединяем колонки 5 и 6
#         self.ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)

#         row += 2

#     def _draw_manager_funnel(self, start_row, managers_data):
#         row = start_row

#         cell = self.ws.cell(row=row, column=2, value="ВОРОНКА ПО МЕНЕДЖЕРАМ: YTD")
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
#         row += 2

#         headers = [
#             "МЕНЕДЖЕР",
#             "ЗАКАЗОВ",
#             "СУММА",
#             "ОПЛАТА",
#             "КОНВ.%",
#             "ОТГРУЗКА",
#             "КОНВ.%",
#             "ЗАКРЫТО",
#             "КОНВ.%",
#             "БЕЗ ОПЛАТЫ",
#             "БЕЗ ОТГРУЗКИ",
#             "НЕ ЗАКРЫТО",
#         ]

#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 35
#         row += 1

#         def format_manager_name(manager_name):
#             if not manager_name:
#                 return ""
#             words = str(manager_name).strip().split()
#             if not words:
#                 return ""
#             first = words[0].upper()
#             if len(words) >= 2 and words[1]:
#                 return f"{first} {words[1][0].upper()}."
#             return first

#         for idx, mgr in enumerate(managers_data[:15]):
#             ytd = mgr["ytd"]
#             fill = FILLS["odd_row"] if idx % 2 == 0 else FILLS["even_row"]
#             formatted_name = format_manager_name(mgr["manager"])

#             cell = self.ws.cell(row=row, column=2, value=formatted_name)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="left", vertical="center")

#             numeric_columns = [
#                 (3, ytd["total_orders"], "number"),
#                 (4, ytd["total_amount"], "money"),
#                 (5, ytd["paid_orders"], "number"),
#                 (7, ytd["shipped_orders"], "number"),
#                 (9, ytd["closed_orders"], "number"),
#                 (11, ytd["stuck_without_payment"], "number_red"),
#                 (12, ytd["stuck_without_shipment"], "number_red"),
#                 (13, ytd["stuck_not_closed"], "number_red"),
#             ]

#             for col, value, kind in numeric_columns:
#                 cell = self.ws.cell(row=row, column=col)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]

#                 if kind == "money":
#                     self._set_money(cell, value)
#                 else:
#                     self._set_number(cell, value)

#                 if kind == "number_red":
#                     cell.font = Font(name="Roboto", size=9, color=COLORS["warning_red"])

#                 cell.alignment = Alignment(horizontal="right", vertical="center")

#             percent_columns = [
#                 (6, ytd["conv_paid"]),
#                 (8, ytd["conv_shipped"]),
#                 (10, ytd["conv_closed"]),
#             ]

#             for col, value in percent_columns:
#                 cell = self.ws.cell(row=row, column=col)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]
#                 self._set_percent(cell, value)
#                 cell.font = Font(
#                     name="Roboto",
#                     size=9,
#                     bold=True,
#                     color=self._get_conv_color(value),
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")

#             self.ws.row_dimensions[row].height = 22
#             row += 1

#         return row

#     def _draw_orders_detail_table(self, start_row, orders_detail):
#         row = start_row

#         cell = self.ws.cell(row=row, column=2, value="ДЕТАЛИЗАЦИЯ ПРОБЛЕМНЫХ ЗАКАЗОВ: YTD")
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
#         row += 1

#         cell = self.ws.cell(
#             row=row,
#             column=2,
#             value="Таблица для фильтрации заказов и поиска проблем: нет оплаты, нет отгрузки, частичная отгрузка, частичная оплата, не закрыт в системе.",
#         )
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
#         row += 2

#         headers = [
#             "ПРОБЛЕМА",
#             "КОММЕНТАРИЙ",
#             "НОМЕР",
#             "ДАТА",
#             "СТАТУС",
#             "КЛИЕНТ",
#             "МЕНЕДЖЕР",
#             "МАГАЗИН",
#             "СУММА ЗАКАЗА",
#             "ОПЛАЧЕНО",
#             "ОТГРУЖЕНО",
#         ]

#         fields = [
#             "problem_status",
#             "problem_comment",
#             "number",
#             "date_from",
#             "status",
#             "client",
#             "manager",
#             "store",
#             "amount_active",
#             "cash_pmts",
#             "total_shiped_amount",
#         ]

#         header_row = row

#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 34
#         row += 1

#         for idx, order in enumerate(orders_detail):
#             fill = FILLS["odd_row"] if idx % 2 == 0 else FILLS["even_row"]

#             for col_idx, field in enumerate(fields, start=2):
#                 value = order.get(field)

#                 cell = self.ws.cell(row=row, column=col_idx)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]
#                 cell.font = Font(name="Roboto", size=8, color=COLORS["text_gray"])

#                 if field == "date_from":
#                     cell.value = self._format_date(value)
#                     cell.alignment = Alignment(horizontal="center", vertical="center")

#                 elif field in ["amount_active", "cash_pmts", "total_shiped_amount"]:
#                     if value:
#                         cell.value = int(round(value))
#                         cell.number_format = '#,##0 ₽'
#                     else:
#                         cell.value = 0
#                         cell.number_format = '#,##0 ₽'
#                     cell.alignment = Alignment(horizontal="right", vertical="center")

#                 else:
#                     cell.value = value
#                     cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

#                 if field == "problem_status":
#                     if value == "Нет оплаты":
#                         cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["warning_red"])
#                     elif value == "Нет отгрузки":
#                         cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["orange"])
#                     elif value == "Частичная отгрузка":
#                         cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["orange"])
#                     elif value == "Частичная оплата":
#                         cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["orange"])
#                     elif value == "Не закрыт в системе":
#                         cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["dark_green"])

#             self.ws.row_dimensions[row].height = 22
#             row += 1

#         last_row = row - 1

#         if last_row > header_row:
#             table_ref = f"B{header_row}:L{last_row}"
#             table = Table(displayName="OrdersDetailTable", ref=table_ref)

#             style = TableStyleInfo(
#                 name="TableStyleMedium2",
#                 showFirstColumn=False,
#                 showLastColumn=False,
#                 showRowStripes=False,
#                 showColumnStripes=False,
#             )

#             table.tableStyleInfo = style
#             self.ws.add_table(table)

#         return row + 2

#     def build(self, funnel_overall, funnel_by_manager, loss_analysis, orders_detail=None):
#         row = 1

#         if orders_detail is None:
#             orders_detail = []

#         draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

#         self.ws.row_dimensions[1].height = 20
#         self.ws.row_dimensions[2].height = 8

#         row = self.sheet_title.draw(
#             row=3,
#             title="ОПЕРАЦИОННАЯ ВОРОНКА ЗАКАЗОВ",
#             subtitle="Заказ → Оплата → Отгрузка → Полная оплата → Закрытие",
#             date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y')}",
#             start_col=2,
#             end_col=9,
#         )

#         row += 2

#         row = self._draw_period_section(row, funnel_overall["ytd"])
#         row += 1

#         row = self._draw_period_section(row, funnel_overall["mtd"])
#         row += 2

#         row = self._draw_manager_funnel(row, funnel_by_manager)
#         row += 3

#         row = self._draw_orders_detail_table(row, orders_detail)

#         col_widths = {
#             "B": 22,
#             "C": 15,
#             "D": 15,
#             "E": 12,
#             "F": 12,
#             "G": 15,
#             "H": 15,
#             "I": 15,
#             "J": 15,
#             "K": 15,
#             "L": 15,
#             "M": 15,
#         }

#         for col, width in col_widths.items():
#             self.ws.column_dimensions[col].width = width

#         self.ws.sheet_view.showGridLines = False



# # orders/reports/sheets/funnel_analysis_sheet.py
# from datetime import datetime
# from openpyxl.styles import Font, PatternFill, Alignment
# from openpyxl.worksheet.table import Table, TableStyleInfo

# from .base_sheet import BaseSheet
# from ..styles.theme import COLORS, BORDERS, FILLS
# from ..styles.helpers import draw_toc_button
# from ..components.footnote import Footnote
# from ..components.sheet_title import create_sheet_title


# class FunnelAnalysisSheet(BaseSheet):
#     def __init__(self, workbook, sheet_number):
#         super().__init__(workbook, sheet_number)
#         self.ws.title = str(sheet_number)
#         self.sheet_title = create_sheet_title(self.ws)

#     def _get_conv_color(self, percent):
#         if percent >= 70:
#             return COLORS["dark_green"]
#         if percent >= 40:
#             return COLORS["orange"]
#         return COLORS["warning_red"]

#     def _number_font(self, bold=False, color=None, size=9):
#         return Font(
#             name="Roboto",
#             size=size,
#             bold=bold,
#             color=color or COLORS["text_gray"],
#         )

#     def _set_number(self, cell, value, bold=False, size=9):
#         cell.value = int(round(value or 0))
#         cell.number_format = '#,##0'
#         cell.font = self._number_font(bold=bold, size=size)

#     def _set_money(self, cell, value, bold=False, size=9):
#         cell.value = int(round(value or 0))
#         cell.number_format = '#,##0 ₽'
#         cell.font = self._number_font(bold=bold, size=size)

#     def _set_percent(self, cell, value):
#         cell.value = int(round(value or 0))
#         cell.number_format = '0"%"'

#     def _format_date(self, value):
#         if not value:
#             return None
#         if hasattr(value, "date"):
#             value = value.date()
#         if value:
#             return value.strftime("%Y-%m-%d")
#         return None

#     def _draw_period_section(self, start_row, funnel_data):
#         row = start_row

#         cell = self.ws.cell(row=row, column=2, value=funnel_data["period_name"])
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
#         row += 1

#         stages = funnel_data["stages"]
#         first = stages[0]
#         last = stages[-1]

#         summary = (
#             f"Создано заказов без учета отмен и возвратов: {int(first['orders']):,}".replace(",", " ")
#             + f" | Сумма: {int(first['amount']):,} ₽".replace(",", " ")
#             + f" | Закрыто: {int(last['orders']):,}".replace(",", " ")
#             + f" | Отменено: {int(funnel_data['canceled_orders']):,}".replace(",", " ")
#             + f" | Возвращено: {int(funnel_data['returned_orders']):,}".replace(",", " ")
#         )

#         cell = self.ws.cell(row=row, column=2, value=summary)
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
#         row += 2

#         headers = [
#             "ЭТАП",
#             "ЗАКАЗОВ",
#             "СУММА",
#             "КОНВ. ОТ ЭТАПА",
#             "КОНВ. ОТ НАЧАЛА",
#             "ЗАВИСЛО ЗАКАЗОВ",
#             "ЗАВИСЛО СУММА",
#             "ЗАВИСЛО %",
#         ]

#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 34
#         row += 1

#         for idx, stage in enumerate(stages):
#             prev = stages[idx - 1] if idx > 0 else None

#             stuck_orders = prev["orders"] - stage["orders"] if prev else 0
#             stuck_amount = prev["amount"] - stage["amount"] if prev else 0
#             stuck_percent = (stuck_orders / prev["orders"] * 100) if prev and prev["orders"] else 0

#             fill = FILLS["odd_row"] if idx % 2 == 0 else FILLS["even_row"]

#             cell = self.ws.cell(row=row, column=2, value=stage["name"])
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
#             cell.alignment = Alignment(horizontal="left", vertical="center")

#             cell = self.ws.cell(row=row, column=3)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_number(cell, stage["orders"], bold=True)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=4)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_money(cell, stage["amount"], bold=True)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=5)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_percent(cell, stage["conversion_from_prev"])
#             cell.font = Font(
#                 name="Roboto",
#                 size=9,
#                 bold=True,
#                 color=self._get_conv_color(stage["conversion_from_prev"]),
#             )
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#             cell = self.ws.cell(row=row, column=6)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_percent(cell, stage["conversion_from_start"])
#             cell.font = Font(
#                 name="Roboto",
#                 size=9,
#                 bold=True,
#                 color=self._get_conv_color(stage["conversion_from_start"]),
#             )
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#             cell = self.ws.cell(row=row, column=7)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_number(cell, stuck_orders)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=8)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_money(cell, stuck_amount)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#             cell = self.ws.cell(row=row, column=9)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             self._set_percent(cell, stuck_percent)
#             cell.font = self._number_font()
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#             self.ws.row_dimensions[row].height = 24
#             row += 1

#         row += 1

#         # Секция отмененных заказов
#         cell = self.ws.cell(row=row, column=2, value="ОТМЕНЕННЫЕ ЗАКАЗЫ")
#         cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["warning_red"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         row += 1

#         cancel_headers = ["ПЕРИОД", "ОТМЕНЕНО ЗАКАЗОВ", "СУММА ОТМЕН", "КОММЕНТАРИЙ"]

#         for col_idx, header in enumerate(cancel_headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["warning_red"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         row += 1

#         # ПЕРИОД
#         cell = self.ws.cell(row=row, column=2, value=funnel_data["period_key"])
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")

#         # ОТМЕНЕНО ЗАКАЗОВ
#         cell = self.ws.cell(row=row, column=3)
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         self._set_number(cell, funnel_data["canceled_orders"])
#         cell.alignment = Alignment(horizontal="right", vertical="center")

#         # СУММА ОТМЕН
#         cell = self.ws.cell(row=row, column=4)
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         self._set_money(cell, funnel_data["canceled_amount"])
#         cell.alignment = Alignment(horizontal="right", vertical="center")

#         # КОММЕНТАРИЙ - объединяем колонки 5 и 6 (E и F)
#         comment_text = "Исключены из основной воронки."
#         cell = self.ws.cell(row=row, column=5, value=comment_text)
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
#         # Объединяем колонки 5 и 6
#         self.ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)

#         row += 2

#         # Секция возвращенных заказов
#         cell = self.ws.cell(row=row, column=2, value="ВОЗВРАЩЕННЫЕ ЗАКАЗЫ")
#         cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["orange"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         row += 1

#         returned_headers = ["ПЕРИОД", "ВОЗВРАЩЕНО ЗАКАЗОВ", "СУММА ВОЗВРАТОВ", "КОММЕНТАРИЙ"]

#         for col_idx, header in enumerate(returned_headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["orange"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         row += 1

#         # ПЕРИОД
#         cell = self.ws.cell(row=row, column=2, value=funnel_data["period_key"])
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")

#         # ВОЗВРАЩЕНО ЗАКАЗОВ
#         cell = self.ws.cell(row=row, column=3)
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         self._set_number(cell, funnel_data["returned_orders"])
#         cell.alignment = Alignment(horizontal="right", vertical="center")

#         # СУММА ВОЗВРАТОВ
#         cell = self.ws.cell(row=row, column=4)
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         self._set_money(cell, funnel_data["returned_amount"])
#         cell.alignment = Alignment(horizontal="right", vertical="center")

#         # КОММЕНТАРИЙ - объединяем колонки 5 и 6 (E и F)
#         comment_text = "Была оплата/отгрузка, но клиент отказался (полный возврат)."
#         cell = self.ws.cell(row=row, column=5, value=comment_text)
#         cell.fill = FILLS["odd_row"]
#         cell.border = BORDERS["thin"]
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
#         # Объединяем колонки 5 и 6
#         self.ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)

#         row += 2
        
#         return row

#     def _draw_manager_funnel(self, start_row, managers_data):
#         row = start_row

#         cell = self.ws.cell(row=row, column=2, value="ВОРОНКА ПО МЕНЕДЖЕРАМ: YTD")
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
#         row += 2

#         headers = [
#             "МЕНЕДЖЕР",
#             "ЗАКАЗОВ",
#             "СУММА",
#             "ОПЛАТА",
#             "КОНВ.%",
#             "ОТГРУЗКА",
#             "КОНВ.%",
#             "ЗАКРЫТО",
#             "КОНВ.%",
#             "БЕЗ ОПЛАТЫ",
#             "БЕЗ ОТГРУЗКИ",
#             "НЕ ЗАКРЫТО",
#         ]

#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 35
#         row += 1

#         def format_manager_name(manager_name):
#             if not manager_name:
#                 return ""
#             words = str(manager_name).strip().split()
#             if not words:
#                 return ""
#             first = words[0].upper()
#             if len(words) >= 2 and words[1]:
#                 return f"{first} {words[1][0].upper()}."
#             return first

#         for idx, mgr in enumerate(managers_data[:15]):
#             ytd = mgr["ytd"]
#             fill = FILLS["odd_row"] if idx % 2 == 0 else FILLS["even_row"]
#             formatted_name = format_manager_name(mgr["manager"])

#             cell = self.ws.cell(row=row, column=2, value=formatted_name)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
#             cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="left", vertical="center")

#             numeric_columns = [
#                 (3, ytd["total_orders"], "number"),
#                 (4, ytd["total_amount"], "money"),
#                 (5, ytd["paid_orders"], "number"),
#                 (7, ytd["shipped_orders"], "number"),
#                 (9, ytd["closed_orders"], "number"),
#                 (11, ytd["stuck_without_payment"], "number_red"),
#                 (12, ytd["stuck_without_shipment"], "number_red"),
#                 (13, ytd["stuck_not_closed"], "number_red"),
#             ]

#             for col, value, kind in numeric_columns:
#                 cell = self.ws.cell(row=row, column=col)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]

#                 if kind == "money":
#                     self._set_money(cell, value)
#                 else:
#                     self._set_number(cell, value)

#                 if kind == "number_red":
#                     cell.font = Font(name="Roboto", size=9, color=COLORS["warning_red"])

#                 cell.alignment = Alignment(horizontal="right", vertical="center")

#             percent_columns = [
#                 (6, ytd["conv_paid"]),
#                 (8, ytd["conv_shipped"]),
#                 (10, ytd["conv_closed"]),
#             ]

#             for col, value in percent_columns:
#                 cell = self.ws.cell(row=row, column=col)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]
#                 self._set_percent(cell, value)
#                 cell.font = Font(
#                     name="Roboto",
#                     size=9,
#                     bold=True,
#                     color=self._get_conv_color(value),
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")

#             self.ws.row_dimensions[row].height = 22
#             row += 1

#         return row

#     def _draw_orders_detail_table(self, start_row, orders_detail):
#         row = start_row

#         cell = self.ws.cell(row=row, column=2, value="ДЕТАЛИЗАЦИЯ ПРОБЛЕМНЫХ ЗАКАЗОВ: YTD")
#         cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
#         row += 1

#         cell = self.ws.cell(
#             row=row,
#             column=2,
#             value="Таблица для фильтрации заказов и поиска проблем: нет оплаты, нет отгрузки, частичная отгрузка, частичная оплата, не закрыт в системе.",
#         )
#         cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
#         cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
#         row += 2

#         headers = [
#             "ПРОБЛЕМА",
#             "КОММЕНТАРИЙ",
#             "НОМЕР",
#             "ДАТА",
#             "СТАТУС",
#             "КЛИЕНТ",
#             "МЕНЕДЖЕР",
#             "МАГАЗИН",
#             "СУММА ЗАКАЗА",
#             "ОПЛАЧЕНО",
#             "ОТГРУЖЕНО",
#         ]

#         fields = [
#             "problem_status",
#             "problem_comment",
#             "number",
#             "date_from",
#             "status",
#             "client",
#             "manager",
#             "store",
#             "amount_active",
#             "cash_pmts",
#             "total_shiped_amount",
#         ]

#         header_row = row

#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["white"])
#             cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 34
#         row += 1

#         for idx, order in enumerate(orders_detail):
#             fill = FILLS["odd_row"] if idx % 2 == 0 else FILLS["even_row"]

#             for col_idx, field in enumerate(fields, start=2):
#                 value = order.get(field)

#                 cell = self.ws.cell(row=row, column=col_idx)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]
#                 cell.font = Font(name="Roboto", size=8, color=COLORS["text_gray"])

#                 if field == "date_from":
#                     cell.value = self._format_date(value)
#                     cell.alignment = Alignment(horizontal="center", vertical="center")

#                 elif field in ["amount_active", "cash_pmts", "total_shiped_amount"]:
#                     if value:
#                         cell.value = int(round(value))
#                         cell.number_format = '#,##0 ₽'
#                     else:
#                         cell.value = 0
#                         cell.number_format = '#,##0 ₽'
#                     cell.alignment = Alignment(horizontal="right", vertical="center")

#                 else:
#                     cell.value = value
#                     cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

#                 if field == "problem_status":
#                     if value == "Нет оплаты":
#                         cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["warning_red"])
#                     elif value == "Нет отгрузки":
#                         cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["orange"])
#                     elif value == "Частичная отгрузка":
#                         cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["orange"])
#                     elif value == "Частичная оплата":
#                         cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["orange"])
#                     elif value == "Не закрыт в системе":
#                         cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["dark_green"])

#             self.ws.row_dimensions[row].height = 22
#             row += 1

#         last_row = row - 1

#         if last_row > header_row:
#             table_ref = f"B{header_row}:L{last_row}"
#             table = Table(displayName="OrdersDetailTable", ref=table_ref)

#             style = TableStyleInfo(
#                 name="TableStyleMedium2",
#                 showFirstColumn=False,
#                 showLastColumn=False,
#                 showRowStripes=False,
#                 showColumnStripes=False,
#             )

#             table.tableStyleInfo = style
#             self.ws.add_table(table)

#         return row + 2

#     def build(self, funnel_overall, funnel_by_manager, loss_analysis, orders_detail=None):
#         row = 1

#         if orders_detail is None:
#             orders_detail = []

#         draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

#         self.ws.row_dimensions[1].height = 20
#         self.ws.row_dimensions[2].height = 8

#         row = self.sheet_title.draw(
#             row=3,
#             title="ОПЕРАЦИОННАЯ ВОРОНКА ЗАКАЗОВ",
#             subtitle="Заказ → Оплата → Отгрузка → Полная оплата → Закрытие",
#             date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y')}",
#             start_col=2,
#             end_col=9,
#         )

#         row += 2

#         row = self._draw_period_section(row, funnel_overall["ytd"])
#         row = self._draw_period_section(row, funnel_overall["mtd"])

#         row = self._draw_manager_funnel(row, funnel_by_manager)

#         row = self._draw_orders_detail_table(row, orders_detail)

#         col_widths = {
#             "B": 22,
#             "C": 15,
#             "D": 15,
#             "E": 12,
#             "F": 12,
#             "G": 15,
#             "H": 15,
#             "I": 15,
#             "J": 15,
#             "K": 15,
#             "L": 15,
#             "M": 15,
#         }

#         for col, width in col_widths.items():
#             self.ws.column_dimensions[col].width = width

#         self.ws.sheet_view.showGridLines = False



from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

from .base_sheet import BaseSheet
from ..styles.theme import COLORS, BORDERS, FILLS
from ..styles.helpers import draw_toc_button
from ..components.footnote import Footnote
from ..components.sheet_title import create_sheet_title


class FunnelAnalysisSheet(BaseSheet):
    def __init__(self, workbook, sheet_number):
        super().__init__(workbook, sheet_number)
        self.ws.title = str(sheet_number)
        self.sheet_title = create_sheet_title(self.ws)

    def _get_conv_color(self, percent):
        if percent >= 70:
            return COLORS["dark_green"]
        if percent >= 40:
            return COLORS["orange"]
        return COLORS["warning_red"]

    def _number_font(self, bold=False, color=None, size=9):
        return Font(
            name="Roboto",
            size=size,
            bold=bold,
            color=color or COLORS["text_gray"],
        )

    def _set_number(self, cell, value, bold=False, size=9):
        cell.value = int(round(value or 0))
        cell.number_format = '#,##0'
        cell.font = self._number_font(bold=bold, size=size)

    def _set_money(self, cell, value, bold=False, size=9):
        cell.value = int(round(value or 0))
        cell.number_format = '#,##0 ₽'
        cell.font = self._number_font(bold=bold, size=size)

    def _set_percent(self, cell, value):
        cell.value = int(round(value or 0))
        cell.number_format = '0"%"'

    def _format_date(self, value):
        if not value:
            return None
        if hasattr(value, "date"):
            value = value.date()
        if value:
            return value.strftime("%Y-%m-%d")
        return None

    def _draw_period_section(self, start_row, funnel_data):
        row = start_row

        cell = self.ws.cell(row=row, column=2, value=funnel_data["period_name"])
        cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
        cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        row += 1

        stages = funnel_data["stages"]
        first = stages[0]
        last = stages[-1]

        summary = (
            f"Создано заказов без учета отмен и возвратов: {int(first['orders']):,}".replace(",", " ")
            + f" | Сумма: {int(first['amount']):,} ₽".replace(",", " ")
            + f" | Закрыто: {int(last['orders']):,}".replace(",", " ")
            + f" | Отменено: {int(funnel_data['canceled_orders']):,}".replace(",", " ")
            + f" | Возвращено: {int(funnel_data['returned_orders']):,}".replace(",", " ")
        )

        cell = self.ws.cell(row=row, column=2, value=summary)
        cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
        cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        row += 2

        headers = [
            "ЭТАП",
            "ЗАКАЗОВ",
            "СУММА",
            "КОНВ. ОТ ЭТАПА",
            "КОНВ. ОТ НАЧАЛА",
            "ЗАВИСЛО ЗАКАЗОВ",
            "ЗАВИСЛО СУММА",
            "ЗАВИСЛО %",
        ]

        for col_idx, header in enumerate(headers, start=2):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
            cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDERS["thin"]

        self.ws.row_dimensions[row].height = 34
        row += 1

        for idx, stage in enumerate(stages):
            prev = stages[idx - 1] if idx > 0 else None

            stuck_orders = prev["orders"] - stage["orders"] if prev else 0
            stuck_amount = prev["amount"] - stage["amount"] if prev else 0
            stuck_percent = (stuck_orders / prev["orders"] * 100) if prev and prev["orders"] else 0

            fill = FILLS["odd_row"] if idx % 2 == 0 else FILLS["even_row"]

            cell = self.ws.cell(row=row, column=2, value=stage["name"])
            cell.fill = fill
            cell.border = BORDERS["thin"]
            cell.font = Font(name="Roboto", size=9, color=COLORS["text_dark"])
            cell.alignment = Alignment(horizontal="left", vertical="center")

            cell = self.ws.cell(row=row, column=3)
            cell.fill = fill
            cell.border = BORDERS["thin"]
            self._set_number(cell, stage["orders"], bold=True)
            cell.alignment = Alignment(horizontal="right", vertical="center")

            cell = self.ws.cell(row=row, column=4)
            cell.fill = fill
            cell.border = BORDERS["thin"]
            self._set_money(cell, stage["amount"], bold=True)
            cell.alignment = Alignment(horizontal="right", vertical="center")

            cell = self.ws.cell(row=row, column=5)
            cell.fill = fill
            cell.border = BORDERS["thin"]
            self._set_percent(cell, stage["conversion_from_prev"])
            cell.font = Font(
                name="Roboto",
                size=9,
                bold=True,
                color=self._get_conv_color(stage["conversion_from_prev"]),
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

            cell = self.ws.cell(row=row, column=6)
            cell.fill = fill
            cell.border = BORDERS["thin"]
            self._set_percent(cell, stage["conversion_from_start"])
            cell.font = Font(
                name="Roboto",
                size=9,
                bold=True,
                color=self._get_conv_color(stage["conversion_from_start"]),
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

            cell = self.ws.cell(row=row, column=7)
            cell.fill = fill
            cell.border = BORDERS["thin"]
            self._set_number(cell, stuck_orders)
            cell.alignment = Alignment(horizontal="right", vertical="center")

            cell = self.ws.cell(row=row, column=8)
            cell.fill = fill
            cell.border = BORDERS["thin"]
            self._set_money(cell, stuck_amount)
            cell.alignment = Alignment(horizontal="right", vertical="center")

            cell = self.ws.cell(row=row, column=9)
            cell.fill = fill
            cell.border = BORDERS["thin"]
            self._set_percent(cell, stuck_percent)
            cell.font = self._number_font()
            cell.alignment = Alignment(horizontal="center", vertical="center")

            self.ws.row_dimensions[row].height = 24
            row += 1

        row += 1

        # Секция отмененных заказов
        cell = self.ws.cell(row=row, column=2, value="ОТМЕНЕННЫЕ ЗАКАЗЫ")
        cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["warning_red"])
        cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        # Заголовки - комментарий будет на 2 колонки (5 и 6)
        cancel_headers = ["ПЕРИОД", "ОТМЕНЕНО ЗАКАЗОВ", "СУММА ОТМЕН", "КОММЕНТАРИЙ"]

        for col_idx, header in enumerate(cancel_headers, start=2):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
            cell.fill = PatternFill("solid", fgColor=COLORS["warning_red"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDERS["thin"]
            
            # Если это заголовок "КОММЕНТАРИЙ" - объединяем колонки 5 и 6
            if header == "КОММЕНТАРИЙ":
                self.ws.merge_cells(start_row=row, start_column=col_idx, end_row=row, end_column=col_idx + 4)

        row += 1

        # ПЕРИОД
        cell = self.ws.cell(row=row, column=2, value=funnel_data["period_key"])
        cell.fill = FILLS["odd_row"]
        cell.border = BORDERS["thin"]
        cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
        cell.alignment = Alignment(horizontal="left", vertical="center")

        # ОТМЕНЕНО ЗАКАЗОВ
        cell = self.ws.cell(row=row, column=3)
        cell.fill = FILLS["odd_row"]
        cell.border = BORDERS["thin"]
        self._set_number(cell, funnel_data["canceled_orders"])
        cell.alignment = Alignment(horizontal="right", vertical="center")

        # СУММА ОТМЕН
        cell = self.ws.cell(row=row, column=4)
        cell.fill = FILLS["odd_row"]
        cell.border = BORDERS["thin"]
        self._set_money(cell, funnel_data["canceled_amount"])
        cell.alignment = Alignment(horizontal="right", vertical="center")

        # КОММЕНТАРИЙ - объединяем колонки 5 и 6
        comment_text = "ИСКЛЮЧЕНЫ ИЗ ОСНОВНОЙ ВОРОНКИ"
        cell = self.ws.cell(row=row, column=5, value=comment_text)
        cell.fill = FILLS["odd_row"]
        cell.border = BORDERS["thin"]
        cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["warning_red"])
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # Объединяем колонки 5 и 6
        self.ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=9)

        row += 2

        # Секция возвращенных заказов
        cell = self.ws.cell(row=row, column=2, value="ВОЗВРАЩЕННЫЕ ЗАКАЗЫ")
        cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["orange"])
        cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        # Заголовки - комментарий будет на 2 колонки (5 и 6)
        returned_headers = ["ПЕРИОД", "ВОЗВРАЩЕНО ЗАКАЗОВ", "СУММА ВОЗВРАТОВ", "КОММЕНТАРИЙ"]

        for col_idx, header in enumerate(returned_headers, start=2):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
            cell.fill = PatternFill("solid", fgColor=COLORS["orange"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDERS["thin"]
            
            # Если это заголовок "КОММЕНТАРИЙ" - объединяем колонки 5 и 6
            if header == "КОММЕНТАРИЙ":
                self.ws.merge_cells(start_row=row, start_column=col_idx, end_row=row, end_column=col_idx + 4)

        row += 1

        # ПЕРИОД
        cell = self.ws.cell(row=row, column=2, value=funnel_data["period_key"])
        cell.fill = FILLS["odd_row"]
        cell.border = BORDERS["thin"]
        cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
        cell.alignment = Alignment(horizontal="left", vertical="center")

        # ВОЗВРАЩЕНО ЗАКАЗОВ
        cell = self.ws.cell(row=row, column=3)
        cell.fill = FILLS["odd_row"]
        cell.border = BORDERS["thin"]
        self._set_number(cell, funnel_data["returned_orders"])
        cell.alignment = Alignment(horizontal="right", vertical="center")

        # СУММА ВОЗВРАТОВ
        cell = self.ws.cell(row=row, column=4)
        cell.fill = FILLS["odd_row"]
        cell.border = BORDERS["thin"]
        self._set_money(cell, funnel_data["returned_amount"])
        cell.alignment = Alignment(horizontal="right", vertical="center")

        # КОММЕНТАРИЙ - объединяем колонки 5 и 6
        comment_text = "БЫЛА ОПЛАТА/ОТГРУЗКА, НО КЛИЕНТ ОТКАЗАЛСЯ (ПОЛНЫЙ ВОЗВРАТ)"
        cell = self.ws.cell(row=row, column=5, value=comment_text)
        cell.fill = FILLS["odd_row"]
        cell.border = BORDERS["thin"]
        cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["orange"])
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # Объединяем колонки 5 и 6
        self.ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=9)

        row += 2
        
        return row

    def _draw_manager_funnel(self, start_row, managers_data):
        row = start_row

        cell = self.ws.cell(row=row, column=2, value="ВОРОНКА ПО МЕНЕДЖЕРАМ: YTD")
        cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
        cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
        row += 2

        headers = [
            "МЕНЕДЖЕР",
            "ЗАКАЗОВ",
            "СУММА",
            "ОПЛАТА",
            "КОНВ.%",
            "ОТГРУЗКА",
            "КОНВ.%",
            "ЗАКРЫТО",
            "КОНВ.%",
            "БЕЗ ОПЛАТЫ",
            "БЕЗ ОТГРУЗКИ",
            "НЕ ЗАКРЫТО",
        ]

        for col_idx, header in enumerate(headers, start=2):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["white"])
            cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDERS["thin"]

        self.ws.row_dimensions[row].height = 35
        row += 1

        def format_manager_name(manager_name):
            if not manager_name:
                return ""
            words = str(manager_name).strip().split()
            if not words:
                return ""
            first = words[0].upper()
            if len(words) >= 2 and words[1]:
                return f"{first} {words[1][0].upper()}."
            return first

        for idx, mgr in enumerate(managers_data[:15]):
            ytd = mgr["ytd"]
            fill = FILLS["odd_row"] if idx % 2 == 0 else FILLS["even_row"]
            formatted_name = format_manager_name(mgr["manager"])

            cell = self.ws.cell(row=row, column=2, value=formatted_name)
            cell.fill = fill
            cell.border = BORDERS["thin"]
            cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
            cell.alignment = Alignment(horizontal="left", vertical="center")

            numeric_columns = [
                (3, ytd["total_orders"], "number"),
                (4, ytd["total_amount"], "money"),
                (5, ytd["paid_orders"], "number"),
                (7, ytd["shipped_orders"], "number"),
                (9, ytd["closed_orders"], "number"),
                (11, ytd["stuck_without_payment"], "number_red"),
                (12, ytd["stuck_without_shipment"], "number_red"),
                (13, ytd["stuck_not_closed"], "number_red"),
            ]

            for col, value, kind in numeric_columns:
                cell = self.ws.cell(row=row, column=col)
                cell.fill = fill
                cell.border = BORDERS["thin"]

                if kind == "money":
                    self._set_money(cell, value)
                else:
                    self._set_number(cell, value)

                if kind == "number_red":
                    cell.font = Font(name="Roboto", size=9, color=COLORS["warning_red"])

                cell.alignment = Alignment(horizontal="right", vertical="center")

            percent_columns = [
                (6, ytd["conv_paid"]),
                (8, ytd["conv_shipped"]),
                (10, ytd["conv_closed"]),
            ]

            for col, value in percent_columns:
                cell = self.ws.cell(row=row, column=col)
                cell.fill = fill
                cell.border = BORDERS["thin"]
                self._set_percent(cell, value)
                cell.font = Font(
                    name="Roboto",
                    size=9,
                    bold=True,
                    color=self._get_conv_color(value),
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

            self.ws.row_dimensions[row].height = 22
            row += 1

        return row

    def _draw_orders_detail_table(self, start_row, orders_detail):
        row = start_row + 2


        cell = self.ws.cell(row=row, column=2, value="ДЕТАЛИЗАЦИЯ ПРОБЛЕМНЫХ ЗАКАЗОВ: YTD")
        cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
        cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
        row += 1

        cell = self.ws.cell(
            row=row,
            column=2,
            value="Таблица для фильтрации заказов и поиска проблем: нет оплаты, нет отгрузки, частичная отгрузка, частичная оплата, не закрыт в системе.",
        )
        cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
        cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
        row += 2

        headers = [
            "ПРОБЛЕМА",
            "КОММЕНТАРИЙ",
            "НОМЕР",
            "ДАТА",
            "СТАТУС",
            "КЛИЕНТ",
            "МЕНЕДЖЕР",
            "МАГАЗИН",
            "СУММА ЗАКАЗА",
            "ОПЛАЧЕНО",
            "ОТГРУЖЕНО",
        ]

        fields = [
            "problem_status",
            "problem_comment",
            "number",
            "date_from",
            "status",
            "client",
            "manager",
            "store",
            "amount_active",
            "cash_pmts",
            "total_shiped_amount",
        ]

        header_row = row

        for col_idx, header in enumerate(headers, start=2):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["white"])
            cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDERS["thin"]

        self.ws.row_dimensions[row].height = 34
        row += 1

        for idx, order in enumerate(orders_detail):
            fill = FILLS["odd_row"] if idx % 2 == 0 else FILLS["even_row"]

            for col_idx, field in enumerate(fields, start=2):
                value = order.get(field)

                cell = self.ws.cell(row=row, column=col_idx)
                cell.fill = fill
                cell.border = BORDERS["thin"]
                cell.font = Font(name="Roboto", size=8, color=COLORS["text_gray"])

                if field == "date_from":
                    cell.value = self._format_date(value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                elif field in ["amount_active", "cash_pmts", "total_shiped_amount"]:
                    if value:
                        cell.value = int(round(value))
                        cell.number_format = '#,##0 ₽'
                    else:
                        cell.value = 0
                        cell.number_format = '#,##0 ₽'
                    cell.alignment = Alignment(horizontal="right", vertical="center")

                else:
                    cell.value = value
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

                if field == "problem_status":
                    if value == "Нет оплаты":
                        cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["warning_red"])
                    elif value == "Нет отгрузки":
                        cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["orange"])
                    elif value == "Частичная отгрузка":
                        cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["orange"])
                    elif value == "Частичная оплата":
                        cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["orange"])
                    elif value == "Не закрыт в системе":
                        cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["dark_green"])

            self.ws.row_dimensions[row].height = 22
            row += 1

        last_row = row - 1

        if last_row > header_row:
            table_ref = f"B{header_row}:L{last_row}"
            table = Table(displayName="OrdersDetailTable", ref=table_ref)

            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=False,
                showColumnStripes=False,
            )

            table.tableStyleInfo = style
            self.ws.add_table(table)

        return row + 2

    def build(self, funnel_overall, funnel_by_manager, loss_analysis, orders_detail=None):
        row = 1

        if orders_detail is None:
            orders_detail = []

        draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 8

        row = self.sheet_title.draw(
            row=3,
            title="ОПЕРАЦИОННАЯ ВОРОНКА ЗАКАЗОВ",
            subtitle="Заказ → Оплата → Отгрузка → Полная оплата → Закрытие",
            date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y')}",
            start_col=2,
            end_col=9,
        )

        row += 2

        row = self._draw_period_section(row, funnel_overall["ytd"])
        row = self._draw_period_section(row, funnel_overall["mtd"])

        row = self._draw_manager_funnel(row, funnel_by_manager)

        row = self._draw_orders_detail_table(row, orders_detail)

        col_widths = {
            "B": 22,
            "C": 15,
            "D": 15,
            "E": 12,
            "F": 12,
            "G": 15,
            "H": 15,
            "I": 15,
            "J": 15,
            "K": 15,
            "L": 15,
            "M": 15,
        }

        for col, width in col_widths.items():
            self.ws.column_dimensions[col].width = width

        self.ws.sheet_view.showGridLines = False