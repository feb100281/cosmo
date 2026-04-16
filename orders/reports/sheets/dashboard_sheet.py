# orders/reports/sheets/dashboard_sheet.py
from datetime import date
from openpyxl.styles import Font, PatternFill
from .base_sheet import BaseSheet
from ..styles.theme import FILLS, FONTS, BORDERS, ALIGNMENTS, COLORS, FORMATS
from ..styles.helpers import draw_toc_button


class DashboardSheet(BaseSheet):
    """Лист с общей аналитикой - дашборд"""

    def __init__(self, workbook, sheet_number):
        super().__init__(workbook, sheet_number)

    def _format_currency(self, value):
        """Форматирует валюту"""
        if value is None:
            return "0 ₽"
        return f"{int(value):,} ₽".replace(",", " ")

    def _format_number(self, value):
        """Форматирует число"""
        if value is None:
            return "0"
        return f"{int(value):,}".replace(",", " ")

    def _draw_kpi_card(self, row, col, title, value, subtitle=None, color=COLORS["dark_green"]):
        """Рисует KPI карточку"""
        # Фон
        for r in range(row, row + 3):
            for c in range(col, col + 2):
                cell = self.ws.cell(row=r, column=c)
                cell.fill = FILLS["summary"]
                cell.border = BORDERS["thin"]
        
        # Заголовок
        title_cell = self.ws.cell(row=row, column=col, value=title)
        title_cell.font = FONTS["bold"]
        title_cell.alignment = ALIGNMENTS["left"]
        
        # Значение
        self.ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
        value_cell = self.ws.cell(row=row+1, column=col, value=value)
        value_cell.font = Font(name="Roboto", size=20, bold=True, color=color)
        value_cell.alignment = ALIGNMENTS["left"]
        
        # Подзаголовок
        if subtitle:
            sub_cell = self.ws.cell(row=row+2, column=col, value=subtitle)
            sub_cell.font = FONTS["subtitle"]
            sub_cell.alignment = ALIGNMENTS["left"]

    def _draw_section_header(self, row, title):
        """Рисует заголовок секции (без эмодзи)"""
        cell = self.ws.cell(row=row, column=2, value=title)
        cell.font = Font(name="Roboto", size=14, bold=True, color=COLORS["dark_green"])
        cell.alignment = ALIGNMENTS["left"]
        return row + 1

    def _draw_table_header(self, row, headers):
        """Рисует заголовок таблицы"""
        for col_idx, header in enumerate(headers, 2):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = FONTS["header_white"]
            cell.fill = FILLS["header"]
            cell.alignment = ALIGNMENTS["center"]
            cell.border = BORDERS["thin"]
        return row + 1

    def build(self, dashboard_data, active_debtors=None, status_summary=None):
        """Построение дашборда"""
        kpi = dashboard_data.get("kpi", {})
        ytd = dashboard_data.get("ytd", {})
        mtd = dashboard_data.get("mtd", {})
        managers_top = dashboard_data.get("managers_top", [])[:5]
        clients_top = dashboard_data.get("clients_top", [])[:10]
        
        if active_debtors is None:
            active_debtors = []
        
        today = date.today()
        report_date = today.strftime("%d.%m.%Y")
        
        row = 1
        
        # Кнопка назад
        draw_toc_button(self.ws, cell="B1", text="← ВЕРНУТЬСЯ В ОГЛАВЛЕНИЕ", target_sheet="TOC")
        
        # ============================================================
        # ЗАГОЛОВОК
        # ============================================================
        self.ws.cell(row=row + 1, column=2, value="КОНТРОЛЬ ЗАКАЗОВ")
        self.ws.cell(row=row + 1, column=2).font = Font(name="Roboto", size=22, bold=True, color=COLORS["dark_green"])
        
        self.ws.cell(row=row + 2, column=2, value="Оперативная аналитика по активным заказам")
        self.ws.cell(row=row + 2, column=2).font = FONTS["subtitle"]
        
        self.ws.cell(row=row + 3, column=2, value=f"Сформировано: {report_date}")
        self.ws.cell(row=row + 3, column=2).font = Font(name="Roboto", size=9, italic=True, color=COLORS["text_gray"])
        
        row = 6
        
        # ============================================================
        # БЛОК 1: ГЛАВНЫЕ KPI
        # ============================================================
        total_amount = kpi.get('active_amount', 0)
        total_paid = kpi.get('active_paid', 0)
        total_debt = kpi.get('active_debt', 0)
        paid_percent = (total_paid / total_amount * 100) if total_amount > 0 else 0
        
        # Карточка 1: Сумма
        self._draw_kpi_card(row, 2, "АКТИВНЫЕ ЗАКАЗЫ", 
                           self._format_currency(total_amount),
                           f"{self._format_number(kpi.get('active_orders', 0))} заказов", 
                           COLORS["dark_green"])
        
        # Карточка 2: Оплачено
        self._draw_kpi_card(row, 4, "ОПЛАЧЕНО", 
                           self._format_currency(total_paid),
                           f"{paid_percent:.0f}% от суммы", 
                           COLORS["blue"])
        
        # Карточка 3: Долг
        if total_debt > 1000000:
            debt_color = "DC3545"
        elif total_debt > 500000:
            debt_color = "FF8C00"
        else:
            debt_color = "28A745"
        
        self._draw_kpi_card(row, 6, "ДОЛГ", 
                           self._format_currency(total_debt),
                           "требует оплаты", 
                           debt_color)
        
        row += 4
        
        
        # ============================================================
        # БЛОК 2: ДИНАМИКА (табличный вид)
        # ============================================================
        row = self._draw_section_header(row, "ДИНАМИКА")
        
        # Подзаголовок
        self.ws.cell(row=row, column=2, value="Показатели текущего месяца и с начала года")
        self.ws.cell(row=row, column=2).font = Font(name="Roboto", size=10, italic=True, color=COLORS["text_gray"])
        row += 1
        
        # Заголовки таблицы динамики
        dyn_headers = ["ПЕРИОД", "ЗАКАЗОВ", "СУММА", "ОПЛАЧЕНО", "ДОЛГ"]
        for col_idx, header in enumerate(dyn_headers, 2):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = FONTS["header_white"]
            cell.fill = FILLS["header"]
            cell.alignment = ALIGNMENTS["center"]
            cell.border = BORDERS["thin"]
        row += 1
        
        # MTD строка
        mtd_amount = mtd.get('amount', 0)
        mtd_paid = mtd.get('paid', 0)
        mtd_debt = mtd.get('debt', 0)
        mtd_orders = mtd.get('orders', 0)
        
        self.ws.cell(row=row, column=2, value="ТЕКУЩИЙ МЕСЯЦ (MTD)")
        self.ws.cell(row=row, column=2).font = FONTS["bold"]
        self.ws.cell(row=row, column=2).border = BORDERS["thin"]
        self.ws.cell(row=row, column=2).alignment = ALIGNMENTS["left"]
        
        self.ws.cell(row=row, column=3, value=mtd_orders)
        self.ws.cell(row=row, column=3).border = BORDERS["thin"]
        self.ws.cell(row=row, column=3).alignment = ALIGNMENTS["center"]
        
        self.ws.cell(row=row, column=4, value=mtd_amount)
        self.ws.cell(row=row, column=4).border = BORDERS["thin"]
        self.ws.cell(row=row, column=4).number_format = FORMATS["money"]
        self.ws.cell(row=row, column=4).alignment = ALIGNMENTS["right"]
        
        self.ws.cell(row=row, column=5, value=mtd_paid)
        self.ws.cell(row=row, column=5).border = BORDERS["thin"]
        self.ws.cell(row=row, column=5).number_format = FORMATS["money"]
        self.ws.cell(row=row, column=5).alignment = ALIGNMENTS["right"]
        
        self.ws.cell(row=row, column=6, value=mtd_debt)
        self.ws.cell(row=row, column=6).border = BORDERS["thin"]
        self.ws.cell(row=row, column=6).number_format = FORMATS["money"]
        self.ws.cell(row=row, column=6).alignment = ALIGNMENTS["right"]
        
        row += 1
        
        # YTD строка
        ytd_amount = ytd.get('amount', 0)
        ytd_paid = ytd.get('paid', 0)
        ytd_debt = ytd.get('debt', 0)
        ytd_orders = ytd.get('orders', 0)
        
        fill = FILLS["alt"]
        
        self.ws.cell(row=row, column=2, value="С НАЧАЛА ГОДА (YTD)")
        self.ws.cell(row=row, column=2).font = FONTS["bold"]
        self.ws.cell(row=row, column=2).fill = fill
        self.ws.cell(row=row, column=2).border = BORDERS["thin"]
        self.ws.cell(row=row, column=2).alignment = ALIGNMENTS["left"]
        
        self.ws.cell(row=row, column=3, value=ytd_orders)
        self.ws.cell(row=row, column=3).fill = fill
        self.ws.cell(row=row, column=3).border = BORDERS["thin"]
        self.ws.cell(row=row, column=3).alignment = ALIGNMENTS["center"]
        
        self.ws.cell(row=row, column=4, value=ytd_amount)
        self.ws.cell(row=row, column=4).fill = fill
        self.ws.cell(row=row, column=4).border = BORDERS["thin"]
        self.ws.cell(row=row, column=4).number_format = FORMATS["money"]
        self.ws.cell(row=row, column=4).alignment = ALIGNMENTS["right"]
        
        self.ws.cell(row=row, column=5, value=ytd_paid)
        self.ws.cell(row=row, column=5).fill = fill
        self.ws.cell(row=row, column=5).border = BORDERS["thin"]
        self.ws.cell(row=row, column=5).number_format = FORMATS["money"]
        self.ws.cell(row=row, column=5).alignment = ALIGNMENTS["right"]
        
        self.ws.cell(row=row, column=6, value=ytd_debt)
        self.ws.cell(row=row, column=6).fill = fill
        self.ws.cell(row=row, column=6).border = BORDERS["thin"]
        self.ws.cell(row=row, column=6).number_format = FORMATS["money"]
        self.ws.cell(row=row, column=6).alignment = ALIGNMENTS["right"]
        
        row += 3
        
        # ============================================================
        # БЛОК 3: КРУПНАЯ ДЕБИТОРСКАЯ ЗАДОЛЖЕННОСТЬ
        # ============================================================
        major_debtors = [d for d in active_debtors if d.get('debt', 0) >= 300000]
        major_debtors = sorted(major_debtors, key=lambda x: x.get('debt', 0), reverse=True)[:5]
        
        if major_debtors:
            row = self._draw_section_header(row, "КРУПНАЯ ДЕБИТОРСКАЯ ЗАДОЛЖЕННОСТЬ")
            
            self.ws.cell(row=row, column=2, value=f"Требуют внимания: {len(major_debtors)} заказов")
            self.ws.cell(row=row, column=2).font = Font(name="Roboto", size=10, italic=True, color=COLORS["text_gray"])
            row += 1
            
            headers = ["КЛИЕНТ", "ЗАКАЗ", "СУММА", "ОПЛАЧЕНО", "ДОЛГ"]
            row = self._draw_table_header(row, headers)
            
            for idx, debtor in enumerate(major_debtors):
                fill = FILLS["alt"] if idx % 2 == 1 else FILLS["none"]
                
                self.ws.cell(row=row, column=2, value=debtor.get('client', '')[:30].upper())
                self.ws.cell(row=row, column=2).fill = fill
                self.ws.cell(row=row, column=2).alignment = ALIGNMENTS["left"]
                self.ws.cell(row=row, column=2).border = BORDERS["thin"]
                
                self.ws.cell(row=row, column=3, value=debtor.get('order_number', ''))
                self.ws.cell(row=row, column=3).fill = fill
                self.ws.cell(row=row, column=3).alignment = ALIGNMENTS["left"]
                self.ws.cell(row=row, column=3).border = BORDERS["thin"]
                
                self.ws.cell(row=row, column=4, value=debtor.get('order_amount', 0))
                self.ws.cell(row=row, column=4).fill = fill
                self.ws.cell(row=row, column=4).number_format = FORMATS["money"]
                self.ws.cell(row=row, column=4).alignment = ALIGNMENTS["right"]
                self.ws.cell(row=row, column=4).border = BORDERS["thin"]
                
                self.ws.cell(row=row, column=5, value=debtor.get('paid_amount', 0))
                self.ws.cell(row=row, column=5).fill = fill
                self.ws.cell(row=row, column=5).number_format = FORMATS["money"]
                self.ws.cell(row=row, column=5).alignment = ALIGNMENTS["right"]
                self.ws.cell(row=row, column=5).border = BORDERS["thin"]
                
                debt_cell = self.ws.cell(row=row, column=6, value=debtor.get('debt', 0))
                debt_cell.font = Font(bold=True, color="DC3545")
                debt_cell.fill = fill
                debt_cell.number_format = FORMATS["money"]
                debt_cell.alignment = ALIGNMENTS["right"]
                debt_cell.border = BORDERS["thin"]
                
                row += 1
            
            row += 2
        
        # ============================================================
        # БЛОК 4: ТОП МЕНЕДЖЕРОВ
        # ============================================================
        if managers_top:
            row = self._draw_section_header(row, "ТОП МЕНЕДЖЕРОВ")
            
            self.ws.cell(row=row, column=2, value="По сумме активных заказов")
            self.ws.cell(row=row, column=2).font = Font(name="Roboto", size=10, italic=True, color=COLORS["text_gray"])
            row += 1
            
            headers = ["МЕНЕДЖЕР", "ЗАКАЗОВ", "СУММА", "СРЕДНИЙ ЧЕК"]
            row = self._draw_table_header(row, headers)
            
            for idx, manager in enumerate(managers_top):
                fill = FILLS["alt"] if idx % 2 == 1 else FILLS["none"]
                
                name = manager.get("manager", "Не назначен")
                if name != "Не назначен" and len(name.split()) >= 2:
                    parts = name.split()
                    name = f"{parts[0]} {parts[1][0]}."

                self.ws.cell(row=row, column=2, value=name.upper())
                self.ws.cell(row=row, column=2).fill = fill
                self.ws.cell(row=row, column=2).alignment = ALIGNMENTS["left"]
                self.ws.cell(row=row, column=2).border = BORDERS["thin"]
                
                self.ws.cell(row=row, column=3, value=manager.get("orders_count", 0))
                self.ws.cell(row=row, column=3).fill = fill
                self.ws.cell(row=row, column=3).alignment = ALIGNMENTS["center"]
                self.ws.cell(row=row, column=3).border = BORDERS["thin"]
                
                self.ws.cell(row=row, column=4, value=manager.get("total_amount", 0))
                self.ws.cell(row=row, column=4).fill = fill
                self.ws.cell(row=row, column=4).number_format = FORMATS["money"]
                self.ws.cell(row=row, column=4).alignment = ALIGNMENTS["right"]
                self.ws.cell(row=row, column=4).border = BORDERS["thin"]
                
                self.ws.cell(row=row, column=5, value=manager.get("avg_check", 0))
                self.ws.cell(row=row, column=5).fill = fill
                self.ws.cell(row=row, column=5).number_format = FORMATS["money"]
                self.ws.cell(row=row, column=5).alignment = ALIGNMENTS["right"]
                self.ws.cell(row=row, column=5).border = BORDERS["thin"]
                
                row += 1
            
            row += 2
        
        # ============================================================
        # БЛОК 5: ТОП КЛИЕНТОВ
        # ============================================================
        if clients_top:
            row = self._draw_section_header(row, "ТОП КЛИЕНТОВ")
            
            self.ws.cell(row=row, column=2, value="По сумме активных заказов")
            self.ws.cell(row=row, column=2).font = Font(name="Roboto", size=10, italic=True, color=COLORS["text_gray"])
            row += 1
            
            headers = ["КЛИЕНТ", "ЗАКАЗОВ", "СУММА", "СРЕДНИЙ ЧЕК"]
            row = self._draw_table_header(row, headers)
            
            for idx, client in enumerate(clients_top):
                fill = FILLS["alt"] if idx % 2 == 1 else FILLS["none"]
                
                self.ws.cell(row=row, column=2, value=client.get("client", "НЕ УКАЗАН")[:35])
                self.ws.cell(row=row, column=2).fill = fill
                self.ws.cell(row=row, column=2).alignment = ALIGNMENTS["left"]
                self.ws.cell(row=row, column=2).border = BORDERS["thin"]
                
                self.ws.cell(row=row, column=3, value=client.get("orders_count", 0))
                self.ws.cell(row=row, column=3).fill = fill
                self.ws.cell(row=row, column=3).alignment = ALIGNMENTS["center"]
                self.ws.cell(row=row, column=3).border = BORDERS["thin"]
                
                self.ws.cell(row=row, column=4, value=client.get("total_amount", 0))
                self.ws.cell(row=row, column=4).fill = fill
                self.ws.cell(row=row, column=4).number_format = FORMATS["money"]
                self.ws.cell(row=row, column=4).alignment = ALIGNMENTS["right"]
                self.ws.cell(row=row, column=4).border = BORDERS["thin"]
                
                self.ws.cell(row=row, column=5, value=client.get("avg_check", 0))
                self.ws.cell(row=row, column=5).fill = fill
                self.ws.cell(row=row, column=5).number_format = FORMATS["money"]
                self.ws.cell(row=row, column=5).alignment = ALIGNMENTS["right"]
                self.ws.cell(row=row, column=5).border = BORDERS["thin"]
                
                row += 1
        
        # ============================================================
        # НАСТРОЙКА ШИРИНЫ КОЛОНОК
        # ============================================================
        self.ws.column_dimensions["A"].width = 3
        self.ws.column_dimensions["B"].width = 28
        self.ws.column_dimensions["C"].width = 14
        self.ws.column_dimensions["D"].width = 18
        self.ws.column_dimensions["E"].width = 18
        self.ws.column_dimensions["F"].width = 18
        self.ws.column_dimensions["G"].width = 14
        self.ws.column_dimensions["H"].width = 14
        self.ws.column_dimensions["I"].width = 14
        self.ws.column_dimensions["J"].width = 14
        self.ws.column_dimensions["K"].width = 14
        self.ws.column_dimensions["L"].width = 14
        self.ws.column_dimensions["M"].width = 14
        self.ws.column_dimensions["N"].width = 14
        
        # Высоты строк
        self.ws.row_dimensions[1].height = 30
        self.ws.row_dimensions[2].height = 25
        self.ws.row_dimensions[3].height = 20
        
        # Скрываем сетку
        self.ws.sheet_view.showGridLines = False