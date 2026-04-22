# # orders/reports/sheets/negative_payments_sheet.py
# from datetime import datetime
# from openpyxl.styles import Font, PatternFill
# from openpyxl.utils import get_column_letter
# from .base_sheet import BaseSheet
# from ..styles.theme import COLORS, ALIGNMENTS, BORDERS, FILLS, Alignment
# from ..components import create_kpi_cards, create_header, create_table
# from ..styles.helpers import draw_toc_button
# from ..components.footnote import Footnote
# from ..components.sheet_title import create_sheet_title


# class NegativePaymentsSheet(BaseSheet):
#     """Лист с заказами, у которых есть только отрицательные оплаты"""

#     def __init__(self, workbook, sheet_number):
#         super().__init__(workbook, sheet_number)
#         self.sheet_title = create_sheet_title(self.ws)
#         self.kpi = create_kpi_cards(self.ws)
#         self.header = create_header(self.ws)
#         self.table = create_table(self.ws)
        
#         # Красный цвет для отрицательных сумм
#         self.NEGATIVE_COLOR = "D32F2F"
#         self.WARNING_COLOR = "F57C00"

#     def _format_currency(self, value):
#         if value is None or value == 0:
#             return "0 ₽"
#         return f"{int(round(abs(value))):,} ₽".replace(",", " ")

#     def _format_negative_currency(self, value):
#         """Форматирует отрицательную сумму с минусом"""
#         if value is None or value == 0:
#             return "0 ₽"
#         sign = "-" if value < 0 else ""
#         return f"{sign}{int(round(abs(value))):,} ₽".replace(",", " ")

#     def _format_number(self, value):
#         if value is None or value == 0:
#             return "0"
#         return f"{int(round(abs(value))):,}".replace(",", " ")

#     def _safe_float(self, value, default=0.0):
#         if value is None:
#             return default
#         try:
#             return float(value)
#         except (ValueError, TypeError):
#             return default

#     def _safe_int(self, value, default=0):
#         if value is None:
#             return default
#         try:
#             return int(round(float(value)))
#         except (ValueError, TypeError):
#             return default

#     def _format_datetime(self, date_value):
#         """Форматирует дату без времени: YYYY-MM-DD"""
#         if not date_value:
#             return ""
#         if hasattr(date_value, 'strftime'):
#             return date_value.strftime('%Y-%m-%d')
#         if isinstance(date_value, str):
#             for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d.%m.%Y']:
#                 try:
#                     dt = datetime.strptime(date_value, fmt)
#                     return dt.strftime('%Y-%m-%d')
#                 except ValueError:
#                     continue
#         return str(date_value)[:10] if len(str(date_value)) > 10 else str(date_value)

#     def _format_manager_name(self, full_name):
#         """Форматирует имя менеджера: Иванов Иван -> ИВАНОВ И."""
#         if not full_name:
#             return ""
#         parts = str(full_name).strip().split()
#         if len(parts) >= 1:
#             last_name = parts[0].upper()
#             if len(parts) >= 2:
#                 first_initial = parts[1][0].upper() if parts[1] else ""
#                 return f"{last_name} {first_initial}."
#             return last_name
#         return str(full_name).upper()

#     def _format_client_name(self, name):
#         """Форматирует имя клиента: ООО Ромашка -> ООО РОМАШКА"""
#         if not name:
#             return ""
#         return str(name).upper()

#     def _format_payment_dates_simple(self, payment_dates):
#         """Форматирует даты оплат, выделяя отрицательные суммы красным"""
#         if not payment_dates:
#             return "—"
        
#         if isinstance(payment_dates, str):
#             import re
#             pattern = r'(\d{4}-\d{2}-\d{2}):\s*([\-\d\.]+)\s*руб'
#             matches = re.findall(pattern, payment_dates, re.IGNORECASE)
            
#             if matches:
#                 formatted = []
#                 for date, amount_str in matches[:10]:
#                     try:
#                         amount = float(amount_str)
#                         if amount < 0:
#                             formatted.append(f"{date}: -{self._format_currency(abs(amount))}")
#                         else:
#                             formatted.append(f"{date}: {self._format_currency(amount)}")
#                     except ValueError:
#                         formatted.append(f"{date}: {amount_str} руб")
                
#                 result = "\n".join(formatted)
#                 if len(matches) > 10:
#                     result += f"\n+ еще {len(matches) - 10}"
#                 return result
        
#         return str(payment_dates)[:100]

#     def build(self, orders_data, summary_data):
#         """Построение листа с заказами, у которых только отрицательные оплаты"""
#         row = 1

#         # ============================================================
#         # КНОПКА НАЗАД
#         # ============================================================
#         draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

#         self.ws.row_dimensions[1].height = 20
#         self.ws.row_dimensions[2].height = 8

#         # ============================================================
#         # ШАПКА
#         # ============================================================
#         row = self.sheet_title.draw(
#             row=3,
#             title="ЗАКАЗЫ С ОТРИЦАТЕЛЬНЫМИ ОПЛАТАМИ (БЕЗ ПОЛОЖИТЕЛЬНЫХ)",
#             subtitle=f"Заказы с возвратами/корректировками, созданные с 01.04.2025 | Всего найдено: {summary_data['total_orders']} заказов",
#             date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
#             start_col=2,
#             end_col=16
#         )

#         row += 1

#         # ============================================================
#         # KPI КАРТОЧКИ
#         # ============================================================
#         row1_cards = [
#             {
#                 'title': 'ЗАКАЗЫ С ВОЗВРАТАМИ',
#                 'value': self._format_number(summary_data['total_orders']),
#                 'subtitle': 'без положительных оплат',
#                 'color': self.WARNING_COLOR,
#                 'width': 1
#             },
#             {
#                 'title': 'ОБЩАЯ СУММА ВОЗВРАТОВ',
#                 'value': self._format_currency(abs(summary_data['total_negative_amount'])),
#                 'subtitle': 'возвращено клиентам',
#                 'color': self.NEGATIVE_COLOR
#             },
#             {
#                 'title': 'СУММА ЗАКАЗОВ',
#                 'value': self._format_currency(summary_data['total_order_amount']),
#                 'subtitle': 'активная сумма заказов',
#                 'color': COLORS["dark_green"],
#                 'width': 1
#             },
#         ]

#         row = self.kpi.draw_two_rows(row, row1_cards, [])
#         row -= 2

#         # Второй ряд KPI
#         row2_cards = [
#             {
#                 'title': 'УНИКАЛЬНЫХ КЛИЕНТОВ',
#                 'value': self._format_number(summary_data['unique_clients']),
#                 'subtitle': 'с проблемными заказами',
#                 'color': COLORS["blue"]
#             },
#             {
#                 'title': 'УНИКАЛЬНЫХ МЕНЕДЖЕРОВ',
#                 'value': self._format_number(summary_data['unique_managers']),
#                 'subtitle': 'вели эти заказы',
#                 'color': COLORS["blue"]
#             },
#             {
#                 'title': 'СР. СУММА ВОЗВРАТА',
#                 'value': self._format_currency(abs(summary_data['avg_negative_per_order'])),
#                 'subtitle': 'на один заказ',
#                 'color': self.WARNING_COLOR,
#                 'width': 1
#             },
#         ]

#         row = self.kpi.draw_two_rows(row, row2_cards, [])
#         row -= 2
#         row += 2

#         # ============================================================
#         # ЗАГОЛОВОК ТАБЛИЦЫ
#         # ============================================================
#         row = self.header.draw(row, "ДЕТАЛЬНАЯ ИНФОРМАЦИЯ ПО ЗАКАЗАМ", start_col=2, end_col=16)
#         row += 1

#         # ============================================================
#         # ЗАГОЛОВКИ ТАБЛИЦЫ
#         # ============================================================
#         headers = [
#             "ЗАКАЗ",
#             "ДАТА\nСОЗДАНИЯ",
#             "СТАТУС",
#             "КЛИЕНТ",
#             "МЕНЕДЖЕР",
#             "МАГАЗИН",
#             "СУММА\nЗАКАЗА",
#             "СУММА\nВОЗВРАТА",
#             "КОЛ-ВО\nВОЗВРАТОВ",
#             "ОПЛАТЫ\n(отрицательные выделены)",
#         ]

#         table_start_row = row

#         # Рисуем заголовки
#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
#             cell.fill = FILLS["header_blue"] if 'header_blue' in FILLS else PatternFill("solid", fgColor=self.NEGATIVE_COLOR)
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 40
#         row += 1

#         # ============================================================
#         # ДАННЫЕ
#         # ============================================================
#         for idx, order in enumerate(orders_data):
#             # Чередование фона
#             if idx % 2 == 0:
#                 fill = FILLS.get("odd_row", PatternFill(fill_type=None))
#             else:
#                 fill = FILLS.get("even_row", PatternFill(fill_type=None))

#             values = [
#                 order.get('order_name', '') or order.get('number', ''),
#                 self._format_datetime(order.get('date_from')),
#                 order.get('status', ''),
#                 self._format_client_name(order.get('client', '')),
#                 self._format_manager_name(order.get('manager', '')),
#                 (order.get('store', '') or '').upper()[:30],
#                 self._safe_float(order.get('amount_active')),
#                 order.get('negative_payments_total', 0),
#                 order.get('negative_payments_count', 0),
#                 self._format_payment_dates_simple(order.get('payment_dates')),
#             ]

#             for col_idx, value in enumerate(values):
#                 col_num = col_idx + 2
#                 cell = self.ws.cell(row=row, column=col_num, value=value)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]
                
#                 # Настройка выравнивания и цвета
#                 if col_idx in [0, 3, 4, 5]:  # ЗАКАЗ, КЛИЕНТ, МЕНЕДЖЕР, МАГАЗИН
#                     cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
#                     cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))
#                 elif col_idx in [1, 2]:  # ДАТА и СТАТУС
#                     cell.alignment = Alignment(horizontal="center", vertical="center")
#                     cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))
#                 elif col_idx == 7:  # СУММА ВОЗВРАТА - выделяем красным
#                     cell.alignment = Alignment(horizontal="right", vertical="center")
#                     font_color = self.NEGATIVE_COLOR
#                     cell.font = Font(name="Roboto", size=9, bold=True, color=font_color)
#                     cell.number_format = '#,##0'
#                 elif col_idx == 8:  # КОЛ-ВО ВОЗВРАТОВ
#                     cell.alignment = Alignment(horizontal="center", vertical="center")
#                     cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))
#                 elif col_idx == 9:  # ДАТЫ ОПЛАТ
#                     cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
#                     cell.font = Font(name="Roboto", size=8)
#                 else:  # Числовые колонки
#                     cell.alignment = Alignment(horizontal="right", vertical="center")
#                     cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))
#                     if isinstance(value, (int, float)):
#                         cell.number_format = '#,##0'

#                 # Специальная обработка статуса
#                 if col_idx == 2:  # СТАТУС
#                     status = value
#                     if "К выполнению" in status or "В резерве" in status:
#                         cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
#                     elif "На согласовании" in status:
#                         cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS.get("warning_orange", "F57C00"))
#                     elif "Закрыт" in status or "Отменен" in status:
#                         cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS.get("text_gray", "757575"))

#             self.ws.row_dimensions[row].height = 28
#             row += 1

#         # ============================================================
#         # ИТОГОВАЯ СТРОКА
#         # ============================================================
#         total_order_amount = sum(self._safe_float(o.get('amount_active')) for o in orders_data)
#         total_negative_amount = sum(self._safe_float(o.get('negative_payments_total')) for o in orders_data)
#         total_negative_count = sum(o.get('negative_payments_count', 0) for o in orders_data)

#         total_row = [
#             "ИТОГО:",
#             "",
#             "",
#             "",
#             "",
#             "",
#             total_order_amount,
#             total_negative_amount,
#             total_negative_count,
#             "",
#         ]

#         fill = FILLS["total"]
#         for col_idx, value in enumerate(total_row):
#             cell = self.ws.cell(row=row, column=col_idx + 2, value=value)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
            
#             if col_idx == 0:
#                 cell.font = Font(name="Roboto", bold=True, size=10)
#                 cell.alignment = Alignment(horizontal="right", vertical="center")
#             elif col_idx == 6 or col_idx == 7:
#                 cell.font = Font(name="Roboto", bold=True, size=10)
#                 cell.alignment = Alignment(horizontal="right", vertical="center")
#                 if isinstance(value, (int, float)):
#                     cell.number_format = '#,##0'
#                 if col_idx == 7:  # Сумма возврата - красным
#                     cell.font = Font(name="Roboto", bold=True, size=10, color=self.NEGATIVE_COLOR)
#             elif col_idx == 8:
#                 cell.font = Font(name="Roboto", bold=True, size=10)
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#             else:
#                 cell.font = Font(name="Roboto", bold=True, size=10)
#                 cell.alignment = Alignment(horizontal="center", vertical="center")

#         self.ws.row_dimensions[row].height = 30
#         row += 2

#         # ============================================================
#         # ПРИМЕЧАНИЕ
#         # ============================================================
#         footnote = Footnote(self.ws)
#         footnote.draw(
#             row=row,
#             text="* Внимание! В этих заказах есть возвраты/корректировки, но нет ни одной положительной оплаты. Рекомендуется проверить корректность проведенных операций."
#         )
#         row += 1
        
#         footnote.draw(
#             row=row,
#             text=f"** Отчет сформирован по заказам, созданным с 01.04.2025. Отрицательные суммы в колонке 'ОПЛАТЫ' выделены красным."
#         )

#         # ============================================================
#         # НАСТРОЙКА КОЛОНОК
#         # ============================================================
#         self.ws.column_dimensions["A"].width = 2
#         self.ws.column_dimensions["B"].width = 28  # ЗАКАЗ
#         self.ws.column_dimensions["C"].width = 12  # ДАТА СОЗДАНИЯ
#         self.ws.column_dimensions["D"].width = 22  # СТАТУС
#         self.ws.column_dimensions["E"].width = 28  # КЛИЕНТ
#         self.ws.column_dimensions["F"].width = 20  # МЕНЕДЖЕР
#         self.ws.column_dimensions["G"].width = 22  # МАГАЗИН
#         self.ws.column_dimensions["H"].width = 16  # СУММА ЗАКАЗА
#         self.ws.column_dimensions["I"].width = 16  # СУММА ВОЗВРАТА
#         self.ws.column_dimensions["J"].width = 14  # КОЛ-ВО ВОЗВРАТОВ
#         self.ws.column_dimensions["K"].width = 30  # ОПЛАТЫ

#         # Заморозка панелей
#         self.ws.freeze_panes = f'C{table_start_row + 1}'

#         # Автофильтр
#         self.ws.auto_filter.ref = f'B{table_start_row}:K{row - 3}'

#         # Скрываем сетку
#         self.ws.sheet_view.showGridLines = False





# # orders/reports/sheets/negative_payments_sheet.py
# from datetime import datetime
# from openpyxl.styles import Font, PatternFill
# from openpyxl.utils import get_column_letter
# from .base_sheet import BaseSheet
# from ..styles.theme import COLORS, ALIGNMENTS, BORDERS, FILLS, Alignment
# from ..components import create_kpi_cards, create_header, create_table
# from ..styles.helpers import draw_toc_button
# from ..components.footnote import Footnote
# from ..components.sheet_title import create_sheet_title


# class NegativePaymentsSheet(BaseSheet):
#     """Лист с заказами, у которых есть только отрицательные оплаты (возвраты ДС)"""

#     def __init__(self, workbook, sheet_number):
#         super().__init__(workbook, sheet_number)
#         self.sheet_title = create_sheet_title(self.ws)
#         self.kpi = create_kpi_cards(self.ws)
#         self.header = create_header(self.ws)
#         self.table = create_table(self.ws)
        
#         # Используем существующий warning_red из theme.py (7B1F3A - более приглушенный бордовый)
#         # Или можно использовать orange для менее агрессивного выделения
#         self.NEGATIVE_COLOR = COLORS.get("warning_red", "7B1F3A")  # Бордовый, не ярко-красный
#         self.WARNING_COLOR = COLORS.get("orange", "E67E22")  # Оранжевый для предупреждений

#     def _format_currency(self, value):
#         if value is None or value == 0:
#             return "0 ₽"
#         return f"{int(round(abs(value))):,} ₽".replace(",", " ")

#     def _format_negative_currency(self, value):
#         """Форматирует отрицательную сумму с минусом"""
#         if value is None or value == 0:
#             return "0 ₽"
#         sign = "-" if value < 0 else ""
#         return f"{sign}{int(round(abs(value))):,} ₽".replace(",", " ")

#     def _format_number(self, value):
#         if value is None or value == 0:
#             return "0"
#         return f"{int(round(abs(value))):,}".replace(",", " ")

#     def _safe_float(self, value, default=0.0):
#         if value is None:
#             return default
#         try:
#             return float(value)
#         except (ValueError, TypeError):
#             return default

#     def _safe_int(self, value, default=0):
#         if value is None:
#             return default
#         try:
#             return int(round(float(value)))
#         except (ValueError, TypeError):
#             return default

#     def _format_datetime(self, date_value):
#         """Форматирует дату без времени: YYYY-MM-DD"""
#         if not date_value:
#             return ""
#         if hasattr(date_value, 'strftime'):
#             return date_value.strftime('%Y-%m-%d')
#         if isinstance(date_value, str):
#             for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d.%m.%Y']:
#                 try:
#                     dt = datetime.strptime(date_value, fmt)
#                     return dt.strftime('%Y-%m-%d')
#                 except ValueError:
#                     continue
#         return str(date_value)[:10] if len(str(date_value)) > 10 else str(date_value)

#     def _format_manager_name(self, full_name):
#         """Форматирует имя менеджера: Иванов Иван -> ИВАНОВ И."""
#         if not full_name:
#             return ""
#         parts = str(full_name).strip().split()
#         if len(parts) >= 1:
#             last_name = parts[0].upper()
#             if len(parts) >= 2:
#                 first_initial = parts[1][0].upper() if parts[1] else ""
#                 return f"{last_name} {first_initial}."
#             return last_name
#         return str(full_name).upper()

#     def _format_client_name(self, name):
#         """Форматирует имя клиента: ООО Ромашка -> ООО РОМАШКА"""
#         if not name:
#             return ""
#         return str(name).upper()

#     def _format_payment_dates_simple(self, payment_dates):
#         """Форматирует даты оплат, выделяя отрицательные суммы цветом"""
#         if not payment_dates:
#             return "—"
        
#         if isinstance(payment_dates, str):
#             import re
#             pattern = r'(\d{4}-\d{2}-\d{2}):\s*([\-\d\.]+)\s*руб'
#             matches = re.findall(pattern, payment_dates, re.IGNORECASE)
            
#             if matches:
#                 formatted = []
#                 for date, amount_str in matches[:10]:
#                     try:
#                         amount = float(amount_str)
#                         if amount < 0:
#                             formatted.append(f"{date}: -{self._format_currency(abs(amount))}")
#                         else:
#                             formatted.append(f"{date}: {self._format_currency(amount)}")
#                     except ValueError:
#                         formatted.append(f"{date}: {amount_str} руб")
                
#                 result = "\n".join(formatted)
#                 if len(matches) > 10:
#                     result += f"\n+ еще {len(matches) - 10}"
#                 return result
        
#         return str(payment_dates)[:100]

#     def build(self, orders_data, summary_data):
#         """Построение листа с заказами, у которых только отрицательные оплаты"""
#         row = 1

#         # ============================================================
#         # КНОПКА НАЗАД
#         # ============================================================
#         draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

#         self.ws.row_dimensions[1].height = 20
#         self.ws.row_dimensions[2].height = 8

#         # ============================================================
#         # ШАПКА
#         # ============================================================
#         row = self.sheet_title.draw(
#             row=3,
#             title="ЗАКАЗЫ С ВОЗВРАТОМ ДЕНЕЖНЫХ СРЕДСТВ (БЕЗ ОПЛАТ)",
#             subtitle=f"Заказы, в которых были возвраты/корректировки ДС, но нет положительных оплат | Созданы с 01.04.2025 | Найдено: {summary_data['total_orders']} заказов",
#             date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
#             start_col=2,
#             end_col=16
#         )

#         row += 1
#         # ============================================================
#         # KPI КАРТОЧКИ (только первый ряд)
#         # ============================================================
#         row1_cards = [
#             {
#                 'title': 'ЗАКАЗЫ С ВОЗВРАТОМ ДС',
#                 'value': self._format_number(summary_data['total_orders']),
#                 'subtitle': 'без положительных оплат',
#                 'color': self.WARNING_COLOR,
#                 'width': 1
#             },
#             {
#                 'title': 'ВОЗВРАЩЕНО ДЕНЕЖНЫХ СРЕДСТВ',
#                 'value': self._format_currency(abs(summary_data['total_negative_amount'])),
#                 'subtitle': 'сумма возвратов клиентам',
#                 'color': self.NEGATIVE_COLOR
#             },
#             {
#                 'title': 'СУММА ЗАКАЗОВ',
#                 'value': self._format_currency(summary_data['total_order_amount']),
#                 'subtitle': 'активная сумма заказов',
#                 'color': COLORS["dark_green"],
#                 'width': 1
#             },
#         ]

#         # Рисуем только один ряд карточек
#         row = self.kpi.draw_row(row, row1_cards)  # ← используем draw_row вместо draw_two_rows
#         row += 2  # Отступ после карточек

#         # ============================================================
#         # ЗАГОЛОВОК ТАБЛИЦЫ
#         # ============================================================
#         row = self.header.draw(row, "ДЕТАЛЬНАЯ ИНФОРМАЦИЯ ПО ЗАКАЗАМ", start_col=2, end_col=16)
#         row += 1

#         # ============================================================
#         # ЗАГОЛОВКИ ТАБЛИЦЫ
#         # ============================================================
#         headers = [
#             "ЗАКАЗ",
#             "ДАТА\nСОЗДАНИЯ",
#             "СТАТУС",
#             "КЛИЕНТ",
#             "МЕНЕДЖЕР",
#             "МАГАЗИН",
#             "СУММА\nЗАКАЗА",
#             "ВОЗВРАЩЕНО\nДС",  # ← уточнили: возвращено денежных средств
#             "КОЛ-ВО\nВОЗВРАТОВ",
#             "ДАТЫ ОПЛАТ\n(отрицательные - возвраты ДС)",
#         ]

#         table_start_row = row

#         # Рисуем заголовки
#         for col_idx, header in enumerate(headers, start=2):
#             cell = self.ws.cell(row=row, column=col_idx, value=header)
#             cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
#             cell.fill = FILLS["header"] if "header" in FILLS else PatternFill("solid", fgColor=COLORS["dark_green"])
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = BORDERS["thin"]

#         self.ws.row_dimensions[row].height = 40
#         row += 1

#         # ============================================================
#         # ДАННЫЕ
#         # ============================================================
#         for idx, order in enumerate(orders_data):
#             # Чередование фона
#             if idx % 2 == 0:
#                 fill = FILLS.get("odd_row", PatternFill(fill_type=None))
#             else:
#                 fill = FILLS.get("even_row", PatternFill(fill_type=None))

#             values = [
#                 order.get('order_name', '') or order.get('number', ''),
#                 self._format_datetime(order.get('date_from')),
#                 order.get('status', ''),
#                 self._format_client_name(order.get('client', '')),
#                 self._format_manager_name(order.get('manager', '')),
#                 (order.get('store', '') or '').upper()[:30],
#                 self._safe_float(order.get('amount_active')),
#                 order.get('negative_payments_total', 0),
#                 order.get('negative_payments_count', 0),
#                 self._format_payment_dates_simple(order.get('payment_dates')),
#             ]

#             for col_idx, value in enumerate(values):
#                 col_num = col_idx + 2
#                 cell = self.ws.cell(row=row, column=col_num, value=value)
#                 cell.fill = fill
#                 cell.border = BORDERS["thin"]
                
#                 # Настройка выравнивания и цвета
#                 if col_idx in [0, 3, 4, 5]:  # ЗАКАЗ, КЛИЕНТ, МЕНЕДЖЕР, МАГАЗИН
#                     cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
#                     cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))
#                 elif col_idx in [1, 2]:  # ДАТА и СТАТУС
#                     cell.alignment = Alignment(horizontal="center", vertical="center")
#                     cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))
#                 elif col_idx == 7:  # ВОЗВРАЩЕНО ДС - выделяем приглушенным бордовым
#                     cell.alignment = Alignment(horizontal="right", vertical="center")
#                     font_color = self.NEGATIVE_COLOR
#                     cell.font = Font(name="Roboto", size=9, bold=True, color=font_color)
#                     cell.number_format = '#,##0'
#                 elif col_idx == 8:  # КОЛ-ВО ВОЗВРАТОВ
#                     cell.alignment = Alignment(horizontal="center", vertical="center")
#                     cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))
#                 elif col_idx == 9:  # ДАТЫ ОПЛАТ
#                     cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
#                     cell.font = Font(name="Roboto", size=8)
#                 else:  # Числовые колонки
#                     cell.alignment = Alignment(horizontal="right", vertical="center")
#                     cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))
#                     if isinstance(value, (int, float)):
#                         cell.number_format = '#,##0'

#                 # Специальная обработка статуса
#                 if col_idx == 2:  # СТАТУС
#                     status = value
#                     if "К выполнению" in status or "В резерве" in status:
#                         cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
#                     elif "На согласовании" in status:
#                         cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS.get("orange", "E67E22"))
#                     elif "Закрыт" in status or "Отменен" in status:
#                         cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS.get("text_gray", "757575"))

#             self.ws.row_dimensions[row].height = 28
#             row += 1

#         # ============================================================
#         # ИТОГОВАЯ СТРОКА
#         # ============================================================
#         total_order_amount = sum(self._safe_float(o.get('amount_active')) for o in orders_data)
#         total_negative_amount = sum(self._safe_float(o.get('negative_payments_total')) for o in orders_data)
#         total_negative_count = sum(o.get('negative_payments_count', 0) for o in orders_data)

#         total_row = [
#             "ИТОГО:",
#             "",
#             "",
#             "",
#             "",
#             "",
#             total_order_amount,
#             total_negative_amount,
#             total_negative_count,
#             "",
#         ]

#         fill = FILLS["total"]
#         for col_idx, value in enumerate(total_row):
#             cell = self.ws.cell(row=row, column=col_idx + 2, value=value)
#             cell.fill = fill
#             cell.border = BORDERS["thin"]
            
#             if col_idx == 0:
#                 cell.font = Font(name="Roboto", bold=True, size=10)
#                 cell.alignment = Alignment(horizontal="right", vertical="center")
#             elif col_idx == 6 or col_idx == 7:
#                 cell.font = Font(name="Roboto", bold=True, size=10)
#                 cell.alignment = Alignment(horizontal="right", vertical="center")
#                 if isinstance(value, (int, float)):
#                     cell.number_format = '#,##0'
#                 if col_idx == 7:  # Возвращено ДС - бордовым цветом
#                     cell.font = Font(name="Roboto", bold=True, size=10, color=self.NEGATIVE_COLOR)
#             elif col_idx == 8:
#                 cell.font = Font(name="Roboto", bold=True, size=10)
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#             else:
#                 cell.font = Font(name="Roboto", bold=True, size=10)
#                 cell.alignment = Alignment(horizontal="center", vertical="center")

#         self.ws.row_dimensions[row].height = 30
#         row += 2

#         # ============================================================
#         # ПРИМЕЧАНИЕ
#         # ============================================================
#         footnote = Footnote(self.ws)
#         footnote.draw(
#             row=row,
#             text="* Внимание! В этих заказах есть возвраты/корректировки денежных средств, но нет ни одной положительной оплаты. Рекомендуется проверить корректность проведенных операций."
#         )
#         row += 1
        
#         footnote.draw(
#             row=row,
#             text=f"** Отчет сформирован по заказам, созданным с 01.04.2025. Отрицательные суммы (возвраты ДС) в колонке 'ДАТЫ ОПЛАТ' выделены красным цветом."
#         )

#         # ============================================================
#         # НАСТРОЙКА КОЛОНОК
#         # ============================================================
#         self.ws.column_dimensions["A"].width = 2
#         self.ws.column_dimensions["B"].width = 28  # ЗАКАЗ
#         self.ws.column_dimensions["C"].width = 12  # ДАТА СОЗДАНИЯ
#         self.ws.column_dimensions["D"].width = 22  # СТАТУС
#         self.ws.column_dimensions["E"].width = 28  # КЛИЕНТ
#         self.ws.column_dimensions["F"].width = 20  # МЕНЕДЖЕР
#         self.ws.column_dimensions["G"].width = 22  # МАГАЗИН
#         self.ws.column_dimensions["H"].width = 16  # СУММА ЗАКАЗА
#         self.ws.column_dimensions["I"].width = 18  # ВОЗВРАЩЕНО ДС (шире, чтобы влезло)
#         self.ws.column_dimensions["J"].width = 14  # КОЛ-ВО ВОЗВРАТОВ
#         self.ws.column_dimensions["K"].width = 35  # ДАТЫ ОПЛАТ (шире для лучшей читаемости)

#         # Заморозка панелей
#         self.ws.freeze_panes = f'C{table_start_row + 1}'

#         # Автофильтр
#         self.ws.auto_filter.ref = f'B{table_start_row}:K{row - 3}'

#         # Скрываем сетку
#         self.ws.sheet_view.showGridLines = False




# orders/reports/sheets/negative_payments_sheet.py
from datetime import datetime
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from .base_sheet import BaseSheet
from ..styles.theme import COLORS, ALIGNMENTS, BORDERS, FILLS, Alignment
from ..components import create_kpi_cards, create_header, create_table
from ..styles.helpers import draw_toc_button
from ..components.footnote import Footnote
from ..components.sheet_title import create_sheet_title


class NegativePaymentsSheet(BaseSheet):
    """Лист с заказами, у которых есть только отрицательные оплаты (возвраты ДС)"""

    def __init__(self, workbook, sheet_number):
        super().__init__(workbook, sheet_number)
        self.sheet_title = create_sheet_title(self.ws)
        self.kpi = create_kpi_cards(self.ws)
        self.header = create_header(self.ws)
        self.table = create_table(self.ws)
        
        # Используем существующий warning_red из theme.py (7B1F3A - более приглушенный бордовый)
        self.NEGATIVE_COLOR = COLORS.get("warning_red", "7B1F3A")
        self.WARNING_COLOR = COLORS.get("orange", "E67E22")
        self.SUCCESS_COLOR = COLORS.get("dark_green", "2F6656")

    def _format_currency(self, value):
        if value is None or value == 0:
            return "0 ₽"
        return f"{int(round(abs(value))):,} ₽".replace(",", " ")

    def _format_negative_currency(self, value):
        """Форматирует отрицательную сумму с минусом"""
        if value is None or value == 0:
            return "0 ₽"
        sign = "-" if value < 0 else ""
        return f"{sign}{int(round(abs(value))):,} ₽".replace(",", " ")

    def _format_number(self, value):
        if value is None or value == 0:
            return "0"
        return f"{int(round(abs(value))):,}".replace(",", " ")

    def _format_percent(self, value):
        """Форматирует процент"""
        if value is None:
            return "0%"
        return f"{value:.1f}%"

    def _safe_float(self, value, default=0.0):
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default

    def _safe_int(self, value, default=0):
        if value is None:
            return default
        try:
            return int(round(float(value)))
        except (ValueError, TypeError):
            return default

    def _format_datetime(self, date_value):
        """Форматирует дату без времени: YYYY-MM-DD"""
        if not date_value:
            return ""
        if hasattr(date_value, 'strftime'):
            return date_value.strftime('%Y-%m-%d')
        if isinstance(date_value, str):
            for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d.%m.%Y']:
                try:
                    dt = datetime.strptime(date_value, fmt)
                    return dt.strftime('%Y-%m-%d')
                except ValueError:
                    continue
        return str(date_value)[:10] if len(str(date_value)) > 10 else str(date_value)

    def _format_manager_name(self, full_name):
        """Форматирует имя менеджера: Иванов Иван -> ИВАНОВ И."""
        if not full_name:
            return ""
        parts = str(full_name).strip().split()
        if len(parts) >= 1:
            last_name = parts[0].upper()
            if len(parts) >= 2:
                first_initial = parts[1][0].upper() if parts[1] else ""
                return f"{last_name} {first_initial}."
            return last_name
        return str(full_name).upper()

    def _format_client_name(self, name):
        """Форматирует имя клиента: ООО Ромашка -> ООО РОМАШКА"""
        if not name:
            return ""
        return str(name).upper()

    def _format_payment_dates_simple(self, payment_dates):
        """Форматирует даты оплат, выделяя отрицательные суммы цветом"""
        if not payment_dates:
            return "—"
        
        if isinstance(payment_dates, str):
            import re
            pattern = r'(\d{4}-\d{2}-\d{2}):\s*([\-\d\.]+)\s*руб'
            matches = re.findall(pattern, payment_dates, re.IGNORECASE)
            
            if matches:
                formatted = []
                for date, amount_str in matches[:10]:
                    try:
                        amount = float(amount_str)
                        if amount < 0:
                            formatted.append(f"{date}: -{self._format_currency(abs(amount))}")
                        else:
                            formatted.append(f"{date}: {self._format_currency(amount)}")
                    except ValueError:
                        formatted.append(f"{date}: {amount_str} руб")
                
                result = "\n".join(formatted)
                if len(matches) > 10:
                    result += f"\n+ еще {len(matches) - 10}"
                return result
        
        return str(payment_dates)[:100]

    def build(self, orders_data, summary_data):
        """Построение листа с заказами, у которых только отрицательные оплаты"""
        row = 1

        # ============================================================
        # КНОПКА НАЗАД
        # ============================================================
        draw_toc_button(self.ws, cell="B1", text="← Оглавление", target_sheet="TOC")

        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 8

        # ============================================================
        # ШАПКА
        # ============================================================
        row = self.sheet_title.draw(
            row=3,
            title="ЗАКАЗЫ С ВОЗВРАТОМ ДЕНЕЖНЫХ СРЕДСТВ (БЕЗ ОПЛАТ)",
            subtitle=f"Заказы, в которых были возвраты/корректировки ДС, но нет положительных оплат | Созданы с 01.04.2025 | Найдено: {summary_data['total_orders']} заказов",
            date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            start_col=2,
            end_col=18
        )

        row += 1
        
        # ============================================================
        # KPI КАРТОЧКИ (первый ряд)
        # ============================================================
        row1_cards = [
            {
                'title': 'ЗАКАЗЫ С ВОЗВРАТОМ ДС',
                'value': self._format_number(summary_data['total_orders']),
                'subtitle': 'без положительных оплат',
                'color': self.WARNING_COLOR,
                'width': 1
            },
            {
                'title': 'ВОЗВРАЩЕНО ДС',
                'value': self._format_currency(abs(summary_data['total_negative_amount'])),
                'subtitle': 'сумма возвратов клиентам',
                'color': self.NEGATIVE_COLOR
            },
            {
                'title': 'ОТГРУЖЕНО НА СУММУ',
                'value': self._format_currency(summary_data['total_shipped_amount']),
                'subtitle': f'в {summary_data["orders_with_shipment"]} заказах',
                'color': self.SUCCESS_COLOR,
                'width': 1
            },
        ]

        row = self.kpi.draw_row(row, row1_cards)
        row += 2

        # ============================================================
        # ЗАГОЛОВОК ТАБЛИЦЫ
        # ============================================================
        row = self.header.draw(row, "ДЕТАЛЬНАЯ ИНФОРМАЦИЯ ПО ЗАКАЗАМ", start_col=2, end_col=18)
        row += 1

        # ============================================================
        # ЗАГОЛОВКИ ТАБЛИЦЫ (расширенные)
        # ============================================================
        headers = [
            "ЗАКАЗ",
            "ДАТА\nСОЗДАНИЯ",
            "СТАТУС",
            "КЛИЕНТ",
            "МЕНЕДЖЕР",
            "МАГАЗИН",
            "СУММА\nЗАКАЗА",
            "ОТГРУЖЕНО\nВСЕГО",
            "В Т.Ч.\nДОСТАВКА",
            "ОТКЛОНЕНИЕ",
            "%",
            "ВОЗВРАЩЕНО\nДС",
            "КОЛ-ВО\nВОЗВРАТОВ",
            "ДАТЫ ОПЛАТ",
        ]

        table_start_row = row

        # Рисуем заголовки
        for col_idx, header in enumerate(headers, start=2):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")
            cell.fill = FILLS["header"] if "header" in FILLS else PatternFill("solid", fgColor=COLORS["dark_green"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDERS["thin"]

        self.ws.row_dimensions[row].height = 40
        row += 1

        # ============================================================
        # ДАННЫЕ
        # ============================================================
        for idx, order in enumerate(orders_data):
            # Чередование фона
            if idx % 2 == 0:
                fill = FILLS.get("odd_row", PatternFill(fill_type=None))
            else:
                fill = FILLS.get("even_row", PatternFill(fill_type=None))

            # Получаем данные с отклонениями
            shipment_variance = order.get('shipment_variance', 0)
            shipment_percent = order.get('shipment_percent', 0)
            
            values = [
                order.get('order_name', '') or order.get('number', ''),
                self._format_datetime(order.get('date_from')),
                order.get('status', ''),
                self._format_client_name(order.get('client', '')),
                self._format_manager_name(order.get('manager', '')),
                (order.get('store', '') or '').upper()[:30],
                self._safe_float(order.get('amount_active')),
                self._safe_float(order.get('total_shiped_amount')),
                self._safe_float(order.get('shiped_delivery_amount')),
                shipment_variance,
                shipment_percent,
                order.get('negative_payments_total', 0),
                order.get('negative_payments_count', 0),
                self._format_payment_dates_simple(order.get('payment_dates')),
            ]

            for col_idx, value in enumerate(values):
                col_num = col_idx + 2
                cell = self.ws.cell(row=row, column=col_num, value=value)
                cell.fill = fill
                cell.border = BORDERS["thin"]
                
                # Настройка выравнивания и цвета
                if col_idx in [0, 3, 4, 5]:  # ЗАКАЗ, КЛИЕНТ, МЕНЕДЖЕР, МАГАЗИН
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))
                elif col_idx in [1, 2]:  # ДАТА и СТАТУС
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))
                elif col_idx == 9:  # ОТКЛОНЕНИЕ
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '#,##0'
                    if value > 0:
                        cell.font = Font(name="Roboto", size=9, color=self.WARNING_COLOR)
                    elif value < 0:
                        cell.font = Font(name="Roboto", size=9, color=self.NEGATIVE_COLOR)
                    else:
                        cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_gray", "666666"))
                elif col_idx == 10:  # %
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '0.0"%"' if value else '0.0"%"' 
                    if value > 100:
                        cell.font = Font(name="Roboto", size=9, color=self.WARNING_COLOR)
                    elif value < 100 and value > 0:
                        cell.font = Font(name="Roboto", size=9, color=self.NEGATIVE_COLOR)
                    else:
                        cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_gray", "666666"))
                elif col_idx == 11:  # ВОЗВРАЩЕНО ДС
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    font_color = self.NEGATIVE_COLOR
                    cell.font = Font(name="Roboto", size=9, bold=True, color=font_color)
                    cell.number_format = '#,##0'
                elif col_idx == 12:  # КОЛ-ВО ВОЗВРАТОВ
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))
                elif col_idx == 13:  # ДАТЫ ОПЛАТ
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.font = Font(name="Roboto", size=8)
                else:  # Числовые колонки
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.font = Font(name="Roboto", size=9, color=COLORS.get("text_dark", "1F1F1F"))
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0'

                # Специальная обработка статуса
                if col_idx == 2:  # СТАТУС
                    status = value
                    if "К выполнению" in status or "В резерве" in status:
                        cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
                    elif "На согласовании" in status:
                        cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS.get("orange", "E67E22"))
                    elif "Закрыт" in status or "Отменен" in status:
                        cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS.get("text_gray", "757575"))

            self.ws.row_dimensions[row].height = 28
            row += 1

        # ============================================================
        # ИТОГОВАЯ СТРОКА
        # ============================================================
        total_order_amount = sum(self._safe_float(o.get('amount_active')) for o in orders_data)
        total_shipped_amount = sum(self._safe_float(o.get('total_shiped_amount')) for o in orders_data)
        total_delivery_amount = sum(self._safe_float(o.get('shiped_delivery_amount')) for o in orders_data)
        total_variance = total_shipped_amount - total_order_amount
        total_negative_amount = sum(self._safe_float(o.get('negative_payments_total')) for o in orders_data)
        total_negative_count = sum(o.get('negative_payments_count', 0) for o in orders_data)

        total_row = [
            "ИТОГО:",
            "",
            "",
            "",
            "",
            "",
            total_order_amount,
            total_shipped_amount,
            total_delivery_amount,
            total_variance,
            "",
            total_negative_amount,
            total_negative_count,
            "",
        ]

        fill = FILLS["total"]
        for col_idx, value in enumerate(total_row):
            cell = self.ws.cell(row=row, column=col_idx + 2, value=value)
            cell.fill = fill
            cell.border = BORDERS["thin"]
            
            if col_idx == 0:
                cell.font = Font(name="Roboto", bold=True, size=10)
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [6, 7, 8, 9, 11]:
                cell.font = Font(name="Roboto", bold=True, size=10)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0'
                if col_idx == 11:  # Возвращено ДС
                    cell.font = Font(name="Roboto", bold=True, size=10, color=self.NEGATIVE_COLOR)
                if col_idx == 9 and value > 0:  # Отклонение положительное
                    cell.font = Font(name="Roboto", bold=True, size=10, color=self.WARNING_COLOR)
                elif col_idx == 9 and value < 0:
                    cell.font = Font(name="Roboto", bold=True, size=10, color=self.NEGATIVE_COLOR)
            elif col_idx == 12:
                cell.font = Font(name="Roboto", bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = Font(name="Roboto", bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        self.ws.row_dimensions[row].height = 30
        row += 2

        # ============================================================
        # ПРИМЕЧАНИЕ
        # ============================================================
        footnote = Footnote(self.ws)
        footnote.draw(
            row=row,
            text="* Внимание! В этих заказах есть возвраты/корректировки денежных средств, но нет ни одной положительной оплаты."
        )
        row += 1
        
        footnote.draw(
            row=row,
            text="** Отклонение = Отгружено всего - Сумма заказа. Положительное отклонение = переотгрузка, отрицательное = недогруз."
        )
        row += 1
        
        footnote.draw(
            row=row,
            text=f"*** Отчет сформирован по заказам, созданным с 01.04.2025."
        )

        # ============================================================
        # НАСТРОЙКА КОЛОНОК
        # ============================================================
        self.ws.column_dimensions["A"].width = 2
        self.ws.column_dimensions["B"].width = 28  # ЗАКАЗ
        self.ws.column_dimensions["C"].width = 12  # ДАТА СОЗДАНИЯ
        self.ws.column_dimensions["D"].width = 22  # СТАТУС
        self.ws.column_dimensions["E"].width = 28  # КЛИЕНТ
        self.ws.column_dimensions["F"].width = 20  # МЕНЕДЖЕР
        self.ws.column_dimensions["G"].width = 22  # МАГАЗИН
        self.ws.column_dimensions["H"].width = 16  # СУММА ЗАКАЗА
        self.ws.column_dimensions["I"].width = 16  # ОТГРУЖЕНО ВСЕГО
        self.ws.column_dimensions["J"].width = 14  # В Т.Ч. ДОСТАВКА
        self.ws.column_dimensions["K"].width = 14  # ОТКЛОНЕНИЕ
        self.ws.column_dimensions["L"].width = 10  # %
        self.ws.column_dimensions["M"].width = 18  # ВОЗВРАЩЕНО ДС
        self.ws.column_dimensions["N"].width = 14  # КОЛ-ВО ВОЗВРАТОВ
        self.ws.column_dimensions["O"].width = 35  # ДАТЫ ОПЛАТ

        # Заморозка панелей
        self.ws.freeze_panes = f'D{table_start_row + 1}'

        # Автофильтр
        self.ws.auto_filter.ref = f'B{table_start_row}:O{row - 4}'

        # Скрываем сетку
        self.ws.sheet_view.showGridLines = False