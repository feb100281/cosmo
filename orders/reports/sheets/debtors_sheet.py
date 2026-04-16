# orders/reports/sheets/debtors_sheet.py
from datetime import date
from openpyxl.styles import Font, PatternFill
from .base_sheet import BaseSheet
from ..styles.theme import FILLS, FONTS, BORDERS, ALIGNMENTS, COLORS


class ActiveDebtorsSheet(BaseSheet):
    """Активная дебиторка - активные заказы с долгом"""

    def __init__(self, workbook, sheet_number):
        super().__init__(workbook, sheet_number)

    def build(self, debtors):
        today = date.today().strftime("%d.%m.%Y")
        row = 1
        
        # Анализируем переплаты (отрицательный долг) и нормальные долги
        overpayments = [d for d in debtors if d.get('debt', 0) < 0]
        total_overpayment = sum(abs(d.get('debt', 0)) for d in overpayments)
        normal_debtors = [d for d in debtors if d.get('debt', 0) > 0]
        total_normal_debt = sum(d.get('debt', 0) for d in normal_debtors)
        zero_debtors = [d for d in debtors if d.get('debt', 0) == 0]

        # Ссылка в оглавление
        toc_link = self.ws.cell(row=row, column=2, value="← ВЕРНУТЬСЯ В ОГЛАВЛЕНИЕ")
        toc_link.font = Font(name="Roboto", size=9, bold=True, color=COLORS["blue"], underline="single")
        toc_link.alignment = ALIGNMENTS["left"]
        toc_link.hyperlink = "#'TOC'!A1"

        self.ws.cell(row=row + 1, column=2, value="АКТИВНАЯ ДЕБИТОРСКАЯ ЗАДОЛЖЕННОСТЬ")
        self.ws.cell(row=row + 1, column=2).font = Font(name="Roboto", size=16, bold=True, color=COLORS["dark_green"])

        self.ws.cell(row=row + 2, column=2, value=f"Активные заказы с долгом на дату отчёта: {today}")
        self.ws.cell(row=row + 2, column=2).font = Font(name="Roboto", size=9, italic=True, color=COLORS["text_gray"])

        row = 6
        
        # ============================================================
        # БЛОК САММАРИ: ПЕРЕПЛАТЫ И ОБЩАЯ СТАТИСТИКА
        # ============================================================
        if overpayments:
            # Желтый фон для предупреждения
            warning_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            
            # Заголовок предупреждения
            warn_cell = self.ws.cell(row=row, column=2, value="⚠ ВНИМАНИЕ: ОБНАРУЖЕНЫ ПЕРЕПЛАТЫ (ОТРИЦАТЕЛЬНЫЙ ДОЛГ)")
            warn_cell.font = Font(name="Roboto", size=11, bold=True, color="856404")
            warn_cell.fill = warning_fill
            warn_cell.alignment = ALIGNMENTS["left"]
            self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            row += 1
            
            # Информация о переплатах
            info_cell = self.ws.cell(row=row, column=2, value=f"• Количество заказов с переплатой: {len(overpayments)}")
            info_cell.font = Font(name="Roboto", size=10)
            info_cell.fill = warning_fill
            info_cell.alignment = ALIGNMENTS["left"]
            self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            row += 1
            
            overpay_cell = self.ws.cell(row=row, column=2, value=f"• Общая сумма переплаты: {total_overpayment:,.0f} ₽".replace(",", " "))
            overpay_cell.font = Font(name="Roboto", size=10, bold=True, color="DC3545")
            overpay_cell.fill = warning_fill
            overpay_cell.alignment = ALIGNMENTS["left"]
            self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            row += 1
            
            row += 1  # Дополнительный отступ
        
        # Общая статистика по долгам
        stats_row = row
        stats_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        
        # Количество должников
        self.ws.cell(row=stats_row, column=2, value="Всего должников:")
        self.ws.cell(row=stats_row, column=2).font = Font(name="Roboto", size=10, bold=True)
        self.ws.cell(row=stats_row, column=2).fill = stats_fill
        self.ws.cell(row=stats_row, column=2).alignment = ALIGNMENTS["left"]
        self.ws.cell(row=stats_row, column=2).border = BORDERS["thin"]
        
        self.ws.cell(row=stats_row, column=3, value=f"{len(normal_debtors)}")
        self.ws.cell(row=stats_row, column=3).font = Font(name="Roboto", size=10)
        self.ws.cell(row=stats_row, column=3).fill = stats_fill
        self.ws.cell(row=stats_row, column=3).alignment = ALIGNMENTS["left"]
        self.ws.cell(row=stats_row, column=3).border = BORDERS["thin"]
        
        # Общая сумма долга
        self.ws.cell(row=stats_row, column=5, value="Общая сумма долга:")
        self.ws.cell(row=stats_row, column=5).font = Font(name="Roboto", size=10, bold=True)
        self.ws.cell(row=stats_row, column=5).fill = stats_fill
        self.ws.cell(row=stats_row, column=5).alignment = ALIGNMENTS["right"]
        self.ws.cell(row=stats_row, column=5).border = BORDERS["thin"]
        
        total_debt_cell = self.ws.cell(row=stats_row, column=6, value=total_normal_debt)
        total_debt_cell.font = Font(name="Roboto", size=10, bold=True, color="CD5C5C")
        total_debt_cell.fill = stats_fill
        total_debt_cell.alignment = ALIGNMENTS["right"]
        total_debt_cell.border = BORDERS["thin"]
        total_debt_cell.number_format = '#,##0.00'
        
        row += 2

        # ============================================================
        # ОСНОВНАЯ ТАБЛИЦА (только с положительным долгом)
        # ============================================================
        headers = ["КЛИЕНТ", "ЗАКАЗ", "ДАТА", "СТАТУС", "СУММА", "ОПЛАЧЕНО", "ДОЛГ"]
        for col_idx, header in enumerate(headers, 2):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = FONTS["header_white"]
            cell.alignment = ALIGNMENTS["center"]
            cell.fill = FILLS["header"]
            cell.border = BORDERS["thin"]
        row += 1

        # Показываем только заказы с положительным долгом
        display_debtors = normal_debtors
        
        for idx, debtor in enumerate(display_debtors):
            fill = FILLS["alt"] if idx % 2 == 1 else FILLS["none"]

            client = debtor.get("client", "НЕ УКАЗАН")
            order_number = debtor.get("order_number", "")
            order_date = debtor.get("order_date", "")
            status = debtor.get("status", "")
            amount = debtor.get("order_amount", 0)
            paid = debtor.get("paid_amount", 0)
            debt = debtor.get("debt", 0)

            # Клиент - заглавными буквами
            c1 = self.ws.cell(row=row, column=2, value=str(client).upper()[:40])
            c1.font = FONTS["normal"]
            c1.alignment = ALIGNMENTS["left"]
            c1.border = BORDERS["thin"]
            c1.fill = fill

            # Заказ - с гиперссылкой
            c2 = self.ws.cell(row=row, column=3, value=order_number)
            c2.font = Font(name="Roboto", size=9, color=COLORS["blue"], underline="single")
            c2.alignment = ALIGNMENTS["left"]
            c2.border = BORDERS["thin"]
            c2.fill = fill

            # Дата
            date_str = order_date.strftime("%d.%m.%Y") if hasattr(order_date, "strftime") else str(order_date)
            c3 = self.ws.cell(row=row, column=4, value=date_str)
            c3.font = FONTS["normal"]
            c3.alignment = ALIGNMENTS["center"]
            c3.border = BORDERS["thin"]
            c3.fill = fill

            # Статус
            c4 = self.ws.cell(row=row, column=5, value=status)
            c4.font = FONTS["normal"]
            c4.alignment = ALIGNMENTS["left"]
            c4.border = BORDERS["thin"]
            c4.fill = fill

            # Сумма
            c5 = self.ws.cell(row=row, column=6, value=amount)
            c5.font = FONTS["normal"]
            c5.alignment = ALIGNMENTS["right"]
            c5.border = BORDERS["thin"]
            c5.fill = fill
            c5.number_format = '#,##0.00'

            # Оплачено
            c6 = self.ws.cell(row=row, column=7, value=paid)
            c6.font = FONTS["normal"]
            c6.alignment = ALIGNMENTS["right"]
            c6.border = BORDERS["thin"]
            c6.fill = fill
            c6.number_format = '#,##0.00'

            # Долг
            c7 = self.ws.cell(row=row, column=8, value=debt)
            c7.font = Font(name="Roboto", size=9, bold=True, color="CD5C5C")
            c7.alignment = ALIGNMENTS["right"]
            c7.border = BORDERS["thin"]
            c7.fill = fill
            c7.number_format = '#,##0.00'

            row += 1

        # ИТОГО
        total_label = self.ws.cell(row=row, column=7, value="ИТОГО:")
        total_label.font = Font(name="Roboto", size=10, bold=True)
        total_label.alignment = ALIGNMENTS["right"]
        total_label.border = BORDERS["thin"]
        total_label.fill = FILLS["total"]

        total_cell = self.ws.cell(row=row, column=8, value=total_normal_debt)
        total_cell.font = Font(name="Roboto", size=10, bold=True, color="CD5C5C")
        total_cell.alignment = ALIGNMENTS["right"]
        total_cell.border = BORDERS["thin"]
        total_cell.fill = FILLS["total"]
        total_cell.number_format = '#,##0.00'

        # Настройка ширины колонок
        self.ws.column_dimensions["A"].width = 3
        self.ws.column_dimensions["B"].width = 32
        self.ws.column_dimensions["C"].width = 20
        self.ws.column_dimensions["D"].width = 14
        self.ws.column_dimensions["E"].width = 20
        self.ws.column_dimensions["F"].width = 18
        self.ws.column_dimensions["G"].width = 18
        self.ws.column_dimensions["H"].width = 18

        self.ws.freeze_panes = "A7"
        self.ws.sheet_view.showGridLines = False


class ClosedDebtorsSheet(BaseSheet):
    """Закрытая дебиторка - отгружено, но не оплачено"""

    def __init__(self, workbook, sheet_number):
        super().__init__(workbook, sheet_number)

    def build(self, debtors):
        today = date.today().strftime("%d.%m.%Y")
        row = 1
        
        # Анализируем переплаты (отрицательный долг) и нормальные долги
        overpayments = [d for d in debtors if d.get('debt', 0) < 0]
        total_overpayment = sum(abs(d.get('debt', 0)) for d in overpayments)
        normal_debtors = [d for d in debtors if d.get('debt', 0) > 0]
        total_normal_debt = sum(d.get('debt', 0) for d in normal_debtors)
        zero_debtors = [d for d in debtors if d.get('debt', 0) == 0]

        # Ссылка в оглавление
        toc_link = self.ws.cell(row=row, column=2, value="← ВЕРНУТЬСЯ В ОГЛАВЛЕНИЕ")
        toc_link.font = Font(name="Roboto", size=9, bold=True, color=COLORS["blue"], underline="single")
        toc_link.alignment = ALIGNMENTS["left"]
        toc_link.hyperlink = "#'TOC'!A1"

        self.ws.cell(row=row + 1, column=2, value="⚠️ ЗАКРЫТАЯ ДЕБИТОРСКАЯ ЗАДОЛЖЕННОСТЬ")
        self.ws.cell(row=row + 1, column=2).font = Font(name="Roboto", size=16, bold=True, color="FF6B6B")

        self.ws.cell(row=row + 2, column=2, value=f"Отгружено, но не оплачено на дату отчёта: {today}")
        self.ws.cell(row=row + 2, column=2).font = Font(name="Roboto", size=9, italic=True, color=COLORS["text_gray"])

        self.ws.cell(row=row + 3, column=2, value="‼️ Требуют срочного внимания - товар отгружен, деньги не получены")
        self.ws.cell(row=row + 3, column=2).font = Font(name="Roboto", size=9, bold=True, color="FF6B6B")

        row = 7
        
        # ============================================================
        # БЛОК САММАРИ: ПЕРЕПЛАТЫ И ОБЩАЯ СТАТИСТИКА
        # ============================================================
        if overpayments:
            # Желтый фон для предупреждения
            warning_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            
            # Заголовок предупреждения
            warn_cell = self.ws.cell(row=row, column=2, value="⚠ ВНИМАНИЕ: ОБНАРУЖЕНЫ ПЕРЕПЛАТЫ (ОТРИЦАТЕЛЬНЫЙ ДОЛГ)")
            warn_cell.font = Font(name="Roboto", size=11, bold=True, color="856404")
            warn_cell.fill = warning_fill
            warn_cell.alignment = ALIGNMENTS["left"]
            self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            row += 1
            
            # Информация о переплатах
            info_cell = self.ws.cell(row=row, column=2, value=f"• Количество заказов с переплатой: {len(overpayments)}")
            info_cell.font = Font(name="Roboto", size=10)
            info_cell.fill = warning_fill
            info_cell.alignment = ALIGNMENTS["left"]
            self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            row += 1
            
            overpay_cell = self.ws.cell(row=row, column=2, value=f"• Общая сумма переплаты: {total_overpayment:,.0f} ₽".replace(",", " "))
            overpay_cell.font = Font(name="Roboto", size=10, bold=True, color="DC3545")
            overpay_cell.fill = warning_fill
            overpay_cell.alignment = ALIGNMENTS["left"]
            self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            row += 1
            
            row += 1  # Дополнительный отступ
        
        # Общая статистика по долгам
        stats_row = row
        stats_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        
        # Количество должников
        self.ws.cell(row=stats_row, column=2, value="Всего должников:")
        self.ws.cell(row=stats_row, column=2).font = Font(name="Roboto", size=10, bold=True)
        self.ws.cell(row=stats_row, column=2).fill = stats_fill
        self.ws.cell(row=stats_row, column=2).alignment = ALIGNMENTS["left"]
        self.ws.cell(row=stats_row, column=2).border = BORDERS["thin"]
        
        self.ws.cell(row=stats_row, column=3, value=f"{len(normal_debtors)}")
        self.ws.cell(row=stats_row, column=3).font = Font(name="Roboto", size=10)
        self.ws.cell(row=stats_row, column=3).fill = stats_fill
        self.ws.cell(row=stats_row, column=3).alignment = ALIGNMENTS["left"]
        self.ws.cell(row=stats_row, column=3).border = BORDERS["thin"]
        
        # Общая сумма долга
        self.ws.cell(row=stats_row, column=5, value="Общая сумма долга:")
        self.ws.cell(row=stats_row, column=5).font = Font(name="Roboto", size=10, bold=True)
        self.ws.cell(row=stats_row, column=5).fill = stats_fill
        self.ws.cell(row=stats_row, column=5).alignment = ALIGNMENTS["right"]
        self.ws.cell(row=stats_row, column=5).border = BORDERS["thin"]
        
        total_debt_cell = self.ws.cell(row=stats_row, column=6, value=total_normal_debt)
        total_debt_cell.font = Font(name="Roboto", size=10, bold=True, color="FF6B6B")
        total_debt_cell.fill = stats_fill
        total_debt_cell.alignment = ALIGNMENTS["right"]
        total_debt_cell.border = BORDERS["thin"]
        total_debt_cell.number_format = '#,##0.00'
        
        row += 2

        # ============================================================
        # ОСНОВНАЯ ТАБЛИЦА (только с положительным долгом)
        # ============================================================
        headers = ["КЛИЕНТ", "ЗАКАЗ", "ДАТА", "СТАТУС", "СУММА", "ОПЛАЧЕНО", "ДОЛГ"]
        for col_idx, header in enumerate(headers, 2):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = FONTS["header_white"]
            cell.alignment = ALIGNMENTS["center"]
            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell.border = BORDERS["thin"]
        row += 1

        # Показываем только заказы с положительным долгом
        display_debtors = normal_debtors
        
        for idx, debtor in enumerate(display_debtors):
            fill = FILLS["alt"] if idx % 2 == 1 else FILLS["none"]

            client = debtor.get("client", "НЕ УКАЗАН")
            order_number = debtor.get("order_number", "")
            order_date = debtor.get("order_date", "")
            status = debtor.get("status", "")
            amount = debtor.get("order_amount", 0)
            paid = debtor.get("paid_amount", 0)
            debt = debtor.get("debt", 0)

            # Клиент - заглавными буквами
            c1 = self.ws.cell(row=row, column=2, value=str(client).upper()[:40])
            c1.font = FONTS["normal"]
            c1.alignment = ALIGNMENTS["left"]
            c1.border = BORDERS["thin"]
            c1.fill = fill

            # Заказ - с гиперссылкой
            c2 = self.ws.cell(row=row, column=3, value=order_number)
            c2.font = Font(name="Roboto", size=9, color=COLORS["blue"], underline="single")
            c2.alignment = ALIGNMENTS["left"]
            c2.border = BORDERS["thin"]
            c2.fill = fill

            # Дата
            date_str = order_date.strftime("%d.%m.%Y") if hasattr(order_date, "strftime") else str(order_date)
            c3 = self.ws.cell(row=row, column=4, value=date_str)
            c3.font = FONTS["normal"]
            c3.alignment = ALIGNMENTS["center"]
            c3.border = BORDERS["thin"]
            c3.fill = fill

            # Статус
            c4 = self.ws.cell(row=row, column=5, value=status)
            c4.font = FONTS["normal"]
            c4.alignment = ALIGNMENTS["left"]
            c4.border = BORDERS["thin"]
            c4.fill = fill

            # Сумма
            c5 = self.ws.cell(row=row, column=6, value=amount)
            c5.font = FONTS["normal"]
            c5.alignment = ALIGNMENTS["right"]
            c5.border = BORDERS["thin"]
            c5.fill = fill
            c5.number_format = '#,##0.00'

            # Оплачено
            c6 = self.ws.cell(row=row, column=7, value=paid)
            c6.font = FONTS["normal"]
            c6.alignment = ALIGNMENTS["right"]
            c6.border = BORDERS["thin"]
            c6.fill = fill
            c6.number_format = '#,##0.00'

            # Долг
            c7 = self.ws.cell(row=row, column=8, value=debt)
            c7.font = Font(name="Roboto", size=9, bold=True, color="FF6B6B")
            c7.alignment = ALIGNMENTS["right"]
            c7.border = BORDERS["thin"]
            c7.fill = fill
            c7.number_format = '#,##0.00'

            row += 1

        # ИТОГО
        total_label = self.ws.cell(row=row, column=7, value="ИТОГО:")
        total_label.font = Font(name="Roboto", size=10, bold=True)
        total_label.alignment = ALIGNMENTS["right"]
        total_label.border = BORDERS["thin"]
        total_label.fill = FILLS["total"]

        total_cell = self.ws.cell(row=row, column=8, value=total_normal_debt)
        total_cell.font = Font(name="Roboto", size=10, bold=True, color="FF6B6B")
        total_cell.alignment = ALIGNMENTS["right"]
        total_cell.border = BORDERS["thin"]
        total_cell.fill = FILLS["total"]
        total_cell.number_format = '#,##0.00'

        # Настройка ширины колонок
        self.ws.column_dimensions["A"].width = 3
        self.ws.column_dimensions["B"].width = 32
        self.ws.column_dimensions["C"].width = 20
        self.ws.column_dimensions["D"].width = 14
        self.ws.column_dimensions["E"].width = 20
        self.ws.column_dimensions["F"].width = 18
        self.ws.column_dimensions["G"].width = 18
        self.ws.column_dimensions["H"].width = 18

        self.ws.freeze_panes = "A8"
        self.ws.sheet_view.showGridLines = False