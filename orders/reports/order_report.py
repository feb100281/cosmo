# orders/reports/order_report.py
from django.http import HttpResponse
from openpyxl import Workbook
from datetime import datetime
import io
from django.db.models import Sum, Count, Q, Avg
from django.db.models.functions import TruncMonth

from .queries.dashboard_queries import DashboardQueries
from .queries.executive_queries import ExecutiveQueries
from .queries.orders_detail_queries import OrdersDetailQueries
from .queries.delivery_analysis_query import DeliveryAnalysisQueries  
from .queries.manager_queries import ManagerQueries
from .queries.client_analysis_query import ClientAnalysisQueries
from .queries.payments_analysis_query import PaymentsAnalysisQueries 
from .queries.negative_payments_query import NegativePaymentsQueries
from .queries.unpaid_shipments_query import UnpaidShipmentsQueries
from .queries.completed_orders_wrong_status_query import CompletedOrdersWrongStatusQueries
from .queries.cancelled_analysis_query import CancelledAnalysisQueries
from .queries.funnel_analysis_query import FunnelAnalysisQueries

from .sheets.toc_sheet import TOCSheet
from .sheets.executive_sheet import ExecutiveSheet
from .sheets.orders_detail_sheet import OrdersDetailSheet
from .sheets.manager_analysis_sheet import ManagerAnalysisSheet
from .sheets.client_analysis_sheet import ClientAnalysisSheet
from .sheets.payments_analysis_sheet import PaymentsAnalysisSheet
from .sheets.daily_payments_sheet import DailyPaymentsSheet
from .sheets.delivery_analysis_sheet import DeliveryAnalysisSheet
from .sheets.negative_payments_sheet import NegativePaymentsSheet
from .sheets.unpaid_shipments_sheet import UnpaidShipmentsSheet
from .sheets.completed_orders_wrong_status_sheet import CompletedOrdersWrongStatusSheet
from .sheets.cancelled_analysis_sheet import CancelledAnalysisSheet
from .sheets.funnel_analysis_sheet import FunnelAnalysisSheet



def generate_orders_report(request):
    """Главная функция генерации отчета"""
    
    wb = Workbook()
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    sheets_info = []
    sheet_counter = 1
    
    # ============================================================
    # 1. Executive Summary
    # ============================================================
    exec_queries = ExecutiveQueries(request)
    exec_sheet = ExecutiveSheet(wb, str(sheet_counter), request=request)
    exec_sheet.build()
    sheets_info.append({
        "number": sheet_counter,
        "name": "Executive Summary",
        "description": "Ключевые показатели, тренды, топ-менеджеры"
    })
    sheet_counter += 1
    
    # ============================================================
    # 2. Детализация АКТИВНЫХ заказов
    # ============================================================
    orders_detail_queries = OrdersDetailQueries()
    active_orders_data = orders_detail_queries.get_active_orders_detail()
    orders_sheet = OrdersDetailSheet(wb, str(sheet_counter))
    orders_sheet.build(active_orders_data)
    sheets_info.append({
        "number": sheet_counter,
        "name": "Детализация заказов",
        "description": f"Активные заказы ({len(active_orders_data)} шт.)"
    })
    sheet_counter += 1
    
    # ============================================================
    # 3. АНАЛИЗ ДОСТАВКИ
    # ============================================================
    delivery_analysis_queries = DeliveryAnalysisQueries(request)
    delivery_analysis_sheet = DeliveryAnalysisSheet(wb, str(sheet_counter), request=request)
    delivery_analysis_sheet.build()
    sheets_info.append({
        "number": sheet_counter,
        "name": "Анализ доставки",
        "description": "Аналитика по услугам доставки"
    })
    sheet_counter += 1
    
    # ============================================================
    # 4. DashboardQueries
    # ============================================================
    queries = DashboardQueries(request)
    
    # ============================================================
    # 5. Анализ по менеджерам
    # ============================================================
    manager_queries = ManagerQueries(request)
    all_managers_data = manager_queries.get_all_managers_summary()
    
    enhanced_managers_data = []
    for manager in all_managers_data:
        manager_name = manager.get('manager')
        if manager_name:
            details = manager_queries.get_manager_details(manager_name)
            enhanced_manager = {
                **manager,
                'remaining_to_ship': details.get('remaining_to_ship', 0),
                'top_clients': details.get('top_clients', []),
                'oldest_unpaid_invoice': details.get('oldest_unpaid_invoice'),
            }
            enhanced_managers_data.append(enhanced_manager)
        else:
            enhanced_managers_data.append(manager)
    
    manager_sheet = ManagerAnalysisSheet(wb, str(sheet_counter))
    manager_sheet.build(enhanced_managers_data)
    sheets_info.append({
        "number": sheet_counter,
        "name": "Анализ по менеджерам",
        "description": f"Портфолио менеджеров ({len(enhanced_managers_data)} чел.)"
    })
    sheet_counter += 1
    
    # ============================================================
    # 6. Анализ клиентской базы 
    # ============================================================
    client_analysis_queries = ClientAnalysisQueries(request)
    client_analysis_sheet = ClientAnalysisSheet(wb, str(sheet_counter), request=request)
    client_analysis_sheet.build()
    sheets_info.append({
        "number": sheet_counter,
        "name": "Анализ клиентов",
        "description": "ABC-анализ, топ-клиенты, риски"
    })
    sheet_counter += 1
    
    # ============================================================
    # 7. АНАЛИЗ ОПЛАТ
    # ============================================================
    payments_analysis_queries = PaymentsAnalysisQueries(request)
    payments_analysis_sheet = PaymentsAnalysisSheet(wb, str(sheet_counter), request=request)
    payments_analysis_sheet.build()
    sheets_info.append({
        "number": sheet_counter,
        "name": "Анализ оплат",
        "description": "Полная аналитика по платежам"
    })
    sheet_counter += 1
    
    # ============================================================
    # 8. ПОДНЕВНАЯ ДИНАМИКА ОПЛАТ
    # ============================================================
    daily_payments_data = payments_analysis_queries.get_daily_payments_by_store()
    daily_payments_sheet = DailyPaymentsSheet(wb, str(sheet_counter), request=request)
    daily_payments_sheet.build(daily_payments_data)
    sheets_info.append({
        "number": sheet_counter,
        "name": "Подневная динамика оплат",
        "description": f"Детальная разбивка оплат по дням"
    })
    sheet_counter += 1
    
    # ============================================================
    # 9. ЗАКАЗЫ С ОТРИЦАТЕЛЬНЫМИ ОПЛАТАМИ
    # ============================================================
    negative_payments_queries = NegativePaymentsQueries()
    negative_orders = negative_payments_queries.get_orders_with_negative_payments_only(start_date='2025-04-01')
    negative_summary = negative_payments_queries.get_negative_payments_summary(negative_orders)
    negative_payments_sheet = NegativePaymentsSheet(wb, str(sheet_counter))
    negative_payments_sheet.build(negative_orders, negative_summary)
    sheets_info.append({
        "number": sheet_counter,
        "name": "Проблемные возвраты",
        "description": f"Заказы с возвратами без оплат"
    })
    sheet_counter += 1
    
    # ============================================================
    # 10. ОТГРУЗКИ БЕЗ ОПЛАТЫ
    # ============================================================
    unpaid_shipments_queries = UnpaidShipmentsQueries()
    unpaid_orders = unpaid_shipments_queries.get_orders_with_unpaid_shipments()
    unpaid_summary = unpaid_shipments_queries.get_unpaid_shipments_summary(unpaid_orders)
    unpaid_sheet = UnpaidShipmentsSheet(wb, str(sheet_counter))
    unpaid_sheet.build(unpaid_orders, unpaid_summary)
    sheets_info.append({
        "number": sheet_counter,
        "name": "Отгрузки без оплаты",
        "description": f"Заказы с отгрузкой, но оплата меньше суммы"
    })
    sheet_counter += 1
    
    # ============================================================
    # 11. ЗАКАЗЫ К ЗАКРЫТИЮ
    # ============================================================
    wrong_status_queries = CompletedOrdersWrongStatusQueries()
    wrong_status_orders = wrong_status_queries.get_orders_completed_wrong_status()
    wrong_status_summary = wrong_status_queries.get_summary(wrong_status_orders)
    wrong_status_sheet = CompletedOrdersWrongStatusSheet(wb, str(sheet_counter))
    wrong_status_sheet.build(wrong_status_orders, wrong_status_summary)
    sheets_info.append({
        "number": sheet_counter,
        "name": "Заказы к закрытию",
        "description": f"Полностью выполненные заказы со статусом не 'Закрыт'"
    })
    sheet_counter += 1
    
    # ============================================================
    # 12. АНАЛИЗ ОТМЕНЁННЫХ ЗАКАЗОВ
    # # ============================================================
    cancelled_queries = CancelledAnalysisQueries(request)
    cancelled_orders = cancelled_queries.get_cancelled_orders_detail(start_date='2025-01-01')
    cancelled_summary = cancelled_queries.get_cancelled_summary(cancelled_orders)
    cancelled_by_reason = cancelled_queries.get_cancelled_by_reason(cancelled_orders)
    cancelled_by_manager = cancelled_queries.get_cancelled_by_manager(cancelled_orders)
    top_cancelled_orders = cancelled_queries.get_top_cancelled_orders(cancelled_orders, limit=50)
    cancelled_sheet = CancelledAnalysisSheet(wb, str(sheet_counter))
    cancelled_sheet.build(cancelled_orders, cancelled_summary, cancelled_by_reason, cancelled_by_manager, top_cancelled_orders)
    sheets_info.append({
        "number": sheet_counter,
        "name": "Анализ отмен",
        "description": f"Детальный анализ отмен"
    })
    sheet_counter += 1
    
    # ============================================================
    # 13. ВОРОНКА ПРОДАЖ
    # ============================================================
    funnel_queries = FunnelAnalysisQueries()

    funnel_overall = funnel_queries.get_funnel_overall()
    funnel_by_manager = funnel_queries.get_funnel_by_manager(limit=20)
    loss_analysis = funnel_queries.get_loss_analysis()
    orders_detail = funnel_queries.get_orders_detail_table("YTD")

    funnel_sheet = FunnelAnalysisSheet(wb, str(sheet_counter))
    funnel_sheet.build(
        funnel_overall=funnel_overall,
        funnel_by_manager=funnel_by_manager,
        loss_analysis=loss_analysis,
        orders_detail=orders_detail,
    )

    sheets_info.append({
        "number": sheet_counter,
        "name": "Воронка продаж",
        "description": f"Конверсия заказ→оплата→отгрузка→закрытие; детализация заказов: {len(orders_detail)} шт."
    })
    sheet_counter += 1
    

    
    # ============================================================
    # ОГЛАВЛЕНИЕ
    # ============================================================
    toc_sheet = TOCSheet(wb)
    toc_sheet.build(sheets_info, request=request)
    
    toc_index = wb.sheetnames.index("TOC")
    wb.move_sheet(wb["TOC"], offset=-toc_index)
    
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f'Orders_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    
    with io.BytesIO() as output:
        wb.save(output)
        response.write(output.getvalue())
    
    return response