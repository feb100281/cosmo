# orders/reports/components/kpi_cards.py
from openpyxl.styles import Font, PatternFill
from ..styles.theme import COLORS, BORDERS, ALIGNMENTS


class KPICards:
    """Класс для отрисовки KPI карточек"""

    def __init__(self, worksheet):
        self.ws = worksheet

    def draw_compact_card(self, row, col, title, value, subtitle=None, color=None, width=2):
        """
        Рисует одну компактную KPI карточку

        row, col - позиция карточки
        width - ширина карточки в колонках
        """
        bg_color = COLORS.get("light_gray", "F8F9FA")

        # Основные 2 строки карточки
        for r in range(row, row + 2):
            for c in range(col, col + width):
                cell = self.ws.cell(row=r, column=c)
                cell.fill = PatternFill("solid", fgColor=bg_color)
                cell.border = BORDERS["thin"]

        # Значение
        if width > 1:
            self.ws.merge_cells(
                start_row=row,
                start_column=col,
                end_row=row,
                end_column=col + width - 1
            )
        value_cell = self.ws.cell(row=row, column=col, value=value)
        value_cell.font = Font(
            name="Roboto",
            size=14,
            bold=True,
            color=color or COLORS["dark_green"]
        )
        value_cell.alignment = ALIGNMENTS["center"]

        # Заголовок
        if width > 1:
            self.ws.merge_cells(
                start_row=row + 1,
                start_column=col,
                end_row=row + 1,
                end_column=col + width - 1
            )
        title_cell = self.ws.cell(row=row + 1, column=col, value=title)
        title_cell.font = Font(
            name="Roboto",
            size=8,
            color=COLORS["text_gray"]
        )
        title_cell.alignment = ALIGNMENTS["center"]

        # Подзаголовок
        if subtitle:
            for c in range(col, col + width):
                cell = self.ws.cell(row=row + 2, column=c)
                cell.fill = PatternFill("solid", fgColor=bg_color)
                cell.border = BORDERS["thin"]

            if width > 1:
                self.ws.merge_cells(
                    start_row=row + 2,
                    start_column=col,
                    end_row=row + 2,
                    end_column=col + width - 1
                )

            sub_cell = self.ws.cell(row=row + 2, column=col, value=subtitle)
            sub_cell.font = Font(
                name="Roboto",
                size=8,
                bold=True,
                color=color or COLORS["dark_green"]
            )
            sub_cell.alignment = ALIGNMENTS["center"]
            return 3

        return 2

    def draw_row(self, start_row, cards):
        """Рисует строку карточек"""
        current_col = 2
        max_height = 2

        for card in cards:
            width = card.get("width", 2)

            height = self.draw_compact_card(
                row=start_row,
                col=current_col,
                title=card.get("title", ""),
                value=card.get("value", ""),
                subtitle=card.get("subtitle"),
                color=card.get("color"),
                width=width
            )
            max_height = max(max_height, height)
            current_col += width

        return start_row + max_height

    def draw_two_rows(self, start_row, row1_cards, row2_cards):
        """Рисует два ряда карточек (финансовые + операционные)"""
        row = self.draw_row(start_row, row1_cards)
        row = self.draw_row(row, row2_cards)
        return row + 1


def create_kpi_cards(worksheet):
    """Создает экземпляр KPICards"""
    return KPICards(worksheet)