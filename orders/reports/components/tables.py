# orders/reports/components/tables.py
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ..styles.theme import COLORS, ALIGNMENTS, BORDERS, FILLS


class TableComponent:
    """Класс для отрисовки таблиц"""
    
    def __init__(self, worksheet):
        self.ws = worksheet
    
    def draw(self, start_row, headers, data_rows, start_col=2, number_format='#,##0', highlight_cols=None):
        row = start_row
        
        # Заголовки
        for col_idx, header in enumerate(headers):
            cell = self.ws.cell(row=row, column=start_col + col_idx, value=header)
            cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["white"])
            cell.fill = FILLS["header"]
            cell.alignment = ALIGNMENTS["center"]
            cell.border = BORDERS["thin"]
        row += 1
        
        # Данные
        for idx, data_row in enumerate(data_rows):
            fill = FILLS["alt"] if idx % 2 == 0 else FILLS["none"]
            is_last_row = (idx == len(data_rows) - 1)
            
            for col_idx, value in enumerate(data_row):
                cell = self.ws.cell(row=row, column=start_col + col_idx, value=value)
                cell.fill = fill
                
                # Для всех строк используем одинаковые границы
                thin_side = Side(border_style="thin", color=COLORS["border_gray"])
                cell.border = Border(
                    top=thin_side,
                    left=thin_side,
                    right=thin_side,
                    bottom=thin_side  # Добавляем нижнюю границу для всех строк
                )
                
                if col_idx == 0:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                
                cell.font = Font(name="Roboto", size=10, color=COLORS["text_gray"])
                
                if col_idx >= 1 and isinstance(value, (int, float)):
                    cell.number_format = number_format
                
                if highlight_cols and col_idx in highlight_cols:
                    if isinstance(value, (int, float)) and value > 0:
                        cell.font = Font(
                            name="Roboto",
                            size=10,
                            bold=True,
                            color=COLORS.get("warning_red", "C62828")
                        )
            row += 1
        
        return row


def create_table(worksheet):
    """Создает экземпляр TableComponent"""
    return TableComponent(worksheet)