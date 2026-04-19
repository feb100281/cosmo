# # orders/reports/components/manager_profile.py
# from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
# from ..styles.theme import COLORS, FILLS, FONTS, BORDERS, ALIGNMENTS


# class ManagerProfile:
#     """Компонент для отображения профиля менеджера (улучшенный деловой стиль)"""
    
#     def __init__(self, worksheet):
#         self.ws = worksheet
#         self.colors = COLORS
#         self.fills = FILLS
#         self.fonts = FONTS
#         self.borders = BORDERS
#         self.alignments = ALIGNMENTS
    
#     def _format_currency(self, value):
#         if value is None or value == 0:
#             return "0 ₽"
#         return f"{int(round(value)):,} ₽".replace(",", " ")
    
#     def _format_number(self, value):
#         if value is None or value == 0:
#             return "0"
#         return f"{int(round(value)):,}".replace(",", " ")
    
#     def _format_percent(self, value):
#         if value is None:
#             return "0%"
#         return f"{value:.1f}%"
    
#     def _create_mini_bar(self, percent, width=10):
#         """Создает текстовую мини-гистограмму"""
#         filled = int(round(percent / 100 * width))
#         empty = width - filled
#         return "▰" * filled + "▱" * empty
    
#     def _draw_metric_row(self, row, col, title, main_value, 
#                          secondary_value=None, is_currency=True,
#                          highlight_negative=False, debt_warning=False,
#                          mini_bar=False):
#         """
#         Рисует строку метрики с улучшенным визуалом
#         """
#         # Заголовок
#         title_cell = self.ws.cell(row=row, column=col, value=title)
#         title_cell.font = Font(name="Roboto", size=8, color=self.colors["text_gray"])
#         title_cell.alignment = self.alignments["right"]  # Выравнивание вправо
#         title_cell.border = Border(right=Side(style='thin', color=self.colors["border_gray"]))
        
#         # Основное значение
#         if is_currency:
#             formatted_value = self._format_currency(main_value)
#         else:
#             formatted_value = self._format_number(main_value)
        
#         value_cell = self.ws.cell(row=row + 1, column=col, value=formatted_value)
        
#         # Фон для предупреждения о долге
#         if debt_warning and main_value > 0:
#             value_cell.fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
        
#         # Определяем цвет основного значения
#         if highlight_negative and main_value < 0:
#             value_cell.font = Font(name="Roboto", size=12, bold=True, color=self.colors["warning_red"])
#         elif is_currency and main_value > 0:
#             value_cell.font = Font(name="Roboto", size=12, bold=True, color=self.colors["dark_green"])
#         else:
#             value_cell.font = Font(name="Roboto", size=12, bold=True, color=self.colors["black"])
        
#         value_cell.alignment = self.alignments["right"]
#         value_cell.border = Border(right=Side(style='thin', color=self.colors["border_gray"]))
        
#         # Дополнительное значение (процент)
#         if secondary_value is not None:
#             percent_cell = self.ws.cell(row=row + 2, column=col, value=f"({self._format_percent(secondary_value)})")
#             percent_cell.font = Font(name="Roboto", size=8, color=self.colors["text_gray"])
#             percent_cell.alignment = self.alignments["right"]
#             percent_cell.border = Border(right=Side(style='thin', color=self.colors["border_gray"]))
            
#             # Мини-гистограмма
#             if mini_bar:
#                 bar_cell = self.ws.cell(row=row + 2, column=col + 1, value=self._create_mini_bar(secondary_value))
#                 bar_cell.font = Font(name="Roboto", size=8, color=self.colors["dark_green"])
#                 bar_cell.alignment = self.alignments["left"]
            
#             return 3
#         else:
#             return 2
    
#     def draw_two_periods(self, row, manager_name, mtd_data, ytd_data):
#         """
#         Рисует профиль менеджера с двумя периодами (улучшенная версия)
#         """
#         # ============================================================
#         # ЗАГОЛОВОК ПРОФИЛЯ
#         # ============================================================
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
#         header_cell = self.ws.cell(row=row, column=2, value=f"👤 {manager_name}")
#         header_cell.font = Font(name="Roboto", size=13, bold=True, color=self.colors["dark_green"])
#         header_cell.alignment = self.alignments["left"]
#         header_cell.fill = self.fills["section"]
#         self.ws.row_dimensions[row].height = 30
#         row += 1
        
#         # Разделитель
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
#         sep_cell = self.ws.cell(row=row, column=2, value="")
#         sep_cell.border = self.borders["bottom_medium"]
#         self.ws.row_dimensions[row].height = 3
#         row += 1
        
#         # Отступ
#         self.ws.row_dimensions[row].height = 6
#         row += 1
        
#         # ============================================================
#         # ЗАГОЛОВКИ ПЕРИОДОВ
#         # ============================================================
#         # MTD (с легким фоном)
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
#         mtd_title = self.ws.cell(row=row, column=2, value="МЕСЯЦ (MTD)")
#         mtd_title.font = Font(name="Roboto", size=10, bold=True, color=self.colors["blue"])
#         mtd_title.alignment = self.alignments["center"]
#         mtd_title.fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
        
#         # YTD
#         self.ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
#         ytd_title = self.ws.cell(row=row, column=5, value="ГОД (YTD)")
#         ytd_title.font = Font(name="Roboto", size=10, bold=True, color=self.colors["dark_green"])
#         ytd_title.alignment = self.alignments["center"]
#         ytd_title.fill = self.fills["heat_light"]
        
#         self.ws.row_dimensions[row].height = 26
#         row += 1
        
#         # Вертикальный разделитель между периодами
#         for r in range(row, row + 18):
#             divider = self.ws.cell(row=r, column=4, value="")
#             divider.border = Border(left=Side(style='thin', color=self.colors["border_gray"]))
        
#         # Отступ
#         self.ws.row_dimensions[row].height = 6
#         row += 1
        
#         # ============================================================
#         # ГРУППА 1: ОСНОВНЫЕ ПОКАЗАТЕЛИ
#         # ============================================================
#         # Подзаголовок группы
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
#         group1 = self.ws.cell(row=row, column=2, value="▸ ОСНОВНЫЕ ПОКАЗАТЕЛИ")
#         group1.font = Font(name="Roboto", size=9, bold=True, color=self.colors["text_gray"])
#         group1.alignment = self.alignments["left"]
#         row += 1
        
#         self.ws.row_dimensions[row].height = 4
#         row += 1
        
#         # Заказы
#         self._draw_metric_row(row, 2, "Заказы", 
#                              mtd_data.get('orders_count', 0), 
#                              is_currency=False)
#         self._draw_metric_row(row, 5, "Заказы", 
#                              ytd_data.get('orders_count', 0), 
#                              is_currency=False)
#         row += 3
        
#         # Сумма заказов
#         self._draw_metric_row(row, 2, "Сумма заказов", 
#                              mtd_data.get('total_amount', 0), 
#                              is_currency=True)
#         self._draw_metric_row(row, 5, "Сумма заказов", 
#                              ytd_data.get('total_amount', 0), 
#                              is_currency=True)
#         row += 3
        
#         # Средний чек
#         self._draw_metric_row(row, 2, "Средний чек", 
#                              mtd_data.get('avg_check', 0), 
#                              is_currency=True)
#         self._draw_metric_row(row, 5, "Средний чек", 
#                              ytd_data.get('avg_check', 0), 
#                              is_currency=True)
#         row += 3
        
#         # Разделитель групп
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
#         sep1 = self.ws.cell(row=row, column=2, value="")
#         sep1.border = Border(top=Side(style='thin', color=self.colors["border_gray"]))
#         self.ws.row_dimensions[row].height = 6
#         row += 1
        
#         # ============================================================
#         # ГРУППА 2: ВЫПОЛНЕНИЕ
#         # ============================================================
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
#         group2 = self.ws.cell(row=row, column=2, value="▸ ВЫПОЛНЕНИЕ")
#         group2.font = Font(name="Roboto", size=9, bold=True, color=self.colors["text_gray"])
#         group2.alignment = self.alignments["left"]
#         row += 1
        
#         self.ws.row_dimensions[row].height = 4
#         row += 1
        
#         # Оплачено (с мини-гистограммой)
#         self._draw_metric_row(row, 2, "Оплачено", 
#                              mtd_data.get('paid_amount', 0), 
#                              secondary_value=mtd_data.get('paid_percent', 0),
#                              is_currency=True, mini_bar=True)
#         self._draw_metric_row(row, 5, "Оплачено", 
#                              ytd_data.get('paid_amount', 0), 
#                              secondary_value=ytd_data.get('paid_percent', 0),
#                              is_currency=True, mini_bar=True)
#         row += 3
        
#         # Отгружено (с мини-гистограммой)
#         self._draw_metric_row(row, 2, "Отгружено", 
#                              mtd_data.get('shipped_amount', 0), 
#                              secondary_value=mtd_data.get('shipped_percent', 0),
#                              is_currency=True, mini_bar=True)
#         self._draw_metric_row(row, 5, "Отгружено", 
#                              ytd_data.get('shipped_amount', 0), 
#                              secondary_value=ytd_data.get('shipped_percent', 0),
#                              is_currency=True, mini_bar=True)
#         row += 3
        
#         # Разделитель групп
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
#         sep2 = self.ws.cell(row=row, column=2, value="")
#         sep2.border = Border(top=Side(style='thin', color=self.colors["border_gray"]))
#         self.ws.row_dimensions[row].height = 6
#         row += 1
        
#         # ============================================================
#         # ГРУППА 3: ОБЯЗАТЕЛЬСТВА
#         # ============================================================
#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
#         group3 = self.ws.cell(row=row, column=2, value="▸ ОБЯЗАТЕЛЬСТВА")
#         group3.font = Font(name="Roboto", size=9, bold=True, color=self.colors["text_gray"])
#         group3.alignment = self.alignments["left"]
#         row += 1
        
#         self.ws.row_dimensions[row].height = 4
#         row += 1
        
#         # Долг (с предупреждением)
#         mtd_debt = mtd_data.get('debt_amount', 0)
#         mtd_total = mtd_data.get('total_amount', 1)
#         mtd_debt_warning = (mtd_debt / mtd_total) > 0.1 if mtd_total > 0 else False
        
#         ytd_debt = ytd_data.get('debt_amount', 0)
#         ytd_total = ytd_data.get('total_amount', 1)
#         ytd_debt_warning = (ytd_debt / ytd_total) > 0.1 if ytd_total > 0 else False
        
#         self._draw_metric_row(row, 2, "Долг", 
#                              mtd_debt, 
#                              is_currency=True,
#                              highlight_negative=True,
#                              debt_warning=mtd_debt_warning)
#         self._draw_metric_row(row, 5, "Долг", 
#                              ytd_debt, 
#                              is_currency=True,
#                              highlight_negative=True,
#                              debt_warning=ytd_debt_warning)
#         row += 3
        
#         # Доставка
#         self._draw_metric_row(row, 2, "Доставка", 
#                              mtd_data.get('delivery_amount', 0), 
#                              secondary_value=mtd_data.get('delivery_percent', 0),
#                              is_currency=True)
#         self._draw_metric_row(row, 5, "Доставка", 
#                              ytd_data.get('delivery_amount', 0), 
#                              secondary_value=ytd_data.get('delivery_percent', 0),
#                              is_currency=True)
#         row += 3
        
        
#         row += 1
        
#         return row
    
    
    
    
    
    
    # def draw_monthly_trends_compact(self, row, manager_name, monthly_data):
    #     """
    #     Компактная версия - всё в 2 строки на менеджера
    #     """
    #     months = monthly_data.get('months', [])
    #     amounts = monthly_data.get('amounts', [])
    #     avg_checks = monthly_data.get('avg_checks', [])
        
    #     max_amount = max(amounts) if amounts else 1
        
    #     # Заголовок
    #     self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    #     title_cell = self.ws.cell(row=row, column=2, value="▸ Динамика")
    #     title_cell.font = Font(name="Roboto", size=9, bold=True, color=self.colors["text_gray"])
    #     title_cell.alignment = self.alignments["left"]
        
    #     # Для каждого месяца: цифра + мини-бар
    #     for i, amount in enumerate(amounts):
    #         col = 4 + i
    #         if col <= 8:
    #             # Формат: "1.2M ███"
    #             if amount >= 1_000_000:
    #                 amount_short = f"{amount/1_000_000:.1f}M"
    #             elif amount >= 1_000:
    #                 amount_short = f"{amount/1_000:.0f}K"
    #             else:
    #                 amount_short = str(int(amount))
                
    #             bar_length = int(round(amount / max_amount * 6))
    #             bar = "█" * bar_length
                
    #             value_cell = self.ws.cell(row=row, column=col, value=f"{amount_short} {bar}")
    #             value_cell.font = Font(name="Segoe UI Symbol", size=8)
    #             value_cell.alignment = self.alignments["center"]
    #             value_cell.comment = f"{self._format_currency(amount)}"
    #     row += 1
        
    #     # Строка с месяцами и средним чеком
    #     for i, month in enumerate(months):
    #         col = 4 + i
    #         if col <= 8:
    #             check = avg_checks[i] if i < len(avg_checks) else 0
                
    #             if check >= 1_000:
    #                 check_short = f"{check/1_000:.0f}K"
    #             else:
    #                 check_short = str(int(check))
                
    #             month_cell = self.ws.cell(row=row, column=col, value=f"{month}\n{check_short}")
    #             month_cell.font = Font(name="Roboto", size=7, color=self.colors["text_gray"])
    #             month_cell.alignment = self.alignments["center"]
    #             month_cell.comment = f"Ср.чек: {self._format_currency(check)}"
        
    #     row += 1
    #     self.ws.row_dimensions[row - 2].height = 24
    #     self.ws.row_dimensions[row - 1].height = 20
        
    #     return row
    
    
    
# orders/reports/components/manager_profile.py
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from ..styles.theme import COLORS, FILLS, FONTS, BORDERS, ALIGNMENTS


class ManagerProfile:
    """Компонент для отображения профиля менеджера (улучшенный деловой стиль)"""
    
    def __init__(self, worksheet):
        self.ws = worksheet
        self.colors = COLORS
        self.fills = FILLS
        self.fonts = FONTS
        self.borders = BORDERS
        self.alignments = ALIGNMENTS
    
    def _format_currency(self, value):
        if value is None or value == 0:
            return "0 ₽"
        return f"{int(round(value)):,} ₽".replace(",", " ")
    
    def _format_number(self, value):
        if value is None or value == 0:
            return "0"
        return f"{int(round(value)):,}".replace(",", " ")
    
    def _format_percent(self, value):
        if value is None:
            return "0%"
        return f"{value:.1f}%"
    
    def _create_mini_bar(self, percent, width=10):
        """Создает текстовую мини-гистограмму"""
        filled = int(round(percent / 100 * width))
        empty = width - filled
        return "▰" * filled + "▱" * empty
    
    def _draw_metric_row(self, row, col, title, main_value, 
                         secondary_value=None, is_currency=True,
                         highlight_negative=False, debt_warning=False,
                         mini_bar=False):
        """
        Рисует строку метрики с улучшенным визуалом
        """
        # Заголовок
        title_cell = self.ws.cell(row=row, column=col, value=title)
        title_cell.font = Font(name="Roboto", size=8, color=self.colors["text_gray"])
        title_cell.alignment = self.alignments["right"]
        title_cell.border = Border(right=Side(style='thin', color=self.colors["border_gray"]))
        
        # Основное значение
        if is_currency:
            formatted_value = self._format_currency(main_value)
        else:
            formatted_value = self._format_number(main_value)
        
        value_cell = self.ws.cell(row=row + 1, column=col, value=formatted_value)
        
        # Фон для предупреждения о долге
        if debt_warning and main_value > 0:
            value_cell.fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
        
        # Определяем цвет основного значения
        if highlight_negative and main_value < 0:
            value_cell.font = Font(name="Roboto", size=12, bold=True, color=self.colors["warning_red"])
        elif is_currency and main_value > 0:
            value_cell.font = Font(name="Roboto", size=12, bold=True, color=self.colors["dark_green"])
        else:
            value_cell.font = Font(name="Roboto", size=12, bold=True, color=self.colors["black"])
        
        value_cell.alignment = self.alignments["right"]
        value_cell.border = Border(right=Side(style='thin', color=self.colors["border_gray"]))
        
        # Дополнительное значение (процент)
        if secondary_value is not None:
            percent_cell = self.ws.cell(row=row + 2, column=col, value=f"({self._format_percent(secondary_value)})")
            percent_cell.font = Font(name="Roboto", size=8, color=self.colors["text_gray"])
            percent_cell.alignment = self.alignments["right"]
            percent_cell.border = Border(right=Side(style='thin', color=self.colors["border_gray"]))
            
            # Мини-гистограмма
            if mini_bar:
                bar_cell = self.ws.cell(row=row + 2, column=col + 1, value=self._create_mini_bar(secondary_value))
                bar_cell.font = Font(name="Roboto", size=8, color=self.colors["dark_green"])
                bar_cell.alignment = self.alignments["left"]
            
            return 3
        else:
            return 2
    
    def _draw_top_clients(self, row, col, top_clients):
        """
        Рисует блок с топ-5 клиентами менеджера
        """
        if not top_clients:
            return 0
        
        # Заголовок
        title_cell = self.ws.cell(row=row, column=col, value="▸ ТОП-5 КЛИЕНТОВ")
        title_cell.font = Font(name="Roboto", size=8, bold=True, color=self.colors["text_gray"])
        title_cell.alignment = self.alignments["left"]
        row += 1
        
        # Заголовки таблицы
        header_client = self.ws.cell(row=row, column=col, value="Клиент")
        header_client.font = Font(name="Roboto", size=7, bold=True, color=self.colors["text_gray"])
        header_client.alignment = self.alignments["left"]
        
        header_amount = self.ws.cell(row=row, column=col + 1, value="Сумма")
        header_amount.font = Font(name="Roboto", size=7, bold=True, color=self.colors["text_gray"])
        header_amount.alignment = self.alignments["right"]
        row += 1
        
        # Данные клиентов
        for client in top_clients[:5]:
            client_name = client.get('client', '')[:25]  # Ограничиваем длину
            if len(client.get('client', '')) > 25:
                client_name += "..."
            
            client_cell = self.ws.cell(row=row, column=col, value=client_name)
            client_cell.font = Font(name="Roboto", size=7)
            client_cell.alignment = self.alignments["left"]
            
            amount_cell = self.ws.cell(row=row, column=col + 1, value=self._format_currency(client.get('total_amount', 0)))
            amount_cell.font = Font(name="Roboto", size=7)
            amount_cell.alignment = self.alignments["right"]
            
            row += 1
        
        return 2 + len(top_clients[:5])  # Возвращаем количество занятых строк
    
    def _draw_oldest_invoice_warning(self, row, col, invoice_data):
        """
        Рисует предупреждение о самом старом неоплаченном счете
        """
        if not invoice_data:
            return 0
        
        # Цвет в зависимости от просрочки
        days_overdue = invoice_data.get('days_overdue', 0)
        if days_overdue > 30:
            bg_color = "FFE5E5"  # Красный для сильной просрочки
            icon = "🔴"
        elif days_overdue > 7:
            bg_color = "FFF4E5"  # Оранжевый для средней просрочки
            icon = "🟠"
        else:
            bg_color = "FFFDE5"  # Желтый для небольшой просрочки
            icon = "🟡"
        
        # Заголовок
        title_cell = self.ws.cell(row=row, column=col, value="⚠️ САМЫЙ СТАРЫЙ ЗАКАЗ")
        title_cell.font = Font(name="Roboto", size=8, bold=True, color=self.colors["warning_red"])
        title_cell.alignment = self.alignments["left"]
        row += 1
        
        # Информация о счете
        invoice_number = invoice_data.get('number', '—')
        invoice_date = invoice_data.get('date_from')
        date_str = invoice_date.strftime('%d.%m.%Y') if invoice_date else '—'
        client_name = invoice_data.get('client', '—')[:30]
        remaining_amount = invoice_data.get('remaining_amount', 0)
        
        info_text = f"{icon} Заказ №{invoice_number} от {date_str}\n"
        info_text += f"   Клиент: {client_name}\n"
        info_text += f"   Сумма долга: {self._format_currency(remaining_amount)}"
        
        if days_overdue > 0:
            info_text += f"\n   Просрочка: {days_overdue} дн."
        
        info_cell = self.ws.cell(row=row, column=col, value=info_text)
        info_cell.font = Font(name="Roboto", size=8)
        info_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        info_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        
        # Объединяем ячейки для информации (3 строки)
        self.ws.merge_cells(start_row=row, start_column=col, end_row=row + 2, end_column=col + 1)
        
        return 4  # Возвращаем количество занятых строк (заголовок + 3 строки информации)
    
    def draw_two_periods(self, row, manager_name, mtd_data, ytd_data):
        """
        Рисует профиль менеджера с двумя периодами (улучшенная версия)
        """
        # ============================================================
        # ЗАГОЛОВОК ПРОФИЛЯ
        # ============================================================
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        header_cell = self.ws.cell(row=row, column=2, value=f"👤 {manager_name}")
        header_cell.font = Font(name="Roboto", size=13, bold=True, color=self.colors["dark_green"])
        header_cell.alignment = self.alignments["left"]
        header_cell.fill = self.fills["section"]
        self.ws.row_dimensions[row].height = 30
        row += 1
        
        # Разделитель
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        sep_cell = self.ws.cell(row=row, column=2, value="")
        sep_cell.border = self.borders["bottom_medium"]
        self.ws.row_dimensions[row].height = 3
        row += 1
        
        # Отступ
        self.ws.row_dimensions[row].height = 6
        row += 1
        
        # ============================================================
        # ЗАГОЛОВКИ ПЕРИОДОВ
        # ============================================================
        # MTD (с легким фоном)
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        mtd_title = self.ws.cell(row=row, column=2, value="МЕСЯЦ (MTD)")
        mtd_title.font = Font(name="Roboto", size=10, bold=True, color=self.colors["blue"])
        mtd_title.alignment = self.alignments["center"]
        mtd_title.fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
        
        # YTD
        self.ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
        ytd_title = self.ws.cell(row=row, column=5, value="ГОД (YTD)")
        ytd_title.font = Font(name="Roboto", size=10, bold=True, color=self.colors["dark_green"])
        ytd_title.alignment = self.alignments["center"]
        ytd_title.fill = self.fills["heat_light"]
        
        self.ws.row_dimensions[row].height = 26
        row += 1
        
        # Вертикальный разделитель между периодами
        for r in range(row, row + 25):  # Увеличили диапазон для новых блоков
            divider = self.ws.cell(row=r, column=4, value="")
            divider.border = Border(left=Side(style='thin', color=self.colors["border_gray"]))
        
        # Отступ
        self.ws.row_dimensions[row].height = 6
        row += 1
        
        # ============================================================
        # ГРУППА 1: ОСНОВНЫЕ ПОКАЗАТЕЛИ
        # ============================================================
        # Подзаголовок группы
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        group1 = self.ws.cell(row=row, column=2, value="▸ ОСНОВНЫЕ ПОКАЗАТЕЛИ")
        group1.font = Font(name="Roboto", size=9, bold=True, color=self.colors["text_gray"])
        group1.alignment = self.alignments["left"]
        row += 1
        
        self.ws.row_dimensions[row].height = 4
        row += 1
        
        # Заказы
        self._draw_metric_row(row, 2, "Заказы", 
                             mtd_data.get('orders_count', 0), 
                             is_currency=False)
        self._draw_metric_row(row, 5, "Заказы", 
                             ytd_data.get('orders_count', 0), 
                             is_currency=False)
        row += 3
        
        # Сумма заказов
        self._draw_metric_row(row, 2, "Сумма заказов", 
                             mtd_data.get('total_amount', 0), 
                             is_currency=True)
        self._draw_metric_row(row, 5, "Сумма заказов", 
                             ytd_data.get('total_amount', 0), 
                             is_currency=True)
        row += 3
        
        # Средний чек
        self._draw_metric_row(row, 2, "Средний чек", 
                             mtd_data.get('avg_check', 0), 
                             is_currency=True)
        self._draw_metric_row(row, 5, "Средний чек", 
                             ytd_data.get('avg_check', 0), 
                             is_currency=True)
        row += 3
        
        # Разделитель групп
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        sep1 = self.ws.cell(row=row, column=2, value="")
        sep1.border = Border(top=Side(style='thin', color=self.colors["border_gray"]))
        self.ws.row_dimensions[row].height = 6
        row += 1
        
        # ============================================================
        # ГРУППА 2: ВЫПОЛНЕНИЕ
        # ============================================================
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        group2 = self.ws.cell(row=row, column=2, value="▸ ВЫПОЛНЕНИЕ")
        group2.font = Font(name="Roboto", size=9, bold=True, color=self.colors["text_gray"])
        group2.alignment = self.alignments["left"]
        row += 1
        
        self.ws.row_dimensions[row].height = 4
        row += 1
        
        # Оплачено (с мини-гистограммой)
        self._draw_metric_row(row, 2, "Оплачено", 
                             mtd_data.get('paid_amount', 0), 
                             secondary_value=mtd_data.get('paid_percent', 0),
                             is_currency=True, mini_bar=True)
        self._draw_metric_row(row, 5, "Оплачено", 
                             ytd_data.get('paid_amount', 0), 
                             secondary_value=ytd_data.get('paid_percent', 0),
                             is_currency=True, mini_bar=True)
        row += 3
        
        # Отгружено (с мини-гистограммой)
        self._draw_metric_row(row, 2, "Отгружено", 
                             mtd_data.get('shipped_amount', 0), 
                             secondary_value=mtd_data.get('shipped_percent', 0),
                             is_currency=True, mini_bar=True)
        self._draw_metric_row(row, 5, "Отгружено", 
                             ytd_data.get('shipped_amount', 0), 
                             secondary_value=ytd_data.get('shipped_percent', 0),
                             is_currency=True, mini_bar=True)
        row += 3
        
        # Осталось отгрузить (только для MTD)
        remaining_to_ship = mtd_data.get('remaining_to_ship', 0)
        if remaining_to_ship > 0:
            self._draw_metric_row(row, 2, "Осталось отгрузить", 
                                 remaining_to_ship, 
                                 is_currency=True)
            row += 3
        
        # Разделитель групп
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        sep2 = self.ws.cell(row=row, column=2, value="")
        sep2.border = Border(top=Side(style='thin', color=self.colors["border_gray"]))
        self.ws.row_dimensions[row].height = 6
        row += 1
        
        # ============================================================
        # ГРУППА 3: ОБЯЗАТЕЛЬСТВА
        # ============================================================
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        group3 = self.ws.cell(row=row, column=2, value="▸ ОБЯЗАТЕЛЬСТВА")
        group3.font = Font(name="Roboto", size=9, bold=True, color=self.colors["text_gray"])
        group3.alignment = self.alignments["left"]
        row += 1
        
        self.ws.row_dimensions[row].height = 4
        row += 1
        
        # Долг (с предупреждением)
        mtd_debt = mtd_data.get('debt_amount', 0)
        mtd_total = mtd_data.get('total_amount', 1)
        mtd_debt_warning = (mtd_debt / mtd_total) > 0.1 if mtd_total > 0 else False
        
        ytd_debt = ytd_data.get('debt_amount', 0)
        ytd_total = ytd_data.get('total_amount', 1)
        ytd_debt_warning = (ytd_debt / ytd_total) > 0.1 if ytd_total > 0 else False
        
        self._draw_metric_row(row, 2, "Долг", 
                             mtd_debt, 
                             is_currency=True,
                             highlight_negative=True,
                             debt_warning=mtd_debt_warning)
        self._draw_metric_row(row, 5, "Долг", 
                             ytd_debt, 
                             is_currency=True,
                             highlight_negative=True,
                             debt_warning=ytd_debt_warning)
        row += 3
        
        # Доставка
        self._draw_metric_row(row, 2, "Доставка", 
                             mtd_data.get('delivery_amount', 0), 
                             secondary_value=mtd_data.get('delivery_percent', 0),
                             is_currency=True)
        self._draw_metric_row(row, 5, "Доставка", 
                             ytd_data.get('delivery_amount', 0), 
                             secondary_value=ytd_data.get('delivery_percent', 0),
                             is_currency=True)
        row += 3
        
        # Разделитель перед дополнительными блоками
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        sep3 = self.ws.cell(row=row, column=2, value="")
        sep3.border = Border(top=Side(style='thin', color=self.colors["border_gray"]))
        self.ws.row_dimensions[row].height = 6
        row += 1
        
        # ============================================================
        # ДОПОЛНИТЕЛЬНЫЕ БЛОКИ (справа, под YTD)
        # ============================================================
        # Топ-5 клиентов (справа, под YTD)
        top_clients = mtd_data.get('top_clients', [])
        if top_clients:
            # Сохраняем текущую позицию для левого блока
            clients_row_start = row
            
            # Рисуем топ-клиентов в правой части (начиная с колонки 5)
            clients_height = self._draw_top_clients(row, 5, top_clients)
            
            # Сдвигаем row на высоту блока клиентов (если он выше других блоков)
            row += clients_height
        else:
            row += 1
        
        # Самый старый неоплаченный счет (слева, под MTD)
        oldest_invoice = mtd_data.get('oldest_unpaid_invoice')
        if oldest_invoice:
            # Отступ
            self.ws.row_dimensions[row].height = 6
            row += 1
            
            # Рисуем блок с предупреждением (слева, колонка 2)
            invoice_height = self._draw_oldest_invoice_warning(row, 2, oldest_invoice)
            row += invoice_height
        
        row += 1
        
        return row 
    
    
    