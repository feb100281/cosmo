from openpyxl.styles import Font, Border, Side, Alignment
from ..styles.theme import COLORS, FILLS


class SheetTitle:
    """Большой заголовок листа"""

    def __init__(self, worksheet):
        self.ws = worksheet

    def draw(
        self,
        row,
        title,
        subtitle="",
        date_text="",
        start_col=2,
        end_col=6
    ):
        # Основной блок заголовка
        self.ws.merge_cells(
            start_row=row,
            start_column=start_col,
            end_row=row + 1,
            end_column=end_col
        )

        title_cell = self.ws.cell(row=row, column=start_col, value=f"  {title}")
        title_cell.font = Font(
            name="Roboto",
            size=16,
            bold=True,
            color=COLORS["dark_green"]
        )
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        title_cell.fill = FILLS["section"]

        accent_side = Side(style="medium", color=COLORS["dark_green"])
        thin_side = Side(style="thin", color=COLORS["border_light"])

        for r in range(row, row + 2):
            for c in range(start_col, end_col + 1):
                cell = self.ws.cell(row=r, column=c)
                cell.fill = FILLS["section"]
                cell.border = Border(
                    left=accent_side if c == start_col else Side(style=None),
                    right=Side(style=None),
                    top=Side(style=None),
                    bottom=thin_side if r == row + 1 else Side(style=None)
                )

        # Подзаголовок слева
        if subtitle:
            self.ws.merge_cells(
                start_row=row + 2,
                start_column=start_col,
                end_row=row + 2,
                end_column=end_col - 2
            )
            subtitle_cell = self.ws.cell(
                row=row + 2,
                column=start_col,
                value=f"  {subtitle}"
            )
            subtitle_cell.font = Font(
                name="Roboto",
                size=9,
                color=COLORS["text_gray"]
            )
            subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
            subtitle_cell.fill = FILLS["summary"]

        # Дата справа в merge-диапазоне
        if date_text:
            self.ws.merge_cells(
                start_row=row + 2,
                start_column=end_col - 1,
                end_row=row + 2,
                end_column=end_col
            )
            date_cell = self.ws.cell(
                row=row + 2,
                column=end_col - 1,
                value=date_text
            )
            date_cell.font = Font(
                name="Roboto",
                size=8,
                italic=True,
                color=COLORS["text_gray"]
            )
            date_cell.alignment = Alignment(horizontal="right", vertical="center")
            date_cell.fill = FILLS["summary"]

        # Заливка для всей строки subtitle/date
        for c in range(start_col, end_col + 1):
            cell = self.ws.cell(row=row + 2, column=c)
            cell.fill = FILLS["summary"]

        self.ws.row_dimensions[row].height = 24
        self.ws.row_dimensions[row + 1].height = 20
        self.ws.row_dimensions[row + 2].height = 18

        return row + 4


def create_sheet_title(worksheet):
    return SheetTitle(worksheet)