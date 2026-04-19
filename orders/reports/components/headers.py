# # orders/reports/components/headers.py
# from openpyxl.styles import Font, Border, Side
# from ..styles.theme import COLORS, ALIGNMENTS


# class SectionHeader:
#     """Класс для отрисовки заголовков секций"""
    
#     def __init__(self, worksheet):
#         self.ws = worksheet
    
#     def draw(self, row, title, start_col=2):
#         """
#         Рисует заголовок секции без отдельной длинной линии
#         """
#         cell = self.ws.cell(row=row, column=start_col, value=title)
#         cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["dark_green"])
#         cell.alignment = ALIGNMENTS["left"]

#         return row + 2


# def create_header(worksheet):
#     """Создает экземпляр SectionHeader"""
#     return SectionHeader(worksheet)





from openpyxl.styles import Font, Border, Side
from ..styles.theme import COLORS, ALIGNMENTS, FILLS


class SectionHeader:
    """Класс для отрисовки заголовков секций"""

    def __init__(self, worksheet):
        self.ws = worksheet

    def draw(self, row, title, start_col=2, end_col=6):
        """
        Рисует заголовок секции в виде плашки на ширину таблицы.
        """
        # Объединяем ячейки по ширине таблицы
        self.ws.merge_cells(
            start_row=row,
            start_column=start_col,
            end_row=row,
            end_column=end_col
        )

        cell = self.ws.cell(row=row, column=start_col, value=title)

        # Шрифт
        cell.font = Font(
            name="Roboto",
            size=12,
            bold=True,
            color=COLORS["dark_green"]
        )

        # Выравнивание
        cell.alignment = ALIGNMENTS["left"]

        # Фон
        cell.fill = FILLS["section"]

        # Границы
        accent_side = Side(style="medium", color=COLORS["dark_green"])
        thin_side = Side(style="thin", color=COLORS["border_light"])

        # Проставляем стиль по всему диапазону merge
        for col in range(start_col, end_col + 1):
            current_cell = self.ws.cell(row=row, column=col)
            current_cell.fill = FILLS["section"]

            current_cell.border = Border(
                left=accent_side if col == start_col else Side(style=None),
                right=Side(style=None),
                top=Side(style=None),
                bottom=thin_side
            )

        self.ws.row_dimensions[row].height = 22

        return row + 2


def create_header(worksheet):
    return SectionHeader(worksheet)