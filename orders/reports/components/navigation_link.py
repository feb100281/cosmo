# orders/reports/components/navigation_link.py
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


class NavigationLink:
    """Компонент для создания кликабельной навигационной ссылки"""

    def __init__(self, worksheet):
        self.ws = worksheet

    def draw(self, row, text, target_sheet, target_cell="A1", 
             start_col=2, end_col=6, alignment="left", 
             with_icon=True, icon="🔗"):
        """
        Рисует кликабельную ссылку на другой лист
        
        Args:
            row: номер строки
            text: текст ссылки
            target_sheet: имя целевого листа
            target_cell: ячейка на целевом листе (по умолчанию "A1")
            start_col: начальная колонка для объединения
            end_col: конечная колонка для объединения
            alignment: выравнивание ("left", "center", "right")
            with_icon: показывать иконку
            icon: иконка (по умолчанию "🔗")
        
        Returns:
            int: следующая строка после ссылки
        """
        # Формируем текст с иконкой
        display_text = f"{icon} {text}" if with_icon else text
        
        # Объединяем ячейки для ссылки
        self.ws.merge_cells(
            start_row=row, 
            start_column=start_col, 
            end_row=row, 
            end_column=end_col
        )
        
        # Создаем ячейку с ссылкой
        cell = self.ws.cell(row=row, column=start_col, value=display_text)
        
        # Настройка внешнего вида
        cell.font = Font(
            name="Roboto", 
            size=10, 
            italic=True,
            color="0066CC",  # Синий цвет для ссылки
            underline="single"
        )
        
        # Выравнивание
        if alignment == "left":
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif alignment == "center":
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Добавляем гиперссылку
        cell.hyperlink = f"#'{target_sheet}'!{target_cell}"
        
        # Небольшой отступ сверху и снизу
        self.ws.row_dimensions[row].height = 24
        
        return row + 1
    
    def draw_as_button(self, row, text, target_sheet, target_cell="A1",
                       start_col=2, end_col=4, bg_color="E8F0FE"):
        """
        Рисует ссылку в виде кнопки (с фоном и рамкой)
        
        Args:
            row: номер строки
            text: текст ссылки
            target_sheet: имя целевого листа
            target_cell: ячейка на целевом листе
            start_col: начальная колонка для объединения
            end_col: конечная колонка для объединения
            bg_color: цвет фона (по умолчанию светло-синий)
        
        Returns:
            int: следующая строка
        """
        from openpyxl.styles import PatternFill, Border, Side
        
        # Объединяем ячейки
        self.ws.merge_cells(
            start_row=row, 
            start_column=start_col, 
            end_row=row, 
            end_column=end_col
        )
        
        # Создаем ячейку
        cell = self.ws.cell(row=row, column=start_col, value=f"📎 {text}")
        
        # Стили
        cell.font = Font(
            name="Roboto", 
            size=10, 
            bold=True,
            color="0066CC"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.border = Border(
            left=Side(style='thin', color="CCCCCC"),
            right=Side(style='thin', color="CCCCCC"),
            top=Side(style='thin', color="CCCCCC"),
            bottom=Side(style='thin', color="CCCCCC")
        )
        
        # Гиперссылка
        cell.hyperlink = f"#'{target_sheet}'!{target_cell}"
        
        self.ws.row_dimensions[row].height = 28
        
        return row + 1