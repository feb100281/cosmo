# orders/reports/components/alerts.py
from openpyxl.styles import Font, Alignment
from ..styles.theme import COLORS, ALIGNMENTS
from datetime import datetime, timedelta


class AlertComponent:
    """Класс для отрисовки алертов в Excel отчетах"""
    
    def __init__(self, worksheet):
        self.ws = worksheet
    
    def draw_cancelled_alert(self, row, cancelled_stats, period_days=1):
        """
        Рисует предупреждение об отменах за указанный период
        
        period_days: 1 - вчера, 2 - позавчера, и т.д.
        """
        period_date = (datetime.now() - timedelta(days=period_days)).strftime('%d.%m.%Y')
        count = cancelled_stats.get(f'last_{period_days}_days_count', 0)
        amount = cancelled_stats.get(f'last_{period_days}_days_amount', 0)
        details = cancelled_stats.get(f'last_{period_days}_days_details', [])
        
        if count == 0:
            return row
        
        # Заголовок алерта
        warn_cell = self.ws.cell(row=row, column=2, 
                                value=f"⚠️ ВНИМАНИЕ: За {period_date} отменено {count} заказов на сумму {self._format_currency(amount)}")
        warn_cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS.get("warning_red", "C62828"))
        warn_cell.alignment = ALIGNMENTS["left"]
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        row += 1
        
        # Детализация
        if details:
            for canc in details[:5]:
                order_display = canc.get('order_name', canc.get('number', ''))
                manager = canc.get('manager', '')
                amount = self._format_currency(canc.get('amount_cancelled', 0))
                
                detail_cell = self.ws.cell(row=row, column=3, 
                                        value=f"• {order_display} | Менеджер: {manager} | {amount}")
                detail_cell.font = Font(name="Roboto", size=8, color=COLORS["text_gray"])
                detail_cell.alignment = ALIGNMENTS["left"]
                self.ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
                row += 1
        
        return row + 1
    
    def draw_new_orders_alert(self, row, new_orders_stats, period_days=1):
        """
        Рисует алерт о новых заказах за указанный период
        
        period_days: 1 - вчера, 2 - позавчера, и т.д.
        """
        period_date = (datetime.now() - timedelta(days=period_days)).strftime('%d.%m.%Y')
        count = new_orders_stats.get('count', 0)
        amount = new_orders_stats.get('amount', 0)
        
        if count == 0:
            return row
        
        # Заголовок алерта
        alert_cell = self.ws.cell(row=row, column=2, 
                                 value=f"📋 НОВЫЕ ЗАКАЗЫ за {period_date}: {count} шт на сумму {self._format_currency(amount)}")
        alert_cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["dark_green"])
        alert_cell.alignment = ALIGNMENTS["left"]
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        
        return row + 1
    
    def _format_currency(self, value):
        if value is None or value == 0:
            return "0 ₽"
        return f"{int(round(value)):,} ₽".replace(",", " ")


def create_alert(worksheet):
    """Создает экземпляр AlertComponent"""
    return AlertComponent(worksheet)