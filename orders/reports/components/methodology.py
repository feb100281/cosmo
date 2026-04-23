# orders/reports/components/methodology.py
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from ..styles.theme import COLORS


class Methodology:
    """Компонент для отображения методологии расчетов"""
    
    def __init__(self, worksheet):
        self.ws = worksheet
        self.colors = COLORS
    
    def draw_metrics_methodology(self, row, start_col=2, end_col=6):
        """
        Рисует блок с методологией расчета ключевых метрик клиентского анализа
        
        Args:
            row: строка начала блока
            start_col: начальная колонка
            end_col: конечная колонка для объединения
        """
        # Заголовок блока
        self.ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        title_cell = self.ws.cell(row=row, column=start_col, value="")
        title_cell.font = Font(name="Roboto", size=10, bold=True, color=self.colors["dark_green"])
        title_cell.fill = PatternFill(start_color=self.colors["light_gray"], end_color=self.colors["light_gray"], fill_type="solid")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 1
        
        # Метрики и их описание
        metrics = [
            {
                'name': 'Активные клиенты (MTD/YTD)',
                'description': 'Уникальные клиенты, у которых есть заказы в периоде (MTD - с начала месяца, YTD - с начала года)',
                'formula': 'COUNT(DISTINCT client) WHERE date_from >= start_date AND change_status != "Отменен"'
            },
            {
                'name': 'Клиенты с 2+ заказами',
                'description': 'Клиенты, совершившие 2 и более заказа в указанном периоде',
                'formula': 'COUNT(client WHERE orders_count >= 2)'
            },
            {
                'name': 'Сумма заказа от клиентов',
                'description': 'Общая сумма активных заказов (без учета отмененных)',
                'formula': 'SUM(amount_active) WHERE change_status != "Отменен"'
            },
            {
                'name': 'Средний чек',
                'description': 'Средняя сумма одного заказа',
                'formula': 'SUM(amount_active) / COUNT(order_id)'
            },
            {
                'name': 'ABC-анализ',
                'description': 'Классификация клиентов по правилу Парето 80/20',
                'formula': 'A: 80% выручки, B: 15% выручки, C: 5% выручки'
            },
            {
                'name': 'Долг клиента',
                'description': 'Сумма заказов минус оплаченные средства',
                'formula': 'MAX(0, amount_active - cash_pmts)'
            },
            {
                'name': 'Клиенты с высоким риском',
                'description': 'Клиенты, у которых долг превышает 30% от суммы заказов',
                'formula': 'debt / total_amount > 0.3 AND debt > 0'
            },
        ]
        
        # Рисуем таблицу методологии
        for metric in metrics:
            # Название метрики
            name_cell = self.ws.cell(row=row, column=start_col, value=f"• {metric['name']}")
            name_cell.font = Font(name="Roboto", size=8, bold=True, color=self.colors["dark_green"])
            name_cell.alignment = Alignment(horizontal="left", vertical="top")
            
            # Описание
            desc_cell = self.ws.cell(row=row, column=start_col + 1, value=metric['description'])
            desc_cell.font = Font(name="Roboto", size=7, color=self.colors["text_gray"])
            desc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # Формула (более мелким шрифтом, курсивом)
            formula_cell = self.ws.cell(row=row + 1, column=start_col + 1, value=f"Формула: {metric['formula']}")
            formula_cell.font = Font(name="Roboto", size=6, italic=True, color=self.colors["text_gray"])
            formula_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            row += 2
            
            # Тонкая линия-разделитель между метриками
            if metric != metrics[-1]:
                for col in range(start_col, end_col + 1):
                    sep_cell = self.ws.cell(row=row, column=col, value="")
                    sep_cell.border = Border(top=Side(style='thin', color=self.colors["border_gray"]))
                row += 1
        
        # Примечание
        row += 1
        self.ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        note_cell = self.ws.cell(
            row=row, 
            column=start_col, 
            value="Все расчеты исключают полностью отмененные заказы (change_status = 'Отменен')"
        )
        note_cell.font = Font(name="Roboto", size=7, italic=True, color=self.colors["warning_red"])
        note_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        row += 2
        return row
    
    def draw_compact_methodology(self, row, start_col=2, end_col=6):
        """
        Компактная версия методологии (только основные метрики)
        """
        # Заголовок
        self.ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        title_cell = self.ws.cell(row=row, column=start_col, value="📐 Методология расчетов")
        title_cell.font = Font(name="Roboto", size=9, bold=True, color=self.colors["dark_green"])
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 1
        
        # Компактное описание
        methodology_text = (
            "Активные клиенты: уникальные клиенты с заказами в периоде (MTD/YTD)\n"
            "Клиенты с 2+ заказами: клиенты, совершившие 2 и более заказа\n"
            "Выручка: сумма активных заказов (без отмененных)\n"
            "Средний чек: выручка / количество заказов\n"
            "ABC-анализ: A - 80% выручки, B - 15%, C - 5%\n"
            "Долг: сумма заказов - оплачено\n"
            "Клиенты с высоким риском: долг > 30% от суммы заказов"
        )
        
        text_cell = self.ws.cell(row=row, column=start_col, value=methodology_text)
        text_cell.font = Font(name="Roboto", size=7, color=self.colors["text_gray"])
        text_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        self.ws.merge_cells(start_row=row, start_column=start_col, end_row=row + 6, end_column=end_col)
        
        row += 7
        return row